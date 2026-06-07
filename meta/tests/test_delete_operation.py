import pytest

pytestmark = pytest.mark.integration

# -*- coding: utf-8 -*-
"""
测试删除操作功能
"""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from meta.core.datasource import get_data_source
from meta.core.models import registry
from meta.core.yaml_loader import register_from_directory, get_yaml_schema_dir
from meta.services.import_export_service import ImportExportService
from meta.services.manage_service import ManageService, CreateRequest, DeleteRequest
from meta.services.query_service import QueryService, SearchRequest

def setup():
    schema_dir = get_yaml_schema_dir()
    register_from_directory(schema_dir)
    
    domain_obj = registry.get("domain")
    assert domain_obj is not None, "domain meta object not found in registry after registration"
    
    db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'meta', 'architecture.db')
    print(f"Database path: {db_path}")
    data_source = get_data_source("sqlite", database=db_path)
    return data_source

def test_delete_operation():
    print("=== 测试删除操作功能 ===\n")
    
    data_source = setup()
    if data_source is None:
        return
    manage_service = ManageService(data_source)
    query_service = QueryService(data_source)
    import_export_service = ImportExportService(data_source, manage_service, query_service)
    
    # 1. 创建测试数据
    print("1. 创建测试数据...")
    test_code = "TEST_DELETE_001"
    test_name = "测试删除数据"
    
    # 检查是否已存在
    search_request = SearchRequest(
        object_type="domain",
        conditions=[],
        page=1,
        page_size=1000,
    )
    result = query_service.search(search_request)
    existing = next((r for r in result.data if r.get("code") == test_code), None)
    
    if existing:
        print(f"   测试数据已存在，ID: {existing['id']}")
        test_id = existing['id']
    else:
        # 创建新数据
        create_result = manage_service.create(CreateRequest(
            object_type="domain",
            data={
                "code": test_code,
                "name": test_name,
                "version_id": 2
            }
        ))
        print(f"   创建测试数据成功: {create_result.success}")
        
        # 获取创建的数据
        search_request = SearchRequest(
            object_type="domain",
            conditions=[],
            page=1,
            page_size=1000,
        )
        result = query_service.search(search_request)
        test_data = next(r for r in result.data if r.get("code") == test_code)
        test_id = test_data['id']
        print(f"   测试数据ID: {test_id}")
    
    # 2. 创建删除导入文件
    print("\n2. 创建删除导入文件...")
    upload_dir = os.path.join(os.getcwd(), "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "领域"
    
    # 添加表头
    headers = ["操作模式", "编码", "名称", "ID", "版本"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # 添加删除数据行
    ws.cell(row=2, column=1, value="删除")
    ws.cell(row=2, column=2, value=test_code)
    ws.cell(row=2, column=3, value=test_name)
    ws.cell(row=2, column=4, value=test_id)
    ws.cell(row=2, column=5, value=2)
    
    delete_file = os.path.join(upload_dir, "test_delete.xlsx")
    wb.save(delete_file)
    wb.close()
    print(f"   删除文件已创建: {delete_file}")
    
    # 3. 执行导入
    print("\n3. 执行删除导入...")
    import_result = import_export_service.import_cascade(delete_file, mode="execute", conflict_strategy="upsert")
    
    print(f"   导入结果: success={import_result.success}")
    if import_result.results:
        for obj_type, result in import_result.results.items():
            print(f"   {obj_type}: success={result.get('success', 0)}, deleted={result.get('deleted', 0)}, failed={result.get('failed', 0)}")
            if result.get('errors'):
                for err in result['errors']:
                    print(f"      错误: {err}")
    
    # 4. 验证数据已删除
    print("\n4. 验证数据已删除...")
    search_request = SearchRequest(
        object_type="domain",
        conditions=[],
        page=1,
        page_size=1000,
    )
    result = query_service.search(search_request)
    deleted_data = next((r for r in result.data if r.get("code") == test_code), None)
    
    if deleted_data:
        print(f"   [FAIL] Delete failed: data still exists (ID: {deleted_data['id']})")
        return False
    else:
        print(f"   [PASS] Delete success: data has been deleted")
        return True

if __name__ == "__main__":
    success = test_delete_operation()
    print(f"\n{'Test PASSED' if success else 'Test FAILED'}")
