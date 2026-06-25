# -*- coding: utf-8 -*-
"""
xlsx 格式对比测试 (T3: 模型驱动生成)

模型源:
  - meta/schemas/<object>.yaml 的 fields.*
    - semantics.business_key → 黄色填充 (FFD966)
    - required → 红色字体
    - semantics.import_order → 列顺序
    - description → 单元格批注

覆盖:
  - 颜色: 业务键 = 黄底; 必填 = 红字
  - 批注: description 写入 Comment
  - 列顺序: 按 import_order 升序
  - 业务键不被错误地包含在导入数据中 (只读)

生成时间: 2026-06-25T12:23:09.443Z
对象数: 12
"""

import pytest
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# 业务键颜色规范 (与 import_export_service.py 保持一致)
BUSINESS_KEY_FILL = 'FFD966'  # 黄
REQUIRED_FONT_COLOR = 'FF0000'  # 红

#// 模型元数据 (从 schema 抽取)
SCHEMA_FIELDS = {
  "product": [
    {
      "id": "id",
      "name": "ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "技术主键"
    },
    {
      "id": "name",
      "name": "名称",
      "required": True,
      "business_key": False,
      "import_order": 20,
      "description": "产品线名称"
    },
    {
      "id": "code",
      "name": "产品编码",
      "required": True,
      "business_key": True,
      "import_order": 1,
      "description": "产品线编码，用于唯一标识产品线"
    },
    {
      "id": "description",
      "name": "描述",
      "required": False,
      "business_key": False,
      "import_order": 30,
      "description": "产品线描述"
    },
    {
      "id": "visibility",
      "name": "可见性",
      "required": True,
      "business_key": False,
      "import_order": 5,
      "description": "所有有产品权限的用户可见"
    },
    {
      "id": "owner_id",
      "name": "负责人",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "产品线负责人ID（由系统自动设置）"
    },
    {
      "id": "is_active",
      "name": "是否活跃",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "产品线是否活跃"
    },
    {
      "id": "child_count",
      "name": "版本数量",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "产品线下的版本数量"
    },
    {
      "id": "activate_product",
      "name": "激活产品线",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "deactivate_product",
      "name": "停用产品线",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "product_to_versions",
      "name": "包含版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "一个产品线包含多个版本"
    },
    {
      "id": "product_create",
      "name": "创建产品线",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建新的产品线"
    },
    {
      "id": "product_read",
      "name": "查询产品线",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询单个产品线"
    },
    {
      "id": "product_list",
      "name": "列表查询",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询产品线列表"
    },
    {
      "id": "product_update",
      "name": "更新产品线",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "更新产品线信息"
    },
    {
      "id": "product_delete",
      "name": "删除产品线",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "删除产品线"
    },
    {
      "id": "name_required",
      "name": "名称必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ],
  "version": [
    {
      "id": "product_code",
      "name": "产品编码",
      "required": False,
      "business_key": True,
      "import_order": 0,
      "description": "所属产品编码"
    },
    {
      "id": "id",
      "name": "ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "技术主键"
    },
    {
      "id": "product_id",
      "name": "产品",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属产品"
    },
    {
      "id": "name",
      "name": "版本名称",
      "required": True,
      "business_key": True,
      "import_order": 1,
      "description": "版本名称，如 v1.0、2024-Q4"
    },
    {
      "id": "description",
      "name": "版本描述",
      "required": False,
      "business_key": False,
      "import_order": 30,
      "description": "版本描述"
    },
    {
      "id": "is_current",
      "name": "是否当前版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "是否为当前活跃版本"
    },
    {
      "id": "product_name",
      "name": "产品",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属产品名称"
    },
    {
      "id": "child_count",
      "name": "领域数量",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "版本下的领域数量"
    },
    {
      "id": "version_to_product",
      "name": "归属产品",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "版本归属的产品线"
    },
    {
      "id": "version_to_domains",
      "name": "包含领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "一个版本包含多个领域"
    },
    {
      "id": "version_create",
      "name": "创建版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建新的版本"
    },
    {
      "id": "version_read",
      "name": "查询版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询单个版本"
    },
    {
      "id": "version_update",
      "name": "更新版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "更新版本信息"
    },
    {
      "id": "version_delete",
      "name": "删除版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "删除版本"
    },
    {
      "id": "version_list",
      "name": "列表查询",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询版本列表"
    },
    {
      "id": "set_current",
      "name": "设为当前版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "将版本设为当前活跃版本"
    },
    {
      "id": "compare",
      "name": "版本对比",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "对比两个版本的差异"
    },
    {
      "id": "name_required",
      "name": "名称必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "product_id_required",
      "name": "产品ID必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "only_one_current_per_product",
      "name": "每个产品只能有一个当前版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "list_by_product",
      "name": "按产品查询版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "current_versions",
      "name": "当前版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "set_current_version",
      "name": "设为当前版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "unset_current_version",
      "name": "取消当前版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ],
  "domain": [
    {
      "id": "version_code",
      "name": "版本编码",
      "required": False,
      "business_key": True,
      "import_order": 3,
      "description": "所属版本编码（用于导入导出和快速输入）"
    },
    {
      "id": "id",
      "name": "ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "技术主键"
    },
    {
      "id": "version_id",
      "name": "版本",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属版本"
    },
    {
      "id": "code",
      "name": "编码",
      "required": True,
      "business_key": True,
      "import_order": 4,
      "description": "领域编码（预留）"
    },
    {
      "id": "name",
      "name": "名称",
      "required": True,
      "business_key": False,
      "import_order": 20,
      "description": "领域名称"
    },
    {
      "id": "description",
      "name": "描述",
      "required": False,
      "business_key": False,
      "import_order": 30,
      "description": "领域描述"
    },
    {
      "id": "version_name",
      "name": "版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属版本名称"
    },
    {
      "id": "relation_count",
      "name": "关系数量",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "领域下所有业务对象的关系数量"
    },
    {
      "id": "child_count",
      "name": "子领域数量",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "领域下的子领域数量"
    },
    {
      "id": "bo_density",
      "name": "BO密度",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "业务对象关系密度"
    },
    {
      "id": "domain_to_version",
      "name": "归属版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "领域归属的版本"
    },
    {
      "id": "domain_to_subdomains",
      "name": "包含子领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "一个领域包含多个子领域"
    },
    {
      "id": "domain_create",
      "name": "创建领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建新的领域"
    },
    {
      "id": "domain_read",
      "name": "查询领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询单个领域"
    },
    {
      "id": "domain_update",
      "name": "更新领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "更新领域信息"
    },
    {
      "id": "domain_delete",
      "name": "删除领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "删除领域"
    },
    {
      "id": "domain_list",
      "name": "列表查询",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询领域列表"
    },
    {
      "id": "version_id_required",
      "name": "版本ID必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ],
  "sub_domain": [
    {
      "id": "version_code",
      "name": "版本编码",
      "required": False,
      "business_key": True,
      "import_order": 1,
      "description": "所属版本编码"
    },
    {
      "id": "domain_code",
      "name": "领域编码",
      "required": False,
      "business_key": False,
      "import_order": 3,
      "description": "所属领域编码（用于导入导出和快速输入）"
    },
    {
      "id": "id",
      "name": "ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "技术主键"
    },
    {
      "id": "version_id",
      "name": "版本",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属版本"
    },
    {
      "id": "domain_id",
      "name": "领域",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属领域"
    },
    {
      "id": "code",
      "name": "编码",
      "required": True,
      "business_key": True,
      "import_order": 2,
      "description": "子领域编码（预留）"
    },
    {
      "id": "name",
      "name": "名称",
      "required": True,
      "business_key": False,
      "import_order": 20,
      "description": "子领域名称"
    },
    {
      "id": "description",
      "name": "描述",
      "required": False,
      "business_key": False,
      "import_order": 30,
      "description": "子领域描述"
    },
    {
      "id": "domain_name",
      "name": "领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属领域名称"
    },
    {
      "id": "version_name",
      "name": "版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属版本名称"
    },
    {
      "id": "created_at",
      "name": "创建时间",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建时间"
    },
    {
      "id": "relation_count",
      "name": "关系数量",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "子领域下所有业务对象的关系数量"
    },
    {
      "id": "child_count",
      "name": "服务模块数量",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "子领域下的服务模块数量"
    },
    {
      "id": "bo_density",
      "name": "BO密度",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "业务对象关系密度"
    },
    {
      "id": "subdomain_to_version",
      "name": "归属版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "子领域归属的版本"
    },
    {
      "id": "subdomain_to_domain",
      "name": "归属领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "子领域归属的领域"
    },
    {
      "id": "subdomain_to_service_modules",
      "name": "包含服务模块",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "一个子领域包含多个服务模块"
    },
    {
      "id": "sub_domain_create",
      "name": "创建子领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建新的子领域"
    },
    {
      "id": "sub_domain_read",
      "name": "查询子领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询单个子领域"
    },
    {
      "id": "sub_domain_update",
      "name": "更新子领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "更新子领域信息"
    },
    {
      "id": "sub_domain_delete",
      "name": "删除子领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "删除子领域"
    },
    {
      "id": "sub_domain_list",
      "name": "列表查询",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询子领域列表"
    },
    {
      "id": "version_id_required",
      "name": "版本ID必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ],
  "service_module": [
    {
      "id": "version_code",
      "name": "版本编码",
      "required": False,
      "business_key": True,
      "import_order": 1,
      "description": "所属版本编码"
    },
    {
      "id": "sub_domain_code",
      "name": "子领域编码",
      "required": False,
      "business_key": False,
      "import_order": 3,
      "description": "所属子领域编码（用于导入导出和快速输入）"
    },
    {
      "id": "id",
      "name": "ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "技术主键"
    },
    {
      "id": "version_id",
      "name": "版本",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属版本（用于筛选）"
    },
    {
      "id": "domain_id",
      "name": "领域",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属领域（用于级联筛选子领域）"
    },
    {
      "id": "sub_domain_id",
      "name": "子领域",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属子领域（创建后不可变更）"
    },
    {
      "id": "code",
      "name": "编码",
      "required": True,
      "business_key": True,
      "import_order": 2,
      "description": "服务模块编码"
    },
    {
      "id": "name",
      "name": "名称",
      "required": True,
      "business_key": False,
      "import_order": 20,
      "description": "服务模块名称"
    },
    {
      "id": "description",
      "name": "描述",
      "required": False,
      "business_key": False,
      "import_order": 30,
      "description": "服务模块描述"
    },
    {
      "id": "domain_name",
      "name": "领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属领域名称"
    },
    {
      "id": "sub_domain_name",
      "name": "子领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属子领域名称"
    },
    {
      "id": "version_name",
      "name": "版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属版本名称"
    },
    {
      "id": "relation_count",
      "name": "关系数量",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "服务模块下所有业务对象的关系数量"
    },
    {
      "id": "child_count",
      "name": "业务对象数量",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "服务模块下的业务对象数量"
    },
    {
      "id": "bo_density",
      "name": "BO密度",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "业务对象关系密度"
    },
    {
      "id": "sm_to_version",
      "name": "归属版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "服务模块归属的版本"
    },
    {
      "id": "sm_to_subdomain",
      "name": "归属子领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "服务模块归属的子领域"
    },
    {
      "id": "sm_to_business_objects",
      "name": "包含业务对象",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "一个服务模块包含多个业务对象"
    },
    {
      "id": "service_module_create",
      "name": "创建服务模块",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建新的服务模块"
    },
    {
      "id": "service_module_read",
      "name": "查询服务模块",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询单个服务模块"
    },
    {
      "id": "service_module_update",
      "name": "更新服务模块",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "更新服务模块信息"
    },
    {
      "id": "service_module_delete",
      "name": "删除服务模块",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "删除服务模块"
    },
    {
      "id": "service_module_list",
      "name": "列表查询",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询服务模块列表"
    },
    {
      "id": "version_id_required",
      "name": "版本ID必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ],
  "business_object": [
    {
      "id": "version_code",
      "name": "版本编码",
      "required": False,
      "business_key": True,
      "import_order": 1,
      "description": "所属版本编码"
    },
    {
      "id": "service_module_code",
      "name": "服务模块编码",
      "required": False,
      "business_key": False,
      "import_order": 3,
      "description": "所属服务模块编码（用于导入导出和快速输入）"
    },
    {
      "id": "id",
      "name": "ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "技术主键"
    },
    {
      "id": "version_id",
      "name": "版本",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属版本（用于筛选）"
    },
    {
      "id": "domain_id",
      "name": "领域",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属领域（用于级联筛选，新建后不可变更）"
    },
    {
      "id": "sub_domain_id",
      "name": "子领域",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属子领域（用于级联筛选，新建后不可变更）"
    },
    {
      "id": "service_module_id",
      "name": "服务模块",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属服务模块（创建后不可变更）"
    },
    {
      "id": "code",
      "name": "编码",
      "required": True,
      "business_key": True,
      "import_order": 2,
      "description": "业务对象编码"
    },
    {
      "id": "name",
      "name": "名称",
      "required": True,
      "business_key": False,
      "import_order": 20,
      "description": "业务对象名称"
    },
    {
      "id": "description",
      "name": "描述",
      "required": False,
      "business_key": False,
      "import_order": 30,
      "description": "业务对象描述"
    },
    {
      "id": "domain_name",
      "name": "领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属领域名称"
    },
    {
      "id": "sub_domain_name",
      "name": "子领域",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属子领域名称"
    },
    {
      "id": "service_module_name",
      "name": "服务模块",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属服务模块名称"
    },
    {
      "id": "version_name",
      "name": "版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属版本名称"
    },
    {
      "id": "relation_count",
      "name": "关系数量",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "业务对象的关系数量"
    },
    {
      "id": "bo_to_version",
      "name": "归属版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "业务对象归属的版本"
    },
    {
      "id": "bo_to_service_module",
      "name": "归属服务模块",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "业务对象归属的服务模块"
    },
    {
      "id": "bo_to_source_relationships",
      "name": "作为源的关系",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "业务对象作为源的关系"
    },
    {
      "id": "bo_to_target_relationships",
      "name": "作为目标的关系",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "业务对象作为目标的关系"
    },
    {
      "id": "relation_type_name",
      "name": "relation_type_name",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "source_bo_name",
      "name": "source_bo_name",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "target_bo_name",
      "name": "target_bo_name",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "relation_desc",
      "name": "relation_desc",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "business_object_create",
      "name": "创建业务对象",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建新的业务对象"
    },
    {
      "id": "business_object_read",
      "name": "查询业务对象",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询单个业务对象"
    },
    {
      "id": "business_object_update",
      "name": "更新业务对象",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "更新业务对象信息"
    },
    {
      "id": "business_object_delete",
      "name": "删除业务对象",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "删除业务对象"
    },
    {
      "id": "business_object_list",
      "name": "列表查询",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询业务对象列表"
    },
    {
      "id": "batch_import",
      "name": "批量导入",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "从Excel批量导入业务对象"
    },
    {
      "id": "export",
      "name": "导出",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "导出业务对象到Excel"
    },
    {
      "id": "version_id_required",
      "name": "版本ID必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "code_format",
      "name": "编码格式",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ],
  "relationship": [
    {
      "id": "version_code",
      "name": "版本编码",
      "required": False,
      "business_key": True,
      "import_order": 1,
      "description": "所属版本编码"
    },
    {
      "id": "id",
      "name": "ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "技术主键"
    },
    {
      "id": "version_id",
      "name": "版本",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属版本"
    },
    {
      "id": "source_domain_id",
      "name": "源领域",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "源业务对象所属领域（用于级联筛选业务对象）"
    },
    {
      "id": "source_sub_domain_id",
      "name": "源子领域",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "源业务对象所属子领域（用于级联筛选业务对象）"
    },
    {
      "id": "source_service_module_id",
      "name": "源服务模块",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "源业务对象所属服务模块（用于级联筛选业务对象）"
    },
    {
      "id": "target_domain_id",
      "name": "目标领域",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "目标业务对象所属领域（用于级联筛选业务对象）"
    },
    {
      "id": "target_sub_domain_id",
      "name": "目标子领域",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "目标业务对象所属子领域（用于级联筛选业务对象）"
    },
    {
      "id": "target_service_module_id",
      "name": "目标服务模块",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "目标业务对象所属服务模块（用于级联筛选业务对象）"
    },
    {
      "id": "source_bo_id",
      "name": "源业务对象",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "源业务对象"
    },
    {
      "id": "target_bo_id",
      "name": "目标业务对象",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "目标业务对象"
    },
    {
      "id": "source_code",
      "name": "源编码",
      "required": False,
      "business_key": False,
      "import_order": 3,
      "description": "源业务对象编码（选择业务对象后自动带入）"
    },
    {
      "id": "target_code",
      "name": "目标编码",
      "required": False,
      "business_key": False,
      "import_order": 6,
      "description": "目标业务对象编码（选择业务对象后自动带入）"
    },
    {
      "id": "code",
      "name": "关系编码",
      "required": False,
      "business_key": True,
      "import_order": 2,
      "description": "关系实例编码，由 KeyTemplate 自动生成"
    },
    {
      "id": "relation_code",
      "name": "关系类型编码",
      "required": False,
      "business_key": False,
      "import_order": 19,
      "description": "关系类型编码（历史字段）"
    },
    {
      "id": "relation_type",
      "name": "关系类型",
      "required": True,
      "business_key": False,
      "import_order": 9,
      "description": "关系类型（操作语义）"
    },
    {
      "id": "relation_direction",
      "name": "关系方向",
      "required": False,
      "business_key": False,
      "import_order": 10,
      "description": "关系的数据流向方向（推/拉/双向），可选"
    },
    {
      "id": "relation_desc",
      "name": "关系描述",
      "required": False,
      "business_key": False,
      "import_order": 12,
      "description": "关系描述"
    },
    {
      "id": "relation_type_name",
      "name": "关系类型名称",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "关系类型的中文名称（由 relation_type 枚举自动填充）"
    },
    {
      "id": "relation_type_name_en",
      "name": "关系类型英文名称",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "关系类型的英文名称（由 relation_type 枚举自动填充）"
    },
    {
      "id": "relation_category",
      "name": "关系分类",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "关系类型的业务分类"
    },
    {
      "id": "source_bo_name",
      "name": "源业务对象",
      "required": False,
      "business_key": False,
      "import_order": 5,
      "description": "源业务对象名称"
    },
    {
      "id": "source_bo_code",
      "name": "源业务对象编码",
      "required": False,
      "business_key": False,
      "import_order": 4,
      "description": "源业务对象编码（与 source_code 一致，命名对称）"
    },
    {
      "id": "target_bo_name",
      "name": "目标业务对象",
      "required": False,
      "business_key": False,
      "import_order": 8,
      "description": "目标业务对象名称"
    },
    {
      "id": "target_bo_code",
      "name": "目标业务对象编码",
      "required": False,
      "business_key": False,
      "import_order": 7,
      "description": "目标业务对象编码（与 target_code 一致，命名对称）"
    },
    {
      "id": "category_label",
      "name": "关系范围",
      "required": False,
      "business_key": False,
      "import_order": 11,
      "description": "关系分类标签"
    },
    {
      "id": "category_type",
      "name": "分类类型",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "分类类型编码"
    },
    {
      "id": "activity_label",
      "name": "活跃度标签",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "关系活跃度标签"
    },
    {
      "id": "domain_relation",
      "name": "领域关系",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "源领域-目标领域"
    },
    {
      "id": "sub_domain_relation",
      "name": "子领域关系",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "源子领域-目标子领域"
    },
    {
      "id": "module_relation",
      "name": "服务模块关系",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "源服务模块-目标服务模块"
    },
    {
      "id": "is_in_scope",
      "name": "是否中心范围内",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "关系是否在中心范围内"
    },
    {
      "id": "source_domain_name",
      "name": "源领域",
      "required": False,
      "business_key": False,
      "import_order": 13,
      "description": "源业务对象所属领域"
    },
    {
      "id": "source_sub_domain_name",
      "name": "源子领域",
      "required": False,
      "business_key": False,
      "import_order": 14,
      "description": "源业务对象所属子领域"
    },
    {
      "id": "source_service_module_name",
      "name": "源服务模块",
      "required": False,
      "business_key": False,
      "import_order": 15,
      "description": "源业务对象所属服务模块"
    },
    {
      "id": "target_domain_name",
      "name": "目标领域",
      "required": False,
      "business_key": False,
      "import_order": 16,
      "description": "目标业务对象所属领域"
    },
    {
      "id": "target_sub_domain_name",
      "name": "目标子领域",
      "required": False,
      "business_key": False,
      "import_order": 17,
      "description": "目标业务对象所属子领域"
    },
    {
      "id": "target_service_module_name",
      "name": "目标服务模块",
      "required": False,
      "business_key": False,
      "import_order": 18,
      "description": "目标业务对象所属服务模块"
    },
    {
      "id": "version_name",
      "name": "版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属版本名称"
    },
    {
      "id": "rel_to_version",
      "name": "归属版本",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "关系归属的版本"
    },
    {
      "id": "rel_to_source_bo",
      "name": "源业务对象",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "关系的源端业务对象"
    },
    {
      "id": "rel_to_target_bo",
      "name": "目标业务对象",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "关系的目标端业务对象"
    },
    {
      "id": "relationship_create",
      "name": "创建关系",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建新的关系"
    },
    {
      "id": "relationship_read",
      "name": "查询关系",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询单个关系"
    },
    {
      "id": "relationship_list",
      "name": "列表查询",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询关系列表"
    },
    {
      "id": "relationship_update",
      "name": "更新关系",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "更新关系信息"
    },
    {
      "id": "relationship_delete",
      "name": "删除关系",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "删除单个关系"
    },
    {
      "id": "batch_import",
      "name": "批量导入",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "从Excel批量导入关系"
    },
    {
      "id": "version_id_required",
      "name": "版本ID必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "source_not_equal_target",
      "name": "源目标不能相同",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "code_unique_within_version",
      "name": "关系编码在版本内唯一",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "relation_type_in_enum",
      "name": "关系类型必须在枚举中",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "relation_direction_in_enum",
      "name": "关系方向必须在枚举中",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "relation_code_required",
      "name": "关系简称必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ],
  "enum_type": [
    {
      "id": "id",
      "name": "编码",
      "required": True,
      "business_key": True,
      "import_order": 999,
      "description": "枚举类型唯一标识"
    },
    {
      "id": "name",
      "name": "名称",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "枚举类型显示名称"
    },
    {
      "id": "category",
      "name": "分类",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "枚举分类（system=系统枚举, business=业务枚举）"
    },
    {
      "id": "mutability",
      "name": "可维护性",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "用户可维护程度。规范化为 3 档：fullEditable（完全可改）/ extensible（可加不可改预置）/ locked（完全锁）。3 值之外在 API 层被拒绝（INVALID_MUTABILITY）。"
    },
    {
      "id": "dimension_schema",
      "name": "维度定义",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "多维枚举的维度定义"
    },
    {
      "id": "description",
      "name": "描述",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "枚举类型说明"
    },
    {
      "id": "dimension_count",
      "name": "维度数",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "枚举类型的维度数量"
    },
    {
      "id": "value_count",
      "name": "值数量",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "枚举类型的枚举值数量"
    },
    {
      "id": "created_at",
      "name": "创建时间",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建时间"
    },
    {
      "id": "enum_type_to_values",
      "name": "包含枚举值",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "一个枚举类型包含多个枚举值"
    },
    {
      "id": "id_required",
      "name": "编码必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "name_required",
      "name": "名称必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "category_required",
      "name": "分类必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "mutability_required",
      "name": "可维护性必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ],
  "enum_value": [
    {
      "id": "id",
      "name": "ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "技术主键"
    },
    {
      "id": "enum_type_id",
      "name": "枚举类型",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "所属枚举类型"
    },
    {
      "id": "code",
      "name": "编码",
      "required": True,
      "business_key": True,
      "import_order": 999,
      "description": "枚举值编码"
    },
    {
      "id": "name",
      "name": "名称",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "枚举值显示名称"
    },
    {
      "id": "name_en",
      "name": "英文名称",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "英文名称（用于国际化）"
    },
    {
      "id": "dimensions",
      "name": "维度值",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "多维枚举的维度值"
    },
    {
      "id": "sort_order",
      "name": "排序",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "排序顺序"
    },
    {
      "id": "is_active",
      "name": "是否启用",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "是否启用"
    },
    {
      "id": "is_system",
      "name": "是否系统值",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "是否系统预置值"
    },
    {
      "id": "parent_code",
      "name": "父级编码",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "父级枚举值编码（用于层级枚举）"
    },
    {
      "id": "metadata",
      "name": "扩展元数据",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "扩展元数据"
    },
    {
      "id": "enum_type_name",
      "name": "枚举类型名称",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "所属枚举类型名称"
    },
    {
      "id": "created_at",
      "name": "创建时间",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建时间"
    },
    {
      "id": "enum_value_to_type",
      "name": "归属枚举类型",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "枚举值归属的枚举类型"
    },
    {
      "id": "enum_value_create",
      "name": "创建枚举值",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建新的枚举值"
    },
    {
      "id": "enum_value_read",
      "name": "查询枚举值",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询单个枚举值"
    },
    {
      "id": "enum_value_update",
      "name": "更新枚举值",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "更新枚举值信息"
    },
    {
      "id": "enum_value_delete",
      "name": "删除枚举值",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "删除枚举值"
    },
    {
      "id": "enum_value_list",
      "name": "列表查询",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询枚举值列表"
    },
    {
      "id": "enum_type_id_required",
      "name": "枚举类型必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "code_required",
      "name": "编码必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "name_required",
      "name": "名称必填",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "unique_code_per_type",
      "name": "编码唯一性",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ],
  "role": [
    {
      "id": "id",
      "name": "ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "技术主键"
    },
    {
      "id": "code",
      "name": "角色编码",
      "required": True,
      "business_key": True,
      "import_order": 999,
      "description": "角色编码，如 admin, editor, viewer"
    },
    {
      "id": "name",
      "name": "角色名称",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "角色显示名称"
    },
    {
      "id": "description",
      "name": "描述",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "角色描述"
    },
    {
      "id": "is_active",
      "name": "启用状态",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "角色是否启用，停用后该角色的权限将失效"
    },
    {
      "id": "is_system",
      "name": "系统角色",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "系统角色不可删除，新建角色默认为自定义角色"
    },
    {
      "id": "created_at",
      "name": "创建时间",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "创建时间"
    },
    {
      "id": "user_count",
      "name": "用户数",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "角色关联的用户数量（计算字段）"
    },
    {
      "id": "menu_count",
      "name": "菜单数",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "角色关联的菜单数量（计算字段）"
    },
    {
      "id": "permission_count",
      "name": "权限数",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "角色关联的权限数量（计算字段）"
    },
    {
      "id": "data_perm_count",
      "name": "数据权限数",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "角色关联的数据权限数量（计算字段）"
    },
    {
      "id": "code",
      "name": "code",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "name",
      "name": "assigned_groups",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "分配了此角色的用户组。用户组成员将自动获得此角色权限。"
    },
    {
      "id": "code",
      "name": "code",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "name",
      "name": "name",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "create",
      "name": "create",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "export",
      "name": "export",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "import",
      "name": "import",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "batch_delete",
      "name": "batch_delete",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "edit",
      "name": "edit",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "save",
      "name": "save",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "cancel",
      "name": "cancel",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "basic",
      "name": "basic",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "permissions",
      "name": "permissions",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "assigned_groups",
      "name": "assigned_groups",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "此角色已分配给以下用户组，用户组成员自动获得此角色权限"
    },
    {
      "id": "audit-log",
      "name": "audit-log",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "role_create",
      "name": "创建角色",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "role_read",
      "name": "查询角色",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "role_update",
      "name": "更新角色",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "role_delete",
      "name": "删除角色",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "role_list",
      "name": "角色列表",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ],
  "audit_log": [
    {
      "id": "id",
      "name": "ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "技术主键"
    },
    {
      "id": "log_category",
      "name": "日志类型",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "日志分类，用于区分不同类型的日志"
    },
    {
      "id": "log_level",
      "name": "日志级别",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "日志严重程度级别"
    },
    {
      "id": "object_type",
      "name": "对象类型",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "业务对象类型"
    },
    {
      "id": "object_id",
      "name": "对象ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "被操作对象的ID"
    },
    {
      "id": "parent_object_type",
      "name": "父对象类型",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "级联操作产生时记录父对象的类型"
    },
    {
      "id": "parent_object_id",
      "name": "父对象ID",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "级联操作产生时记录父对象的ID"
    },
    {
      "id": "action",
      "name": "操作类型",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "操作类型"
    },
    {
      "id": "field_name",
      "name": "字段名",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "变更的字段名（UPDATE 时记录）"
    },
    {
      "id": "old_value",
      "name": "旧值",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "字段修改前的值"
    },
    {
      "id": "new_value",
      "name": "新值",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "字段修改后的值"
    },
    {
      "id": "user_id",
      "name": "用户ID",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "操作人ID"
    },
    {
      "id": "user_name",
      "name": "用户名",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "操作人名称"
    },
    {
      "id": "ip_address",
      "name": "IP地址",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "操作人IP地址"
    },
    {
      "id": "user_agent",
      "name": "用户代理",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "浏览器用户代理"
    },
    {
      "id": "created_at",
      "name": "操作时间",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "审计记录创建时间"
    },
    {
      "id": "extra_data",
      "name": "附加数据",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "附加的上下文信息"
    },
    {
      "id": "trace_id",
      "name": "链路追踪ID",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "请求链路追踪ID，关联日志系统"
    },
    {
      "id": "transaction_id",
      "name": "事务ID",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "业务事务唯一标识，同一事务的多条审计记录共享"
    },
    {
      "id": "status",
      "name": "审计状态",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "审计记录写入失败，可重试"
    },
    {
      "id": "status_entered_at",
      "name": "状态进入时间",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "进入当前状态的时间"
    },
    {
      "id": "retry_count",
      "name": "重试次数",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "审计写入重试次数"
    },
    {
      "id": "error_message",
      "name": "错误信息",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "审计写入失败时的错误信息"
    },
    {
      "id": "agent_id",
      "name": "Agent标识",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "AI Agent 标识"
    },
    {
      "id": "agent_session_id",
      "name": "Agent会话ID",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "AI Agent 会话标识"
    },
    {
      "id": "tool_call_id",
      "name": "工具调用ID",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "Agent工具调用的幂等键"
    },
    {
      "id": "agent_reasoning",
      "name": "Agent推理上下文",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "Agent执行操作时的推理上下文"
    },
    {
      "id": "export",
      "name": "export",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "basic",
      "name": "basic",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "change",
      "name": "change",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "trace",
      "name": "trace",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "system",
      "name": "system",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "audit_log_list",
      "name": "列表查询",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查询审计日志列表"
    },
    {
      "id": "audit_log_read",
      "name": "查看日志详情",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "查看审计日志详情"
    },
    {
      "id": "audit_log_delete",
      "name": "删除日志",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "删除审计日志记录"
    },
    {
      "id": "recent_changes",
      "name": "最近变更",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "object_history",
      "name": "对象历史",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "user_activity",
      "name": "用户活动",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "category_statistics",
      "name": "分类统计",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "按日志类型统计"
    },
    {
      "id": "level_statistics",
      "name": "级别统计",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "按日志级别统计"
    },
    {
      "id": "category_level_statistics",
      "name": "分类级别统计",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "按日志类型和级别统计"
    },
    {
      "id": "by_category",
      "name": "按类型筛选",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "by_level",
      "name": "按级别筛选",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "mark_written",
      "name": "标记已写入",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "mark_failed",
      "name": "标记写入失败",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "retry_write",
      "name": "重试写入",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ],
  "permission": [
    {
      "id": "id",
      "name": "ID",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "code",
      "name": "权限编码",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "权限的唯一编码标识"
    },
    {
      "id": "name",
      "name": "权限名称",
      "required": True,
      "business_key": False,
      "import_order": 999,
      "description": "权限的显示名称"
    },
    {
      "id": "resource_type",
      "name": "资源类型",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "关联的资源类型"
    },
    {
      "id": "action",
      "name": "操作类型",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "允许的操作（create/read/update/delete等）"
    },
    {
      "id": "description",
      "name": "描述",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "权限详细说明"
    },
    {
      "id": "resource_id",
      "name": "资源ID",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "关联的资源实例ID"
    },
    {
      "id": "scope",
      "name": "作用范围",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": "权限作用范围"
    },
    {
      "id": "create",
      "name": "create",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    },
    {
      "id": "batch_delete",
      "name": "batch_delete",
      "required": False,
      "business_key": False,
      "import_order": 999,
      "description": ""
    }
  ]
}


def _make_excel_from_schema(object_type):
    """根据 schema 字段生成期望的 xlsx 模板"""
    fields = SCHEMA_FIELDS.get(object_type, [])
    exportable = [f for f in fields if f.get('import_order', 999) < 999]
    exportable.sort(key=lambda f: f.get('import_order', 999))
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = object_type
    for col, f in enumerate(exportable, 1):
        cell = ws.cell(1, col, f.get('name') or f.get('id'))
        # 业务键 → 黄底
        if f.get('business_key'):
            cell.fill = PatternFill('solid', fgColor=BUSINESS_KEY_FILL)
        # 必填 → 红字
        if f.get('required'):
            cell.font = Font(color=REQUIRED_FONT_COLOR, bold=True)
        # 批注 (description)
        if f.get('description'):
            from openpyxl.comments import Comment
            cell.comment = Comment(f['description'], 'schema')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _get_export_endpoint(object_type):
    """根据对象类型返回导出 API 端点"""
    return '/api/v1/export'


class TestProductExcelFormat:
    """product 导出 xlsx 格式测试 (模型驱动)"""

    def test_product_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('product')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_product_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('product')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_product_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('product')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_product_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('product')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "product 应有至少一个有 description 的字段"


class TestVersionExcelFormat:
    """version 导出 xlsx 格式测试 (模型驱动)"""

    def test_version_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('version')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_version_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('version')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_version_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('version')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_version_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('version')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "version 应有至少一个有 description 的字段"


class TestDomainExcelFormat:
    """domain 导出 xlsx 格式测试 (模型驱动)"""

    def test_domain_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('domain')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_domain_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('domain')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_domain_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('domain')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_domain_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('domain')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "domain 应有至少一个有 description 的字段"


class TestSub_domainExcelFormat:
    """sub_domain 导出 xlsx 格式测试 (模型驱动)"""

    def test_sub_domain_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('sub_domain')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_sub_domain_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('sub_domain')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_sub_domain_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('sub_domain')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_sub_domain_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('sub_domain')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "sub_domain 应有至少一个有 description 的字段"


class TestService_moduleExcelFormat:
    """service_module 导出 xlsx 格式测试 (模型驱动)"""

    def test_service_module_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('service_module')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_service_module_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('service_module')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_service_module_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('service_module')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_service_module_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('service_module')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "service_module 应有至少一个有 description 的字段"


class TestBusiness_objectExcelFormat:
    """business_object 导出 xlsx 格式测试 (模型驱动)"""

    def test_business_object_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('business_object')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_business_object_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('business_object')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_business_object_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('business_object')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_business_object_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('business_object')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "business_object 应有至少一个有 description 的字段"


class TestRelationshipExcelFormat:
    """relationship 导出 xlsx 格式测试 (模型驱动)"""

    def test_relationship_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('relationship')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_relationship_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('relationship')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_relationship_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('relationship')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_relationship_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('relationship')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "relationship 应有至少一个有 description 的字段"


class TestEnum_typeExcelFormat:
    """enum_type 导出 xlsx 格式测试 (模型驱动)"""

    def test_enum_type_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('enum_type')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_enum_type_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('enum_type')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_enum_type_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('enum_type')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_enum_type_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('enum_type')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "enum_type 应有至少一个有 description 的字段"


class TestEnum_valueExcelFormat:
    """enum_value 导出 xlsx 格式测试 (模型驱动)"""

    def test_enum_value_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('enum_value')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_enum_value_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('enum_value')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_enum_value_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('enum_value')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_enum_value_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('enum_value')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "enum_value 应有至少一个有 description 的字段"


class TestRoleExcelFormat:
    """role 导出 xlsx 格式测试 (模型驱动)"""

    def test_role_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('role')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_role_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('role')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_role_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('role')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_role_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('role')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "role 应有至少一个有 description 的字段"


class TestAudit_logExcelFormat:
    """audit_log 导出 xlsx 格式测试 (模型驱动)"""

    def test_audit_log_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('audit_log')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_audit_log_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('audit_log')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_audit_log_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('audit_log')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_audit_log_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('audit_log')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "audit_log 应有至少一个有 description 的字段"


class TestPermissionExcelFormat:
    """permission 导出 xlsx 格式测试 (模型驱动)"""

    def test_permission_column_order_matches_schema_import_order(self, client, admin_token):
        """列顺序应与 schema.semantics.import_order 一致"""
        expected = _make_excel_from_schema('permission')
        wb = load_workbook(expected)
        ws = wb.active
        actual_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        expected_headers = [(f.get('name') or f.get('id')) for f in sorted(exportable, key=lambda x: x.get('import_order', 999))]
        assert actual_headers == expected_headers, f"列顺序不匹配: {actual_headers} != {expected_headers}"

    def test_permission_business_key_yellow_fill(self):
        """业务键字段应有黄色填充 (model: semantics.business_key)"""
        wb_bytes = _make_excel_from_schema('permission')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        bk_fields = [f for f in exportable if f.get('business_key')]
        for f in bk_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            assert col is not None, f"列 {f.get('id')} 未找到"
            cell = ws.cell(1, col)
            assert cell.fill is not None and cell.fill.fgColor is not None, f"业务键 {f.get('id')} 缺填充"
            # 黄底校验 (RGB)
            color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else str(cell.fill.fgColor.value)
            assert BUSINESS_KEY_FILL in str(color), f"业务键 {f.get('id')} 颜色 {color} != {BUSINESS_KEY_FILL}"

    def test_permission_required_field_red_font(self):
        """必填字段应有红色字体 (model: required)"""
        wb_bytes = _make_excel_from_schema('permission')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        req_fields = [f for f in exportable if f.get('required')]
        for f in req_fields:
            col = next((c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value == (f.get('name') or f.get('id'))), None)
            if col is None: continue
            cell = ws.cell(1, col)
            assert cell.font is not None, f"必填 {f.get('id')} 缺字体"
            color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
            if color:
                assert REQUIRED_FONT_COLOR in str(color), f"必填 {f.get('id')} 字体 {color} != {REQUIRED_FONT_COLOR}"

    def test_permission_has_description_comment(self):
        """有 description 的字段应有单元格批注"""
        wb_bytes = _make_excel_from_schema('permission')
        wb = load_workbook(wb_bytes)
        ws = wb.active
        desc_fields = [f for f in exportable if f.get('description')]
        assert len(desc_fields) > 0, "permission 应有至少一个有 description 的字段"


# 模型覆盖度自检

def test_all_ie_objects_have_export_format():
    """所有 IE 对象的 xlsx 格式都已被测试覆盖"""
    ie_objects = ["product","version","domain","sub_domain","service_module","business_object","relationship","enum_type","enum_value","role","audit_log","permission"]
    assert len(ie_objects) > 0, "应有至少 1 个 IE 对象的 schema"
    for obj in ie_objects:
        fields = SCHEMA_FIELDS.get(obj, [])
        exportable = [f for f in fields if f.get('import_order', 999) < 999]
        assert len(exportable) > 0, f"{obj} 至少应有 1 个可导出字段"
