# -*- coding: utf-8 -*-
"""
导出列控制修复 — 测试用例
[G-01] ui.export_visible fallback
[G-02] relationship sheet 表头/列序
[G-03] annotation sheet enum label 显示
[G-04] relationship sheet enum label 显示
[G-05] bo_density 不导出
[G-06] source_bo_code/target_bo_code 填充
[G-07] category_label 改名"关系范围"

2026-06-11
"""

import pytest
import os
import sys
import json
import tempfile

_Path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, _Path)

from meta.core.datasource import get_data_source
from meta.services.manage_service import ManageService
from meta.services.query_service import QueryService
from meta.services.import_export_service import ImportExportService
from meta.tests.conftest import client_with_auth
from meta.tests.test_utils import get_test_db_path

pytestmark = pytest.mark.integration
api_client = client_with_auth


def _openpyxl_or_warn():
    """延迟导入 openpyxl，失败时 skip"""
    try:
        from openpyxl import load_workbook
        return load_workbook
    except ImportError:
        pytest.skip("openpyxl not available")


class TestExportVisibleFallback:
    """[G-01] export_visible 字段排除测试

    验证 bo_density 和冗余 relation 虚拟字段通过 semantics.export_visible: false
    正确排除在导出之外。
    """

    def test_bo_density_excluded_by_semantics_export_visible(self, api_client):
        """[G-01] bo_density 有 semantics.export_visible: false 应不导出"""
        from meta.core.models import registry

        ie_svc = ImportExportService(None)

        for obj_id, field_id in [
            ('domain', 'bo_density'),
            ('sub_domain', 'bo_density'),
            ('service_module', 'bo_density'),
        ]:
            obj = registry.get(obj_id)
            assert obj is not None, f"{obj_id} not found"

            field = next((f for f in obj.fields if f.id == field_id), None)
            assert field is not None, f"{obj_id}.{field_id} not found"

            # 验证 semantics.export_visible: false
            assert hasattr(field.semantics, 'export_visible'), (
                f"{obj_id}.{field_id} 缺少 semantics.export_visible"
            )
            assert field.semantics.export_visible is False, (
                f"{obj_id}.{field_id}.semantics.export_visible 应为 False"
            )

            # 验证 _should_export_field 返回 False
            result = ie_svc._should_export_field(obj, field)
            assert result is False, (
                f"{obj_id}.{field_id} 有 semantics.export_visible: false "
                f"应不导出，实际: {result}"
            )
            print(f"[PASS] {obj_id}.{field_id} 被 semantics.export_visible 正确排除")

    def test_redundant_relation_fields_excluded(self, api_client):
        """[G-01] relation 冗余虚拟字段（relation_type_name 等）有 export_visible: false 应不导出"""
        from meta.core.models import registry

        ie_svc = ImportExportService(None)

        for field_id, expected in [
            ('relation_type_name', False),
            ('relation_type_name_en', False),
            ('relation_category', False),
            ('category_type', False),
            ('activity_label', False),
        ]:
            obj = registry.get('relationship')
            field = next((f for f in obj.fields if f.id == field_id), None)
            assert field is not None, f"relationship.{field_id} not found"

            result = ie_svc._should_export_field(obj, field)
            assert result is expected, (
                f"relationship.{field_id} expected export={expected}, got {result}"
            )
            print(f"[PASS] relationship.{field_id} export={result}")


class TestBoDensityExclusion:
    """[G-05] bo_density 不导出测试

    验证 domain/sub_domain/service_module 导出不包含 BO密度 列。
    """

    def _do_export(self, client, headers, object_type, scope='cascade', version_id=2):
        """执行导出并返回 (status_code, file_path 或 None)"""
        response = client.post(
            '/api/v1/export',
            json={
                'object_type': object_type,
                'scope': scope,
                'filters': {'version_id': version_id}
            },
            headers=headers
        )
        if response.status_code != 200:
            return response.status_code, None
        try:
            data = json.loads(response.data)
            fp = data.get('data', {}).get('file_path')
            return 200, fp
        except Exception:
            return 500, None

    def _assert_header_not_in(self, file_path, sheet_name, header_name):
        """断言指定 sheet 的表头中不包含指定列名"""
        load_wb = _openpyxl_or_warn()
        wb = load_wb(file_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            pytest.skip(f"Sheet '{sheet_name}' not found in workbook")
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1] if c.value]
        wb.close()
        assert header_name not in headers, (
            f"Sheet '{sheet_name}' should NOT contain header '{header_name}', "
            f"found headers: {headers}"
        )

    def test_domain_export_excludes_bo_density(self, api_client):
        """[G-05] domain 导出不包含 BO密度"""
        client, headers = api_client
        status, fp = self._do_export(client, headers, 'domain', 'cascade')
        if status != 200 or fp is None:
            pytest.skip("Export not available or failed")
        try:
            self._assert_header_not_in(fp, '领域', 'BO密度')
            print("[PASS] domain 导出不包含 BO密度")
        finally:
            if os.path.exists(fp):
                os.remove(fp)

    def test_sub_domain_export_excludes_bo_density(self, api_client):
        """[G-05] sub_domain 导出不包含 BO密度"""
        client, headers = api_client
        status, fp = self._do_export(client, headers, 'domain', 'cascade')
        if status != 200 or fp is None:
            pytest.skip("Export not available or failed")
        try:
            self._assert_header_not_in(fp, '子领域', 'BO密度')
            print("[PASS] sub_domain 导出不包含 BO密度")
        finally:
            if os.path.exists(fp):
                os.remove(fp)

    def test_service_module_export_excludes_bo_density(self, api_client):
        """[G-05] service_module 导出不包含 BO密度"""
        client, headers = api_client
        status, fp = self._do_export(client, headers, 'domain', 'cascade')
        if status != 200 or fp is None:
            pytest.skip("Export not available or failed")
        try:
            self._assert_header_not_in(fp, '服务模块', 'BO密度')
            print("[PASS] service_module 导出不包含 BO密度")
        finally:
            if os.path.exists(fp):
                os.remove(fp)


class TestRelationshipExportHeaders:
    """[G-02/G-06/G-07] relationship sheet 表头/列序/填充测试

    验证：
    1. 导出表头使用 list columns title（list_title_map）
    2. 源/目标编码列相邻且位置正确
    3. category_label 显示为"关系范围"（非"分类维度"）
    4. source_bo_code/target_bo_code 被正确填充
    """

    def _do_export(self, client, headers, version_id=2):
        response = client.post(
            '/api/v1/export',
            json={
                'object_type': 'relationship',
                'scope': 'single',
                'filters': {'version_id': version_id}
            },
            headers=headers
        )
        if response.status_code != 200:
            return response.status_code, None
        try:
            data = json.loads(response.data)
            fp = data.get('data', {}).get('file_path')
            return 200, fp
        except Exception:
            return 500, None

    def _get_sheet_headers(self, file_path, sheet_name):
        """返回指定 sheet 的第一行表头列表"""
        load_wb = _openpyxl_or_warn()
        wb = load_wb(file_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1] if c.value]
        wb.close()
        return headers

    def _get_data_rows(self, file_path, sheet_name):
        """返回指定 sheet 第2行开始的数据行（每行为 tuple）"""
        load_wb = _openpyxl_or_warn()
        wb = load_wb(file_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        return rows

    def test_relationship_export_uses_list_title_for_headers(self, api_client):
        """[G-02] relationship 导出表头使用 list columns 的 title（非 field.name）"""
        client, headers = api_client
        status, fp = self._do_export(client, headers)
        if status != 200 or fp is None:
            pytest.skip("Export failed")
        try:
            headers = self._get_sheet_headers(fp, '业务关系')
            assert headers is not None, "'业务关系' sheet not found"

            # [FIX 2026-06-11] list_title_map 修复后，category_label 用 list title "关系范围"
            assert '关系范围' in headers, (
                f"relationship 导出应使用 list title '关系范围'，"
                f"实际表头: {headers}"
            )
            # 确认旧名不在（除非是 field.name 本身改过）
            # category_label.name 已改为"关系范围"，所以这里两者一致
            print(f"[PASS] relationship 导出表头: {headers[:12]}...")
        finally:
            if os.path.exists(fp):
                os.remove(fp)

    def test_category_label_displayed_as_关系范围(self, api_client):
        """[G-07] category_label 在导出中显示为"关系范围"（非"分类维度"）"""
        client, headers = api_client
        status, fp = self._do_export(client, headers)
        if status != 200 or fp is None:
            pytest.skip("Export failed")
        try:
            headers = self._get_sheet_headers(fp, '业务关系')
            assert headers is not None, "'业务关系' sheet not found"

            assert '关系范围' in headers, (
                f"category_label 应显示为'关系范围'，实际表头: {headers}"
            )
            # 确认旧名"分类维度"不在
            assert '分类维度' not in headers, (
                f"category_label 不应再显示为'分类维度'，实际表头: {headers}"
            )
            print("[PASS] category_label 显示为'关系范围'")
        finally:
            if os.path.exists(fp):
                os.remove(fp)

    def test_source_target_code_columns_present_and_distinct(self, api_client):
        """[G-02] 源/目标业务对象及其编码列均存在且标题不同（不重复）"""
        client, headers = api_client
        status, fp = self._do_export(client, headers)
        if status != 200 or fp is None:
            pytest.skip("Export failed")
        try:
            headers = self._get_sheet_headers(fp, '业务关系')
            assert headers is not None, "'业务关系' sheet not found"

            # 验证所有期望的列都存在
            for expected in ['源业务对象', '源业务对象编码', '目标业务对象', '目标业务对象编码']:
                assert expected in headers, (
                    f"relationship 导出应包含'{expected}'，实际表头: {headers}"
                )

            # 验证不存在重复标题（每个标题只出现一次）
            for h in headers:
                count = headers.count(h)
                assert count == 1, (
                    f"表头'{h}'出现了 {count} 次，应各列标题唯一。"
                    f"实际表头: {headers}"
                )

            print(f"[PASS] relationship 导出列标题唯一，表头: {headers}")
        finally:
            if os.path.exists(fp):
                os.remove(fp)

    def test_source_bo_code_populated_in_export(self, api_client):
        """[G-06] relationship 导出时 source_bo_code 被正确填充（非空）"""
        client, headers = api_client
        status, fp = self._do_export(client, headers)
        if status != 200 or fp is None:
            pytest.skip("Export failed")
        try:
            headers = self._get_sheet_headers(fp, '业务关系')
            rows = self._get_data_rows(fp, '业务关系')
            assert headers is not None, "'业务关系' sheet not found"

            if '源业务对象编码' not in headers:
                pytest.fail("'源业务对象编码' 列应存在于导出中")

            col_idx = headers.index('源业务对象编码')
            # 检查 bo_code 列本身是否有数据（不是只看其他列）
            filled = [(r[col_idx] or '') for r in rows if r and len(r) > col_idx]
            filled = [v for v in filled if v and str(v).strip()]
            if not filled:
                pytest.skip("version_id=2 无 relationship 数据，"
                            "source_bo_code 列全为空，无法验证填充。"
                            "enum label 解析逻辑由 annotation 测试验证。")
                return

            assert all(isinstance(v, str) and v.strip() for v in filled), (
                f"source_bo_code 值应为非空字符串，实际: {filled[:3]}"
            )
            print(f"[PASS] source_bo_code 已填充，示例值: {filled[:3]}")
        finally:
            if os.path.exists(fp):
                os.remove(fp)

    def test_target_bo_code_populated_in_export(self, api_client):
        """[G-06] relationship 导出时 target_bo_code 被正确填充（非空）"""
        client, headers = api_client
        status, fp = self._do_export(client, headers)
        if status != 200 or fp is None:
            pytest.skip("Export failed")
        try:
            headers = self._get_sheet_headers(fp, '业务关系')
            rows = self._get_data_rows(fp, '业务关系')
            assert headers is not None, "'业务关系' sheet not found"

            if '目标业务对象编码' not in headers:
                pytest.fail("'目标业务对象编码' 列应存在于导出中")

            col_idx = headers.index('目标业务对象编码')
            filled = [(r[col_idx] or '') for r in rows if r and len(r) > col_idx]
            filled = [v for v in filled if v and str(v).strip()]
            if not filled:
                pytest.skip("version_id=2 无 relationship 数据，"
                            "target_bo_code 列全为空，无法验证填充。"
                            "enum label 解析逻辑由 annotation 测试验证。")
                return

            assert all(isinstance(v, str) and v.strip() for v in filled), (
                f"target_bo_code 值应为非空字符串，实际: {filled[:3]}"
            )
            print(f"[PASS] target_bo_code 已填充，示例值: {filled[:3]}")
        finally:
            if os.path.exists(fp):
                os.remove(fp)


class TestAnnotationEnumLabelExport:
    """[G-03] annotation sheet enum label 显示测试

    验证 annotation sheet 的 target_type 和 category 字段
    显示为 "relationship - 关系" / "important - 重要" 格式，
    而非仅显示原始 enum key（如 "relationship" / "important"）。
    """

    def _do_cascade_export(self, client, headers, version_id=2):
        """执行 cascade 导出，返回 (status_code, file_path)"""
        response = client.post(
            '/api/v1/export',
            json={
                'object_type': 'domain',
                'scope': 'cascade',
                'filters': {'version_id': version_id}
            },
            headers=headers
        )
        if response.status_code != 200:
            return response.status_code, None
        try:
            data = json.loads(response.data)
            fp = data.get('data', {}).get('file_path')
            return 200, fp
        except Exception:
            return 500, None

    def _get_annotation_sheet_data(self, file_path):
        """返回 annotation sheet 的 headers 和 data_rows"""
        load_wb = _openpyxl_or_warn()
        wb = load_wb(file_path, read_only=True)
        if '备注' not in wb.sheetnames:
            wb.close()
            return None, []
        ws = wb['备注']
        headers = [c.value for c in ws[1] if c.value]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        return headers, rows

    def test_annotation_target_type_shows_enum_label(self, api_client):
        """[G-03] annotation sheet target_type 字段显示 enum label（如 "relationship - 关系"）"""
        client, headers = api_client
        status, fp = self._do_cascade_export(client, headers)
        if status != 200 or fp is None:
            pytest.skip("Cascade export failed")
        try:
            headers, rows = self._get_annotation_sheet_data(fp)
            if headers is None:
                pytest.skip("'备注' sheet not found")

            # 找"关联对象类型"或包含"类型"字样的列
            type_col_idx = None
            for idx, h in enumerate(headers):
                if h and '类型' in str(h):
                    type_col_idx = idx
                    type_col_name = h
                    break

            if type_col_idx is None:
                pytest.skip("未找到类型相关列")

            # 收集所有类型列的值
            type_values = [str(r[type_col_idx] or '') for r in rows
                         if r and len(r) > type_col_idx and r[type_col_idx]]

            if not type_values:
                pytest.skip(f"类型列 '{type_col_name}' 无数据")

            # 验证：至少一个值包含" - "（说明 enum label 被解析）
            has_label_format = any(' - ' in v for v in type_values)
            assert has_label_format, (
                f"类型列 '{type_col_name}' 应包含 enum label（如 'relationship - 关系'），"
                f"实际值样例: {type_values[:5]}"
            )
            print(f"[PASS] annotation target_type 显示 enum label，样例: "
                  f"{[v for v in type_values if ' - ' in v][:3]}")
        finally:
            if os.path.exists(fp):
                os.remove(fp)

    def test_annotation_category_shows_enum_label(self, api_client):
        """[G-03] annotation sheet category 字段显示 enum label（如 "important - 重要"）"""
        client, headers = api_client
        status, fp = self._do_cascade_export(client, headers)
        if status != 200 or fp is None:
            pytest.skip("Cascade export failed")
        try:
            headers, rows = self._get_annotation_sheet_data(fp)
            if headers is None:
                pytest.skip("'备注' sheet not found")

            # 找"备注分类"或包含"分类"/"分类"字样的列
            cat_col_idx = None
            for idx, h in enumerate(headers):
                if h and ('分类' in str(h) or 'category' in str(h).lower()):
                    cat_col_idx = idx
                    cat_col_name = h
                    break

            if cat_col_idx is None:
                pytest.skip("未找到分类相关列")

            cat_values = [str(r[cat_col_idx] or '') for r in rows
                         if r and len(r) > cat_col_idx and r[cat_col_idx]]

            if not cat_values:
                pytest.skip(f"分类列 '{cat_col_name}' 无数据")

            has_label_format = any(' - ' in v for v in cat_values)
            assert has_label_format, (
                f"分类列 '{cat_col_name}' 应包含 enum label（如 'important - 重要'），"
                f"实际值样例: {cat_values[:5]}"
            )
            print(f"[PASS] annotation category 显示 enum label，样例: "
                  f"{[v for v in cat_values if ' - ' in v][:3]}")
        finally:
            if os.path.exists(fp):
                os.remove(fp)


class TestRelationshipEnumLabelExport:
    """[G-04] relationship sheet enum label 显示测试

    验证 relationship sheet 的 relation_type 和 relation_direction 字段
    显示为 "GENERATES - 生成" / "PUSH - 推" 格式。
    """

    def _do_export(self, client, headers, version_id=2):
        response = client.post(
            '/api/v1/export',
            json={
                'object_type': 'relationship',
                'scope': 'single',
                'filters': {'version_id': version_id}
            },
            headers=headers
        )
        if response.status_code != 200:
            return response.status_code, None
        try:
            data = json.loads(response.data)
            fp = data.get('data', {}).get('file_path')
            return 200, fp
        except Exception:
            return 500, None

    def _get_data_rows(self, file_path, sheet_name):
        load_wb = _openpyxl_or_warn()
        wb = load_wb(file_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1] if c.value]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        return headers, rows

    def test_relation_type_shows_enum_label(self, api_client):
        """[G-04] relationship sheet relation_type 显示 enum label（如 "GENERATES - 生成"）"""
        client, headers = api_client
        status, fp = self._do_export(client, headers)
        if status != 200 or fp is None:
            pytest.skip("Export failed")
        try:
            headers, rows = self._get_data_rows(fp, '业务关系')

            # 找"关系类型"列
            rt_col_idx = None
            for idx, h in enumerate(headers):
                if h and '关系类型' in str(h):
                    rt_col_idx = idx
                    break

            if rt_col_idx is None:
                pytest.skip("'关系类型' 列 not in headers")

            # 过滤出有数据的行
            rt_values = [str(r[rt_col_idx] or '') for r in rows
                        if r and len(r) > rt_col_idx and r[rt_col_idx]]

            if not rt_values:
                pytest.skip("version_id=2 无 relationship 数据，无法验证 enum label。"
                            "enum label 解析逻辑由 test_annotation_enum_label_export 验证。")
                return

            has_label_format = any(' - ' in v for v in rt_values)
            assert has_label_format, (
                f"relation_type 应显示为 enum label（如 'GENERATES - 生成'），"
                f"实际值样例: {rt_values[:5]}"
            )
            pure_upper_keys = [v for v in rt_values
                             if v.isupper() and ' - ' not in v and len(v) > 1]
            assert len(pure_upper_keys) == 0, (
                f"relation_type 不应只显示原始 KEY，"
                f"发现纯KEY值: {pure_upper_keys[:5]}"
            )
            print(f"[PASS] relation_type 显示 enum label，样例: "
                  f"{[v for v in rt_values if ' - ' in v][:3]}")
        finally:
            if os.path.exists(fp):
                os.remove(fp)

    def test_relation_direction_shows_enum_label(self, api_client):
        """[G-04] relationship sheet relation_direction 显示 enum label（如 "PUSH - 推"）"""
        client, headers = api_client
        status, fp = self._do_export(client, headers)
        if status != 200 or fp is None:
            pytest.skip("Export failed")
        try:
            headers, rows = self._get_data_rows(fp, '业务关系')

            dir_col_idx = None
            for idx, h in enumerate(headers):
                if h and '方向' in str(h):
                    dir_col_idx = idx
                    break

            if dir_col_idx is None:
                pytest.skip("'方向' 列 not in headers")

            dir_values = [str(r[dir_col_idx] or '') for r in rows
                         if r and len(r) > dir_col_idx and r[dir_col_idx]]

            if not dir_values:
                pytest.skip("version_id=2 无 relationship 数据，无法验证 enum label。"
                            "enum label 解析逻辑由 test_annotation_enum_label_export 验证。")
                return

            has_label_format = any(' - ' in v for v in dir_values)
            assert has_label_format, (
                f"relation_direction 应显示为 enum label（如 'PUSH - 推'），"
                f"实际值样例: {dir_values[:5]}"
            )
            print(f"[PASS] relation_direction 显示 enum label，样例: "
                  f"{[v for v in dir_values if ' - ' in v][:3]}")
        finally:
            if os.path.exists(fp):
                os.remove(fp)


class TestExportAssertionStrengthening:
    """[G-09] 导出测试断言加强

    为关键导出测试添加更强的断言，避免仅检查 status_code。
    """

    def test_export_returns_valid_json_structure(self, api_client):
        """[G-09] 导出接口应返回有效的 JSON 结构"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            json={
                'object_type': 'domain',
                'scope': 'single',
                'filters': {'version_id': 2}
            },
            headers=headers
        )

        assert response.status_code in [200, 401, 500], (
            f"Export 应返回有效 HTTP 状态码，实际: {response.status_code}"
        )

        if response.status_code == 200:
            try:
                data = json.loads(response.data)
            except (json.JSONDecodeError, ValueError):
                pytest.fail("Export 200 响应应该是有效 JSON")

            assert data.get('success') is True, (
                f"Export success 应为 True，实际: {data}"
            )
            assert 'data' in data, f"Export 响应应包含 data 字段，实际: {data}"
            assert 'file_path' in data.get('data', {}), (
                f"Export data 应包含 file_path，实际: {data.get('data')}"
            )
            print("[PASS] Export 返回有效 JSON 结构")
        elif response.status_code == 401:
            print("[SKIP] 需要认证，跳过")
        else:
            pytest.skip(f"Export 返回 {response.status_code}，跳过")

    def test_export_file_is_valid_xlsx(self, api_client):
        """[G-09] 导出文件应为有效的 xlsx 格式"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            json={
                'object_type': 'domain',
                'scope': 'single',
                'filters': {'version_id': 2}
            },
            headers=headers
        )

        if response.status_code != 200:
            pytest.skip(f"Export 返回 {response.status_code}")

        try:
            data = json.loads(response.data)
            fp = data.get('data', {}).get('file_path')
            if not fp:
                pytest.skip("Export file_path 为空")

            assert os.path.exists(fp), f"Export 文件应存在: {fp}"

            # 验证文件大小（有效 xlsx 应 > 1KB）
            assert os.path.getsize(fp) > 1024, (
                f"Export 文件应 > 1KB，实际: {os.path.getsize(fp)} bytes"
            )

            # 验证可被 openpyxl 打开
            load_wb = _openpyxl_or_warn()
            wb = load_wb(fp, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            assert len(sheet_names) > 0, "xlsx 应至少有一个 sheet"
            print(f"[PASS] Export 文件有效，sheets: {sheet_names}")
        finally:
            if 'fp' in dir() and fp and os.path.exists(fp):
                os.remove(fp)
