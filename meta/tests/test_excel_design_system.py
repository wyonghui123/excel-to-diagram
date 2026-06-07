# -*- coding: utf-8 -*-
"""
SVC-003: excel_design_system 单元测试 (10 用例)

[NEW] 2026-06-07 批次: 补齐 ExcelDesignSystem 工具类测试
- 颜色常量 / 样式常量
- apply_header_style / apply_label_style / apply_value_style
- apply_section_style / apply_readonly_style / apply_required_style
- apply_business_key_style
- auto_column_width (含 min/max 边界)
- create_data_validation_list
"""
import pytest
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

pytestmark = pytest.mark.unit


class TestExcelDesignSystem:
    """ExcelDesignSystem 单元测试 (SVC-003)"""

    def test_color_constants_defined(self):
        """主色调/功能色/中性色常量"""
        from meta.services.excel_design_system import ExcelDesignSystem
        assert ExcelDesignSystem.PRIMARY_COLOR == "1565C0"
        assert ExcelDesignSystem.SUCCESS_COLOR == "2E7D32"
        assert ExcelDesignSystem.WARNING_COLOR == "F57C00"
        assert ExcelDesignSystem.ERROR_COLOR == "C62828"
        assert ExcelDesignSystem.GRAY_500 == "9E9E9E"
        assert ExcelDesignSystem.GRAY_900 == "212121"

    def test_style_constants_defined(self):
        """HEADER_FILL / LABEL_FONT / THIN_BORDER 等"""
        from meta.services.excel_design_system import ExcelDesignSystem
        assert ExcelDesignSystem.HEADER_FILL is not None
        assert ExcelDesignSystem.HEADER_FONT.bold is True
        assert ExcelDesignSystem.HEADER_FONT.color is not None
        assert ExcelDesignSystem.THIN_BORDER is not None

    def test_apply_header_style(self):
        """apply_header_style 设置 fill/font/alignment/border"""
        from meta.services.excel_design_system import ExcelDesignSystem
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Header'
        ExcelDesignSystem.apply_header_style(ws['A1'])
        cell = ws['A1']
        assert cell.font.bold is True
        assert cell.fill.fgColor.rgb is not None
        assert cell.alignment.horizontal == 'center'

    def test_apply_label_and_value_style(self):
        """apply_label_style + apply_value_style 不同样式"""
        from meta.services.excel_design_system import ExcelDesignSystem
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Label'
        ws['B1'] = 'Value'
        ExcelDesignSystem.apply_label_style(ws['A1'])
        ExcelDesignSystem.apply_value_style(ws['B1'])
        assert ws['A1'].font.bold is True
        assert ws['B1'].font.bold is None or ws['B1'].font.bold is False

    def test_apply_section_style(self):
        """apply_section_style 蓝色 section 背景"""
        from meta.services.excel_design_system import ExcelDesignSystem
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Section'
        ExcelDesignSystem.apply_section_style(ws['A1'])
        cell = ws['A1']
        assert cell.font.bold is True
        assert cell.fill.fgColor.rgb is not None

    def test_apply_readonly_required_business_key_styles(self):
        """apply_readonly/required/business_key_style 字段状态"""
        from meta.services.excel_design_system import ExcelDesignSystem
        wb = Workbook()
        ws = wb.active
        for i, style in enumerate([
            ExcelDesignSystem.apply_readonly_style,
            ExcelDesignSystem.apply_required_style,
            ExcelDesignSystem.apply_business_key_style,
        ]):
            ws.cell(row=i+1, column=1, value=f'cell{i}')
            style(ws.cell(row=i+1, column=1))
        # 三种样式都设置 fill
        assert all(ws.cell(row=i+1, column=1).fill.fgColor.rgb is not None
                   for i in range(3))

    def test_auto_column_width(self):
        """auto_column_width 根据内容调整列宽 (clamp 10-50)"""
        from meta.services.excel_design_system import ExcelDesignSystem
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'A'
        ws['A2'] = 'Very long content exceeding the max width'
        ExcelDesignSystem.auto_column_width(ws, min_width=10, max_width=50)
        col_a_width = ws.column_dimensions['A'].width
        # 应被 clamp 到 10-50 之间
        assert 10 <= col_a_width <= 50

    def test_auto_column_width_with_short_content(self):
        """短内容时列宽应 = min_width"""
        from meta.services.excel_design_system import ExcelDesignSystem
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Hi'
        ExcelDesignSystem.auto_column_width(ws, min_width=15, max_width=50)
        assert ws.column_dimensions['A'].width == 15

    def test_create_data_validation_list(self):
        """create_data_validation_list 返回 DataValidation"""
        from meta.services.excel_design_system import ExcelDesignSystem
        dv = ExcelDesignSystem.create_data_validation_list(['A', 'B', 'C'], allow_blank=True)
        assert isinstance(dv, DataValidation)
        # 公式应含 3 个值
        assert 'A' in dv.formula1
        assert 'C' in dv.formula1
        assert dv.allow_blank is True

    def test_create_data_validation_list_no_blank(self):
        """allow_blank=False"""
        from meta.services.excel_design_system import ExcelDesignSystem
        dv = ExcelDesignSystem.create_data_validation_list(['X'], allow_blank=False)
        assert dv.allow_blank is False
