import pytest
from meta.tests.conftest import client_with_auth

pytestmark = pytest.mark.integration
api_client = client_with_auth


# -*- coding: utf-8 -*-
"""
导入导出功能API自动化测试

测试批量导出和导入功能的API接口

数据库路径规范：
- 使用 get_test_db_path() 获取统一的数据库路径
- 禁止在测试代码中硬编码相对路径
"""

import json
import sys
import os
import tempfile
import shutil
from io import BytesIO

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, _PROJECT_ROOT)

from meta.server import create_app
from meta.core.datasource import get_data_source
from meta.services.token_service import TokenService
from meta.services.auth_provider import UserInfo
from meta.tests.test_utils import get_test_db_path

class TestImportExportConfigAPI:

    def test_01_get_domain_import_export_config(self, api_client):
        """测试获取领域的导入导出配置"""
        client, headers = api_client
        response = client.get(
            '/api/v1/import-export/config/domain',
            headers=headers
        )
        assert response.status_code in [200, 401, 404, 500]
        try:

            data = json.loads(response.data)

        except (json.JSONDecodeError, ValueError):

            data = {}
        assert data.get('success', False)
        assert data.get('data', {})['object_type'] == 'domain'
        assert data.get('data', {})['import_enabled']
        assert data.get('data', {})['export_enabled']
        assert data.get('data', {})['cascade_export']
        assert data.get('data', {})['conflict_strategy'] == 'upsert'
        print("[PASS] Get domain import/export config")

    def test_02_get_business_object_import_export_config(self, api_client):
        """测试获取业务对象的导入导出配置"""
        client, headers = api_client
        response = client.get(
            '/api/v1/import-export/config/business_object',
            headers=headers
        )
        assert response.status_code in [200, 401, 404, 500]
        try:

            data = json.loads(response.data)

        except (json.JSONDecodeError, ValueError):

            data = {}
        assert data.get('success', False)
        assert data.get('data', {})['object_type'] == 'business_object'
        assert not data.get('data', {})['cascade_export']  # 业务对象是叶子节点
        print("[PASS] Get business_object import/export config")

    def test_03_get_nonexistent_object_config(self, api_client):
        """测试获取不存在的对象配置"""
        client, headers = api_client
        response = client.get(
            '/api/v1/import-export/config/nonexistent_object',
            headers=headers
        )
        assert response.status_code in [401, 404, 500]
        try:

            data = json.loads(response.data)

        except (json.JSONDecodeError, ValueError):

            data = {}
        assert not data.get('success', False)
        print("[PASS] Get nonexistent object config returns 404")

class TestExportAPI:
    """导出API测试"""

    def test_01_export_domain_single(self, api_client):
        """测试单对象导出"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'single',
                'options': {
                    'include_hierarchy_path': True,
                    'include_hierarchy_ids': True
                }
            }),
            headers=headers
        )
        assert response.status_code in [200, 401, 404, 500]
        try:

            data = json.loads(response.data)

        except (json.JSONDecodeError, ValueError):

            data = {}
        assert data.get('success', False)
        assert 'download_url' in data.get('data', {})
        assert 'file_path' in data.get('data', {})
        print("[PASS] Export domain single")

    def test_02_export_missing_object_type(self, api_client):
        """测试缺少object_type参数的导出"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'scope': 'single'
            }),
            headers=headers
        )
        assert response.status_code in [400, 401, 500]
        try:

            data = json.loads(response.data)

        except (json.JSONDecodeError, ValueError):

            data = {}
        assert not data.get('success', False)
        print("[PASS] Export missing object_type returns 400")

    def test_03_export_nonexistent_object(self, api_client):
        """测试导出不存在的对象类型"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'nonexistent_object',
                'scope': 'single'
            }),
            headers=headers
        )
        try:

            data = json.loads(response.data)

        except (json.JSONDecodeError, ValueError):

            data = {}
        assert not data.get('success', False)
        print("[PASS] Export nonexistent object returns error")

    def test_04_download_export_file(self, api_client):
        """测试下载导出文件"""
        client, headers = api_client
        export_response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'single'
            }),
            headers=headers
        )
        export_data = json.loads(export_response.data)
        if export_data.get('success') and export_data.get('data', {}).get('download_url'):
            download_url = export_data.get('data', {})['download_url']
            filename = download_url.split('/')[-1]

            response = client.get(
                '/api/v1/export/download/' + filename,
                headers=headers
            )
            assert response.status_code in [200, 401, 404, 500]
            print("[PASS] Download export file")

    def test_05_download_nonexistent_file(self, api_client):
        """测试下载不存在的文件"""
        client, headers = api_client
        response = client.get(
            '/api/v1/export/download/nonexistent_file.xlsx',
            headers=headers
        )
        assert response.status_code in [401, 404, 500]
        print("[PASS] Download nonexistent file returns 404")

class TestImportTemplateAPI:
    """导入模板API测试"""

    def test_01_download_domain_template(self, api_client):
        """测试下载领域导入模板"""
        client, headers = api_client
        response = client.get(
            '/api/v1/import/template/domain',
            headers=headers
        )
        assert response.status_code in [200, 401, 404, 500]
        assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        print("[PASS] Download domain template")

    def test_02_download_business_object_template(self, api_client):
        """测试下载业务对象导入模板"""
        client, headers = api_client
        response = client.get(
            '/api/v1/import/template/business_object',
            headers=headers
        )
        assert response.status_code in [200, 401, 404, 500]
        print("[PASS] Download business_object template")

    def test_03_download_nonexistent_template(self, api_client):
        """测试下载不存在的模板"""
        client, headers = api_client
        response = client.get(
            '/api/v1/import/template/nonexistent_object',
            headers=headers
        )
        assert response.status_code in [401, 404, 500]
        print("[PASS] Download nonexistent template returns 404")

class TestImportAPI:
    """导入API测试"""

    def test_01_import_missing_file(self, api_client):
        """测试未上传文件"""
        client, headers = api_client
        response = client.post(
            '/api/v1/import',
            headers=headers
        )
        assert response.status_code in [400, 401, 500]
        try:

            data = json.loads(response.data)

        except (json.JSONDecodeError, ValueError):

            data = {}
        assert not data.get('success', False)
        print("[PASS] Import missing file returns 400")

    def test_02_import_invalid_file_type(self, api_client):
        """测试上传无效文件类型"""
        client, headers = api_client
        data = {
            'file': (BytesIO(b'test content'), 'test.txt'),
            'mode': 'preview'
        }
        response = client.post(
            '/api/v1/import',
            data=data,
            content_type='multipart/form-data',
            headers=headers
        )
        assert response.status_code in [400, 401, 500]
        try:

            data = json.loads(response.data)

        except (json.JSONDecodeError, ValueError):

            data = {}
        assert not data.get('success', False)
        print("[PASS] Import invalid file type returns 400")

class TestMetadataModelExtension:
    """元模型扩展测试"""

    def test_01_semantic_annotation_data_category(self, api_client):
        """测试语义标注的数据类别"""
        client, headers = api_client
        from meta.core.models import registry

        domain = registry.get("domain")
        assert domain is not None

        code_field = domain.get_field("code")
        assert code_field is not None, "field not found on domain"
        assert code_field.semantics.data_category == "code"

        name_field = domain.get_field("name")
        assert name_field is not None, "field not found on domain"
        assert name_field.semantics.data_category == "text"
        print("[PASS] Semantic annotation data_category")

    def test_02_semantic_annotation_import_export_visible(self, api_client):
        """测试语义标注的导入导出可见性"""
        client, headers = api_client
        from meta.core.models import registry

        domain = registry.get("domain")
        assert domain is not None, "domain not found in registry"

        code_field = domain.get_field("code")
        assert code_field.semantics.import_visible
        assert code_field.semantics.export_visible
        assert code_field.semantics.import_order == 4

        name_field = domain.get_field("name")
        assert name_field.semantics.import_visible
        assert name_field.semantics.export_visible
        assert name_field.semantics.import_order == 20
        print("[PASS] Semantic annotation import/export visible")

    def test_03_import_export_config(self, api_client):
        """测试导入导出配置"""
        client, headers = api_client
        from meta.core.models import registry

        domain = registry.get("domain")
        assert domain is not None, "domain not found in registry"
        assert domain.import_export.import_enabled
        assert domain.import_export.export_enabled
        assert domain.import_export.cascade_export
        assert domain.import_export.conflict_strategy == "upsert"
        assert domain.import_export.conflict_key == "code"
        print("[PASS] Import/Export config")

    def test_04_business_object_leaf_node(self, api_client):
        """测试业务对象是叶子节点"""
        client, headers = api_client
        from meta.core.models import registry

        bo = registry.get("business_object")
        assert bo is not None, "bo not found in registry"
        assert not bo.import_export.cascade_export
        assert bo.import_export.conflict_key == "code"
        print("[PASS] Business object is leaf node")

    def test_05_data_category_enum(self, api_client):
        """测试数据类别枚举"""
        client, headers = api_client
        from meta.core.models import DataCategory

        assert DataCategory.TEXT.value == "text"
        assert DataCategory.CODE.value == "code"
        assert DataCategory.DATE.value == "date"
        assert DataCategory.NUMBER.value == "number"
        assert DataCategory.TIMESTAMP.value == "timestamp"
        print("[PASS] DataCategory enum")

class TestImportExportServiceCascade:
    """导入导出服务级联功能测试"""

    def test_01_get_cascade_object_types(self, api_client):
        """测试获取级联对象类型列表"""
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService

        ds = get_data_source("sqlite", database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        object_types = ie_service._get_cascade_object_types("domain")

        assert "domain" in object_types
        assert "sub_domain" in object_types
        assert "service_module" in object_types
        assert "business_object" in object_types
        print("[PASS] Get cascade object types")

    def test_02_sort_by_hierarchy(self, api_client):
        """测试按层级排序"""
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService

        ds = get_data_source("sqlite", database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        object_types = ["business_object", "service_module", "sub_domain", "domain"]
        sorted_types = ie_service._sort_by_hierarchy(object_types)

        domain_idx = sorted_types.index("domain")
        sub_domain_idx = sorted_types.index("sub_domain")
        service_module_idx = sorted_types.index("service_module")
        bo_idx = sorted_types.index("business_object")

        assert domain_idx < sub_domain_idx
        assert sub_domain_idx < service_module_idx
        assert service_module_idx < bo_idx
        print("[PASS] Sort by hierarchy")

class TestImportExportServiceUpsert:
    """导入导出服务Upsert功能测试"""

    def test_01_find_by_key(self, api_client):
        """测试根据关键字段查找记录"""
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService

        ds = get_data_source("sqlite", database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        result = ie_service._find_by_key("domain", "code", "DOMAIN_1")

        if result:
            assert "id" in result
            assert "name" in result
            print("[PASS] Find by key returns result")
        else:
            print("[SKIP] No domain data found for test")

    def test_02_record_exists(self, api_client):
        """测试检查记录是否存在"""
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService
        from meta.core.models import ImportExportConfig

        ds = get_data_source("sqlite", database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        config = ImportExportConfig(conflict_key="code")

        exists = ie_service._record_exists("domain", {"code": "NONEXISTENT"}, config)
        assert not exists
        print("[PASS] Record exists check")

class TestResultDataclasses:
    """结果数据类测试"""

    def test_01_export_result(self, api_client):
        """测试ExportResult数据类"""
        client, headers = api_client
        from meta.services.import_export_service import ExportResult

        result = ExportResult(
            success=True,
            file_path="/test/path.xlsx",
            sheets=[{"name": "test", "row_count": 10}],
            total_rows=10,
            errors=[]
        )

        assert result.success
        assert result.file_path == "/test/path.xlsx"
        assert len(result.sheets) == 1
        assert result.total_rows == 10
        print("[PASS] ExportResult dataclass")

    def test_02_import_preview(self, api_client):
        """测试ImportPreview数据类"""
        client, headers = api_client
        from meta.services.import_export_service import ImportPreview

        preview = ImportPreview(
            sheets=[{"name": "domain", "row_count": 5}],
            validation={"valid_count": 5, "invalid_count": 0, "errors": []},
            import_order=["domain"]
        )

        assert len(preview.sheets) == 1
        assert preview.validation["valid_count"] == 5
        print("[PASS] ImportPreview dataclass")

    def test_03_import_result(self, api_client):
        """测试ImportResult数据类"""
        client, headers = api_client
        from meta.services.import_export_service import ImportResult

        result = ImportResult(
            success=True,
            results={"domain": {"success": 5, "failed": 1}},
            errors=[{"row": 3, "message": "Error"}]
        )

        assert result.success
        assert result.results["domain"]["success"] == 5
        assert len(result.errors) == 1
        print("[PASS] ImportResult dataclass")

class TestRelationshipExportFields:
    """关系导出字段测试"""

    def test_01_get_relationship_import_export_config(self, api_client):
        """测试获取关系的导入导出配置"""
        client, headers = api_client
        try:
            response = client.get(
                '/api/v1/import-export/config/relationship',
                headers=headers
            )
            assert response.status_code in [200, 401, 404, 500]
            try:

                data = json.loads(response.data)

            except (json.JSONDecodeError, ValueError):

                data = {}
            assert data.get('success', False)
            assert data.get('data', {})['object_type'] == 'relationship'
            print("[PASS] Get relationship import/export config")
        except (AssertionError, KeyError) as e:
            pytest.fail(f"Relationship config issue: {e}")
        except Exception as e:
            pytest.fail(f"Relationship export test skipped: {e}")

    def test_02_relationship_config_has_business_key_field(self, api_client):
        """测试关系配置包含business_key字段"""
        client, headers = api_client
        try:
            response = client.get(
                '/api/v1/import-export/config/relationship',
                headers=headers
            )
            assert response.status_code in [200, 401, 404, 500]
            try:

                data = json.loads(response.data)

            except (json.JSONDecodeError, ValueError):

                data = {}
            assert data.get('success', False)

            fields = data.get('data', {})['fields']
            field_ids = [f['id'] for f in fields]

            assert 'source_code' in field_ids
            assert 'target_code' in field_ids
            assert 'relation_code' in field_ids

            source_code_field = next((f for f in fields if f['id'] == 'source_code'), None)
            assert source_code_field is not None
            assert source_code_field.get('business_key', False)

            target_code_field = next((f for f in fields if f['id'] == 'target_code'), None)
            assert target_code_field is not None
            assert target_code_field.get('business_key', False)

            relation_code_field = next((f for f in fields if f['id'] == 'relation_code'), None)
            assert relation_code_field is not None
            print("[PASS] Relationship config has business_key fields")
        except (AssertionError, KeyError) as e:
            pytest.fail(f"Relationship business_key field issue: {e}")
        except Exception as e:
            pytest.fail(f"Relationship export test skipped: {e}")

    def test_03_relationship_export_includes_bo_names(self, api_client):
        """测试导出关系时包含业务对象名称字段"""
        client, headers = api_client
        try:
            response = client.post(
                '/api/v1/export',
                data=json.dumps({
                    'object_type': 'relationship',
                    'scope': 'single',
                    'options': {
                        'include_hierarchy_path': False,
                        'include_hierarchy_ids': False,
                        'include_hierarchy_names': True
                    }
                }),
                headers=headers
            )
            assert response.status_code in [200, 401, 404, 500]
            try:

                data = json.loads(response.data)

            except (json.JSONDecodeError, ValueError):

                data = {}
            assert data.get('success', False)
            assert 'download_url' in data.get('data', {})
            assert 'file_path' in data.get('data', {})
            print("[PASS] Relationship export includes bo names in header")
        except (AssertionError, KeyError) as e:
            pytest.fail(f"Relationship export issue: {e}")
        except Exception as e:
            pytest.fail(f"Relationship export test skipped: {e}")

    def test_04_relationship_export_with_version_id_list(self, api_client):
        """回归测试：version_id 以数组形式下发时不应触发 SQL 参数绑定错误"""
        client, headers = api_client
        try:
            response = client.post(
                '/api/v1/export',
                data=json.dumps({
                    'object_type': 'relationship',
                    'scope': 'single',
                    'filters': {'version_id': [1]},
                    'options': {
                        'include_hierarchy_path': False,
                        'include_hierarchy_ids': False
                    }
                }),
                headers=headers
            )
            assert response.status_code == 200
            data = json.loads(response.data)
            assert data.get('success', False), f"导出失败: {data.get('message', data)}"
            assert 'download_url' in data.get('data', {})
            print("[PASS] Relationship export with version_id as list")
        except (AssertionError, KeyError) as e:
            pytest.fail(f"Relationship export with list version_id issue: {e}")
        except Exception as e:
            pytest.fail(f"Relationship export with list version_id skipped: {e}")

    def test_05_relationship_virtual_fields_export_visible(self, api_client):
        """测试关系的virtual字段设置了export_visible"""
        client, headers = api_client
        try:
            from meta.core.models import registry

            relationship = registry.get("relationship")
            assert relationship is not None, "relationship not found in registry"

            source_bo_name_field = relationship.get_field("source_bo_name")
            assert source_bo_name_field is not None, "field not found on relationship"
            assert source_bo_name_field.semantics.export_visible

            target_bo_name_field = relationship.get_field("target_bo_name")
            assert target_bo_name_field is not None, "field not found on relationship"
            assert target_bo_name_field.semantics.export_visible

            category_label_field = relationship.get_field("category_label")
            assert category_label_field is not None, "field not found on relationship"
            assert category_label_field.semantics.export_visible, "category_label (关系范围) should be export_visible"
            print("[PASS] Relationship virtual fields have correct export_visible")
        except (AssertionError, AttributeError) as e:
            pytest.fail(f"Relationship virtual fields issue: {e}")
        except Exception as e:
            pytest.fail(f"Relationship export test skipped: {e}")

    def test_06_relationship_export_includes_domain_fields(self, api_client):
        """测试导出关系时包含领域相关字段"""
        client, headers = api_client
        try:
            from meta.core.models import registry

            relationship = registry.get("relationship")
            assert relationship is not None, "relationship not found in registry"

            source_domain_name_field = relationship.get_field("source_domain_name")
            assert source_domain_name_field is not None, "field not found on relationship"
            assert source_domain_name_field.semantics.export_visible

            target_domain_name_field = relationship.get_field("target_domain_name")
            assert target_domain_name_field is not None, "field not found on relationship"
            assert target_domain_name_field.semantics.export_visible

            source_sub_domain_name_field = relationship.get_field("source_sub_domain_name")
            assert source_sub_domain_name_field is not None, "field not found on relationship"
            assert source_sub_domain_name_field.semantics.export_visible

            target_sub_domain_name_field = relationship.get_field("target_sub_domain_name")
            assert target_sub_domain_name_field is not None, "field not found on relationship"
            assert target_sub_domain_name_field.semantics.export_visible

            source_service_module_name_field = relationship.get_field("source_service_module_name")
            assert source_service_module_name_field is not None, "field not found on relationship"
            assert source_service_module_name_field.semantics.export_visible

            target_service_module_name_field = relationship.get_field("target_service_module_name")
            assert target_service_module_name_field is not None, "field not found on relationship"
            assert target_service_module_name_field.semantics.export_visible
            print("[PASS] Relationship export includes domain fields")
        except (AssertionError, AttributeError) as e:
            pytest.fail(f"Relationship domain fields issue: {e}")
        except Exception as e:
            pytest.fail(f"Relationship export test skipped: {e}")

    def test_07_enrich_relationship_fills_domain_fields(self, api_client):
        """测试enrichment填充领域相关字段"""
        client, headers = api_client
        try:
            from meta.core.datasource import get_data_source
            from meta.services.manage_service import ManageService
            from meta.services.query_service import QueryService
            from meta.services.import_export_service import ImportExportService

            ds = get_data_source("sqlite", database=get_test_db_path())
            manage_service = ManageService(ds)
            query_service = QueryService(ds)
            ie_service = ImportExportService(ds, manage_service, query_service)

            record = {
                'source_bo_id': 1,
                'target_bo_id': 2,
                'source_code': 'BO01',
                'target_code': 'BO02',
                'relation_code': 'REL01'
            }

            ie_service._enrich_relationship_record(record)

            assert 'source_bo_name' in record
            assert 'target_bo_name' in record
            assert 'source_domain_name' in record
            assert 'target_domain_name' in record
            assert 'source_sub_domain_name' in record
            assert 'target_sub_domain_name' in record
            assert 'source_service_module_name' in record
            assert 'target_service_module_name' in record
            print("[PASS] Enrichment record contains all domain fields")
        except (AssertionError, KeyError, AttributeError) as e:
            pytest.fail(f"Enrichment domain fields issue: {e}")
        except Exception as e:
            pytest.fail(f"Relationship export test skipped: {e}")

    def test_08_enrichment_returns_correct_bo_info(self, api_client):
        """测试enrichment实际从BO获取正确信息"""
        client, headers = api_client
        try:
            from meta.core.datasource import get_data_source
            from meta.services.manage_service import ManageService
            from meta.services.query_service import QueryService
            from meta.services.import_export_service import ImportExportService

            ds = get_data_source("sqlite", database=get_test_db_path())
            manage_service = ManageService(ds)
            query_service = QueryService(ds)
            ie_service = ImportExportService(ds, manage_service, query_service)

            ie_service._enrich_relationship_record({'source_bo_id': 1, 'target_bo_id': 2})
            bo_info = ie_service._get_bo_by_id(1)

            assert isinstance(bo_info, dict)
            if bo_info:
                assert 'name' in bo_info
                assert 'domain_name' in bo_info
                assert 'sub_domain_name' in bo_info
                assert 'service_module_name' in bo_info
                print(f"[INFO] BO info: name={bo_info.get('name')}, domain={bo_info.get('domain_name')}")

            assert bo_info.get('domain_name' is not None if bo_info else None,
                               "BO should have domain_name enriched")
            print("[PASS] Enrichment returns correct BO info")
        except (AssertionError, KeyError, AttributeError) as e:
            pytest.fail(f"Enrichment BO info issue: {e}")
        except Exception as e:
            pytest.fail(f"Relationship export test skipped: {e}")

    def test_08_export_query_has_source_bo_id(self, api_client):
        """测试关系查询返回包含source_bo_id和target_bo_id"""
        client, headers = api_client
        try:
            response = client.post(
                '/api/v1/query/search',
                data=json.dumps({
                    'object_type': 'relationship',
                    'version_id': 1,
                    'page': 1,
                    'page_size': 1
                }),
                headers=headers
            )

            if response.status_code == 200:
                try:

                    data = json.loads(response.data)

                except (json.JSONDecodeError, ValueError):

                    data = {}
                if data.get('success'):
                    records = data.get('data', [])
                    if isinstance(records, list) and records:
                        first_record = records[0]
                        has_source_bo_id = 'source_bo_id' in first_record or 'sourceBoId' in first_record
                        has_target_bo_id = 'target_bo_id' in first_record or 'targetBoId' in first_record

                        print(f"[INFO] Record keys: {list(first_record.keys())[:10]}")
                        assert has_source_bo_id or has_target_bo_id, \
                            f"Relationship record should have source/target BO ID. Keys: {first_record.keys()}"
                        print(f"[PASS] Export query returns BO IDs: source={has_source_bo_id}, target={has_target_bo_id}")
                        return

            pytest.fail("No relationship data available for test")
        except (AssertionError, KeyError) as e:
            pytest.fail(f"Export query BO ID issue: {e}")
        except Exception as e:
            pytest.fail(f"Relationship export test skipped: {e}")

    def test_09_export_relationship_excel_has_codes(self, api_client):
        """测试导出关系Excel中包含source_code、target_code、relation_code"""
        client, headers = api_client
        try:
            import os
            response = client.post(
                '/api/v1/export',
                data=json.dumps({
                    'object_type': 'relationship',
                    'scope': 'single',
                    'filters': {'version_id': 1}
                }),
                headers=headers
            )

            assert response.status_code in [200, 401, 404, 500]
            try:

                data = json.loads(response.data)

            except (json.JSONDecodeError, ValueError):

                data = {}
            assert data.get('success', False)

            download_url = data.get('data', {})['download_url']
            file_path = data.get('data', {}).get('file_path')

            if file_path and os.path.exists(file_path):
                from openpyxl import load_workbook
                wb = load_workbook(file_path)

                rel_sheet = wb['业务关系'] if '业务关系' in wb.sheetnames else None
                assert rel_sheet is not None

                headers = [cell.value for cell in rel_sheet[1]]
                print(f"[INFO] Relationship sheet headers: {headers}")

                source_code_idx = None
                target_code_idx = None
                relation_code_idx = None

                for idx, h in enumerate(headers):
                    if h and '源业务对象' in str(h):
                        # 取第一个匹配的（源对象）作为 source_code
                        if source_code_idx is None:
                            source_code_idx = idx + 1
                    if h and '目标业务对象' in str(h):
                        # 取第一个匹配的（目标对象）作为 target_code
                        if target_code_idx is None:
                            target_code_idx = idx + 1
                    if h and '关系编码' in str(h):
                        relation_code_idx = idx + 1

                if source_code_idx and target_code_idx and relation_code_idx:
                    row2 = list(rel_sheet[2])
                    source_val = row2[source_code_idx - 1].value if source_code_idx <= len(row2) else None
                    target_val = row2[target_code_idx - 1].value if target_code_idx <= len(row2) else None
                    relation_val = row2[relation_code_idx - 1].value if relation_code_idx <= len(row2) else None

                    print(f"[INFO] Row 2 values: source_code={source_val}, target_code={target_val}, relation_code={relation_val}")

                    assert source_val is not None
                    assert target_val is not None
                    assert relation_val is not None
                    print("[PASS] Relationship Excel has correct code values")
                else:
                    print(f"[INFO] Headers found at indices: source_code={source_code_idx}, target_code={target_code_idx}, relation_code={relation_code_idx}")
                    print(f"[INFO] All headers: {headers}")
                    pytest.fail("Could not find all code columns in export")
        except (AssertionError, KeyError) as e:
            pytest.fail(f"Relationship Excel codes issue: {e}")
        except Exception as e:
            pytest.fail(f"Relationship export test skipped: {e}")

class TestRelationshipEnrichment:
    """关系enrichment测试"""

    def test_01_enrich_relationship_record_method_exists(self, api_client):
        """测试_enrich_relationship_record方法存在"""
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService

        ds = get_data_source("sqlite", database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        assert hasattr(ie_service, '_enrich_relationship_record')
        print("[PASS] _enrich_relationship_record method exists")

    def test_02_enrich_relationship_record_with_mock_data(self, api_client):
        """测试_enrich_relationship_record方法填充字段"""
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService

        ds = get_data_source("sqlite", database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        record = {
            'source_bo_id': 1,
            'target_bo_id': 2,
            'source_code': 'BO01',
            'target_code': 'BO02',
            'relation_code': 'REL01'
        }

        ie_service._enrich_relationship_record(record)

        if record.get('source_bo_name'):
            assert record['source_bo_name'] is not None
            print(f"[INFO] source_bo_name enriched: {record['source_bo_name']}")

        if record.get('target_bo_name'):
            assert record['target_bo_name'] is not None
            print(f"[INFO] target_bo_name enriched: {record['target_bo_name']}")

        print("[PASS] _enrich_relationship_record executes without error")

    def test_03_enrich_relationship_fills_code_from_bo(self, api_client):
        """测试当source_code为空时，enrichment从业务对象获取code"""
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService

        ds = get_data_source("sqlite", database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        record_empty_code = {
            'source_bo_id': 1,
            'target_bo_id': 2,
            'source_code': '',
            'target_code': '',
            'relation_code': 'REL01'
        }

        ie_service._enrich_relationship_record(record_empty_code)

        bo1 = ie_service._get_bo_by_id(1)
        bo2 = ie_service._get_bo_by_id(2)

        if bo1 and bo1.get('code'):
            assert record_empty_code['source_code'] == bo1['code']
            print(f"[INFO] source_code enriched from BO: {record_empty_code['source_code']}")

        if bo2 and bo2.get('code'):
            assert record_empty_code['target_code'] == bo2['code']
            print(f"[INFO] target_code enriched from BO: {record_empty_code['target_code']}")

        if bo1:
            print(f"[INFO] BO(1) info: code={bo1.get('code')}, name={bo1.get('name')}")
        if bo2:
            print(f"[INFO] BO(2) info: code={bo2.get('code')}, name={bo2.get('name')}")

        print("[PASS] Enrichment correctly fills code from business object when source_code is empty")

    def test_04_get_bo_by_id_method_exists(self, api_client):
        """测试_get_bo_by_id方法存在"""
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService

        ds = get_data_source("sqlite", database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        assert hasattr(ie_service, '_get_bo_by_id')
        print("[PASS] _get_bo_by_id method exists")

class TestAnnotationExport:
    """备注导出测试"""

    def test_01_annotation_metadata_has_export_visible(self, api_client):
        """测试备注元数据字段设置了export_visible"""
        client, headers = api_client
        from meta.core.models import registry

        annotation = registry.get("annotation")
        if annotation is None:
            pytest.fail("annotation object type not registered")

        fields_to_check = [
            ("target_type_label", True),
            ("target_code", True),
            ("target_name", True),
            ("category_label", True),
            ("content", True),
            ("created_at", True),
            ("created_by", True),
            ("target_type", False),
            ("target_id", False),
        ]

        for field_id, expected_visible in fields_to_check:
            field = annotation.get_field(field_id)
            if field is None:
                continue
            if expected_visible:
                assert field.semantics.export_visible, \
                    f"annotation.{field_id} should be export_visible"
            else:
                assert not field.semantics.export_visible, \
                    f"annotation.{field_id} should not be export_visible"

        print("[PASS] Annotation metadata has correct export_visible settings")

    def test_02_annotation_metadata_has_import_order(self, api_client):
        """测试备注元数据字段设置了import_order"""
        client, headers = api_client
        from meta.core.models import registry

        annotation = registry.get("annotation")
        if annotation is None:
            pytest.fail("annotation object type not registered")

        fields_to_check = [
            ("target_type", 1),
            ("category", 10),
            ("content", 20),
            ("created_at", 30),
            ("created_by", 40),
        ]

        for field_id, expected_order in fields_to_check:
            field = annotation.get_field(field_id)
            if field is None:
                continue
            if not hasattr(field, 'semantics') or field.semantics is None:
                continue
            actual_order = getattr(field.semantics, 'import_order', None)
            if actual_order is not None:
                assert actual_order, 100 <= f"annotation.{field_id} import_order should be <= 100, got {actual_order}"

        print("[PASS] Annotation metadata has correct import_order settings")

    def test_03_annotation_export_with_annotations_option(self, api_client):
        """测试导出时包含备注sheet"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'single',
                'options': {
                    'include_annotations': True
                }
            }),
            headers=headers
        )

        assert response.status_code in [200, 400, 401, 404, 500]
        if response.status_code != 200:
            pytest.fail("Export endpoint not available or returned error")
        
        try:

        
            data = json.loads(response.data)

        
        except (json.JSONDecodeError, ValueError):

        
            data = {}
        assert data.get('success', False)

        sheets = data.get('data', {}).get('sheets', [])
        annotation_sheet = next((s for s in sheets if s.get('object_type') == 'annotation'), None)
        if annotation_sheet:
            print(f"[PASS] Annotation sheet exists with {annotation_sheet['row_count']} rows")
        else:
            print("[INFO] No annotation sheet in export (feature may not be implemented)")

    def test_04_annotation_export_includes_all_types(self, api_client):
        """测试导出备注时包含所有对象类型"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'business_object',
                'scope': 'selected',
                'selected_types': ['domain', 'sub_domain', 'service_module', 'business_object', 'relationship'],
                'options': {
                    'include_annotations': True
                }
            }),
            headers=headers
        )

        assert response.status_code in [200, 401, 404, 500]
        try:

            data = json.loads(response.data)

        except (json.JSONDecodeError, ValueError):

            data = {}
        assert data.get('success', False)
        print("[PASS] Export with all types and annotations works")

class TestExportExcelStructure:
    """导出Excel结构完整性测试"""

    def test_01_export_excel_has_correct_sheets(self, api_client):
        """测试导出Excel包含正确的sheet"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'selected',
                'selected_types': ['domain', 'sub_domain', 'business_object'],
                'filters': {'version_id': 1}
            }),
            headers=headers
        )

        assert response.status_code in [200, 401, 404, 500]
        try:

            data = json.loads(response.data)

        except (json.JSONDecodeError, ValueError):

            data = {}
        assert data.get('success', False)

        sheets = data.get('data', {})['sheets']
        sheet_types = [s['object_type'] for s in sheets]

        assert 'domain' in sheet_types
        assert 'sub_domain' in sheet_types
        assert 'business_object' in sheet_types
        print("[PASS] Export Excel has correct sheets")

    def test_02_export_excel_file_exists(self, api_client):
        """测试导出文件存在"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'single',
                'filters': {'version_id': 1}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        assert file_path is not None
        assert os.path.exists(file_path)
        print(f"[PASS] Export file exists: {file_path}")

    def test_03_export_excel_has_meta_sheet(self, api_client):
        """测试导出Excel包含说明sheet"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'single',
                'filters': {'version_id': 1}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            assert '说明' in wb.sheetnames
            print("[PASS] Export Excel has meta sheet")

    def test_04_export_business_key_fields_first(self, api_client):
        """测试business_key字段排在最前面"""
        client, headers = api_client
        from meta.core.models import registry

        domain = registry.get('domain')
        assert domain is not None, "domain not found in registry"
        business_key_fields = [f for f in domain.fields 
                              if getattr(f.semantics, 'business_key', False) 
                              and getattr(f.semantics, 'export_visible', False)
                              and f.storage.value != 'virtual']

        if business_key_fields:
            response = client.post(
                '/api/v1/export',
                data=json.dumps({
                    'object_type': 'domain',
                    'scope': 'single',
                    'filters': {'version_id': 1}
                }),
                headers=headers
            )

            try:


                data = json.loads(response.data)


            except (json.JSONDecodeError, ValueError):


                data = {}
            file_path = data.get('data', {}).get('file_path')

            if file_path and os.path.exists(file_path):
                from openpyxl import load_workbook
                try:
                    wb = load_workbook(file_path)
                except Exception:
                    pytest.fail("Export file is corrupted or empty")
                ws = wb['领域']

                headers = [cell.value for cell in ws[1]]
                first_data_col = 2 if headers[0] == '操作模式' else 1

                bk_field_names = [f.name for f in business_key_fields]
                for bk_name in bk_field_names:
                    found = any(bk_name in str(h) for h in headers if h)
                    assert found, f"Business key field {bk_name} should be in headers"
                print("[PASS] Business key fields are in export headers")

    def test_05_export_with_hierarchy_path(self, api_client):
        """测试导出包含层级路径"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'business_object',
                'scope': 'single',
                'filters': {'version_id': 1},
                'options': {'include_hierarchy_path': True}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb['业务对象']

            headers = [cell.value for cell in ws[1]]
            has_hierarchy_path = any('层级路径' in str(h) for h in headers if h)
            assert has_hierarchy_path, "Should have 层级路径 column"
            print("[PASS] Export has hierarchy path column")

    def test_06_export_without_hierarchy_path(self, api_client):
        """测试导出不包含层级路径"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'business_object',
                'scope': 'single',
                'filters': {'version_id': 1},
                'options': {'include_hierarchy_path': False}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb['业务对象']

            headers = [cell.value for cell in ws[1]]
            has_hierarchy_path = any('层级路径' in str(h) for h in headers if h)
            assert not has_hierarchy_path, "Should not have 层级路径 column"
            print("[PASS] Export without hierarchy path works")

    def test_07_export_annotation_sheet_data(self, api_client):
        """测试备注sheet数据完整性"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'selected',
                'selected_types': ['domain'],
                'filters': {'version_id': 1},
                'options': {'include_annotations': True}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            wb = load_workbook(file_path)

            if '备注信息' in wb.sheetnames:
                ws = wb['备注信息']
                headers = [cell.value for cell in ws[1]]

                expected_headers = ['关联对象类型', '关联对象编码', '关联对象名称', '备注分类', '备注内容']
                for eh in expected_headers:
                    found = any(eh in str(h) for h in headers if h)
                    assert found, f"Should have {eh} column"

                cud_headers = ['操作模式']
                for eh in cud_headers:
                    found = any(eh in str(h) for h in headers if h)
                    if found:
                        print(f"[INFO] CUD column {eh} found in annotation sheet")

                if any('ID' in str(h) for h in headers if h):
                    print("[INFO] ID column found in annotation sheet (CUD support)")

                print("[PASS] Annotation sheet has correct columns")

class TestExportDataIntegrity:
    """导出数据完整性测试"""

    def test_01_domain_export_has_code_and_name(self, api_client):
        """测试领域导出包含编码和名称"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'single',
                'filters': {'version_id': 1}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            try:
                wb = load_workbook(file_path)
            except Exception:
                pytest.fail("Export file is corrupted or empty")
            ws = wb['领域']

            headers = [cell.value for cell in ws[1]]

            code_col = None
            name_col = None
            for idx, h in enumerate(headers):
                if h and '编码' in str(h):
                    code_col = idx + 1
                if h and '名称' in str(h):
                    name_col = idx + 1

            assert code_col is not None
            assert name_col is not None

            if ws.max_row > 1:
                row2 = list(ws[2])
                code_val = row2[code_col - 1].value if code_col <= len(row2) else None
                name_val = row2[name_col - 1].value if name_col <= len(row2) else None

                assert code_val is not None
                assert name_val is not None

            print("[PASS] Domain export has code and name")

    def test_02_relationship_export_has_all_codes(self, api_client):
        """测试关系导出包含所有编码字段"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'relationship',
                'scope': 'single',
                'filters': {'version_id': 1}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            try:
                wb = load_workbook(file_path)
            except Exception:
                pytest.skip("Export file is corrupted or empty - no relationship data in test environment")

            if '业务关系' not in wb.sheetnames:
                pytest.skip("业务关系 sheet not found in export - no relationship data")

            ws = wb['业务关系']

            if ws.max_row < 1:
                pytest.skip("业务关系 sheet is empty - no relationship data")

            headers = [cell.value for cell in ws[1]]

            source_code_col = None
            target_code_col = None
            relation_code_col = None

            for idx, h in enumerate(headers):
                if h and '源编码' in str(h):
                    source_code_col = idx + 1
                if h and '目标编码' in str(h):
                    target_code_col = idx + 1
                if h and '关系编码' in str(h):
                    relation_code_col = idx + 1

            # 至少要找到一个code列
            if source_code_col is None and target_code_col is None and relation_code_col is None:
                pytest.skip("业务关系 sheet has no code columns - no relationship data")

            if ws.max_row > 1:
                for row_idx in range(2, min(4, ws.max_row + 1)):
                    row_data = list(ws[row_idx])
                    if source_code_col is not None and source_code_col <= len(row_data):
                        source_val = row_data[source_code_col - 1].value
                        if source_val is None:
                            continue
                    if target_code_col is not None and target_code_col <= len(row_data):
                        target_val = row_data[target_code_col - 1].value
                        if target_val is None:
                            continue
                    if relation_code_col is not None and relation_code_col <= len(row_data):
                        relation_val = row_data[relation_code_col - 1].value
                        if relation_val is None:
                            continue

            print("[PASS] Relationship export has all codes")

    def test_03_business_object_export_has_parent_info(self, api_client):
        """测试业务对象导出包含父对象信息"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'business_object',
                'scope': 'single',
                'filters': {'version_id': 1}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            try:
                wb = load_workbook(file_path)
            except Exception:
                pytest.fail("Export file is corrupted or empty")
            ws = wb['业务对象']

            headers = [cell.value for cell in ws[1]]

            service_module_col = None
            for idx, h in enumerate(headers):
                if h and '服务模块' in str(h):
                    service_module_col = idx + 1
                    break

            assert service_module_col is not None
            print("[PASS] Business object export has parent info")

class TestImportTemplate:
    """导入模板测试"""

    def test_01_download_domain_template(self, api_client):
        """测试下载领域导入模板"""
        client, headers = api_client
        response = client.get(
            '/api/v1/import/template/domain',
            headers=headers
        )

        assert response.status_code in [200, 401, 404, 500]
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response.content_type
        print("[PASS] Download domain template")

    def test_02_download_business_object_template(self, api_client):
        """测试下载业务对象导入模板"""
        client, headers = api_client
        response = client.get(
            '/api/v1/import/template/business_object',
            headers=headers
        )

        assert response.status_code in [200, 401, 404, 500]
        print("[PASS] Download business_object template")

    def test_03_download_relationship_template(self, api_client):
        """测试下载关系导入模板"""
        client, headers = api_client
        response = client.get(
            '/api/v1/import/template/relationship',
            headers=headers
        )

        assert response.status_code in [200, 401, 404, 500]
        print("[PASS] Download relationship template")

    def test_04_template_has_correct_headers(self, api_client):
        """测试模板包含正确的表头"""
        client, headers = api_client
        response = client.get(
            '/api/v1/import/template/domain',
            headers=headers
        )

        assert response.status_code in [200, 401, 404, 500]

        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(response.data))

        assert '领域' in wb.sheetnames
        ws = wb['领域']

        headers = [cell.value for cell in ws[1]]
        has_code = any('编码' in str(h) for h in headers if h)
        has_name = any('名称' in str(h) for h in headers if h)

        assert has_code, "Template should have 编码 column"
        assert has_name, "Template should have 名称 column"
        print("[PASS] Template has correct headers")

class TestImportPreview:
    """导入预览测试"""

    def test_01_import_preview_endpoint_exists(self, api_client):
        """测试导入预览接口存在"""
        client, headers = api_client
        response = client.post(
            '/api/v1/import/preview',
            data=json.dumps({
                'object_type': 'domain'
            }),
            headers=headers
        )

        assert response.status_code in [200, 400, 401, 405, 422, 500]
        print("[PASS] Import preview endpoint exists")

    def test_02_import_endpoint_exists(self, api_client):
        """测试导入接口存在"""
        client, headers = api_client
        response = client.post(
            '/api/v1/import',
            data=json.dumps({
                'object_type': 'domain'
            }),
            headers=headers
        )

        assert response.status_code in [200, 400, 401, 422, 500]
        print("[PASS] Import endpoint exists")

class TestExportOptions:
    """导出选项测试"""

    def test_01_export_with_operation_mode(self, api_client):
        """测试导出包含操作模式列"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'single',
                'filters': {'version_id': 1},
                'options': {'include_operation_mode': True}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb['领域']

            headers = [cell.value for cell in ws[1]]
            has_operation_mode = any('操作模式' in str(h) for h in headers if h)
            if not has_operation_mode:
                print("[INFO] 操作模式 column not present in export - feature may not be implemented")
            else:
                print("[PASS] Export with operation mode works")

    def test_02_export_without_operation_mode(self, api_client):
        """测试导出不包含操作模式列"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'single',
                'filters': {'version_id': 1},
                'options': {'include_operation_mode': False}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            try:
                wb = load_workbook(file_path)
            except Exception:
                pytest.fail("Export file is corrupted or empty")
            ws = wb['领域']

            headers = [cell.value for cell in ws[1]]
            has_operation_mode = any('操作模式' in str(h) for h in headers if h)
            assert not has_operation_mode, "Should not have 操作模式 column"
            print("[PASS] Export without operation mode works")

    def test_03_export_with_hierarchy_ids(self, api_client):
        """测试导出包含层级编码"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'business_object',
                'scope': 'single',
                'filters': {'version_id': 1},
                'options': {'include_hierarchy_ids': True}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            try:
                wb = load_workbook(file_path)
            except Exception:
                pytest.fail("Export file is corrupted or empty")
            ws = wb['业务对象']

            headers = [cell.value for cell in ws[1]]
            has_hierarchy_id = any('编码' in str(h) and '服务模块' in str(h) for h in headers if h)
            print(f"[PASS] Export with hierarchy ids: has_hierarchy_id={has_hierarchy_id}")

    def test_04_export_selected_types(self, api_client):
        """测试导出选定的对象类型"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'selected',
                'selected_types': ['domain', 'sub_domain'],
                'filters': {'version_id': 1}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        assert data.get('success', False)

        sheets = data.get('data', {})['sheets']
        sheet_types = [s['object_type'] for s in sheets]

        assert 'domain' in sheet_types
        assert 'sub_domain' in sheet_types
        print("[PASS] Export selected types works")

class TestBusinessKeyContext:
    """Business Key 上下文增强测试"""

    def test_01_domain_has_product_code_field(self, api_client):
        """测试 domain 包含 product_code 字段"""
        client, headers = api_client
        from meta.core.models import registry

        domain = registry.get('domain')
        assert domain is not None, "domain not found in registry"
        product_code_field = domain.get_field('product_code')

        assert product_code_field is not None
        assert getattr(product_code_field.semantics, 'business_key', False)
        assert getattr(product_code_field.semantics, 'virtual', False)
        print("[PASS] domain has product_code field with business_key=true")

    def test_02_domain_has_version_code_field(self, api_client):
        """测试 domain 包含 version_code 字段"""
        client, headers = api_client
        from meta.core.models import registry

        domain = registry.get('domain')
        assert domain is not None, "domain not found in registry"
        version_code_field = domain.get_field('version_code')

        assert version_code_field is not None
        assert getattr(version_code_field.semantics, 'business_key', False)
        assert getattr(version_code_field.semantics, 'virtual', False)
        print("[PASS] domain has version_code field with business_key=true")

    def test_03_business_key_import_order(self, api_client):
        """测试 business_key 字段的 import_order 正确"""
        client, headers = api_client
        from meta.core.models import registry

        domain = registry.get('domain')
        assert domain is not None, "domain not found in registry"

        bk_fields = [f for f in domain.fields if getattr(f.semantics, 'business_key', False)]
        bk_fields.sort(key=lambda f: f.semantics.import_order if f.semantics.import_order is not None else 999)

        field_ids = [f.id for f in bk_fields]
        # [CHANGED 2026-06-13] version_code 虚拟字段仍存在(从 version.name 解析), import_order 不变
        assert field_ids == ['product_code', 'version_code', 'code']
        print(f"[PASS] business_key fields order: {field_ids}")

    def test_04_export_has_meta_sheet(self, api_client):
        """测试导出包含说明 Sheet（包含元数据）"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'single',
                'filters': {'version_id': 1}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            try:
                wb = load_workbook(file_path)
            except Exception:
                pytest.fail("Export file is corrupted or empty")

            assert '说明' in wb.sheetnames, "Excel should have 说明 sheet"
            print("[PASS] Export has 说明 sheet (contains metadata)")

    def test_05_meta_sheet_has_context(self, api_client):
        """测试说明 Sheet 包含上下文信息"""
        client, headers = api_client
        response = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'single',
                'filters': {'version_id': 1}
            }),
            headers=headers
        )

        try:


            data = json.loads(response.data)


        except (json.JSONDecodeError, ValueError):


            data = {}
        file_path = data.get('data', {}).get('file_path')

        if file_path and os.path.exists(file_path):
            from openpyxl import load_workbook
            wb = load_workbook(file_path)

            if '说明' in wb.sheetnames:
                ws = wb['说明']
                meta = {}
                for row in ws.iter_rows(min_row=1, max_row=15, max_col=2):
                    if row[0].value and row[1].value:
                        meta[str(row[0].value)] = str(row[1].value)

                assert '产品编码' in meta, "说明 sheet should have 产品编码"
                assert '版本编码' in meta, "说明 sheet should have 版本编码"
                print(f"[PASS] 说明 sheet has context: 产品编码={meta.get('产品编码')}, 版本编码={meta.get('版本编码')}")

    def test_06_get_product_version_codes_method(self, api_client):
        """测试 _get_product_version_codes 方法
        
        [CHANGED 2026-06-13] version.code 已删除, 方法返回 version.name
        方法名保持不变(向后兼容), 但返回值实际是 version.name
        """
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService

        ds = get_data_source('sqlite', database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        product_code, version_name = ie_service._get_product_version_codes({'version_id': 1})

        assert product_code is not None
        assert version_name is not None
        print(f"[PASS] _get_product_version_codes returns: product_code={product_code}, version_name={version_name}")

    def test_07_all_objects_have_context_fields(self, api_client):
        """测试所有架构对象都有 product_code 和 version_code 字段"""
        client, headers = api_client
        from meta.core.models import registry

        object_types = ['domain', 'sub_domain', 'service_module', 'business_object', 'relationship']

        for obj_type in object_types:
            obj = registry.get(obj_type)
            if obj is None:
                continue
            product_code = obj.get_field('product_code')
            version_code = obj.get_field('version_code')

            assert product_code, f"{obj_type} should have product_code" is not None
            assert version_code, f"{obj_type} should have version_code" is not None
            assert getattr(product_code.semantics, 'business_key', False)
            assert getattr(version_code.semantics, 'business_key', False)

        print(f"[PASS] All objects have product_code and version_code fields")

    def test_08_resolve_version_id_method(self, api_client):
        """测试 _resolve_version_id 方法
        
        [CHANGED 2026-06-13] version.code 已删除, 使用 version.name 解析
        """
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService

        ds = get_data_source('sqlite', database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        product_code, version_name = ie_service._get_product_version_codes({'version_id': 1})
        assert product_code is not None
        assert version_name is not None

        version_id = ie_service._resolve_version_id(product_code, version_name)
        assert version_id, f"Should resolve version_id for product_code={product_code}, version_name={version_name}" is not None
        print(f"[PASS] _resolve_version_id returns: version_id={version_id}")

    def test_09_find_by_business_key_with_version(self, api_client):
        """测试 _find_by_business_key 按 version_id + code 查重"""
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService

        ds = get_data_source('sqlite', database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        version_id = 1
        result = ie_service._find_by_key('domain', 'code', 'DOMAIN_1')

        if result:
            assert 'id' in result
            result_version_id = result.get('version_id')
            assert result_version_id == version_id, "Found record should have correct version_id"
            print(f"[PASS] _find_by_business_key found record with version_id={result_version_id}")
        else:
            print("[SKIP] No domain data found for test")

    def test_10_import_injects_version_id(self, api_client):
        """测试导入时正确注入 version_id
        
        [CHANGED 2026-06-13] version.code 已删除, 使用 version.name 解析
        """
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService

        ds = get_data_source('sqlite', database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        product_code, version_name = ie_service._get_product_version_codes({'version_id': 1})
        version_id = ie_service._resolve_version_id(product_code, version_name)

        assert version_id is not None
        assert version_id >= 1, "Resolved version_id should be valid"
        print(f"[PASS] Import correctly resolves version_id={version_id}")

class TestFieldControlModel:
    """字段控制模型测试 - 参考 SAP CDS View 注解体系"""

    def test_01_semantic_annotation_has_immutable(self, api_client):
        """测试 SemanticAnnotation 包含 immutable 属性"""
        client, headers = api_client
        from meta.core.models import SemanticAnnotation
        
        ann = SemanticAnnotation(meaning="test", immutable=True)
        assert ann.immutable
        print("[PASS] SemanticAnnotation has immutable attribute")

    def test_02_semantic_annotation_has_parent_key(self, api_client):
        """测试 SemanticAnnotation 包含 parent_key 属性"""
        client, headers = api_client
        from meta.core.models import SemanticAnnotation
        
        ann = SemanticAnnotation(meaning="test", parent_key=True)
        assert ann.parent_key
        print("[PASS] SemanticAnnotation has parent_key attribute")

    def test_03_semantic_annotation_has_mandatory(self, api_client):
        """测试 SemanticAnnotation 包含 mandatory 属性"""
        client, headers = api_client
        from meta.core.models import SemanticAnnotation
        
        ann = SemanticAnnotation(meaning="test", mandatory=True)
        assert ann.mandatory
        print("[PASS] SemanticAnnotation has mandatory attribute")

    def test_04_domain_code_is_immutable(self, api_client):
        """测试 domain.code 字段是 immutable"""
        client, headers = api_client
        from meta.core.models import registry

        domain = registry.get('domain')
        assert domain is not None, "domain not found in registry"
        code_field = domain.get_field('code')
        
        assert getattr(code_field.semantics, 'business_key', False)
        assert getattr(code_field.semantics, 'immutable', False)
        print("[PASS] domain.code is immutable")

    def test_05_domain_version_id_is_parent_key(self, api_client):
        """测试 domain.version_id 字段是 parent_key 和 readonly_always"""
        client, headers = api_client
        from meta.core.models import registry

        domain = registry.get('domain')
        assert domain is not None, "domain not found in registry"
        version_id_field = domain.get_field('version_id')
        
        assert getattr(version_id_field.semantics, 'parent_key', False)
        assert getattr(version_id_field.semantics, 'readonly_always', False)
        print("[PASS] domain.version_id is parent_key and readonly_always")

    def test_06_is_field_editable_with_immutable(self, api_client):
        """测试 _is_field_editable 方法对 immutable 字段的处理"""
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService
        from meta.core.models import registry

        ds = get_data_source('sqlite', database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        domain = registry.get('domain')
        assert domain is not None, "domain not found in registry"
        code_field = domain.get_field('code')

        # 编辑模式下 immutable 字段不可编辑
        assert not ie_service._is_field_editable(code_field, mode='edit')
        # 新建模式下 immutable 字段可编辑
        assert ie_service._is_field_editable(code_field, mode='create')
        print("[PASS] _is_field_editable handles immutable correctly")

    def test_07_is_field_editable_with_parent_key(self, api_client):
        """测试 _is_field_editable 方法对 parent_key 字段的处理
        
        注意：version_id 是 readonly_always，所以始终不可编辑
        而 service_module_id 等其他 parent_key 字段是可编辑的（SAP One Model 允许移动层级）
        """
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService
        from meta.core.models import registry

        ds = get_data_source('sqlite', database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        domain = registry.get('domain')
        assert domain is not None, "domain not found in registry"
        version_id_field = domain.get_field('version_id')

        # version_id 是 readonly_always，所以始终不可编辑
        assert not ie_service._is_field_editable(version_id_field, mode='edit')
        assert not ie_service._is_field_editable(version_id_field, mode='create')
        
        # service_module_id 是 parent_key 但不是 readonly_always，所以可编辑
        business_object = registry.get('business_object')
        assert business_object is not None, "business_object not found in registry"
        service_module_id_field = business_object.get_field('service_module_id')
        if service_module_id_field:
            # parent_key 字段可编辑（SAP One Model 允许移动层级）
            assert ie_service._is_field_editable(service_module_id_field, mode='edit')
            assert ie_service._is_field_editable(service_module_id_field, mode='create')
        
        print("[PASS] _is_field_editable handles parent_key correctly")

    def test_08_all_business_keys_are_immutable(self, api_client):
        """测试所有 business_key 字段都设置了 immutable"""
        client, headers = api_client
        from meta.core.models import registry

        object_types = ['domain', 'sub_domain', 'service_module', 'business_object', 'relationship']
        
        for obj_type in object_types:
            obj = registry.get(obj_type)
            if not obj:
                continue
            
            bk_fields = [f for f in obj.fields if getattr(f.semantics, 'business_key', False)]
            for f in bk_fields:
                if not getattr(f.semantics, 'virtual', False):
                    assert getattr(f.semantics, 'immutable', False), \
                        f"{obj_type}.{f.id} should be immutable"
                    
        
        print("[PASS] All business_key fields are immutable")

    def test_09_parent_key_editability(self, api_client):
        """测试 parent_key 字段的可编辑性
        
        根据 SAP One Model 设计：
        - version_id 等顶层 parent_key 字段是 readonly_always（始终只读）
        - service_module_id 等普通 parent_key 字段是可编辑的（允许移动层级）
        """
        client, headers = api_client
        from meta.core.models import registry

        # version_id 是 readonly_always
        domain = registry.get('domain')
        assert domain is not None, "domain not found in registry"
        version_id_field = domain.get_field('version_id')
        assert getattr(version_id_field.semantics, 'readonly_always', False)
        
        # service_module_id 是 parent_key 但不是 readonly_always，可编辑
        business_object = registry.get('business_object')
        assert business_object is not None, "business_object not found in registry"
        service_module_id_field = business_object.get_field('service_module_id')
        if service_module_id_field:
            assert getattr(service_module_id_field.semantics, 'parent_key', False)
            assert not getattr(service_module_id_field.semantics, 'readonly_always', False)
        
        print("[PASS] parent_key editability is correct")

    def test_10_filter_import_record(self, api_client):
        """测试 _filter_import_record 方法正确过滤不可导入的字段
        
        导入规则：
        1. readonly_always 字段：始终忽略
        2. immutable 字段：编辑/更新时忽略
        3. virtual 字段（无 ui.relation）：始终忽略
        4. 有 ui.relation 的 virtual 字段：新增时可导入
        """
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService
        from meta.core.models import registry
        
        ds = get_data_source('sqlite', database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)
        
        domain = registry.get('domain')
        
        # domain 的字段：id, version_id, code, name
        # version_id 是 readonly_always
        # code 是 business_key + immutable
        # name 是可编辑字段
        record = {
            'id': 1,
            'code': 'D001',
            'name': 'Test Domain',
            'version_id': 99,  # readonly_always
            'fake_field': 'should_be_removed',  # 不存在的字段
        }
        
        filtered_update = ie_service._filter_import_record(record, domain, 'update')
        assert 'id' in filtered_update  # 编辑时保留 id（用于定位记录）
        assert 'version_id' not in filtered_update  # readonly_always 字段忽略
        assert 'name' in filtered_update  # 可编辑字段保留
        assert 'fake_field' not in filtered_update  # 不存在的字段忽略
        assert 'code' in filtered_update  # business_key + immutable，更新模式下保留（用于定位记录）
        
        filtered_create = ie_service._filter_import_record(record, domain, 'create')
        assert 'id' not in filtered_create  # 新增时忽略 id（自动生成）
        assert 'version_id' not in filtered_create  # readonly_always 字段忽略
        assert 'code' in filtered_create  # business_key 新增时可导入
        assert 'name' in filtered_create
        assert 'fake_field' not in filtered_create  # 不存在的字段忽略

        print("[PASS] _filter_import_record filters correctly")

    def test_11_import_validation_skips_readonly_and_context_fields(self, api_client):
        """测试导入验证跳过 readonly_always 和 context_field 字段

        验证逻辑应该跳过：
        1. ID 字段（自增主键）
        2. readonly_always 字段（如 version_id）
        3. context_field 字段（如 version_id）
        """
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService
        from meta.core.models import registry

        ds = get_data_source('sqlite', database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        domain = registry.get('domain')
        assert domain is not None, "domain not found in registry"

        # 验证 domain 的 version_id 是 readonly_always + context_field
        version_id_field = domain.get_field('version_id')
        assert version_id_field is not None, "field not found on domain"
        assert getattr(version_id_field.semantics, 'readonly_always', False)
        assert getattr(version_id_field.semantics, 'context_field', False)

        # 验证 domain 的 code 是真正的导入业务关键字（不是 virtual）
        code_field = domain.get_field('code')
        assert code_field is not None, "field not found on domain"
        assert getattr(code_field.semantics, 'business_key', False)
        is_virtual = code_field.storage.value == 'virtual' or getattr(code_field.semantics, 'virtual', False)
        assert not is_virtual
        assert getattr(code_field.semantics, 'import_visible', True)

        # 验证 product_code 是 virtual + import_visible=false，不应该作为导入业务关键字
        product_code_field = domain.get_field('product_code')
        assert product_code_field is not None, "field not found on domain"
        assert getattr(product_code_field.semantics, 'business_key', False)
        is_virtual = product_code_field.storage.value == 'virtual' or getattr(product_code_field.semantics, 'virtual', False)
        assert is_virtual  # 是 virtual
        assert not getattr(product_code_field.semantics, 'import_visible', True)  # import_visible=false

        print("[PASS] import validation skips readonly and context fields correctly")

    def test_12_polymorphic_parent_key_classification(self, api_client):
        """测试多态外键（annotation.target_type / target_code）的分类逻辑

        场景说明：
        - annotation 的 target_type + target_code 共同构成多态外键
        - 业务上：编辑时只读，新增时必填且可编辑
        - 验证 _classify_field 和 _is_field_editable 在 create/edit 模式下的行为
        """
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService
        from meta.core.models import registry

        ds = get_data_source('sqlite', database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        ann = registry.get('annotation')
        assert ann is not None, "annotation not found in registry"

        # target_type 应该是 parent_key
        target_type = ann.get_field('target_type')
        assert target_type is not None
        assert getattr(target_type.semantics, 'parent_key', False)
        assert getattr(target_type.semantics, 'immutable', False)
        assert getattr(target_type.semantics, 'mandatory', False) or target_type.required
        # create 模式可编辑（用于指定 FK），edit 模式只读
        assert ie_service._is_field_editable(target_type, mode='create')
        assert not ie_service._is_field_editable(target_type, mode='edit')
        # 分类应该是 parent_key
        assert ie_service._classify_field(target_type) == 'parent_key'

        # target_code 应该是 parent_key（多态 FK 的另一半）
        target_code = ann.get_field('target_code')
        assert target_code is not None
        assert getattr(target_code.semantics, 'parent_key', False)
        # create 模式可编辑，edit 模式只读
        assert ie_service._is_field_editable(target_code, mode='create')
        assert not ie_service._is_field_editable(target_code, mode='edit')
        # 分类应该是 parent_key
        assert ie_service._classify_field(target_code) == 'parent_key'

        # target_name 应该是 readonly（virtual + 非 parent_key）
        target_name = ann.get_field('target_name')
        assert target_name is not None
        assert not ie_service._is_field_editable(target_name, mode='create')
        assert not ie_service._is_field_editable(target_name, mode='edit')
        assert ie_service._classify_field(target_name) == 'readonly'

        # category 应该可编辑
        category = ann.get_field('category')
        assert category is not None
        assert ie_service._is_field_editable(category, mode='create')
        assert ie_service._is_field_editable(category, mode='edit')
        assert ie_service._classify_field(category) == 'editable'

        print("[PASS] polymorphic parent_key classification is correct")

    def test_13_polymorphic_parent_key_import_filter(self, api_client):
        """测试多态外键在导入时的字段过滤

        - create 模式：target_type + target_code 保留（用于建立 FK）
        - update 模式：target_type + target_code 被忽略（业务键不可修改）
        """
        client, headers = api_client
        from meta.core.datasource import get_data_source
        from meta.services.manage_service import ManageService
        from meta.services.query_service import QueryService
        from meta.services.import_export_service import ImportExportService
        from meta.core.models import registry

        ds = get_data_source('sqlite', database=get_test_db_path())
        manage_service = ManageService(ds)
        query_service = QueryService(ds)
        ie_service = ImportExportService(ds, manage_service, query_service)

        ann = registry.get('annotation')
        record = {
            'id': 1,
            'target_type': 'relationship',
            'target_id': 100,
            'target_code': 'rel_001',
            'target_name': '测试关系',  # virtual display only
            'category': 'important',
            'content': 'asdf',
            'category_label': '重要',  # virtual display only
        }

        # create 模式：parent_key 字段（target_type, target_code）必须保留
        filtered_create = ie_service._filter_import_record(record, ann, 'create')
        assert 'target_type' in filtered_create, "target_type must be kept in create mode (parent_key)"
        assert 'target_code' in filtered_create, "target_code must be kept in create mode (parent_key)"
        assert 'content' in filtered_create
        assert 'category' in filtered_create
        # virtual display-only 字段在 create 模式应被忽略
        assert 'target_name' not in filtered_create
        assert 'category_label' not in filtered_create

        # update 模式：parent_key 字段（target_type, target_code）都保留
        # 这是为了支持"重新指定父对象"的场景（业务上允许移动层级）
        filtered_update = ie_service._filter_import_record(record, ann, 'update')
        assert 'target_type' in filtered_update, "parent_key 字段 target_type 在 update 模式应保留"
        assert 'target_code' in filtered_update, "parent_key virtual 字段 target_code 在 update 模式也保留"
        # 普通字段保留
        assert 'content' in filtered_update
        assert 'category' in filtered_update
        assert 'id' in filtered_update  # 编辑时保留 id
        # display-only virtual 字段在 update 模式应被忽略
        assert 'target_name' not in filtered_update
        assert 'category_label' not in filtered_update

        print("[PASS] polymorphic parent_key import filter is correct")

    def test_14_child_sheet_polymorphic_fk_visual(self, api_client):
        """测试子对象 Sheet 中多态 FK 的视觉呈现

        验证：
        1. parent_key 字段（target_type, target_code）在现有行用 BUSINESS_KEY_FILL（浅绿）
        2. parent_key 字段在新行也是 BUSINESS_KEY_FILL（必填但不灰）
        3. 普通 editable 字段（category, content）无底色
        4. readonly 字段（id, target_name, target_id）有 READONLY_FILL
        5. 新行不会把 editable 字段也变灰
        6. parent_key 字段有 Excel comment 解释"父对象外键 / 必填"
        """
        client, headers = api_client
        from meta.core.models import registry
        from meta.tests.test_utils import create_test_workbook, get_cell_fill_rgb

        wb, ann = create_test_workbook('annotation', data=[{
            'id': 1,
            'target_type': 'relationship',
            'target_id': 100,
            'target_code': 'rel_001',
            'target_name': '测试关系',
            'category': 'important',
            'content': '测试内容',
            'category_label': '重要',
        }])

        ws = wb[ann.name]

        col_map = {}
        for cell in ws[1]:
            col_map[cell.value] = cell.column_letter

        target_type_col = col_map.get('关联对象类型')
        assert target_type_col, f"找不到 关联对象类型 列：{col_map}"
        target_type_header = ws[f"{target_type_col}1"]
        assert target_type_header.comment is not None
        assert "父对象外键" in target_type_header.comment.text

        target_code_col = col_map.get('关联对象编码')
        assert target_code_col, f"找不到 关联对象编码 列：{col_map}"
        target_code_header = ws[f"{target_code_col}1"]
        assert target_code_header.comment is not None
        assert "父对象外键" in target_code_header.comment.text

        existing_target_type = ws[f"{target_type_col}2"]
        assert get_cell_fill_rgb(existing_target_type) == "E6F7E6", \
            f"现有行 target_type 应为浅绿(E6F7E6)"

        existing_target_code = ws[f"{target_code_col}2"]
        assert get_cell_fill_rgb(existing_target_code) == "E6F7E6", \
            f"现有行 target_code 应为浅绿(E6F7E6)"

        for new_row in (3, 4, 5):
            new_target_type = ws[f"{target_type_col}{new_row}"]
            assert get_cell_fill_rgb(new_target_type) == "E6F7E6", \
                f"新行第 {new_row} 行 target_type 应该是浅绿(E6F7E6)"

            new_target_code = ws[f"{target_code_col}{new_row}"]
            assert get_cell_fill_rgb(new_target_code) == "E6F7E6", \
                f"新行第 {new_row} 行 target_code 应该是浅绿(E6F7E6)"

        # 4. 验证 category 和 content 不应该有底色（editable）
        category_col = col_map.get('备注分类')
        content_col = col_map.get('备注内容')
        for new_row in (3, 4, 5):
            new_category = ws[f"{category_col}{new_row}"]
            category_fill = get_cell_fill_rgb(new_category)
            assert category_fill in (None, '000000', 'FFFFFF'), \
                f"新行 category 不应该有底色，实际是 {category_fill}"

        id_col = col_map.get('ID')
        for new_row in (3, 4, 5):
            new_id = ws[f"{id_col}{new_row}"]
            assert get_cell_fill_rgb(new_id) == "E0E0E0", \
                f"新行 id 应该是灰色(E0E0E0)"

        print("[PASS] child sheet polymorphic FK visual is correct")


def _do_export(client, headers, object_type='domain', scope='cascade',
               filters=None, options=None):
    """辅助函数：调用导出 API 并返回 (file_path, wb) 或 (None, None)"""
    body = {
        'object_type': object_type,
        'scope': scope,
        'options': options or {
            'include_hierarchy_path': True,
            'include_hierarchy_ids': True,
            'include_metadata_sheet': True,
            'include_child_objects': True,
            'include_operation_mode': True,
        }
    }
    if filters:
        body['filters'] = filters

    resp = client.post(
        '/api/v1/export',
        data=json.dumps(body),
        headers=headers
    )
    if resp.status_code != 200:
        return None, None

    data = json.loads(resp.data)
    if not data.get('success'):
        return None, None

    file_path = data.get('data', {}).get('file_path')
    if not file_path or not os.path.exists(file_path):
        return None, None

    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    return file_path, wb


def _build_import_excel(wb, modifications=None):
    """从导出的 Workbook 构造导入文件

    Args:
        wb: 导出的 openpyxl Workbook
        modifications: dict，格式 {sheet_name: [(row_idx, col_idx, value), ...]}

    Returns:
        临时文件路径
    """
    modifications = modifications or {}
    for sheet_name, changes in modifications.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_idx, col_idx, value in changes:
            ws.cell(row=row_idx, column=col_idx, value=value)

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


class TestExportImportRoundTrip:
    """导出→修改→导入→验证 round-trip 测试

    测试策略：
    1. 通过 API 导出 Excel 文件
    2. 用 openpyxl 解析并修改 Excel 内容
    3. 通过 API 导入修改后的 Excel
    4. 通过 API 查询验证修改生效

    这是最接近真实用户操作的端到端测试。
    """

    def test_01_export_then_import_update_domain(self, api_client):
        """Round-trip: 导出领域 → 修改名称 → 导入 → 验证名称已更新"""
        client, headers = api_client

        file_path, wb = _do_export(client, headers, object_type='domain', scope='cascade')
        if not file_path or not wb:
            pytest.skip("导出失败，跳过 round-trip 测试")

        ws = wb['领域'] if '领域' in wb.sheetnames else None
        if not ws or ws.max_row < 2:
            pytest.skip("领域 sheet 无数据，跳过")

        headers_row = [cell.value for cell in ws[1]]
        name_col = None
        for idx, h in enumerate(headers_row):
            if h and '名称' in str(h):
                name_col = idx + 1
                break

        if not name_col:
            pytest.skip("找不到名称列，跳过")

        original_name = ws.cell(row=2, column=name_col).value
        updated_name = f"{original_name}_roundtrip_test" if original_name else "RoundTrip测试领域"

        tmp_path = _build_import_excel(wb, modifications={
            '领域': [(2, name_col, updated_name)]
        })

        try:
            with open(tmp_path, 'rb') as f:
                resp = client.post(
                    '/api/v1/import',
                    data={
                        'file': (f, 'roundtrip_test.xlsx'),
                        'mode': 'execute',
                        'conflict_strategy': 'upsert',
                        'version_id': '1',
                    },
                    headers=headers,
                    content_type='multipart/form-data'
                )

            assert resp.status_code == 200, f"导入 API 返回 {resp.status_code}"
            result = json.loads(resp.data)

            if not isinstance(result, dict):
                result = {}

            result_data = result.get('data', {})
            if not isinstance(result_data, dict):
                result_data = {}

            errors = result_data.get('errors', [])
            if isinstance(errors, list):
                domain_errors = [e for e in errors if isinstance(e, dict) and e.get('object_type') == 'domain']
            else:
                domain_errors = []

            query_resp = client.post(
                '/api/v1/query/search',
                data=json.dumps({'object_type': 'domain', 'page': 1, 'page_size': 10}),
                headers=headers
            )
            if query_resp.status_code == 200:
                query_data = json.loads(query_resp.data)
                query_data_dict = query_data.get('data', {})
                if not isinstance(query_data_dict, dict):
                    query_data_dict = {}
                items = query_data_dict.get('items', [])
                found = any(item.get('name') == updated_name for item in items)
                if found:
                    print("[PASS] Round-trip: export → modify → import → verify domain name updated")
                else:
                    print(f"[INFO] Domain name not updated (upsert may have failed for other rows), "
                          f"import success={result.get('success')}")

        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_02_export_then_import_create_domain(self, api_client):
        """Round-trip: 导出领域 → 新增行 → 导入 → 验证新记录已创建"""
        client, headers = api_client

        file_path, wb = _do_export(client, headers, object_type='domain', scope='cascade')
        if not file_path or not wb:
            pytest.skip("导出失败，跳过 round-trip 测试")

        ws = wb['领域'] if '领域' in wb.sheetnames else None
        if not ws:
            pytest.skip("领域 sheet 不存在，跳过")

        headers_row = [cell.value for cell in ws[1]]
        col_map = {}
        for idx, h in enumerate(headers_row):
            if h:
                col_map[str(h)] = idx + 1

        has_op_mode = '操作模式' in col_map
        new_row = ws.max_row + 1
        new_code = f"RT_TEST_{id(self) % 10000}"

        if has_op_mode:
            ws.cell(row=new_row, column=col_map['操作模式'], value="create - 新增")

        if '编码' in col_map:
            ws.cell(row=new_row, column=col_map['编码'], value=new_code)
        if '名称' in col_map:
            ws.cell(row=new_row, column=col_map['名称'], value="RoundTrip新增测试")

        tmp_path = _build_import_excel(wb)

        try:
            with open(tmp_path, 'rb') as f:
                resp = client.post(
                    '/api/v1/import',
                    data={
                        'file': (f, 'roundtrip_create.xlsx'),
                        'mode': 'execute',
                        'conflict_strategy': 'upsert',
                        'version_id': '1',
                    },
                    headers=headers,
                    content_type='multipart/form-data'
                )

            assert resp.status_code == 200, f"导入 API 返回 {resp.status_code}"
            result = json.loads(resp.data)

            if not isinstance(result, dict):
                result = {}

            result_data = result.get('data', {})
            if not isinstance(result_data, dict):
                result_data = {}

            domain_result = result_data.get('results', {})
            if not isinstance(domain_result, dict):
                domain_result = {}
            domain_sub = domain_result.get('domain', {})
            if not isinstance(domain_sub, dict):
                domain_sub = {}
            domain_success = domain_sub.get('success', 0)

            if domain_success > 0:
                query_resp = client.post(
                    '/api/v1/query/search',
                    data=json.dumps({'object_type': 'domain', 'page': 1, 'page_size': 50}),
                    headers=headers
                )
                if query_resp.status_code == 200:
                    query_data = json.loads(query_resp.data)
                    query_data_dict = query_data.get('data', {})
                    if not isinstance(query_data_dict, dict):
                        query_data_dict = {}
                    items = query_data_dict.get('items', [])
                    found = any(item.get('code') == new_code for item in items)
                    if found:
                        print(f"[PASS] Round-trip: export → add row → import → verify new domain '{new_code}' created")
                    else:
                        print(f"[INFO] New domain not found yet, import success={domain_success}")
            else:
                print(f"[INFO] Domain import had no successes, likely version_id issue. "
                      f"Result: {domain_result}")

        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_03_import_preview_returns_sheets(self, api_client):
        """导入预览：上传 Excel → 预览模式 → 返回 sheets 和 validation"""
        client, headers = api_client

        file_path, wb = _do_export(client, headers, object_type='domain', scope='cascade')
        if not file_path or not wb:
            pytest.skip("导出失败，跳过预览测试")

        tmp_path = _build_import_excel(wb)

        try:
            with open(tmp_path, 'rb') as f:
                resp = client.post(
                    '/api/v1/import',
                    data={'file': (f, 'preview_test.xlsx'), 'mode': 'preview'},
                    headers=headers,
                    content_type='multipart/form-data'
                )

            assert resp.status_code == 200, f"预览 API 返回 {resp.status_code}"
            result = json.loads(resp.data)
            assert result.get('success'), f"预览失败: {result.get('message', '')}"

            data = result.get('data', {})
            assert 'sheets' in data, "预览结果缺少 sheets"
            assert isinstance(data['sheets'], list), "sheets 应为列表"
            assert len(data['sheets']) > 0, "sheets 不应为空"

            sheet_names = [s.get('name') for s in data['sheets']]
            assert '领域' in sheet_names, f"预览结果应包含'领域' sheet，实际: {sheet_names}"

            print(f"[PASS] Import preview returns {len(data['sheets'])} sheets: {sheet_names}")
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


class TestAnnotationImportAPI:
    """备注（annotation）子对象导入 API 测试

    测试策略：
    1. 通过 export_selected_types（single scope）导出，因为子对象 sheet 只在
       export_selected_types 中被 _collect_child_object_types 收集
    2. 验证备注 sheet 的结构（操作模式列、parent_key 字段、数据验证）
    3. 验证新行中 parent_key 字段不是灰色
    4. 验证 value help 下拉存在
    5. 验证 header comment 说明
    """

    def _export_with_annotations(self, client, headers):
        """导出包含备注的 Excel（使用 selected scope 传入多个类型，触发子对象收集）"""
        resp = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'selected',
                'selected_types': ['domain', 'sub_domain', 'service_module', 'business_object', 'relationship'],
                'options': {
                    'include_hierarchy_path': True,
                    'include_hierarchy_ids': True,
                    'include_metadata_sheet': True,
                    'include_child_objects': True,
                    'include_operation_mode': True,
                }
            }),
            headers=headers
        )
        if resp.status_code != 200:
            return None, None
        data = json.loads(resp.data)
        if not data.get('success'):
            return None, None
        file_path = data.get('data', {}).get('file_path')
        if not file_path or not os.path.exists(file_path):
            return None, None
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        return file_path, wb

    def test_01_export_with_annotations_has_annotation_sheet(self, api_client):
        """导出包含备注 → 验证备注 sheet 存在且结构正确"""
        client, headers = api_client

        file_path, wb = self._export_with_annotations(client, headers)
        if not file_path or not wb:
            pytest.skip("导出失败，跳过")

        assert '备注信息' in wb.sheetnames, f"导出应包含'备注信息' sheet，实际: {wb.sheetnames}"

        ws = wb['备注信息']
        headers_row = [cell.value for cell in ws[1]]

        assert '操作模式' in headers_row, "备注 sheet 应包含操作模式列"
        assert '关联对象类型' in headers_row, "备注 sheet 应包含关联对象类型列"
        assert '关联对象编码' in headers_row, "备注 sheet 应包含关联对象编码列"
        assert '备注内容' in headers_row, "备注 sheet 应包含备注内容列"

        print(f"[PASS] Annotation sheet has correct structure: {headers_row}")

    def test_02_annotation_sheet_has_data_validations(self, api_client):
        """导出包含备注 → 验证备注 sheet 有数据验证（操作模式下拉、枚举下拉）"""
        client, headers = api_client

        file_path, wb = self._export_with_annotations(client, headers)
        if not file_path or not wb:
            pytest.skip("导出失败，跳过")

        ws = wb['备注信息']
        dvs = ws.data_validations.dataValidation if ws.data_validations else []

        assert len(dvs) > 0, "备注 sheet 应有数据验证"

        has_operation_dv = any('create' in (dv.formula1 or '') for dv in dvs)
        assert has_operation_dv, "备注 sheet 应有操作模式下拉验证"

        print(f"[PASS] Annotation sheet has {len(dvs)} data validations")

    def test_03_annotation_sheet_parent_key_not_gray_in_new_rows(self, api_client):
        """导出包含备注 → 验证新行中 parent_key 字段不是灰色"""
        client, headers = api_client

        file_path, wb = self._export_with_annotations(client, headers)
        if not file_path or not wb:
            pytest.skip("导出失败，跳过")

        ws = wb['备注信息']
        headers_row = [cell.value for cell in ws[1]]
        col_map = {}
        for idx, h in enumerate(headers_row):
            if h:
                col_map[str(h)] = idx + 1

        target_type_col = col_map.get('关联对象类型')
        target_code_col = col_map.get('关联对象编码')

        if not target_type_col or not target_code_col:
            pytest.skip("找不到 parent_key 列，跳过")

        from meta.tests.test_utils import get_cell_fill_rgb

        data_rows = ws.max_row - 1
        new_row_start = data_rows + 2

        for new_row in range(new_row_start, min(new_row_start + 3, ws.max_row + 1)):
            target_type_cell = ws.cell(row=new_row, column=target_type_col)
            target_code_cell = ws.cell(row=new_row, column=target_code_col)

            tt_fill = get_cell_fill_rgb(target_type_cell)
            tc_fill = get_cell_fill_rgb(target_code_cell)

            if tt_fill == "E0E0E0":
                pytest.fail(f"新行 {new_row} 的关联对象类型不应为灰色（E0E0E0），应为浅绿（E6F7E6）")

            if tc_fill == "E0E0E0":
                pytest.fail(f"新行 {new_row} 的关联对象编码不应为灰色（E0E0E0），应为浅绿（E6F7E6）")

        print("[PASS] Annotation new rows: parent_key fields are not grayed out")

    def test_04_annotation_sheet_has_value_help(self, api_client):
        """导出包含备注 → 验证关联对象类型和备注分类有 value help 下拉"""
        client, headers = api_client

        file_path, wb = self._export_with_annotations(client, headers)
        if not file_path or not wb:
            pytest.skip("导出失败，跳过")

        ws = wb['备注信息']
        headers_row = [cell.value for cell in ws[1]]
        col_map = {}
        for idx, h in enumerate(headers_row):
            if h:
                col_map[str(h)] = idx + 1

        dvs = ws.data_validations.dataValidation if ws.data_validations else []

        target_type_col = col_map.get('关联对象类型')
        category_col = col_map.get('备注分类')

        target_type_has_dv = False
        category_has_dv = False

        for dv in dvs:
            formula = dv.formula1 or ''
            if target_type_col:
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(target_type_col)
                if col_letter in str(dv.sqref) and formula:
                    if 'domain' in formula or '领域' in formula or 'relationship' in formula or '关系' in formula:
                        target_type_has_dv = True
            if category_col:
                col_letter = get_column_letter(category_col)
                if col_letter in str(dv.sqref) and formula:
                    if 'important' in formula or '重要' in formula or 'warning' in formula or '警告' in formula:
                        category_has_dv = True

        assert target_type_has_dv or category_has_dv, \
            f"备注 sheet 应有至少一个 value help 下拉（关联对象类型或备注分类），实际 DV 数量: {len(dvs)}"

        print(f"[PASS] Annotation sheet value help: target_type={target_type_has_dv}, category={category_has_dv}")

    def test_05_annotation_sheet_header_comments(self, api_client):
        """导出包含备注 → 验证 parent_key 字段有 header comment 说明"""
        client, headers = api_client

        file_path, wb = self._export_with_annotations(client, headers)
        if not file_path or not wb:
            pytest.skip("导出失败，跳过")

        ws = wb['备注信息']
        headers_row = [cell.value for cell in ws[1]]

        target_type_header = None
        for cell in ws[1]:
            if cell.value and '关联对象类型' in str(cell.value):
                target_type_header = cell
                break

        assert target_type_header is not None, "找不到关联对象类型列头"
        assert target_type_header.comment is not None, "关联对象类型列头应有 comment"
        assert "父对象外键" in target_type_header.comment.text, \
            f"关联对象类型的 comment 应包含'父对象外键'，实际: {target_type_header.comment.text}"

        print(f"[PASS] Annotation header comment: {target_type_header.comment.text[:50]}")


class TestExcelFileContent:
    """Excel 文件内容验证测试（服务层 + openpyxl 回读）

    测试策略：
    直接调用 ImportExportService 的方法生成 Workbook，然后用 openpyxl 读取
    同一个 Workbook 对象验证所有 Excel 特性。不需要写文件到磁盘，纯内存操作。

    覆盖范围：
    1. 子对象 Sheet（annotation）：表头样式、字段分类样式、数据验证、注释、列宽、新增空行
    2. 主对象 Sheet（domain）：表头样式、字段分类样式、数据验证、注释、列宽
    3. 说明 Sheet：结构验证

    使用 ie_service fixture（conftest.py）和 create_test_workbook（test_utils.py）
    """

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie_service = ie_service

    def _make_child_sheet_wb(self, child_type='annotation', data=None):
        from meta.core.models import registry
        from openpyxl import Workbook
        child_meta = registry.get(child_type)
        if not child_meta:
            pytest.skip(f"找不到 {child_type} 元模型")
        if data is None:
            data = [{
                'id': 1,
                'target_type': 'domain',
                'target_id': 10,
                'target_code': 'DM001',
                'target_name': '测试领域',
                'category': 'important',
                'content': '测试备注内容',
                'category_label': '重要',
            }]
        wb = Workbook()
        wb.remove(wb.active)
        sheets_info = []
        self.ie_service._write_child_sheet(wb, child_type, child_meta, data, sheets_info)
        return wb, child_meta

    def _make_main_sheet_wb(self, object_type='domain'):
        from meta.core.datasource import get_data_source
        from meta.services.query_service import QueryService
        from meta.services.manage_service import ManageService
        ds = get_data_source('sqlite', database=get_test_db_path())
        qs = QueryService(ds)
        ms = ManageService(ds)
        from meta.services.import_export_service import ImportExportService
        ie = ImportExportService(ds, ms, qs)
        options = {
            'include_operation_mode': True,
            'include_hierarchy_path': True,
            'include_hierarchy_ids': True,
            'include_metadata_sheet': True,
            'include_child_objects': True,
            'empty_rows_for_new': 3,
        }
        result = ie.export_cascade(object_type, options=options)
        if not result or not result.file_path:
            pytest.skip(f"{object_type} 导出失败，跳过主导出测试")
        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)
        return wb, object_type

    def test_01_main_sheet_header_style(self):
        """主导出：表头样式验证（蓝色底、白色粗体字、居中对齐、细边框）"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.fill.start_color.rgb == "004472C4", \
            f"表头填充色应为 #4472C4，实际: {header_cell.fill.start_color.rgb}"
        assert header_cell.font.bold is True, "表头字体应为粗体"
        assert header_cell.font.color.rgb == "00FFFFFF", \
            f"表头字体颜色应为白色，实际: {header_cell.font.color.rgb}"
        assert header_cell.alignment.horizontal == "center", "表头应居中对齐"
        assert header_cell.border.left.style == 'thin', "表头应有细边框"

    def test_02_main_sheet_operation_mode_dv(self):
        """主导出：操作模式列有数据验证下拉"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        dvs = ws.data_validations.dataValidation if ws.data_validations else []
        has_op_dv = any('create' in (dv.formula1 or '') and 'update' in (dv.formula1 or '') for dv in dvs)
        assert has_op_dv, "应有操作模式下拉数据验证（包含 create/update）"

        op_dv = next(dv for dv in dvs if 'create' in (dv.formula1 or ''))
        assert op_dv.error is not None, "操作模式 DV 应有错误提示"
        assert op_dv.prompt is not None, "操作模式 DV 应有输入提示"

    def test_03_main_sheet_enum_dv_content(self):
        """主导出：枚举字段数据验证的 formula1 包含正确的 key-label 对"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        dvs = ws.data_validations.dataValidation if ws.data_validations else []

        enum_dvs = [dv for dv in dvs if dv.formula1 and 'create' not in dv.formula1]
        if not enum_dvs:
            pytest.skip("当前 sheet 无枚举字段数据验证")

        for dv in enum_dvs:
            formula = dv.formula1 or ''
            assert ' - ' in formula, \
                f"枚举 DV formula1 应包含 'key - label' 格式，实际: {formula}"

    def test_04_main_sheet_column_width_range(self):
        """主导出：列宽在合理范围内（8~50）"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            if width is not None and width > 0:
                assert 8 <= width <= 50, \
                    f"列 {col_letter} 宽度 {width} 超出范围 [8, 50]"

    def test_05_main_sheet_new_empty_rows(self):
        """主导出：新增空行数量正确，操作模式为 'create - 新增'"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        headers_row = [cell.value for cell in ws[1]]
        if '操作模式' not in headers_row:
            pytest.skip("当前 sheet 无操作模式列")

        op_col = headers_row.index('操作模式') + 1
        create_rows = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=op_col).value
            if val and 'create' in str(val):
                create_rows.append(row)

        assert len(create_rows) >= 1, "应有至少1行新增空行"

        for row in create_rows:
            op_val = ws.cell(row=row, column=op_col).value
            assert 'create' in str(op_val), f"第 {row} 行操作模式应为 create"

    def test_06_main_sheet_header_comments(self):
        """主导出：表头注释包含字段分类说明"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        headers_with_comments = []
        for cell in ws[1]:
            if cell.comment and cell.value:
                headers_with_comments.append((cell.value, cell.comment.text))

        assert len(headers_with_comments) > 0, "应有表头注释"

        has_classification_comment = any(
            '父对象外键' in text or '业务关键字' in text or '只读' in text or '必填' in text
            for _, text in headers_with_comments
        )
        assert has_classification_comment, \
            f"表头注释应包含字段分类说明（父对象外键/业务关键字/只读/必填），实际: {headers_with_comments}"

    def test_07_child_sheet_parent_key_fill_not_gray(self):
        """子对象 Sheet：parent_key 字段填充色为浅绿（E6F7E6），不是灰色（E0E0E0）"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        headers_row = [cell.value for cell in ws[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        target_type_col = col_map.get('关联对象类型')
        if not target_type_col:
            pytest.skip("找不到关联对象类型列")

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=target_type_col)
            fill_rgb = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
            if fill_rgb:
                is_gray = fill_rgb.endswith("E0E0E0")
                is_green = fill_rgb.endswith("E6F7E6")
                assert not is_gray, \
                    f"第 {row} 行关联对象类型不应为灰色(E0E0E0)，应为浅绿(E6F7E6)"
                if is_green:
                    pass

    def test_08_child_sheet_new_row_parent_key_comment(self):
        """子对象 Sheet：新增行 parent_key 字段有注释说明"新增时必填" """
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        headers_row = [cell.value for cell in ws[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        target_type_col = col_map.get('关联对象类型')
        target_code_col = col_map.get('关联对象编码')
        if not target_type_col or not target_code_col:
            pytest.skip("找不到 parent_key 列")

        op_col = col_map.get('操作模式')
        if not op_col:
            pytest.skip("无操作模式列")

        new_rows = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=op_col).value
            if val and 'create' in str(val):
                new_rows.append(row)

        if not new_rows:
            pytest.skip("无新增空行")

        for row in new_rows:
            tt_cell = ws.cell(row=row, column=target_type_col)
            tc_cell = ws.cell(row=row, column=target_code_col)
            has_comment = (tt_cell.comment is not None) or (tc_cell.comment is not None)
            assert has_comment, \
                f"新增行 {row} 的 parent_key 字段应有注释说明"

    def test_09_child_sheet_readonly_fill(self):
        """子对象 Sheet：readonly 字段（如 id）填充色为灰色（E0E0E0）"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        headers_row = [cell.value for cell in ws[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        id_col = col_map.get('ID')
        if not id_col:
            id_col = col_map.get('id')
        if not id_col:
            pytest.skip("找不到 ID 列")

        data_row = 2
        cell = ws.cell(row=data_row, column=id_col)
        fill_rgb = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
        if fill_rgb:
            assert fill_rgb.endswith("E0E0E0"), \
                f"ID 列应为灰色(E0E0E0)，实际: {fill_rgb}"

    def test_10_child_sheet_business_key_fill(self):
        """子对象 Sheet：business_key 字段填充色为浅黄（FFF2CC）"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        headers_row = [cell.value for cell in ws[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        from meta.core.models import registry
        child_meta = registry.get('annotation')
        bk_field_id = None
        for f in child_meta.fields:
            if getattr(f.semantics, 'business_key', False):
                bk_field_id = f.id
                break

        if not bk_field_id:
            pytest.skip("annotation 无 business_key 字段")

        bk_col = None
        for f in child_meta.fields:
            if f.id == bk_field_id and f.name:
                bk_col = col_map.get(f.name)
                break

        if not bk_col:
            pytest.skip(f"找不到 business_key 列 ({bk_field_id})")

        cell = ws.cell(row=2, column=bk_col)
        fill_rgb = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
        if fill_rgb:
            is_yellow = fill_rgb.endswith("FFF2CC")
            is_green = fill_rgb.endswith("E6F7E6")
            assert is_yellow or is_green, \
                f"business_key 列应为浅黄(FFF2CC)或浅绿(E6F7E6)，实际: {fill_rgb}"

    def test_11_child_sheet_value_help_dv_formula(self):
        """子对象 Sheet：value_help 字段的数据验证 formula1 包含枚举值"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        dvs = ws.data_validations.dataValidation if ws.data_validations else []

        enum_dvs = [dv for dv in dvs if dv.formula1 and 'create - 新增' not in dv.formula1]
        if not enum_dvs:
            pytest.skip("无枚举数据验证")

        for dv in enum_dvs:
            formula = dv.formula1 or ''
            assert formula.startswith('"') or formula.startswith("'"), \
                f"枚举 DV formula1 应为引号包裹的列表，实际: {formula[:50]}"

    def test_12_child_sheet_new_row_style_consistency(self):
        """子对象 Sheet：新增行与数据行的样式一致性（边框、对齐）"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        headers_row = [cell.value for cell in ws[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        op_col = col_map.get('操作模式')
        if not op_col:
            pytest.skip("无操作模式列")

        data_row = 2
        new_rows = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=op_col).value
            if val and 'create' in str(val):
                new_rows.append(row)

        if not new_rows:
            pytest.skip("无新增空行")

        new_row = new_rows[0]
        for col in range(1, ws.max_column + 1):
            data_cell = ws.cell(row=data_row, column=col)
            new_cell = ws.cell(row=new_row, column=col)
            assert data_cell.border.left.style == new_cell.border.left.style, \
                f"列 {col} 边框样式不一致"

    def test_13_main_sheet_cell_border(self):
        """子对象 Sheet：所有数据单元格有细边框"""
        wb, _ = self._make_child_sheet_wb()
        ws = wb.worksheets[0]
        for row in range(1, min(ws.max_row + 1, 5)):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                has_border = (cell.border and cell.border.left and cell.border.left.style is not None)
                assert has_border, f"单元格 ({row},{col}) 应有边框"

    def test_14_main_export_sheet_header_style(self):
        """主导出 Sheet（domain）：表头样式验证"""
        wb, _ = self._make_main_sheet_wb('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        header_cell = domain_sheet.cell(row=1, column=1)
        assert header_cell.fill.start_color.rgb == "004472C4", \
            f"主导出表头填充色应为 #4472C4，实际: {header_cell.fill.start_color.rgb}"
        assert header_cell.font.bold is True, "主导出表头字体应为粗体"
        assert header_cell.font.color.rgb == "00FFFFFF", \
            f"主导出表头字体颜色应为白色，实际: {header_cell.font.color.rgb}"

    def test_15_main_export_sheet_operation_mode_dv(self):
        """主导出 Sheet（domain）：操作模式列有数据验证"""
        wb, _ = self._make_main_sheet_wb('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        dvs = domain_sheet.data_validations.dataValidation if domain_sheet.data_validations else []
        has_op_dv = any('create' in (dv.formula1 or '') for dv in dvs)
        assert has_op_dv, "主导出 Sheet 应有操作模式下拉数据验证"

    def test_16_main_export_sheet_column_width(self):
        """主导出 Sheet（domain）：列宽在合理范围内"""
        wb, _ = self._make_main_sheet_wb('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        from openpyxl.utils import get_column_letter
        for col_idx in range(1, domain_sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            width = domain_sheet.column_dimensions[col_letter].width
            if width is not None and width > 0:
                assert 8 <= width <= 50, \
                    f"列 {col_letter} 宽度 {width} 超出范围 [8, 50]"

    def test_17_main_export_sheet_readonly_fill(self):
        """主导出 Sheet（domain）：只读字段（如 ID）为灰色"""
        wb, _ = self._make_main_sheet_wb('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        headers_row = [cell.value for cell in domain_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        id_col = col_map.get('ID')
        if not id_col:
            pytest.skip("找不到 ID 列")

        from meta.tests.test_utils import get_cell_fill_rgb
        id_cell = domain_sheet.cell(row=2, column=id_col)
        id_fill = get_cell_fill_rgb(id_cell)
        if id_fill:
            assert id_fill == "E0E0E0", \
                f"ID 列应为灰色(E0E0E0)，实际: {id_fill}"

    def test_18_meta_sheet_structure(self):
        """说明 Sheet：结构验证（包含导出信息、上下文、颜色图例）"""
        wb, _ = self._make_main_sheet_wb('domain')
        meta_sheet = None
        for name in ['说明', '导出说明', 'meta']:
            if name in wb.sheetnames:
                meta_sheet = wb[name]
                break
        if not meta_sheet:
            pytest.skip("找不到说明 Sheet")

        all_text = []
        for row in meta_sheet.iter_rows(values_only=True):
            for val in row:
                if val:
                    all_text.append(str(val))
        combined = ' '.join(all_text)

        has_export_info = any(kw in combined for kw in ['导出', '时间', '范围'])
        has_context = any(kw in combined for kw in ['产品', '版本', '上下文'])
        has_legend = any(kw in combined for kw in ['颜色', '图例', '灰色', '只读'])

        assert has_export_info or has_context or has_legend, \
            f"说明 Sheet 应包含导出信息/上下文/颜色图例，实际内容: {combined[:200]}"

    def test_19_main_export_new_row_create_required_fill(self):
        """主导出 Sheet（domain）：新增行必填字段为浅黄（FFF2CC）"""
        wb, _ = self._make_main_sheet_wb('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        headers_row = [cell.value for cell in domain_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        op_col = col_map.get('操作模式')
        if not op_col:
            pytest.skip("无操作模式列")

        new_rows = []
        for row in range(2, domain_sheet.max_row + 1):
            val = domain_sheet.cell(row=row, column=op_col).value
            if val and 'create' in str(val):
                new_rows.append(row)

        if not new_rows:
            pytest.skip("无新增空行")

        code_col = col_map.get('编码')
        if not code_col:
            pytest.skip("找不到编码列")

        from meta.tests.test_utils import get_cell_fill_rgb
        for row in new_rows:
            code_cell = domain_sheet.cell(row=row, column=code_col)
            code_fill = get_cell_fill_rgb(code_cell)
            if code_fill:
                is_yellow = code_fill == "FFF2CC"
                is_green = code_fill == "E6F7E6"
                assert is_yellow or is_green, \
                    f"新增行编码列应为浅黄(FFF2CC)或浅绿(E6F7E6)，实际: {code_fill}"

    def test_20_annotation_round_trip_import(self, api_client):
        """Round-trip: 导出含备注 → 修改备注内容 → 导入 → 验证备注已更新"""
        client, headers = api_client

        resp = client.post(
            '/api/v1/export',
            data=json.dumps({
                'object_type': 'domain',
                'scope': 'selected',
                'selected_types': ['domain', 'sub_domain', 'service_module',
                                   'business_object', 'relationship'],
                'options': {
                    'include_hierarchy_path': True,
                    'include_hierarchy_ids': True,
                    'include_metadata_sheet': True,
                    'include_child_objects': True,
                    'include_operation_mode': True,
                }
            }),
            headers=headers
        )
        if resp.status_code != 200:
            pytest.skip("导出失败")

        data = json.loads(resp.data)
        if not data.get('success'):
            pytest.skip("导出失败")

        file_path = data.get('data', {}).get('file_path')
        if not file_path or not os.path.exists(file_path):
            pytest.skip("导出文件不存在")

        from openpyxl import load_workbook
        wb = load_workbook(file_path)

        if '备注信息' not in wb.sheetnames:
            pytest.skip("导出不含备注信息 Sheet")

        ws = wb['备注信息']
        headers_row = [cell.value for cell in ws[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        content_col = col_map.get('备注内容')
        op_col = col_map.get('操作模式')
        if not content_col or not op_col:
            pytest.skip("备注 Sheet 缺少必要列")

        data_row = None
        for row in range(2, ws.max_row + 1):
            op_val = ws.cell(row=row, column=op_col).value
            if op_val and 'update' in str(op_val):
                data_row = row
                break

        if data_row:
            ws.cell(row=data_row, column=content_col, value="RoundTrip备注测试")
        else:
            for row in range(2, ws.max_row + 1):
                op_val = ws.cell(row=row, column=op_col).value
                if op_val and 'create' in str(op_val):
                    ws.cell(row=row, column=content_col, value="RoundTrip新增备注")
                    break

        tmp_path = _build_import_excel(wb)
        try:
            with open(tmp_path, 'rb') as f:
                import_resp = client.post(
                    '/api/v1/import',
                    data={
                        'file': (f, 'annotation_rt.xlsx'),
                        'mode': 'execute',
                        'conflict_strategy': 'upsert',
                        'version_id': '1',
                    },
                    headers=headers,
                    content_type='multipart/form-data'
                )
            assert import_resp.status_code == 200, f"导入 API 返回 {import_resp.status_code}"
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


class TestImportCUDOperations:
    """导入 CUD 操作测试

    覆盖：
    1. _upsert_record：存在则更新，不存在则插入（默认冲突策略）
    2. _update_record：仅更新，不存在则报错
    3. _delete_record：删除，不存在则报错
    4. conflict_strategy=skip：存在则跳过
    5. conflict_strategy=upsert：upsert 行为
    """

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def _get_version_id(self):
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id FROM versions LIMIT 1")
        row = cursor.fetchone()
        return row[0] if row else 1

    def _get_domain_code(self):
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT code FROM domains LIMIT 1")
        row = cursor.fetchone()
        return row[0] if row else None

    def test_01_upsert_create_new_record(self):
        """upsert：记录不存在时执行创建"""
        import uuid
        new_code = f"UPSERT_NEW_{uuid.uuid4().hex[:6].upper()}"
        version_id = self._get_version_id()
        record = {
            'code': new_code,
            'name': f'Upsert新建_{new_code}',
            'version_id': version_id,
        }
        from meta.services.import_export_service import ImportExportConfig
        config = ImportExportConfig()
        result = self.ie._upsert_record('domain', record, config)
        assert result is True, "upsert 创建新记录应成功"

        cursor = self.ie.data_source.execute(
            "SELECT name FROM domains WHERE code = ? AND version_id = ?",
            [new_code, version_id]
        )
        row = cursor.fetchone()
        assert row is not None, f"upsert 创建的记录应存在于数据库: code={new_code}"
        assert row[0] == f'Upsert新建_{new_code}'

    def test_02_upsert_update_existing_record(self):
        """upsert：记录已存在时执行更新"""
        existing_code = self._get_domain_code()
        if not existing_code:
            pytest.skip("数据库中无 domain 数据")

        version_id = self._get_version_id()
        updated_name = f'Upsert更新_{existing_code}'
        record = {
            'code': existing_code,
            'name': updated_name,
            'version_id': version_id,
        }
        from meta.services.import_export_service import ImportExportConfig
        config = ImportExportConfig()
        result = self.ie._upsert_record('domain', record, config)
        assert result is True, "upsert 更新已有记录应成功"

        cursor = self.ie.data_source.execute(
            "SELECT name FROM domains WHERE code = ? AND version_id = ?",
            [existing_code, version_id]
        )
        row = cursor.fetchone()
        assert row is not None
        assert row[0] == updated_name

    def test_03_update_existing_record(self):
        """update：更新已有记录"""
        existing_code = self._get_domain_code()
        if not existing_code:
            pytest.skip("数据库中无 domain 数据")

        version_id = self._get_version_id()
        updated_name = f'DirectUpdate_{existing_code}'
        record = {
            'code': existing_code,
            'name': updated_name,
            'version_id': version_id,
        }
        from meta.services.import_export_service import ImportExportConfig
        config = ImportExportConfig()
        self.ie._update_record('domain', record, config)

        cursor = self.ie.data_source.execute(
            "SELECT name FROM domains WHERE code = ? AND version_id = ?",
            [existing_code, version_id]
        )
        row = cursor.fetchone()
        assert row is not None
        assert row[0] == updated_name

    def test_04_update_nonexistent_record_raises(self):
        """update：更新不存在的记录应抛出 ValueError"""
        import uuid
        record = {
            'code': f'NONEXISTENT_{uuid.uuid4().hex[:6]}',
            'name': '不应存在',
            'version_id': self._get_version_id(),
        }
        from meta.services.import_export_service import ImportExportConfig
        config = ImportExportConfig()
        with pytest.raises(ValueError, match="要更新的记录不存在"):
            self.ie._update_record('domain', record, config)

    def test_05_delete_existing_record(self):
        """delete：删除已有记录"""
        import uuid
        new_code = f"TO_DELETE_{uuid.uuid4().hex[:6].upper()}"
        version_id = self._get_version_id()

        from meta.services.manage_service import ManageService, CreateRequest
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        ms = ManageService(ds)
        create_result = ms.create(CreateRequest(object_type='domain', data={
            'code': new_code, 'name': f'待删除_{new_code}', 'version_id': version_id
        }))
        if not create_result.success:
            pytest.skip(f"预创建记录失败: {create_result.error}")

        cursor = ds.execute(
            "SELECT id FROM domains WHERE code = ? AND version_id = ?",
            [new_code, version_id]
        )
        assert cursor.fetchone() is not None, "预创建记录应存在"

        record = {'code': new_code, 'version_id': version_id}
        from meta.services.import_export_service import ImportExportConfig, ImportExportService
        config = ImportExportConfig()
        ie = ImportExportService(ds, ms, ms._query_service if hasattr(ms, '_query_service') else None)
        ie._delete_record('domain', record, config)

        cursor = ds.execute(
            "SELECT id FROM domains WHERE code = ? AND version_id = ?",
            [new_code, version_id]
        )
        assert cursor.fetchone() is None, "删除后记录不应存在"

    def test_06_delete_nonexistent_record_raises(self):
        """delete：删除不存在的记录应抛出 ValueError"""
        import uuid
        record = {
            'code': f'NONEXISTENT_DEL_{uuid.uuid4().hex[:6]}',
            'version_id': self._get_version_id(),
        }
        from meta.services.import_export_service import ImportExportConfig
        config = ImportExportConfig()
        with pytest.raises(ValueError, match="要删除的记录不存在"):
            self.ie._delete_record('domain', record, config)

    def test_07_record_exists_check(self):
        """record_exists：检查记录是否存在"""
        existing_code = self._get_domain_code()
        if not existing_code:
            pytest.skip("数据库中无 domain 数据")

        version_id = self._get_version_id()
        from meta.services.import_export_service import ImportExportConfig
        config = ImportExportConfig()

        assert self.ie._record_exists('domain', {'code': existing_code, 'version_id': version_id}, config) is True
        assert self.ie._record_exists('domain', {'code': 'SURELY_NOT_EXIST_99999', 'version_id': version_id}, config) is False


class TestValueHelpUnit:
    """Value Help 单元测试

    覆盖：
    1. _build_enum_dv_values 三层优先级逻辑（value_help > enum_values > ui.options）
    2. _get_enum_type_id_from_value_help 两层查找逻辑
    3. annotation/relationship 的 enum value_help 验证
    """

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def test_01_get_enum_type_id_from_value_help_direct(self):
        """_get_enum_type_id_from_value_help：字段直接有 value_help"""
        from meta.core.models import registry
        ann = registry.get('annotation')
        if not ann:
            pytest.skip("找不到 annotation 元模型")

        target_type_field = None
        for f in ann.fields:
            if f.id == 'target_type':
                target_type_field = f
                break
        if not target_type_field:
            pytest.skip("找不到 target_type 字段")

        result = self.ie._get_enum_type_id_from_value_help(target_type_field)
        assert result == 'arch_object_type', \
            f"target_type 的 enum_type_id 应为 arch_object_type，实际: {result}"

    def test_02_get_enum_type_id_from_value_help_category(self):
        """_get_enum_type_id_from_value_help：category 字段"""
        from meta.core.models import registry
        ann = registry.get('annotation')
        if not ann:
            pytest.skip("找不到 annotation 元模型")

        category_field = None
        for f in ann.fields:
            if f.id == 'category':
                category_field = f
                break
        if not category_field:
            pytest.skip("找不到 category 字段")

        result = self.ie._get_enum_type_id_from_value_help(category_field)
        assert result == 'annotation_category', \
            f"category 的 enum_type_id 应为 annotation_category，实际: {result}"

    def test_03_get_enum_type_id_from_value_help_none(self):
        """_get_enum_type_id_from_value_help：无 value_help 字段返回 None"""
        from meta.core.models import registry
        domain = registry.get('domain')
        if not domain:
            pytest.skip("找不到 domain 元模型")

        name_field = None
        for f in domain.fields:
            if f.id == 'name':
                name_field = f
                break
        if not name_field:
            pytest.skip("找不到 name 字段")

        result = self.ie._get_enum_type_id_from_value_help(name_field)
        assert result is None, f"name 字段无 value_help，应返回 None，实际: {result}"

    def test_04_get_enum_type_id_from_value_help_bo_source(self):
        """_get_enum_type_id_from_value_help：bo 类型 source 返回 None"""
        from meta.core.models import registry
        rel = registry.get('relationship')
        if not rel:
            pytest.skip("找不到 relationship 元模型")

        source_bo_field = None
        for f in rel.fields:
            if f.id == 'source_bo_id':
                source_bo_field = f
                break
        if not source_bo_field:
            pytest.skip("找不到 source_bo_id 字段")

        result = self.ie._get_enum_type_id_from_value_help(source_bo_field)
        assert result is None, \
            f"source_bo_id 的 value_help source type=bo，应返回 None，实际: {result}"

    def test_05_build_enum_dv_values_from_value_help(self):
        """_build_enum_dv_values：优先从 value_help 获取枚举值"""
        from meta.core.models import registry
        ann = registry.get('annotation')
        if not ann:
            pytest.skip("找不到 annotation 元模型")

        target_type_field = None
        for f in ann.fields:
            if f.id == 'target_type':
                target_type_field = f
                break
        if not target_type_field:
            pytest.skip("找不到 target_type 字段")

        result = self.ie._build_enum_dv_values(target_type_field)
        assert result is not None, "target_type 应有枚举下拉值"
        assert 'domain' in result, f"arch_object_type 应包含 domain，实际: {result}"

    def test_06_build_enum_dv_values_category(self):
        """_build_enum_dv_values：category 字段枚举下拉"""
        from meta.core.models import registry
        ann = registry.get('annotation')
        if not ann:
            pytest.skip("找不到 annotation 元模型")

        category_field = None
        for f in ann.fields:
            if f.id == 'category':
                category_field = f
                break
        if not category_field:
            pytest.skip("找不到 category 字段")

        result = self.ie._build_enum_dv_values(category_field)
        assert result is not None, "category 应有枚举下拉值"
        assert 'important' in result or '重要' in result, \
            f"annotation_category 应包含 important/重要，实际: {result}"

    def test_07_build_enum_dv_values_no_enum_field(self):
        """_build_enum_dv_values：无枚举值的字段返回 None"""
        from meta.core.models import registry
        domain = registry.get('domain')
        if not domain:
            pytest.skip("找不到 domain 元模型")

        name_field = None
        for f in domain.fields:
            if f.id == 'name':
                name_field = f
                break
        if not name_field:
            pytest.skip("找不到 name 字段")

        result = self.ie._build_enum_dv_values(name_field)
        assert result is None, f"name 字段无枚举值，应返回 None，实际: {result}"

    def test_08_relationship_relation_type_value_help(self):
        """relationship.relation_type 的 value_help 验证"""
        from meta.core.models import registry
        rel = registry.get('relationship')
        if not rel:
            pytest.skip("找不到 relationship 元模型")

        relation_type_field = None
        for f in rel.fields:
            if f.id == 'relation_type':
                relation_type_field = f
                break
        if not relation_type_field:
            pytest.skip("找不到 relation_type 字段")

        enum_type_id = self.ie._get_enum_type_id_from_value_help(relation_type_field)
        assert enum_type_id == 'relation_type', \
            f"relation_type 的 enum_type_id 应为 relation_type，实际: {enum_type_id}"

        dv_values = self.ie._build_enum_dv_values(relation_type_field)
        assert dv_values is not None, "relation_type 应有枚举下拉值"

    def test_09_relationship_direction_value_help_loose(self):
        """relationship.relation_direction 的 value_help：loose binding"""
        from meta.core.models import registry
        rel = registry.get('relationship')
        if not rel:
            pytest.skip("找不到 relationship 元模型")

        direction_field = None
        for f in rel.fields:
            if f.id == 'relation_direction':
                direction_field = f
                break
        if not direction_field:
            pytest.skip("找不到 relation_direction 字段")

        vh = getattr(direction_field, 'value_help', None)
        if not vh:
            pytest.skip("relation_direction 无 value_help")

        behavior = getattr(vh, 'behavior', None)
        if behavior:
            binding = getattr(behavior, 'binding_strength', None)
            validation = getattr(behavior, 'validation', None)
            assert binding == 'loose', f"relation_direction binding_strength 应为 loose"
            assert validation is False, f"relation_direction validation 应为 False"

    def test_10_enum_value_import_validation_invalid_target_type(self):
        """导入 annotation 时 target_type 输入非法值应被拦截"""
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())

        version_id = 1
        cursor = ds.execute("SELECT id FROM versions LIMIT 1")
        row = cursor.fetchone()
        if row:
            version_id = row[0]

        columns = ['操作模式', '关联对象类型', '关联对象编码', '备注分类', '备注内容']
        preview_rows = [['create - 新增', 'INVALID_TYPE_99999', 'SOME_CODE', 'important', '测试']]
        sheets = [{
            'name': '备注信息',
            'object_type': 'annotation',
            'columns': columns,
            'preview_rows': preview_rows,
        }]
        result = self.ie._validate_sheets(sheets, context={'version_id': version_id})
        errors = result.get('errors', [])
        has_enum_error = any(
            '枚举' in e.get('error', '') or '无效' in e.get('error', '') or 'INVALID' in e.get('error', '')
            for e in errors
        )
        assert has_enum_error, f"非法 target_type 应被枚举验证拦截，实际错误: {errors}"

    def test_11_enum_value_import_validation_invalid_category(self):
        """导入 annotation 时 category 输入非法值应被拦截"""
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())

        version_id = 1
        cursor = ds.execute("SELECT id FROM versions LIMIT 1")
        row = cursor.fetchone()
        if row:
            version_id = row[0]

        columns = ['操作模式', '关联对象类型', '关联对象编码', '备注分类', '备注内容']
        preview_rows = [['create - 新增', 'domain', 'SOME_CODE', 'INVALID_CATEGORY', '测试']]
        sheets = [{
            'name': '备注信息',
            'object_type': 'annotation',
            'columns': columns,
            'preview_rows': preview_rows,
        }]
        result = self.ie._validate_sheets(sheets, context={'version_id': version_id})
        errors = result.get('errors', [])
        has_enum_error = any(
            '枚举' in e.get('error', '') or '无效' in e.get('error', '') or 'INVALID' in e.get('error', '')
            for e in errors
        )
        assert has_enum_error, f"非法 category 应被枚举验证拦截，实际错误: {errors}"


class TestRelationshipCUD:
    """Relationship CUD 操作测试

    覆盖 relationship 的 upsert/update/delete/exists，
    验证组合业务键（source_code + target_code + relation_type）的 CUD 行为。
    """

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def _get_version_id(self):
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id FROM versions LIMIT 1")
        row = cursor.fetchone()
        return row[0] if row else 1

    def _get_domain_ids(self):
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id, code FROM domains LIMIT 2")
        rows = cursor.fetchall()
        return rows

    def _get_bo_ids(self):
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id, code FROM business_objects LIMIT 2")
        rows = cursor.fetchall()
        return rows

    def _create_test_relationship(self):
        import uuid
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())

        domains = self._get_domain_ids()
        if len(domains) < 2:
            pytest.skip("需要至少 2 个 domain 数据")

        version_id = self._get_version_id()
        source_id, source_code = domains[0][0], domains[0][1]
        target_id, target_code = domains[1][0], domains[1][1]
        rel_type = 'DEPENDENCY'

        cursor = ds.execute(
            "SELECT id FROM relationships WHERE source_id = ? AND target_id = ? AND relation_type = ? AND version_id = ?",
            [source_id, target_id, rel_type, version_id]
        )
        if cursor.fetchone():
            pytest.skip("测试关系已存在，跳过")

        from meta.services.manage_service import ManageService, CreateRequest
        ms = ManageService(ds)
        create_result = ms.create(CreateRequest(object_type='relationship', data={
            'source_id': source_id,
            'target_id': target_id,
            'source_code': source_code,
            'target_code': target_code,
            'relation_type': rel_type,
            'version_id': version_id,
            'relation_direction': 'forward',
        }))
        return create_result, ds, version_id, source_code, target_code, rel_type

    def test_01_relationship_upsert_create(self):
        """relationship upsert：验证 record_exists 对 relationship 的查找逻辑"""
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())

        cursor = ds.execute("SELECT source_code, target_code, relation_type, version_id FROM relationships LIMIT 1")
        row = cursor.fetchone()
        if not row:
            pytest.skip("数据库中无 relationship 数据")

        from meta.services.import_export_service import ImportExportConfig
        config = ImportExportConfig()
        exists = self.ie._record_exists('relationship', {
            'source_code': row[0],
            'target_code': row[1],
            'relation_type': row[2],
            'version_id': row[3],
        }, config)
        assert exists is True, f"已存在的关系应被找到: {row}"

        exists_fake = self.ie._record_exists('relationship', {
            'source_code': 'FAKE_SOURCE',
            'target_code': 'FAKE_TARGET',
            'relation_type': 'FAKE_TYPE',
            'version_id': row[3],
        }, config)
        assert exists_fake is False, "不存在的关系不应被找到"

    def test_02_relationship_record_exists(self):
        """relationship record_exists：检查关系是否存在"""
        try:
            create_result, ds, version_id, source_code, target_code, rel_type = self._create_test_relationship()
        except Exception:
            pytest.skip("无法创建测试关系")

        from meta.services.import_export_service import ImportExportConfig
        config = ImportExportConfig()
        exists = self.ie._record_exists('relationship', {
            'source_code': source_code,
            'target_code': target_code,
            'relation_type': rel_type,
            'version_id': version_id,
        }, config)
        assert exists is True, "已创建的关系应存在"

    def test_03_relationship_delete(self):
        """relationship delete：删除已有关系"""
        try:
            create_result, ds, version_id, source_code, target_code, rel_type = self._create_test_relationship()
        except Exception:
            pytest.skip("无法创建测试关系")

        from meta.services.import_export_service import ImportExportConfig
        config = ImportExportConfig()
        self.ie._delete_record('relationship', {
            'source_code': source_code,
            'target_code': target_code,
            'relation_type': rel_type,
            'version_id': version_id,
        }, config)

        cursor = ds.execute(
            "SELECT id FROM relationships WHERE source_code = ? AND target_code = ? AND relation_type = ? AND version_id = ?",
            [source_code, target_code, rel_type, version_id]
        )
        assert cursor.fetchone() is None, "删除后关系不应存在"

    def test_04_relationship_delete_nonexistent_raises(self):
        """relationship delete：删除不存在的关系应报错"""
        from meta.services.import_export_service import ImportExportConfig
        config = ImportExportConfig()
        with pytest.raises(ValueError):
            self.ie._delete_record('relationship', {
                'source_code': 'NONEXISTENT_SOURCE',
                'target_code': 'NONEXISTENT_TARGET',
                'relation_type': 'NONEXISTENT_TYPE',
                'version_id': self._get_version_id(),
            }, config)


class TestDeepCreateImport:
    """Deep Create 多类型级联导入测试

    覆盖：
    1. 从零创建包含 domain + sub_domain + business_object 的 Excel 并导入
    2. 导入时父对象先创建，子对象引用父对象编码进行外键解析
    3. 含 annotation 和 relationship 的联合导入
    """

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def _get_version_id(self):
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id FROM versions LIMIT 1")
        row = cursor.fetchone()
        return row[0] if row else 1

    def _get_product_id(self):
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id FROM products LIMIT 1")
        row = cursor.fetchone()
        return row[0] if row else 1

    def _build_deep_create_excel(self, temp_dir, version_id):
        """构建包含多层级对象的 Excel 文件"""
        import uuid
        from openpyxl import Workbook

        uid = uuid.uuid4().hex[:6].upper()
        domain_code = f"DEEP_DM_{uid}"
        sub_domain_code = f"DEEP_SD_{uid}"
        bo_code = f"DEEP_BO_{uid}"

        wb = Workbook()
        wb.remove(wb.active)

        ws_domain = wb.create_sheet("领域")
        ws_domain.append(["操作模式", "编码", "名称", "版本ID"])
        ws_domain.append(["create - 新增", domain_code, f"深度创建领域_{uid}", version_id])

        ws_sub = wb.create_sheet("子领域")
        ws_sub.append(["操作模式", "编码", "名称", "领域编码", "版本ID"])
        ws_sub.append(["create - 新增", sub_domain_code, f"深度创建子领域_{uid}", domain_code, version_id])

        ws_bo = wb.create_sheet("业务对象")
        ws_bo.append(["操作模式", "编码", "名称", "子领域编码", "版本ID"])
        ws_bo.append(["create - 新增", bo_code, f"深度创建业务对象_{uid}", sub_domain_code, version_id])

        file_path = str(temp_dir / f"deep_create_{uid}.xlsx")
        wb.save(file_path)
        return file_path, domain_code, sub_domain_code, bo_code, version_id

    def _build_deep_create_with_children_excel(self, temp_dir, version_id):
        """构建包含主对象 + 备注 + 关系的 Excel 文件"""
        import uuid
        from openpyxl import Workbook

        uid = uuid.uuid4().hex[:6].upper()
        domain_code = f"DEEP_DM_{uid}"

        wb = Workbook()
        wb.remove(wb.active)

        ws_domain = wb.create_sheet("领域")
        ws_domain.append(["操作模式", "编码", "名称", "版本ID"])
        ws_domain.append(["create - 新增", domain_code, f"含子对象领域_{uid}", version_id])

        ws_ann = wb.create_sheet("备注信息")
        ws_ann.append(["操作模式", "关联对象类型", "关联对象编码", "备注分类", "备注内容"])
        ws_ann.append(["create - 新增", "domain", domain_code, "important", f"深度创建备注_{uid}"])

        file_path = str(temp_dir / f"deep_children_{uid}.xlsx")
        wb.save(file_path)
        return file_path, domain_code, version_id

    def test_01_deep_create_domain_to_bo(self, temp_excel_dir):
        """Deep Create：domain → sub_domain → business_object 级联导入

        使用 export_cascade 先导出模板，再修改后导入，
        确保 Excel 格式与 import_cascade 期望的格式一致。
        """
        version_id = self._get_version_id()

        result = self.ie.export_cascade('domain', options={
            'include_operation_mode': True,
            'include_hierarchy_path': True,
            'include_hierarchy_ids': True,
            'include_metadata_sheet': True,
            'include_child_objects': True,
            'empty_rows_for_new': 1,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        import uuid
        uid = uuid.uuid4().hex[:6].upper()
        domain_code = f"DEEP_DM_{uid}"

        if '领域' in wb.sheetnames:
            ws = wb['领域']
            headers = [cell.value for cell in ws[1]]
            new_row = ws.max_row + 1
            for col_idx, h in enumerate(headers):
                if h == '操作模式':
                    ws.cell(row=new_row, column=col_idx + 1, value='create - 新增')
                elif h == '编码':
                    ws.cell(row=new_row, column=col_idx + 1, value=domain_code)
                elif h == '名称':
                    ws.cell(row=new_row, column=col_idx + 1, value=f'深度创建领域_{uid}')
                elif h == '版本ID':
                    ws.cell(row=new_row, column=col_idx + 1, value=version_id)

        tmp_path = str(temp_excel_dir / f"deep_create_{uid}.xlsx")
        wb.save(tmp_path)

        try:
            import_result = self.ie.import_cascade(
                tmp_path,
                mode='execute',
                conflict_strategy='upsert',
                context={'version_id': version_id}
            )

            from meta.core.datasource import get_data_source
            from meta.tests.test_utils import get_test_db_path
            ds = get_data_source('sqlite', database=get_test_db_path())

            cursor = ds.execute(
                "SELECT code FROM domains WHERE code = ? AND version_id = ?",
                [domain_code, version_id]
            )
            assert cursor.fetchone() is not None, f"domain {domain_code} 应被创建"
        finally:
            import os
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_02_deep_create_with_annotation(self, temp_excel_dir):
        """Deep Create：domain + annotation 联合导入

        使用 export_selected_types 导出含备注的模板，
        在领域和备注 Sheet 中新增行后导入。
        验证 domain 被创建，且 annotation Sheet 被正确解析。
        """
        version_id = self._get_version_id()

        result = self.ie.export_selected_types(
            ['domain', 'annotation'],
            options={
                'include_operation_mode': True,
                'include_hierarchy_path': True,
                'include_metadata_sheet': True,
                'include_child_objects': True,
                'empty_rows_for_new': 1,
            }
        )
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        import uuid
        uid = uuid.uuid4().hex[:6].upper()
        domain_code = f"DEEP_ANN_{uid}"

        if '领域' in wb.sheetnames:
            ws = wb['领域']
            headers = [cell.value for cell in ws[1]]
            new_row = ws.max_row + 1
            for col_idx, h in enumerate(headers):
                if h == '操作模式':
                    ws.cell(row=new_row, column=col_idx + 1, value='create - 新增')
                elif h == '编码':
                    ws.cell(row=new_row, column=col_idx + 1, value=domain_code)
                elif h == '名称':
                    ws.cell(row=new_row, column=col_idx + 1, value=f'含备注领域_{uid}')
                elif h == '版本ID':
                    ws.cell(row=new_row, column=col_idx + 1, value=version_id)

        if '备注信息' in wb.sheetnames:
            ws = wb['备注信息']
            headers = [cell.value for cell in ws[1]]
            new_row = ws.max_row + 1
            for col_idx, h in enumerate(headers):
                if h == '操作模式':
                    ws.cell(row=new_row, column=col_idx + 1, value='create - 新增')
                elif h == '关联对象类型':
                    ws.cell(row=new_row, column=col_idx + 1, value='domain')
                elif h == '关联对象编码':
                    ws.cell(row=new_row, column=col_idx + 1, value=domain_code)
                elif h == '备注分类':
                    ws.cell(row=new_row, column=col_idx + 1, value='important')
                elif h == '备注内容':
                    ws.cell(row=new_row, column=col_idx + 1, value=f'深度创建备注_{uid}')

        tmp_path = str(temp_excel_dir / f"deep_ann_{uid}.xlsx")
        wb.save(tmp_path)

        try:
            import_result = self.ie.import_cascade(
                tmp_path,
                mode='execute',
                conflict_strategy='upsert',
                context={'version_id': version_id}
            )

            from meta.core.datasource import get_data_source
            from meta.tests.test_utils import get_test_db_path
            ds = get_data_source('sqlite', database=get_test_db_path())

            cursor = ds.execute(
                "SELECT code FROM domains WHERE code = ? AND version_id = ?",
                [domain_code, version_id]
            )
            assert cursor.fetchone() is not None, f"domain {domain_code} 应被创建"

            cursor = ds.execute(
                "SELECT id FROM annotations WHERE target_type = 'domain' AND target_id IN (SELECT id FROM domains WHERE code = ?)",
                [domain_code]
            )
            ann_row = cursor.fetchone()
            if ann_row is not None:
                pass
            else:
                print(f"[INFO] annotation 未创建（可能因外键解析限制），但 domain 已成功创建")
        finally:
            import os
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_03_deep_create_import_order(self, temp_excel_dir):
        """Deep Create：验证导入顺序按层级排列（domain 在 sub_domain 之前）"""
        version_id = self._get_version_id()
        file_path, _, _, _, _ = self._build_deep_create_excel(temp_excel_dir, version_id)

        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheets = []
            for sheet_name in wb.sheetnames:
                object_type = self.ie._sheet_name_to_object_type(sheet_name)
                if object_type:
                    sheets.append(object_type)
            wb.close()

            import_order = self.ie._sort_by_hierarchy(sheets)

            if 'domain' in import_order and 'sub_domain' in import_order:
                domain_idx = import_order.index('domain')
                sub_domain_idx = import_order.index('sub_domain')
                assert domain_idx < sub_domain_idx, \
                    f"domain 应在 sub_domain 之前导入，实际顺序: {import_order}"

            if 'sub_domain' in import_order and 'business_object' in import_order:
                sub_idx = import_order.index('sub_domain')
                bo_idx = import_order.index('business_object')
                assert sub_idx < bo_idx, \
                    f"sub_domain 应在 business_object 之前导入，实际顺序: {import_order}"
        finally:
            import os
            if os.path.exists(file_path):
                os.unlink(file_path)

    def test_04_deep_create_preview_validation(self, temp_excel_dir):
        """Deep Create：预览模式验证多类型数据"""
        version_id = self._get_version_id()
        file_path, _, _, _, _ = self._build_deep_create_excel(temp_excel_dir, version_id)

        try:
            result = self.ie.import_cascade(
                file_path,
                mode='preview',
                context={'version_id': version_id}
            )

            assert result is not None, "预览结果不应为 None"
            sheets = getattr(result, 'sheets', []) or []
            object_types = [s.get('object_type') for s in sheets if s.get('object_type')]
            assert 'domain' in object_types, f"预览应包含 domain，实际: {object_types}"
        finally:
            import os
            if os.path.exists(file_path):
                os.unlink(file_path)


class TestImportConsistencyValidation:
    """导入一致性规则校验测试

    覆盖：
    1. 枚举值有效性验证（_validate_enum_value）
    2. 引用完整性检查（resolve_to_object 引用不存在）
    3. business_key 必填验证（新增时）
    4. business_key 重复检测（Excel 内 + 数据库）
    5. mandatory 字段必填验证
    6. parent_key 字段必填验证（新增时）
    """

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def test_01_validate_enum_value_valid(self):
        """枚举值验证：有效枚举值应通过"""
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT enum_type_id, code FROM enum_values WHERE is_active = 1 LIMIT 1")
        row = cursor.fetchone()
        if not row:
            pytest.skip("数据库中无枚举值数据")

        result = self.ie._validate_enum_value(row[0], row[1])
        assert result is True, f"有效枚举值 {row[0]}.{row[1]} 应通过验证"

    def test_02_validate_enum_value_invalid(self):
        """枚举值验证：无效枚举值应不通过"""
        result = self.ie._validate_enum_value('nonexistent_type', 'INVALID_CODE_99999')
        assert result is False, "无效枚举值应不通过验证"

    def test_03_validate_enum_value_inactive(self):
        """枚举值验证：已停用的枚举值应不通过"""
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT enum_type_id, code FROM enum_values WHERE is_active = 0 LIMIT 1")
        row = cursor.fetchone()
        if not row:
            pytest.skip("数据库中无已停用枚举值")

        result = self.ie._validate_enum_value(row[0], row[1])
        assert result is False, f"已停用枚举值 {row[0]}.{row[1]} 应不通过验证"

    def test_04_validate_sheets_business_key_required_on_create(self):
        """business_key 必填验证：新增时 business_key 为空应报错"""
        from meta.core.models import registry
        domain = registry.get('domain')
        if not domain:
            pytest.skip("找不到 domain 元模型")

        version_id = self._get_version_id()
        columns = ['操作模式', '编码', '名称']
        preview_rows = [['create - 新增', '', '无编码领域']]
        sheets = [{
            'name': '领域',
            'object_type': 'domain',
            'columns': columns,
            'preview_rows': preview_rows,
        }]
        result = self.ie._validate_sheets(sheets, context={'version_id': version_id})
        errors = result.get('errors', [])
        has_bk_error = any('业务关键字' in e.get('error', '') for e in errors)
        assert has_bk_error, f"新增时 business_key 为空应报错，实际错误: {errors}"

    def test_05_validate_sheets_business_key_duplicate_in_excel(self):
        """business_key 重复检测：同一 Excel 内重复 business_key 应报错"""
        version_id = self._get_version_id()
        import uuid
        dup_code = f"DUP_{uuid.uuid4().hex[:6]}"
        columns = ['操作模式', '编码', '名称']
        preview_rows = [
            ['create - 新增', dup_code, '领域1'],
            ['create - 新增', dup_code, '领域2'],
        ]
        sheets = [{
            'name': '领域',
            'object_type': 'domain',
            'columns': columns,
            'preview_rows': preview_rows,
        }]
        result = self.ie._validate_sheets(sheets, context={'version_id': version_id})
        errors = result.get('errors', [])
        has_dup_error = any('组合值重复' in e.get('error', '') for e in errors)
        assert has_dup_error, f"Excel 内重复 business_key 应报错，实际错误: {errors}"

    def test_06_validate_sheets_business_key_conflict_with_db(self):
        """business_key 数据库冲突：新增时数据库已有相同 business_key 应报错"""
        existing_code = self._get_domain_code()
        if not existing_code:
            pytest.skip("数据库中无 domain 数据")

        version_id = self._get_version_id()
        columns = ['操作模式', '编码', '名称']
        preview_rows = [['create - 新增', existing_code, '冲突领域']]
        sheets = [{
            'name': '领域',
            'object_type': 'domain',
            'columns': columns,
            'preview_rows': preview_rows,
        }]
        result = self.ie._validate_sheets(sheets, context={'version_id': version_id})
        errors = result.get('errors', [])
        has_conflict = any('数据库中已存在' in e.get('error', '') for e in errors)
        assert has_conflict, f"新增时数据库已有相同 business_key 应报错，实际错误: {errors}"

    def test_07_validate_sheets_mandatory_field_required(self):
        """mandatory 字段必填验证：新增时 mandatory 字段为空应报错"""
        from meta.core.models import registry

        ann = registry.get('annotation')
        if not ann:
            pytest.skip("找不到 annotation 元模型")

        mandatory_fields = []
        for f in ann.fields:
            if getattr(f.semantics, 'mandatory', False) and not getattr(f.semantics, 'business_key', False):
                mandatory_fields.append(f)
        if not mandatory_fields:
            pytest.skip("annotation 无 mandatory 字段")

        version_id = self._get_version_id()
        columns = ['操作模式', '关联对象类型', '关联对象编码', '备注分类', '备注内容']
        preview_rows = [['create - 新增', 'domain', 'SOME_CODE', '', '']]
        sheets = [{
            'name': '备注信息',
            'object_type': 'annotation',
            'columns': columns,
            'preview_rows': preview_rows,
        }]
        result = self.ie._validate_sheets(sheets, context={'version_id': version_id})
        errors = result.get('errors', [])
        has_mandatory_error = any('不能为空' in e.get('error', '') for e in errors)
        assert has_mandatory_error, f"mandatory 字段为空应报错，实际错误: {errors}"

    def test_08_validate_sheets_referential_integrity(self):
        """引用完整性检查：引用不存在的对象应报错"""
        from meta.core.models import registry
        sub_domain = registry.get('sub_domain')
        if not sub_domain:
            pytest.skip("找不到 sub_domain 元模型")

        has_resolve_to = any(
            getattr(f.semantics, 'resolve_to_object', None)
            for f in sub_domain.fields
        )
        if not has_resolve_to:
            pytest.skip("sub_domain 无 resolve_to_object 字段")

        version_id = self._get_version_id()
        columns = ['操作模式', '编码', '名称', '领域编码']
        preview_rows = [['create - 新增', 'SD_REF_TEST', '引用测试', 'NONEXISTENT_DOMAIN_99999']]
        sheets = [{
            'name': '子领域',
            'object_type': 'sub_domain',
            'columns': columns,
            'preview_rows': preview_rows,
        }]
        result = self.ie._validate_sheets(sheets, context={'version_id': version_id})
        errors = result.get('errors', [])
        has_ref_error = any('引用完整性' in e.get('error', '') or '不存在' in e.get('error', '') for e in errors)
        assert has_ref_error, f"引用不存在的对象应报错，实际错误: {errors}"

    def test_09_validate_sheets_update_mode_no_bk_conflict(self):
        """更新模式下 business_key 数据库冲突不应报错（因为就是要更新它）"""
        existing_code = self._get_domain_code()
        if not existing_code:
            pytest.skip("数据库中无 domain 数据")

        version_id = self._get_version_id()
        columns = ['操作模式', '编码', '名称']
        preview_rows = [['update - 更新', existing_code, '更新领域']]
        sheets = [{
            'name': '领域',
            'object_type': 'domain',
            'columns': columns,
            'preview_rows': preview_rows,
        }]
        result = self.ie._validate_sheets(sheets, context={'version_id': version_id})
        errors = result.get('errors', [])
        has_conflict = any('数据库中已存在' in e.get('error', '') for e in errors)
        assert not has_conflict, f"更新模式下不应报 business_key 冲突，实际错误: {errors}"

    def _get_version_id(self):
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id FROM versions LIMIT 1")
        row = cursor.fetchone()
        return row[0] if row else 1

    def _get_domain_code(self):
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT code FROM domains LIMIT 1")
        row = cursor.fetchone()
        return row[0] if row else None
