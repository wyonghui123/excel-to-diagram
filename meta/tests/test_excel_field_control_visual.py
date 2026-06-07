# -*- coding: utf-8 -*-
"""Excel 字段控制视觉指示测试模块

此文件包含 TestExcelFieldControlVisual 测试类，验证 mandatory、readonly、
操作权限等在 Excel 导出中的视觉反映。
"""

import pytest


class TestExcelFieldControlVisual:
    """Excel 字段控制视觉指示测试

    覆盖：
    1. Sheet 保护（protect_sheet）：locked/unlocked 单元格 + SheetProtection
    2. mandatory 字段表头注释包含【必填】
    3. readonly_always 业务字段灰色填充
    4. 操作模式列与字段填充色关系
    5. 子对象 Sheet 的 protect_sheet 缺失问题
    """

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def _make_main_sheet_wb_with_protection(self, object_type='domain'):
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        from meta.services.query_service import QueryService
        from meta.services.manage_service import ManageService
        ds = get_data_source('sqlite', database=get_test_db_path())
        qs = QueryService(ds)
        ms = ManageService(ds)
        from meta.services.import_export_service import ImportExportService
        ie = ImportExportService(ds, ms, qs)
        options = {
            'include_operation_mode': True,
            'include_hierarchy_path': True,
            'include_hierarchy_ids': True,
            'include_metadata_sheet': True,
            'include_child_objects': True,
            'empty_rows_for_new': 3,
            'protect_sheet': True,
        }
        result = ie.export_cascade(object_type, options=options)
        if not result or not result.file_path:
            pytest.skip(f"{object_type} 导出失败")
        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)
        return wb, object_type

    def test_01_sheet_protection_enabled(self):
        """Sheet 保护：protect_sheet=True 时工作表应受保护"""
        wb, _ = self._make_main_sheet_wb_with_protection('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        assert domain_sheet.protection.sheet is True, \
            "protect_sheet=True 时工作表应受保护"

    def test_02_sheet_protection_readonly_cell_locked(self):
        """Sheet 保护：只读字段单元格应为 locked"""
        wb, _ = self._make_main_sheet_wb_with_protection('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        headers_row = [cell.value for cell in domain_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        id_col = col_map.get('ID')
        if not id_col:
            pytest.skip("找不到 ID 列")

        id_cell = domain_sheet.cell(row=2, column=id_col)
        assert id_cell.protection.locked is True, \
            "ID 列（只读）应为 locked=True"

    def test_03_sheet_protection_editable_cell_unlocked(self):
        """Sheet 保护：可编辑字段单元格应为 unlocked"""
        wb, _ = self._make_main_sheet_wb_with_protection('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        headers_row = [cell.value for cell in domain_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        name_col = col_map.get('名称')
        if not name_col:
            pytest.skip("找不到名称列")

        name_cell = domain_sheet.cell(row=2, column=name_col)
        assert name_cell.protection.locked is False, \
            "名称列（可编辑）应为 locked=False"

    def test_04_sheet_protection_business_key_unlocked(self):
        """Sheet 保护：业务关键字列应为 unlocked（可编辑用于新增）"""
        wb, _ = self._make_main_sheet_wb_with_protection('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        headers_row = [cell.value for cell in domain_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        code_col = col_map.get('编码')
        if not code_col:
            pytest.skip("找不到编码列")

        code_cell = domain_sheet.cell(row=2, column=code_col)
        assert code_cell.protection.locked is False, \
            "编码列（业务关键字）应为 locked=False"

    def test_05_sheet_protection_operation_mode_unlocked(self):
        """Sheet 保护：操作模式列应为 unlocked"""
        wb, _ = self._make_main_sheet_wb_with_protection('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        headers_row = [cell.value for cell in domain_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        op_col = col_map.get('操作模式')
        if not op_col:
            pytest.skip("找不到操作模式列")

        op_cell = domain_sheet.cell(row=2, column=op_col)
        assert op_cell.protection.locked is False, \
            "操作模式列应为 locked=False"

    def test_06_sheet_protection_settings(self):
        """Sheet 保护：保护策略正确（允许格式化/排序/插入行，禁止删除行列/插入列）"""
        wb, _ = self._make_main_sheet_wb_with_protection('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        prot = domain_sheet.protection
        assert prot.formatCells is True, "应允许格式化单元格"
        assert prot.formatColumns is True, "应允许格式化列"
        assert prot.formatRows is True, "应允许格式化行"
        assert prot.insertRows is True, "应允许插入行"
        assert prot.sort is True, "应允许排序"
        assert prot.deleteColumns is False, "应禁止删除列"
        assert prot.deleteRows is False, "应禁止删除行"
        assert prot.insertColumns is False, "应禁止插入列"

    def test_07_no_protection_by_default(self):
        """默认不启用 Sheet 保护"""
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        from meta.services.query_service import QueryService
        from meta.services.manage_service import ManageService
        ds = get_data_source('sqlite', database=get_test_db_path())
        qs = QueryService(ds)
        ms = ManageService(ds)
        from meta.services.import_export_service import ImportExportService
        ie = ImportExportService(ds, ms, qs)
        result = ie.export_cascade('domain', options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        assert domain_sheet.protection.sheet is False, \
            "默认不启用 Sheet 保护"

    def test_08_mandatory_field_header_comment(self):
        """mandatory 字段表头注释包含【必填】"""
        from meta.core.models import registry
        ann = registry.get('annotation')
        if not ann:
            pytest.skip("找不到 annotation 元模型")

        mandatory_fields = []
        for f in ann.fields:
            if getattr(f.semantics, 'mandatory', False) and not getattr(f.semantics, 'business_key', False):
                mandatory_fields.append(f)
        if not mandatory_fields:
            pytest.skip("annotation 无 mandatory 字段")

        from meta.tests.test_utils import create_test_workbook
        wb, _ = create_test_workbook('annotation')
        ws = wb.worksheets[0]

        headers_row = [cell.value for cell in ws[1]]
        for mf in mandatory_fields:
            header_name = None
            if hasattr(mf, 'ui') and hasattr(mf.ui, 'label'):
                header_name = mf.ui.label
            if not header_name:
                for h in headers_row:
                    if h and mf.id in str(h).lower():
                        header_name = h
                        break
            if not header_name:
                continue

            col_idx = None
            for idx, h in enumerate(headers_row):
                if h == header_name:
                    col_idx = idx + 1
                    break
            if col_idx is None:
                continue

            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.comment:
                comment_text = header_cell.comment.text
                assert "必填" in comment_text, \
                    f"mandatory 字段 {mf.id} 的表头注释应包含'必填'，实际: {comment_text}"

    def test_09_readonly_always_business_field_fill(self):
        """readonly_always 业务字段（如 version_id）应为灰色"""
        from meta.core.models import registry
        domain = registry.get('domain')
        if not domain:
            pytest.skip("找不到 domain 元模型")

        readonly_always_fields = []
        for f in domain.fields:
            if getattr(f.semantics, 'readonly_always', False):
                readonly_always_fields.append(f)
        if not readonly_always_fields:
            pytest.skip("domain 无 readonly_always 字段")

        wb, _ = self._make_main_sheet_wb_with_protection('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        from meta.tests.test_utils import get_cell_fill_rgb
        headers_row = [cell.value for cell in domain_sheet[1]]
        for rf in readonly_always_fields:
            header_name = None
            if hasattr(rf, 'ui') and hasattr(rf.ui, 'label'):
                header_name = rf.ui.label
            if not header_name:
                continue

            col_idx = None
            for idx, h in enumerate(headers_row):
                if h == header_name:
                    col_idx = idx + 1
                    break
            if col_idx is None:
                continue

            cell = domain_sheet.cell(row=2, column=col_idx)
            fill = get_cell_fill_rgb(cell)
            assert fill == "E0E0E0", \
                f"readonly_always 字段 {rf.id} 应为灰色(E0E0E0)，实际: {fill}"

    def test_10_immutable_field_edit_mode_fill(self):
        """immutable 字段在编辑行应为灰色（只读）"""
        from meta.core.models import registry
        from meta.tests.test_utils import create_test_workbook, get_cell_fill_rgb

        ann = registry.get('annotation')
        if not ann:
            pytest.skip("找不到 annotation 元模型")

        immutable_fields = []
        for f in ann.fields:
            if getattr(f.semantics, 'immutable', False) and not getattr(f.semantics, 'parent_key', False):
                immutable_fields.append(f)
        if not immutable_fields:
            pytest.skip("annotation 无 immutable 非 parent_key 字段")

        wb, _ = create_test_workbook('annotation')
        ws = wb.worksheets[0]

        headers_row = [cell.value for cell in ws[1]]
        for imf in immutable_fields:
            header_name = None
            if hasattr(imf, 'ui') and hasattr(imf.ui, 'label'):
                header_name = imf.ui.label
            if not header_name:
                continue

            col_idx = None
            for idx, h in enumerate(headers_row):
                if h == header_name:
                    col_idx = idx + 1
                    break
            if col_idx is None:
                continue

            cell = ws.cell(row=2, column=col_idx)
            fill = get_cell_fill_rgb(cell)
            assert fill == "E0E0E0", \
                f"immutable 字段 {imf.id} 在编辑行应为灰色(E0E0E0)，实际: {fill}"

    def test_11_operation_mode_create_row_required_fill(self):
        """操作模式为 create 的行：create_required 字段为浅黄"""
        wb, _ = self._make_main_sheet_wb_with_protection('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        headers_row = [cell.value for cell in domain_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        op_col = col_map.get('操作模式')
        code_col = col_map.get('编码')
        if not op_col or not code_col:
            pytest.skip("缺少操作模式或编码列")

        from meta.tests.test_utils import get_cell_fill_rgb
        for row in range(2, domain_sheet.max_row + 1):
            op_val = domain_sheet.cell(row=row, column=op_col).value
            if op_val and 'create' in str(op_val):
                code_cell = domain_sheet.cell(row=row, column=code_col)
                code_fill = get_cell_fill_rgb(code_cell)
                assert code_fill == "FFF2CC", \
                    f"create 行编码列应为浅黄(FFF2CC)，实际: {code_fill}"
                break

    def test_12_operation_mode_update_row_immutable_readonly(self):
        """操作模式为 update 的行：immutable 字段为灰色"""
        wb, _ = self._make_main_sheet_wb_with_protection('domain')
        domain_sheet = None
        for name in ['领域', 'domain']:
            if name in wb.sheetnames:
                domain_sheet = wb[name]
                break
        if not domain_sheet:
            pytest.skip("找不到领域 Sheet")

        headers_row = [cell.value for cell in domain_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        op_col = col_map.get('操作模式')
        id_col = col_map.get('ID')
        if not op_col or not id_col:
            pytest.skip("缺少操作模式或 ID 列")

        from meta.tests.test_utils import get_cell_fill_rgb
        for row in range(2, domain_sheet.max_row + 1):
            op_val = domain_sheet.cell(row=row, column=op_col).value
            if op_val and 'update' in str(op_val):
                id_cell = domain_sheet.cell(row=row, column=id_col)
                id_fill = get_cell_fill_rgb(id_cell)
                assert id_fill == "E0E0E0", \
                    f"update 行 ID 列应为灰色(E0E0E0)，实际: {id_fill}"
                break

    def test_13_child_sheet_no_protection(self):
        """子对象 Sheet（annotation）当前不支持 protect_sheet（已知限制）"""
        from meta.tests.test_utils import create_test_workbook
        wb, _ = create_test_workbook('annotation')
        ws = wb.worksheets[0]

        assert ws.protection.sheet is False, \
            "子对象 Sheet 当前不支持 protect_sheet（_write_child_sheet 未实现）"


class TestProductVersionUserRoleReadonlyMandatory:
    """product/version/user/role/user_group 的 readonly/mandatory 测试"""

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def _export_and_get_sheet(self, object_type, sheet_names):
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        from meta.services.query_service import QueryService
        from meta.services.manage_service import ManageService
        ds = get_data_source('sqlite', database=get_test_db_path())
        qs = QueryService(ds)
        ms = ManageService(ds)
        from meta.services.import_export_service import ImportExportService
        ie = ImportExportService(ds, ms, qs)
        result = ie.export_cascade(object_type, options={
            'include_operation_mode': True,
            'protect_sheet': True,
        })
        if not result or not result.file_path:
            return None
        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)
        for name in sheet_names:
            if name in wb.sheetnames:
                return wb[name]
        return None

    def test_01_product_code_immutable(self):
        """product.code: immutable 字段验证"""
        from meta.core.models import registry
        product = registry.get('product')
        if not product:
            pytest.skip("找不到 product 元模型")

        code_field = None
        for f in product.fields:
            if f.id == 'code':
                code_field = f
                break
        if not code_field:
            pytest.skip("找不到 product.code 字段")

        immutable = getattr(code_field.semantics, 'immutable', False)
        assert immutable is True, "product.code 应为 immutable"

    def test_02_version_code_immutable(self):
        """version.code: immutable 字段验证"""
        from meta.core.models import registry
        version = registry.get('version')
        if not version:
            pytest.skip("找不到 version 元模型")

        code_field = None
        for f in version.fields:
            if f.id == 'code':
                code_field = f
                break
        if not code_field:
            pytest.skip("找不到 version.code 字段")

        immutable = getattr(code_field.semantics, 'immutable', False)
        assert immutable is True, "version.code 应为 immutable"

    def test_03_role_code_immutable(self):
        """role.code: immutable 字段验证"""
        from meta.core.models import registry
        role = registry.get('role')
        if not role:
            pytest.skip("找不到 role 元模型")

        code_field = None
        for f in role.fields:
            if f.id == 'code':
                code_field = f
                break
        if not code_field:
            pytest.skip("找不到 role.code 字段")

        immutable = getattr(code_field.semantics, 'immutable', False)
        assert immutable is True, "role.code 应为 immutable"

    def test_04_role_is_system_readonly_always(self):
        """role.is_system: readonly_always 字段验证"""
        from meta.core.models import registry
        role = registry.get('role')
        if not role:
            pytest.skip("找不到 role 元模型")

        is_system_field = None
        for f in role.fields:
            if f.id == 'is_system':
                is_system_field = f
                break
        if not is_system_field:
            pytest.skip("找不到 role.is_system 字段")

        readonly_always = getattr(is_system_field.semantics, 'readonly_always', False)
        assert readonly_always is True, "role.is_system 应为 readonly_always"

    def test_05_user_group_code_immutable(self):
        """user_group.code: immutable 字段验证"""
        from meta.core.models import registry
        user_group = registry.get('user_group')
        if not user_group:
            pytest.skip("找不到 user_group 元模型")

        code_field = None
        for f in user_group.fields:
            if f.id == 'code':
                code_field = f
                break
        if not code_field:
            pytest.skip("找不到 user_group.code 字段")

        immutable = getattr(code_field.semantics, 'immutable', False)
        assert immutable is True, "user_group.code 应为 immutable"

    def test_06_product_export_code_locked_in_update(self):
        """导出产品：code 列在 update 行应为灰色（immutable）"""
        ws = self._export_and_get_sheet('product', ['产品', 'product'])
        if not ws:
            pytest.skip("找不到产品 Sheet")

        headers_row = [cell.value for cell in ws[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        op_col = col_map.get('操作模式')
        code_col = col_map.get('编码') or col_map.get('产品编码')
        if not op_col or not code_col:
            pytest.skip("缺少操作模式或编码列")

        from meta.tests.test_utils import get_cell_fill_rgb
        for row in range(2, ws.max_row + 1):
            op_val = ws.cell(row=row, column=op_col).value
            if op_val and 'update' in str(op_val):
                code_cell = ws.cell(row=row, column=code_col)
                code_fill = get_cell_fill_rgb(code_cell)
                assert code_fill == "E0E0E0", \
                    f"update 行编码列应为灰色(E0E0E0)，实际: {code_fill}"
                break

    def test_07_role_export_is_system_readonly(self):
        """导出角色：is_system 列应为灰色（readonly_always）"""
        from meta.core.datasource import get_data_source
        from meta.tests.test_utils import get_test_db_path
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id FROM roles WHERE is_system = 1 LIMIT 1")
        if not cursor.fetchone():
            pytest.skip("数据库中无系统角色")

        ws = self._export_and_get_sheet('role', ['角色', 'role'])
        if not ws:
            pytest.skip("找不到角色 Sheet")

        headers_row = [cell.value for cell in ws[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        is_system_col = col_map.get('系统角色') or col_map.get('是否系统')
        if not is_system_col:
            pytest.skip("找不到系统角色列")

        from meta.tests.test_utils import get_cell_fill_rgb
        for row in range(2, ws.max_row + 1):
            is_system_cell = ws.cell(row=row, column=is_system_col)
            if is_system_cell.value == 1 or is_system_cell.value == True:
                fill = get_cell_fill_rgb(is_system_cell)
                assert fill == "E0E0E0", \
                    f"系统角色列应为灰色(E0E0E0)，实际: {fill}"
                break
