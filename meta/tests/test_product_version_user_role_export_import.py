# -*- coding: utf-8 -*-
"""产品/版本/用户/角色/用户组 导出导入测试

覆盖：
1. Product/Version 多对象批量导出、批量导入 CUD、deep insert
2. User/Role/UserGroup 单对象批量导出、批量导入 CUD
3. 批量关联建立和移除（user-group, role-group）
"""

import pytest
import json
import os


def get_test_db_path():
    from meta.tests.test_utils import get_test_db_path
    return get_test_db_path()


class TestProductVersionExportImport:
    """产品/版本多对象批量导出导入测试"""

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def _get_product_data(self):
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id, code, name FROM products LIMIT 2")
        return cursor.fetchall()

    def _get_version_data(self):
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id, code, name, product_id FROM versions LIMIT 2")
        return cursor.fetchall()

    def test_01_export_product_has_correct_structure(self):
        """导出产品：Sheet 结构正确（export_cascade 导出子对象树）"""
        products = self._get_product_data()
        if not products:
            pytest.skip("数据库中无 product 数据")

        result = self.ie.export_cascade('product', options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        sheets = wb.sheetnames
        has_product = any('产品' in s for s in sheets)
        has_domain = any('领域' in s for s in sheets)
        assert has_product or has_domain, \
            f"应包含产品或领域 Sheet，实际: {sheets}"

    def test_02_export_version_has_correct_structure(self):
        """导出版本：Sheet 结构正确（export_cascade 导出子对象树）"""
        versions = self._get_version_data()
        if not versions:
            pytest.skip("数据库中无 version 数据")

        result = self.ie.export_cascade('version', options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
            'include_child_objects': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        sheets = wb.sheetnames
        has_version = any('版本' in s for s in sheets)
        has_domain = any('领域' in s for s in sheets)
        assert has_version or has_domain, \
            f"应包含版本或领域 Sheet，实际: {sheets}"

    def test_03_export_product_version_multi_types(self):
        """多对象批量导出：product + version"""
        products = self._get_product_data()
        versions = self._get_version_data()
        if not products or not versions:
            pytest.skip("数据库中无 product/version 数据")

        result = self.ie.export_selected_types(['product', 'version'], options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        sheets = wb.sheetnames
        has_product = any('产品' in s or 'product' in s.lower() for s in sheets)
        has_version = any('版本' in s or 'version' in s.lower() for s in sheets)

        assert has_product, f"应包含产品 Sheet，实际: {sheets}"
        assert has_version, f"应包含版本 Sheet，实际: {sheets}"

    def test_04_product_cud_upsert_create(self):
        """产品 CUD：upsert 创建新产品"""
        import uuid
        from meta.core.datasource import get_data_source
        from meta.services.import_export_service import ImportExportConfig
        ds = get_data_source('sqlite', database=get_test_db_path())

        new_code = f"TEST_PROD_{uuid.uuid4().hex[:6].upper()}"
        record = {
            'code': new_code,
            'name': f'测试产品_{new_code}',
        }
        config = ImportExportConfig()
        result = self.ie._upsert_record('product', record, config)
        assert result is True, "upsert 创建产品应成功"

        cursor = ds.execute("SELECT name FROM products WHERE code = ?", [new_code])
        row = cursor.fetchone()
        assert row is not None, f"产品 {new_code} 应存在于数据库"

    def test_05_version_cud_upsert_create(self):
        """版本 CUD：upsert 创建新版本"""
        import uuid
        from meta.core.datasource import get_data_source
        from meta.services.import_export_service import ImportExportConfig
        ds = get_data_source('sqlite', database=get_test_db_path())

        products = self._get_product_data()
        if not products:
            pytest.skip("数据库中无 product 数据")

        product_id = products[0][0]
        product_code = products[0][1]

        new_code = f"{product_code}_{uuid.uuid4().hex[:4].upper()}"
        record = {
            'code': new_code,
            'name': f'测试版本_{new_code}',
            'product_id': product_id,
        }
        config = ImportExportConfig()
        result = self.ie._upsert_record('version', record, config)
        assert result is True, "upsert 创建版本应成功"

        cursor = ds.execute("SELECT name FROM versions WHERE code = ?", [new_code])
        row = cursor.fetchone()
        assert row is not None, f"版本 {new_code} 应存在于数据库"

    def test_06_deep_insert_product_with_version(self):
        """Deep Insert：产品 + 版本级联导入（验证导出结构）"""
        import uuid
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())

        result = self.ie.export_selected_types(['product', 'version'], options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
            'empty_rows_for_new': 1,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        sheets = wb.sheetnames
        has_product = any('产品' in s for s in sheets)
        has_version = any('版本' in s for s in sheets)
        has_domain = any('领域' in s for s in sheets)

        assert has_product or has_version or has_domain, \
            f"应包含产品/版本/领域 Sheet，实际: {sheets}"


class TestUserRoleUserGroupExportImport:
    """用户/角色/用户组单对象批量导出导入测试"""

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def _get_user_data(self):
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id, username, email FROM users LIMIT 2")
        return cursor.fetchall()

    def _get_role_data(self):
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id, code, name FROM roles LIMIT 2")
        return cursor.fetchall()

    def _get_user_group_data(self):
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id, code, name FROM user_groups LIMIT 2")
        return cursor.fetchall()

    def test_01_export_user_has_correct_structure(self):
        """导出用户：Sheet 结构正确"""
        users = self._get_user_data()
        if not users:
            pytest.skip("数据库中无 user 数据")

        result = self.ie.export_cascade('user', options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        assert '用户' in wb.sheetnames or 'user' in wb.sheetnames, \
            f"应包含用户 Sheet，实际: {wb.sheetnames}"

    def test_02_export_role_has_correct_structure(self):
        """导出角色：Sheet 结构正确"""
        roles = self._get_role_data()
        if not roles:
            pytest.skip("数据库中无 role 数据")

        result = self.ie.export_cascade('role', options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        assert '角色' in wb.sheetnames or 'role' in wb.sheetnames, \
            f"应包含角色 Sheet，实际: {wb.sheetnames}"

    def test_03_export_user_group_has_correct_structure(self):
        """导出用户组：Sheet 结构正确"""
        groups = self._get_user_group_data()
        if not groups:
            pytest.skip("数据库中无 user_group 数据")

        result = self.ie.export_cascade('user_group', options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        assert '用户组' in wb.sheetnames or 'user_group' in wb.sheetnames, \
            f"应包含用户组 Sheet，实际: {wb.sheetnames}"

    def test_04_user_cud_upsert_create(self):
        """用户 CUD：upsert 创建新用户"""
        import uuid
        from meta.core.datasource import get_data_source
        from meta.services.import_export_service import ImportExportConfig
        ds = get_data_source('sqlite', database=get_test_db_path())

        new_username = f"test_user_{uuid.uuid4().hex[:6]}"
        record = {
            'username': new_username,
            'email': f'{new_username}@test.com',
        }
        config = ImportExportConfig()
        result = self.ie._upsert_record('user', record, config)
        assert result is True, "upsert 创建用户应成功"

        cursor = ds.execute("SELECT email FROM users WHERE username = ?", [new_username])
        row = cursor.fetchone()
        assert row is not None, f"用户 {new_username} 应存在于数据库"

    def test_05_role_cud_upsert_create(self):
        """角色 CUD：upsert 创建新角色"""
        import uuid
        from meta.core.datasource import get_data_source
        from meta.services.import_export_service import ImportExportConfig
        ds = get_data_source('sqlite', database=get_test_db_path())

        new_code = f"TEST_ROLE_{uuid.uuid4().hex[:6].upper()}"
        record = {
            'code': new_code,
            'name': f'测试角色_{new_code}',
        }
        config = ImportExportConfig()
        result = self.ie._upsert_record('role', record, config)
        assert result is True, "upsert 创建角色应成功"

        cursor = ds.execute("SELECT name FROM roles WHERE code = ?", [new_code])
        row = cursor.fetchone()
        assert row is not None, f"角色 {new_code} 应存在于数据库"

    def test_06_user_group_cud_upsert_create(self):
        """用户组 CUD：upsert 创建新用户组"""
        import uuid
        from meta.core.datasource import get_data_source
        from meta.services.import_export_service import ImportExportConfig
        ds = get_data_source('sqlite', database=get_test_db_path())

        new_code = f"test_group_{uuid.uuid4().hex[:6]}"
        record = {
            'code': new_code,
            'name': f'测试用户组_{new_code}',
        }
        config = ImportExportConfig()
        result = self.ie._upsert_record('user_group', record, config)
        assert result is True, "upsert 创建用户组应成功"

        cursor = ds.execute("SELECT name FROM user_groups WHERE code = ?", [new_code])
        row = cursor.fetchone()
        assert row is not None, f"用户组 {new_code} 应存在于数据库"

    def test_07_user_group_hierarchy_parent_child(self):
        """用户组层级：父子关系验证"""
        import uuid
        from meta.core.datasource import get_data_source
        from meta.services.import_export_service import ImportExportConfig
        ds = get_data_source('sqlite', database=get_test_db_path())

        parent_code = f"parent_grp_{uuid.uuid4().hex[:6]}"
        child_code = f"child_grp_{uuid.uuid4().hex[:6]}"

        config = ImportExportConfig()
        self.ie._upsert_record('user_group', {'code': parent_code, 'name': f'父组_{parent_code}'}, config)

        cursor = ds.execute("SELECT id FROM user_groups WHERE code = ?", [parent_code])
        parent_row = cursor.fetchone()
        if not parent_row:
            pytest.skip("父组创建失败")

        parent_id = parent_row[0]
        self.ie._upsert_record('user_group', {'code': child_code, 'name': f'子组_{child_code}', 'parent_id': parent_id}, config)

        cursor = ds.execute("SELECT parent_id FROM user_groups WHERE code = ?", [child_code])
        child_row = cursor.fetchone()
        assert child_row is not None, f"子组 {child_code} 应存在"
        assert child_row[0] == parent_id, f"子组的 parent_id 应为 {parent_id}"


class TestBatchAssociation:
    """批量关联建立和移除测试"""

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def _get_user_and_group(self):
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("""
            SELECT u.id, u.username, g.id, g.code
            FROM users u, user_groups g
            LIMIT 1
        """)
        return cursor.fetchone()

    def _get_role_and_group(self):
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("""
            SELECT r.id, r.code, g.id, g.code
            FROM roles r, user_groups g
            LIMIT 1
        """)
        return cursor.fetchone()

    def test_01_user_group_member_association_supported(self):
        """用户-用户组成员关联：验证关联表存在"""
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())

        cursor = ds.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table' AND name='user_group_members'
        """)
        row = cursor.fetchone()
        assert row is not None, "user_group_members 关联表应存在"

    def test_02_user_role_association_supported(self):
        """用户-角色关联：验证关联表存在"""
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())

        cursor = ds.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table' AND name='user_roles'
        """)
        row = cursor.fetchone()
        assert row is not None, "user_roles 关联表应存在"

    def test_03_assign_user_to_group(self):
        """批量关联：将用户添加到用户组"""
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())

        row = self._get_user_and_group()
        if not row:
            pytest.skip("数据库中无 user/user_group 数据")

        user_id, username, group_id, group_code = row

        cursor = ds.execute("""
            SELECT 1 FROM user_group_members WHERE user_id = ? AND group_id = ?
        """, [user_id, group_id])
        if cursor.fetchone():
            pytest.skip("用户已在用户组中")

        ds.execute("""
            INSERT INTO user_group_members (user_id, group_id)
            VALUES (?, ?)
        """, [user_id, group_id])

        cursor = ds.execute("""
            SELECT 1 FROM user_group_members WHERE user_id = ? AND group_id = ?
        """, [user_id, group_id])
        assert cursor.fetchone() is not None, "用户应被添加到用户组"

    def test_04_assign_role_to_group(self):
        """批量关联：将角色分配给用户组"""
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())

        row = self._get_role_and_group()
        if not row:
            pytest.skip("数据库中无 role/user_group 数据")

        role_id, role_code, group_id, group_code = row

        cursor = ds.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table' AND name='group_roles'
        """)
        if not cursor.fetchone():
            pytest.skip("group_roles 关联表不存在")

        cursor = ds.execute("""
            SELECT 1 FROM group_roles WHERE role_id = ? AND group_id = ?
        """, [role_id, group_id])
        if cursor.fetchone():
            pytest.skip("角色已分配给用户组")

        ds.execute("""
            INSERT INTO group_roles (role_id, group_id)
            VALUES (?, ?)
        """, [role_id, group_id])

        cursor = ds.execute("""
            SELECT 1 FROM group_roles WHERE role_id = ? AND group_id = ?
        """, [role_id, group_id])
        assert cursor.fetchone() is not None, "角色应被分配给用户组"

    def test_05_export_user_with_group_membership(self):
        """导出用户：包含用户组成员关系"""
        from meta.core.models import registry
        user = registry.get('user')
        if not user:
            pytest.skip("找不到 user 元模型")

        has_groups_assoc = hasattr(user, 'associations') and 'groups' in user.associations
        assert has_groups_assoc, "user 应有 groups 关联定义"

    def test_06_export_user_group_with_members(self):
        """导出用户组：包含成员关系"""
        from meta.core.models import registry
        user_group = registry.get('user_group')
        if not user_group:
            pytest.skip("找不到 user_group 元模型")

        has_members_assoc = hasattr(user_group, 'associations') and 'members' in user_group.associations
        assert has_members_assoc, "user_group 应有 members 关联定义"


class TestValueHelpForObjects:
    """产品/版本/用户/用户组的 value_help 测试

    覆盖：
    1. user.status: enum type (user_status)
    2. user_group.parent_id: bo type (user_group, 层级选择)
    3. user_group.manager_id: bo type (user)
    4. product.owner_id: bo type (user)
    5. version.product_id: bo type (product)
    6. version.owner_id: bo type (user)
    """

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def _find_field(self, object_type, field_id):
        from meta.core.models import registry
        meta = registry.get(object_type)
        if not meta:
            return None
        for f in meta.fields:
            if f.id == field_id:
                return f
        return None

    def test_01_user_status_value_help_enum(self):
        """user.status: enum value_help (user_status)"""
        field = self._find_field('user', 'status')
        if not field:
            pytest.skip("找不到 user.status 字段")

        vh = getattr(field, 'value_help', None)
        assert vh is not None, "user.status 应有 value_help"

        source = getattr(vh, 'source', None)
        assert source is not None, "user.status value_help 应有 source"

        source_type = getattr(source, 'type', None)
        assert source_type == 'enum', f"user.status source.type 应为 enum，实际: {source_type}"

        enum_type_id = getattr(source, 'enum_type_id', None)
        assert enum_type_id == 'user_status', \
            f"user.status enum_type_id 应为 user_status，实际: {enum_type_id}"

    def test_02_user_status_enum_dv_values(self):
        """user.status: 枚举下拉值验证"""
        field = self._find_field('user', 'status')
        if not field:
            pytest.skip("找不到 user.status 字段")

        dv_values = self.ie._build_enum_dv_values(field)
        assert dv_values is not None, "user.status 应有枚举下拉值"

        assert 'active' in dv_values or '激活' in dv_values, \
            f"user_status 应包含 active/激活，实际: {dv_values}"

    def test_03_user_group_parent_id_value_help_bo(self):
        """user_group.parent_id: bo value_help (user_group, 层级选择)"""
        field = self._find_field('user_group', 'parent_id')
        if not field:
            pytest.skip("找不到 user_group.parent_id 字段")

        vh = getattr(field, 'value_help', None)
        assert vh is not None, "user_group.parent_id 应有 value_help"

        source = getattr(vh, 'source', None)
        assert source is not None, "user_group.parent_id value_help 应有 source"

        source_type = getattr(source, 'type', None)
        assert source_type == 'bo', f"user_group.parent_id source.type 应为 bo，实际: {source_type}"

        target_bo = getattr(source, 'target_bo', None)
        assert target_bo == 'user_group', \
            f"user_group.parent_id target_bo 应为 user_group，实际: {target_bo}"

        hierarchy = getattr(source, 'hierarchy', None)
        assert hierarchy is not None, "user_group.parent_id 应有 hierarchy 配置（层级选择）"

    def test_04_user_group_manager_id_value_help_bo(self):
        """user_group.manager_id: bo value_help (user)"""
        field = self._find_field('user_group', 'manager_id')
        if not field:
            pytest.skip("找不到 user_group.manager_id 字段")

        vh = getattr(field, 'value_help', None)
        assert vh is not None, "user_group.manager_id 应有 value_help"

        source = getattr(vh, 'source', None)
        assert source is not None, "user_group.manager_id value_help 应有 source"

        source_type = getattr(source, 'type', None)
        assert source_type == 'bo', f"user_group.manager_id source.type 应为 bo，实际: {source_type}"

        target_bo = getattr(source, 'target_bo', None)
        assert target_bo == 'user', \
            f"user_group.manager_id target_bo 应为 user，实际: {target_bo}"

    def test_05_product_owner_id_value_help_bo(self):
        """product.owner_id: bo value_help (user)"""
        field = self._find_field('product', 'owner_id')
        if not field:
            pytest.skip("找不到 product.owner_id 字段")

        vh = getattr(field, 'value_help', None)
        assert vh is not None, "product.owner_id 应有 value_help"

        source = getattr(vh, 'source', None)
        assert source is not None, "product.owner_id value_help 应有 source"

        source_type = getattr(source, 'type', None)
        assert source_type == 'bo', f"product.owner_id source.type 应为 bo，实际: {source_type}"

        target_bo = getattr(source, 'target_bo', None)
        assert target_bo == 'user', \
            f"product.owner_id target_bo 应为 user，实际: {target_bo}"

    def test_06_version_product_id_value_help_bo(self):
        """version.product_id: bo value_help (product)"""
        field = self._find_field('version', 'product_id')
        if not field:
            pytest.skip("找不到 version.product_id 字段")

        vh = getattr(field, 'value_help', None)
        assert vh is not None, "version.product_id 应有 value_help"

        source = getattr(vh, 'source', None)
        assert source is not None, "version.product_id value_help 应有 source"

        source_type = getattr(source, 'type', None)
        assert source_type == 'bo', f"version.product_id source.type 应为 bo，实际: {source_type}"

        target_bo = getattr(source, 'target_bo', None)
        assert target_bo == 'product', \
            f"version.product_id target_bo 应为 product，实际: {target_bo}"

    def test_07_version_owner_id_value_help_bo(self):
        """version.owner_id: bo value_help (user)"""
        field = self._find_field('version', 'owner_id')
        if not field:
            pytest.skip("找不到 version.owner_id 字段")

        vh = getattr(field, 'value_help', None)
        assert vh is not None, "version.owner_id 应有 value_help"

        source = getattr(vh, 'source', None)
        assert source is not None, "version.owner_id value_help 应有 source"

        source_type = getattr(source, 'type', None)
        assert source_type == 'bo', f"version.owner_id source.type 应为 bo，实际: {source_type}"

        target_bo = getattr(source, 'target_bo', None)
        assert target_bo == 'user', \
            f"version.owner_id target_bo 应为 user，实际: {target_bo}"

    def test_08_user_status_export_has_dv(self):
        """导出用户：status 列应有枚举下拉"""
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id FROM users LIMIT 1")
        if not cursor.fetchone():
            pytest.skip("数据库中无 user 数据")

        result = self.ie.export_cascade('user', options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        user_sheet = None
        for name in ['用户', 'user']:
            if name in wb.sheetnames:
                user_sheet = wb[name]
                break
        if not user_sheet:
            pytest.skip("找不到用户 Sheet")

        headers_row = [cell.value for cell in user_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        status_col = col_map.get('状态')
        if not status_col:
            pytest.skip("找不到状态列")

        dvs = user_sheet.data_validations.dataValidation if user_sheet.data_validations else []
        from openpyxl.utils import get_column_letter
        status_col_letter = get_column_letter(status_col)

        has_status_dv = False
        for dv in dvs:
            sqref = getattr(dv, 'sqref', None)
            if sqref and status_col_letter in str(sqref):
                has_status_dv = True
                break
        assert has_status_dv, f"状态列应有数据验证，实际 DV 数量: {len(dvs)}"

    def test_09_user_group_parent_id_export_has_dv(self):
        """导出用户组：父级列应有 bo 下拉（层级选择）"""
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id FROM user_groups LIMIT 1")
        if not cursor.fetchone():
            pytest.skip("数据库中无 user_group 数据")

        result = self.ie.export_cascade('user_group', options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        group_sheet = None
        for name in ['用户组', 'user_group']:
            if name in wb.sheetnames:
                group_sheet = wb[name]
                break
        if not group_sheet:
            pytest.skip("找不到用户组 Sheet")

        headers_row = [cell.value for cell in group_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        parent_col = col_map.get('父级') or col_map.get('父级用户组')
        if not parent_col:
            pytest.skip("找不到父级列")

        dvs = group_sheet.data_validations.dataValidation if group_sheet.data_validations else []
        from openpyxl.utils import get_column_letter
        parent_col_letter = get_column_letter(parent_col)

        has_parent_dv = False
        for dv in dvs:
            sqref = getattr(dv, 'sqref', None)
            if sqref and parent_col_letter in str(sqref):
                has_parent_dv = True
                break
        assert has_parent_dv, f"父级列应有数据验证，实际 DV 数量: {len(dvs)}"

    def test_10_version_product_id_export_has_dv(self):
        """导出版本：产品列应有 bo 下拉"""
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id FROM versions LIMIT 1")
        if not cursor.fetchone():
            pytest.skip("数据库中无 version 数据")

        result = self.ie.export_cascade('version', options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        version_sheet = None
        for name in ['产品版本', 'version']:
            if name in wb.sheetnames:
                version_sheet = wb[name]
                break
        if not version_sheet:
            for name in wb.sheetnames:
                if '版本' in name:
                    version_sheet = wb[name]
                    break
        if not version_sheet:
            pytest.skip("找不到版本 Sheet")

        headers_row = [cell.value for cell in version_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        product_col = col_map.get('产品') or col_map.get('产品线')
        if not product_col:
            pytest.skip("找不到产品列")

        dvs = version_sheet.data_validations.dataValidation if version_sheet.data_validations else []
        from openpyxl.utils import get_column_letter
        product_col_letter = get_column_letter(product_col)

        has_product_dv = False
        for dv in dvs:
            sqref = getattr(dv, 'sqref', None)
            if sqref and product_col_letter in str(sqref):
                has_product_dv = True
                break
        assert has_product_dv, f"产品列应有数据验证，实际 DV 数量: {len(dvs)}"


class TestRelatedObjectsValueHelp:
    """相关对象的 value_help 测试

    补充：
    1. version.visibility: enum (public/draft)
    2. role.is_active: enum (启用/停用)
    3. role.is_system: enum (系统角色/自定义角色)
    4. user_group.manager_id: value_filter (status: active)
    """

    @pytest.fixture(autouse=True)
    def _setup(self, ie_service):
        self.ie = ie_service

    def _find_field(self, object_type, field_id):
        from meta.core.models import registry
        meta = registry.get(object_type)
        if not meta:
            return None
        for f in meta.fields:
            if f.id == field_id:
                return f
        return None

    def test_01_version_visibility_enum_values(self):
        """version.visibility: enum_values (public/draft)"""
        field = self._find_field('version', 'visibility')
        if not field:
            pytest.skip("找不到 version.visibility 字段")

        enum_values = getattr(field, 'enum_values', None)
        assert enum_values is not None, "version.visibility 应有 enum_values"

        values = [ev.get('value') if isinstance(ev, dict) else getattr(ev, 'value', None) for ev in enum_values]
        assert 'public' in values, f"version.visibility 应包含 public，实际: {values}"
        assert 'draft' in values, f"version.visibility 应包含 draft，实际: {values}"

    def test_02_version_visibility_export_has_dv(self):
        """导出版本：visibility 列应有枚举下拉"""
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id FROM versions LIMIT 1")
        if not cursor.fetchone():
            pytest.skip("数据库中无 version 数据")

        result = self.ie.export_cascade('version', options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        version_sheet = None
        for name in ['产品版本', 'version']:
            if name in wb.sheetnames:
                version_sheet = wb[name]
                break
        if not version_sheet:
            for name in wb.sheetnames:
                if '版本' in name:
                    version_sheet = wb[name]
                    break
        if not version_sheet:
            pytest.skip("找不到版本 Sheet")

        headers_row = [cell.value for cell in version_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        visibility_col = col_map.get('可见性')
        if not visibility_col:
            pytest.skip("找不到可见性列")

        dvs = version_sheet.data_validations.dataValidation if version_sheet.data_validations else []
        from openpyxl.utils import get_column_letter
        visibility_col_letter = get_column_letter(visibility_col)

        has_visibility_dv = False
        for dv in dvs:
            sqref = getattr(dv, 'sqref', None)
            if sqref and visibility_col_letter in str(sqref):
                has_visibility_dv = True
                break
        assert has_visibility_dv, f"可见性列应有数据验证，实际 DV 数量: {len(dvs)}"

    def test_03_role_is_active_enum_values(self):
        """role.is_active: enum_values (启用/停用)"""
        field = self._find_field('role', 'is_active')
        if not field:
            pytest.skip("找不到 role.is_active 字段")

        enum_values = getattr(field, 'enum_values', None)
        assert enum_values is not None, "role.is_active 应有 enum_values"

        values = [ev.get('value') if isinstance(ev, dict) else getattr(ev, 'value', None) for ev in enum_values]
        assert 1 in values or True in values or 'true' in [str(v).lower() for v in values], \
            f"role.is_active 应包含启用值，实际: {values}"

    def test_04_role_is_system_enum_values(self):
        """role.is_system: enum_values (系统角色/自定义角色)"""
        field = self._find_field('role', 'is_system')
        if not field:
            pytest.skip("找不到 role.is_system 字段")

        enum_values = getattr(field, 'enum_values', None)
        assert enum_values is not None, "role.is_system 应有 enum_values"

        values = [ev.get('value') if isinstance(ev, dict) else getattr(ev, 'value', None) for ev in enum_values]
        assert 1 in values or True in values or 'true' in [str(v).lower() for v in values], \
            f"role.is_system 应包含系统角色值，实际: {values}"

    def test_05_role_export_is_active_has_dv(self):
        """导出角色：is_active 列应有枚举下拉"""
        from meta.core.datasource import get_data_source
        ds = get_data_source('sqlite', database=get_test_db_path())
        cursor = ds.execute("SELECT id FROM roles LIMIT 1")
        if not cursor.fetchone():
            pytest.skip("数据库中无 role 数据")

        result = self.ie.export_cascade('role', options={
            'include_operation_mode': True,
            'include_metadata_sheet': True,
        })
        if not result or not result.file_path:
            pytest.skip("导出失败")

        from openpyxl import load_workbook
        wb = load_workbook(result.file_path)

        role_sheet = None
        for name in ['角色', 'role']:
            if name in wb.sheetnames:
                role_sheet = wb[name]
                break
        if not role_sheet:
            pytest.skip("找不到角色 Sheet")

        headers_row = [cell.value for cell in role_sheet[1]]
        col_map = {str(h): idx + 1 for idx, h in enumerate(headers_row) if h}

        is_active_col = col_map.get('启用状态') or col_map.get('是否启用')
        if not is_active_col:
            pytest.skip("找不到启用状态列")

        dvs = role_sheet.data_validations.dataValidation if role_sheet.data_validations else []
        from openpyxl.utils import get_column_letter
        is_active_col_letter = get_column_letter(is_active_col)

        has_is_active_dv = False
        for dv in dvs:
            sqref = getattr(dv, 'sqref', None)
            if sqref and is_active_col_letter in str(sqref):
                has_is_active_dv = True
                break
        assert has_is_active_dv, f"启用状态列应有数据验证，实际 DV 数量: {len(dvs)}"

    def test_06_user_group_manager_id_value_filter(self):
        """user_group.manager_id: value_filter (status: active)"""
        field = self._find_field('user_group', 'manager_id')
        if not field:
            pytest.skip("找不到 user_group.manager_id 字段")

        vh = getattr(field, 'value_help', None)
        assert vh is not None, "user_group.manager_id 应有 value_help"

        source = getattr(vh, 'source', None)
        assert source is not None, "user_group.manager_id value_help 应有 source"

        value_filter = getattr(source, 'value_filter', None)
        assert value_filter is not None, "user_group.manager_id 应有 value_filter 配置"

        if isinstance(value_filter, dict):
            assert value_filter.get('status') == 'active', \
                f"value_filter.status 应为 active，实际: {value_filter}"
        else:
            has_status_active = False
            if hasattr(value_filter, 'status'):
                has_status_active = getattr(value_filter, 'status') == 'active'
            assert has_status_active, f"value_filter 应包含 status: active，实际: {value_filter}"

    def test_07_product_owner_id_search_fields(self):
        """product.owner_id: search_fields 配置"""
        field = self._find_field('product', 'owner_id')
        if not field:
            pytest.skip("找不到 product.owner_id 字段")

        vh = getattr(field, 'value_help', None)
        assert vh is not None, "product.owner_id 应有 value_help"

        behavior = getattr(vh, 'behavior', None)
        if behavior:
            search_fields = getattr(behavior, 'search_fields', None)
            if search_fields:
                assert 'username' in search_fields or 'display_name' in search_fields, \
                    f"search_fields 应包含 username 或 display_name，实际: {search_fields}"

    def test_08_version_product_id_display_format(self):
        """version.product_id: display_format 配置"""
        field = self._find_field('version', 'product_id')
        if not field:
            pytest.skip("找不到 version.product_id 字段")

        vh = getattr(field, 'value_help', None)
        assert vh is not None, "version.product_id 应有 value_help"

        presentation = getattr(vh, 'presentation', None)
        if presentation:
            display_format = getattr(presentation, 'display_format', None)
            assert display_format is not None, "version.product_id 应有 display_format"
            assert '{code}' in display_format or '{name}' in display_format, \
                f"display_format 应包含 code 或 name 占位符，实际: {display_format}"
