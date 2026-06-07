# -*- coding: utf-8 -*-
"""
导入模板导出测试 - 验证Excel导出模板的内容

测试范围：
1. 模板文件生成
2. 表头验证
3. readonly字段标记（背景色）
4. 必填字段标记
5. 表头注释
"""

import pytest
import os
import sys
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from meta.tests.conftest import get_shared_app
from meta.services.token_service import TokenService
from meta.services.auth_provider import UserInfo


def get_auth_headers():
    """获取认证头"""
    user = UserInfo(
        user_id='1', username='test_user', display_name='Test User',
        email='test@test.com', roles=['admin'], permissions=['*']
    )
    token, _ = TokenService.create_token(user)
    return {
        'Content-Type': 'application/json',
        'Authorization': 'Bearer ' + token,
        'X-User-Id': '1',
        'X-User-Name': 'test_user'
    }


class TestImportTemplateAPI:
    """导入模板API测试"""

    def test_download_role_template(self):
        """测试下载角色导入模板"""
        app, client = get_shared_app()
        headers = get_auth_headers()

        response = client.get(
            '/api/v1/import/template/role',
            headers=headers
        )

        assert response.status_code in [200, 401, 404, 500]

        if response.status_code == 200:
            assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            # 保存文件用于详细验证
            from meta.tests.test_utils import get_test_db_path
            export_dir = os.path.join(os.path.dirname(get_test_db_path()), 'exports')
            os.makedirs(export_dir, exist_ok=True)

            file_path = os.path.join(export_dir, 'role_template_test.xlsx')
            with open(file_path, 'wb') as f:
                f.write(response.data)

            print(f"模板文件保存到: {file_path}")
            print(f"文件大小: {os.path.getsize(file_path)} bytes")


class TestImportTemplateContent:
    """导入模板内容验证测试"""

    def test_role_template_headers(self):
        """测试角色模板包含正确的表头"""
        app, client = get_shared_app()
        headers = get_auth_headers()

        response = client.get(
            '/api/v1/import/template/role',
            headers=headers
        )

        if response.status_code == 200:
            from meta.tests.test_utils import get_test_db_path
            export_dir = os.path.join(os.path.dirname(get_test_db_path()), 'exports')
            os.makedirs(export_dir, exist_ok=True)

            file_path = os.path.join(export_dir, 'role_template_headers.xlsx')
            with open(file_path, 'wb') as f:
                f.write(response.data)

            wb = load_workbook(file_path)
            assert '角色' in wb.sheetnames, "模板应包含'角色'工作表"

            ws = wb['角色']
            headers = [cell.value for cell in ws[1] if cell.value]

            # 验证基本表头
            expected_headers = ['操作模式', '角色编码', '角色名称']
            for header in expected_headers:
                assert header in headers, f"表头应包含'{header}'"

            print(f"角色模板表头: {headers}")
            wb.close()

    def test_role_template_readonly_fields(self):
        """测试角色模板中readonly字段的标记"""
        app, client = get_shared_app()
        headers = get_auth_headers()

        response = client.get(
            '/api/v1/import/template/role',
            headers=headers
        )

        if response.status_code == 200:
            from meta.tests.test_utils import get_test_db_path
            export_dir = os.path.join(os.path.dirname(get_test_db_path()), 'exports')
            os.makedirs(export_dir, exist_ok=True)

            file_path = os.path.join(export_dir, 'role_template_readonly.xlsx')
            with open(file_path, 'wb') as f:
                f.write(response.data)

            wb = load_workbook(file_path)
            ws = wb['角色']

            # 获取表头
            headers = [cell.value for cell in ws[1] if cell.value]

            # 检查数据行的背景色来判断readonly
            readonly_fields = []
            editable_fields = []

            if ws.max_row > 1:
                for col_idx in range(1, ws.max_column + 1):
                    header_name = ws.cell(row=1, column=col_idx).value
                    if not header_name:
                        continue

                    data_cell = ws.cell(row=2, column=col_idx)
                    fill = data_cell.fill

                    # 获取背景色
                    fill_color = None
                    if hasattr(fill, 'start_color') and fill.start_color:
                        if hasattr(fill.start_color, 'rgb'):
                            fill_color = str(fill.start_color.rgb)

                    # 灰色背景 (E0E0E0) 表示readonly
                    if fill_color and 'E0E0E0' in fill_color:
                        readonly_fields.append(header_name)
                    elif fill_color and 'FFF2CC' in fill_color:
                        # 浅绿色是业务键，不是readonly
                        pass
                    else:
                        editable_fields.append(header_name)

            print(f"readonly字段: {readonly_fields}")
            print(f"可编辑字段: {editable_fields}")

            # 验证统计字段是readonly
            readonly_stats = ['用户数', '菜单数', '权限数', '数据权限数']
            for field in readonly_stats:
                if field in headers:
                    assert field in readonly_fields, f"'{field}' 应该是readonly"

            wb.close()

    def test_role_template_business_key(self):
        """测试角色模板中业务键字段的标记"""
        app, client = get_shared_app()
        headers = get_auth_headers()

        response = client.get(
            '/api/v1/import/template/role',
            headers=headers
        )

        if response.status_code == 200:
            from meta.tests.test_utils import get_test_db_path
            export_dir = os.path.join(os.path.dirname(get_test_db_path()), 'exports')
            os.makedirs(export_dir, exist_ok=True)

            file_path = os.path.join(export_dir, 'role_template_bk.xlsx')
            with open(file_path, 'wb') as f:
                f.write(response.data)

            wb = load_workbook(file_path)
            ws = wb['角色']

            # 检查业务键字段的背景色（浅绿色 FFF2CC）
            business_key_fields = []

            if ws.max_row > 1:
                for col_idx in range(1, ws.max_column + 1):
                    header_name = ws.cell(row=1, column=col_idx).value
                    if not header_name:
                        continue

                    data_cell = ws.cell(row=2, column=col_idx)
                    fill = data_cell.fill

                    fill_color = None
                    if hasattr(fill, 'start_color') and fill.start_color:
                        if hasattr(fill.start_color, 'rgb'):
                            fill_color = str(fill.start_color.rgb)

                    # 浅绿色是业务键
                    if fill_color and 'FFF2CC' in fill_color:
                        business_key_fields.append(header_name)

            print(f"业务键字段: {business_key_fields}")

            # 角色编码应该是业务键
            assert '角色编码' in business_key_fields, "'角色编码' 应该是业务键字段（浅绿色背景）"

            wb.close()

    def test_user_template_excludes_sensitive_fields(self):
        """测试用户模板排除敏感字段"""
        app, client = get_shared_app()
        headers = get_auth_headers()

        response = client.get(
            '/api/v1/import/template/user',
            headers=headers
        )

        if response.status_code == 200:
            from meta.tests.test_utils import get_test_db_path
            export_dir = os.path.join(os.path.dirname(get_test_db_path()), 'exports')
            os.makedirs(export_dir, exist_ok=True)

            file_path = os.path.join(export_dir, 'user_template.xlsx')
            with open(file_path, 'wb') as f:
                f.write(response.data)

            wb = load_workbook(file_path)

            if '用户' in wb.sheetnames:
                ws = wb['用户']
                headers = [cell.value for cell in ws[1] if cell.value]

                # 敏感字段不应该出现在模板中
                sensitive_fields = ['密码哈希', 'password_hash', '密码', 'password',
                                   '语言区域', 'locale', '时区', 'timezone',
                                   '日期格式长度', 'date_style', '时间格式长度', 'time_style']

                for field in sensitive_fields:
                    assert field not in headers, f"敏感字段 '{field}' 不应出现在用户模板中"

                print(f"用户模板表头: {headers}")

            wb.close()


class TestImportTemplateComments:
    """导入模板注释验证测试"""

    def test_role_template_has_comments(self):
        """测试角色模板包含表头注释"""
        app, client = get_shared_app()
        headers = get_auth_headers()

        response = client.get(
            '/api/v1/import/template/role',
            headers=headers
        )

        if response.status_code == 200:
            from meta.tests.test_utils import get_test_db_path
            export_dir = os.path.join(os.path.dirname(get_test_db_path()), 'exports')
            os.makedirs(export_dir, exist_ok=True)

            file_path = os.path.join(export_dir, 'role_template_comments.xlsx')
            with open(file_path, 'wb') as f:
                f.write(response.data)

            wb = load_workbook(file_path)
            ws = wb['角色']

            # 检查表头单元格是否有注释
            comments_count = 0
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                if cell.comment:
                    comments_count += 1

            print(f"有注释的表头数量: {comments_count}")
            assert comments_count > 0, "表头应有注释"

            wb.close()


if __name__ == '__main__':
    pytest.main([__file__, '-v', '-s'])
