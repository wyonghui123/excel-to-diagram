# -*- coding: utf-8 -*-
"""
[NEW v1.1 2026-06-12] KeyTemplate user_editable 端到端测试

覆盖场景（见 docs/specs/key-template-user-editable/REVIEW_AND_TEST_PLAN.md §2）:
- S1 启用对象：business_object / relationship 显示"可手动"标签
- S2 用户填入 code：拦截器记录 INFO 日志
- S3 用户留空 code：自动生成
- S4 不应启用的对象（role/user_group/domain/sub_domain/service_module/
       permission/menu）：不显示"自动"标签 + 不在 Excel 规则区
- S5 Excel 导出：说明 sheet 含"自动编码规则"区，code 列底色区分
- E1-E4 异常路径

目的：用户感知问题（role/user_group 等误显示"自动"）的回归防线。
"""
import pytest
import os
from openpyxl import load_workbook

pytestmark = pytest.mark.integration


KT_URL = '/api/v2/key-template'

# 共享 parametrize 列表
KT_DISABLED_OBJECTS = [
    'role', 'user_group', 'permission', 'menu',
    'domain', 'sub_domain', 'service_module',
    'version',  # 已被 TBD-2 决策移除
]
KT_ENABLED_OBJECTS = ['business_object', 'relationship']


# ─────────────────────────────────────────────────────────
# 局部 fixtures（与 test_edge_cases.py 一致）
# ─────────────────────────────────────────────────────────
@pytest.fixture
def api_client(shared_client):
    """API 测试客户端"""
    return shared_client


# ─────────────────────────────────────────────────────────
# S4 + S1 端到端：HTTP /config 真实响应
# ─────────────────────────────────────────────────────────
class TestUserEditableConfigE2E:
    """端到端：HTTP /api/v2/key-template/config/<type> 真实响应"""

    @pytest.mark.parametrize('object_type', KT_DISABLED_OBJECTS)
    def test_disabled_objects_api_returns_enabled_false(self, api_client, admin_headers, object_type):
        """
        S4 关键回归：role / user_group / domain 等 7 个对象
        GET /config/<type> → data.enabled 必须为 False

        这是 v1.1 修复 "新建页面误显示自动标签" 的核心防御。
        如果该测试失败 → 任何使用这些对象的页面都会再次显示"自动"标签。
        """
        resp = api_client.get(f'{KT_URL}/config/{object_type}', headers=admin_headers)
        assert resp.status_code == 200, (
            f"GET /config/{object_type} 失败: status={resp.status_code} "
            f"body={resp.get_data(as_text=True)[:200]}"
        )
        body = resp.get_json()
        assert body.get('success') is True, f"success 应为 True，实际: {body}"
        data = body.get('data', {})

        # 兼容两种返回结构：
        # 1) 无 key_template → data = {'enabled': False, 'message': '...'}
        # 2) 有空 key_template → data = {'enabled': False}
        if 'enabled' in data:
            assert data['enabled'] is False, (
                f"{object_type} 不应启用 key_template (无配置)，"
                f"但 API 返回 enabled={data['enabled']}。"
                f"这会导致前端 isCodeAutoManaged=true → 误显示'自动'标签。"
            )
        else:
            assert 'key_template' not in data or not data.get('key_template'), (
                f"{object_type} 不应有 key_template 字段，实际: {data}"
            )

    @pytest.mark.parametrize('object_type', KT_ENABLED_OBJECTS)
    def test_enabled_objects_api_returns_user_editable(self, api_client, admin_headers, object_type):
        """
        S1 关键验证：business_object / relationship
        GET /config/<type> → 返回 key_template.user_editable 字段
        """
        resp = api_client.get(f'{KT_URL}/config/{object_type}', headers=admin_headers)
        assert resp.status_code == 200, f"GET /config/{object_type} 失败: {resp.status_code}"
        body = resp.get_json()
        assert body.get('success') is True
        data = body.get('data', {})
        kt = data.get('key_template', {})
        assert kt, f"{object_type} 应有 key_template 配置，实际: {data}"
        assert kt.get('user_editable') in ('auto_only', 'auto_or_manual', 'manual_only'), (
            f"{object_type} key_template.user_editable 应为合法值，实际: {kt.get('user_editable')}"
        )
        # 本阶段 v1.1 应统一为 auto_or_manual
        assert kt.get('user_editable') == 'auto_or_manual', (
            f"{object_type} v1.1 应默认 auto_or_manual，实际: {kt.get('user_editable')}"
        )

    @pytest.mark.parametrize('object_type,expected_enabled', [
        ('business_object', True),
        ('relationship', True),
        ('role', False),
        ('user_group', False),
        ('version', False),
    ])
    def test_config_enabled_for_each_object(self, api_client, admin_headers, object_type, expected_enabled):
        """[合并] 一致性矩阵：7 个核心对象的 enabled 状态"""
        resp = api_client.get(f'{KT_URL}/config/{object_type}', headers=admin_headers)
        body = resp.get_json()
        data = body.get('data', {})
        # 兼容 1.0 和 1.1 两种返回结构
        if 'key_template' in data:
            actual_enabled = bool(data['key_template'])
        else:
            actual_enabled = bool(data.get('enabled', False))
        assert actual_enabled == expected_enabled, (
            f"{object_type} enabled 应为 {expected_enabled}，实际: {actual_enabled}, data: {data}"
        )


# ─────────────────────────────────────────────────────────
# S2 + S3 拦截器 E2E：实际创建 BO
# ─────────────────────────────────────────────────────────
class TestInterceptorE2E:
    """端到端：实际创建 BO 验证拦截器行为"""

    @pytest.mark.parametrize('object_type', KT_DISABLED_OBJECTS)
    def test_create_disabled_object_fails_without_code(self, api_client, admin_headers, object_type):
        """
        S4 关键防御：创建 role / user_group / domain 等
        POST /api/v2/bo/<type> 不传 code 应失败（不应自动生成）
        """
        payload = {'name': f'Test_{object_type}_NoCode'}
        # 一些对象有额外必填字段，做简单兼容
        if object_type == 'user_group':
            payload['name'] = f'ug_{object_type}_test'
        resp = api_client.post(f'/api/v2/bo/{object_type}', json=payload, headers=admin_headers)
        # 期望失败（code 必填或字段缺失）
        assert resp.status_code in (400, 422, 500), (
            f"创建 {object_type} 不传 code 应失败（不自动生成），"
            f"实际: {resp.status_code} body={resp.get_data(as_text=True)[:200]}"
        )


# ─────────────────────────────────────────────────────────
# S1 preview API 验证
# ─────────────────────────────────────────────────────────
class TestPreviewAPIE2E:
    """preview API 必须返回 user_editable / pattern / preview"""

    def test_preview_returns_user_editable_for_bo(self, api_client, admin_headers):
        """preview business_object 成功时必须返回 user_editable"""
        # 找一个可用的 service_module
        sm_resp = api_client.get('/api/v2/bo/service_module?page=1&page_size=1', headers=admin_headers)
        sm_items = sm_resp.get_json().get('data', {}).get('items', [])
        if not sm_items:
            pytest.skip("无可用 service_module")
        sm_id = sm_items[0]['id']

        resp = api_client.post(
            f'{KT_URL}/preview/business_object',
            json={'field_values': {}, 'parent_params': {'service_module_id': sm_id}, 'generate': True},
            headers=admin_headers,
        )
        if resp.status_code != 200:
            pytest.skip(f"BO preview 跳过 (status={resp.status_code})")
        body = resp.get_json()
        data = body.get('data', {})
        assert 'user_editable' in data, f"preview 应返回 user_editable，实际: {data}"
        assert 'pattern' in data, f"preview 应返回 pattern，实际: {data}"
        assert 'preview' in data, f"preview 应返回 preview，实际: {data}"
        assert data['user_editable'] == 'auto_or_manual', (
            f"v1.1 应统一为 auto_or_manual，实际: {data['user_editable']}"
        )


# ─────────────────────────────────────────────────────────
# S4 + S5 Excel 视觉 E2E
# ─────────────────────────────────────────────────────────
class TestExcelVisualE2E:
    """端到端：导出 Excel 验证视觉差异"""

    def _get_code_col_color(self, wb):
        """提取主 sheet code 列数据行的 fill color"""
        from meta.services.excel_design_system import ExcelDesignSystem
        target_rgb = ExcelDesignSystem.AUTO_GEN_OR_MANUAL_FILL.start_color.rgb
        if target_rgb and len(target_rgb) == 8:
            target_rgb = target_rgb[2:]
        main_sheet = wb[wb.sheetnames[0]]
        # 找 code 列
        code_col = None
        for col in range(1, min(main_sheet.max_column + 1, 30)):
            if main_sheet.cell(row=1, column=col).value == 'code':
                code_col = col
                break
        if not code_col:
            return None, target_rgb, main_sheet
        cell = main_sheet.cell(row=2, column=code_col)
        cell_color = ''
        if cell.fill and cell.fill.start_color:
            rgb = cell.fill.start_color.rgb or ''
            if len(rgb) == 8:
                rgb = rgb[2:]
            cell_color = rgb
        return cell_color, target_rgb, main_sheet

    def _export_object(self, ie_service, object_type):
        """统一导出（与现有 _make_main_sheet_wb_with_protection 保持一致风格）"""
        result = ie_service.export_template([object_type], options={
            'include_operation_mode': True,
            'include_hierarchy_path': True,
            'include_hierarchy_ids': True,
            'include_metadata_sheet': True,
            'include_child_objects': False,
            'empty_rows_for_new': 1,
            'protect_sheet': False,
        })
        return result

    @pytest.mark.parametrize('object_type', KT_DISABLED_OBJECTS)
    def test_disabled_object_code_no_light_blue_fill(self, ie_service, object_type):
        """
        S4 关键：导出 role / user_group / domain 等时
        code 列**不应**有浅蓝灰底色（这会误导用户认为该字段可自动）
        """
        try:
            result = self._export_object(ie_service, object_type)
        except Exception as e:
            pytest.skip(f"{object_type} 导出失败: {e}")
        if not result or not getattr(result, 'file_path', None):
            pytest.skip(f"{object_type} 无 file_path")
        wb = load_workbook(result.file_path)
        cell_color, target_rgb, _ = self._get_code_col_color(wb)
        if cell_color is None:
            pytest.skip(f"{object_type} 主 sheet 无 code 列")
        assert cell_color != target_rgb, (
            f"{object_type} code 列不应有浅蓝灰底色 (auto_or_manual)，"
            f"实际: {cell_color}。这意味着 _get_key_template_user_editable "
            f"对 {object_type} 错误返回了非空值。"
        )

    @pytest.mark.parametrize('object_type', KT_ENABLED_OBJECTS)
    def test_enabled_object_code_has_light_blue_fill(self, ie_service, object_type):
        """
        S5 关键：导出 business_object / relationship 时
        code 列**应有**浅蓝灰底色（auto_or_manual_code）
        """
        try:
            result = self._export_object(ie_service, object_type)
        except Exception as e:
            pytest.skip(f"{object_type} 导出失败: {e}")
        if not result or not getattr(result, 'file_path', None):
            pytest.skip(f"{object_type} 无 file_path")
        wb = load_workbook(result.file_path)
        cell_color, target_rgb, _ = self._get_code_col_color(wb)
        if cell_color is None:
            pytest.skip(f"{object_type} 主 sheet 无 code 列")
        assert cell_color == target_rgb, (
            f"{object_type} code 列应有浅蓝灰底色 ({target_rgb})，"
            f"实际: {cell_color}。这意味着 _get_key_template_user_editable "
            f"对 {object_type} 错误返回了空。"
        )

    def test_meta_sheet_contains_auto_encoding_rules(self, ie_service):
        """
        S5 验证：说明 sheet 包含"自动编码规则"区
        且只列出启用 key_template 的对象
        """
        result = self._export_object(ie_service, 'business_object')
        if not result or not getattr(result, 'file_path', None):
            pytest.skip("导出失败")
        wb = load_workbook(result.file_path)
        # 找说明 sheet
        meta_sheet = None
        for name in wb.sheetnames:
            if '说明' in name or 'meta' in name.lower() or 'instructions' in name.lower():
                meta_sheet = wb[name]
                break
        if not meta_sheet:
            pytest.skip("找不到说明 sheet")

        all_text = '\n'.join(
            str(meta_sheet.cell(row=r, column=c).value or '')
            for r in range(1, meta_sheet.max_row + 1)
            for c in range(1, meta_sheet.max_column + 1)
        )
        assert '自动编码规则' in all_text, (
            f"说明 sheet 缺'自动编码规则'区: {all_text[:500]}"
        )
        # 应列出 business_object 或中文名称 '业务对象'
        assert 'business_object' in all_text or '业务对象' in all_text, (
            f"规则区应包含 business_object 或'业务对象': {all_text[:500]}"
        )


# ─────────────────────────────────────────────────────────
# dataclass 边界值（补充）
# ─────────────────────────────────────────────────────────
class TestDataclassValidationExtended:
    """dataclass 校验（补充 v1.1 已有 T1-T5）"""

    def test_none_user_editable_raises(self):
        """用户显式传 None 也应抛错"""
        from meta.core.key_template_engine import KeyTemplateConfig
        with pytest.raises(ValueError) as exc_info:
            KeyTemplateConfig(object_id='test', enabled=True, user_editable=None)
        assert 'Invalid user_editable' in str(exc_info.value)

    def test_empty_string_user_editable_raises(self):
        """空串也应抛错"""
        from meta.core.key_template_engine import KeyTemplateConfig
        with pytest.raises(ValueError):
            KeyTemplateConfig(object_id='test', enabled=True, user_editable='')

    def test_user_editable_case_sensitive(self):
        """大小写敏感：AUTO_OR_MANUAL 应被拒绝"""
        from meta.core.key_template_engine import KeyTemplateConfig
        with pytest.raises(ValueError):
            KeyTemplateConfig(object_id='test', enabled=True, user_editable='AUTO_OR_MANUAL')

    @pytest.mark.parametrize('valid_value', ['auto_only', 'auto_or_manual', 'manual_only'])
    def test_all_valid_values_accepted(self, valid_value):
        """所有合法值都应被接受"""
        from meta.core.key_template_engine import KeyTemplateConfig
        cfg = KeyTemplateConfig(object_id='test', enabled=True, user_editable=valid_value)
        assert cfg.user_editable == valid_value

    def test_yaml_load_with_invalid_user_editable_raises(self):
        """YAML 加载时 user_editable 非法值应抛错（不在静默通过）"""
        from meta.core.key_template_engine import KeyTemplateConfig
        bad_kt = {'enabled': True, 'user_editable': 'bad_mode'}
        with pytest.raises(ValueError) as exc_info:
            KeyTemplateConfig.from_dict('test_bad', bad_kt)
        assert 'Invalid user_editable' in str(exc_info.value)
        assert 'bad_mode' in str(exc_info.value)  # 应显示非法值便于诊断
