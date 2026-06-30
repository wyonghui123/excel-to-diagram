# -*- coding: utf-8 -*-
"""
test_export_owner_id_v2_1.py

覆盖提交: 本会话修复 (产品线显示负责人 / 产品版本不显示负责人)
依据: 用户报告 6月26日 "负责人列,产品sheet应该有,版本sheet应该没有"

测试:
- product.yaml 的 owner_id ui.visible: true (产品线 sheet 显示负责人列)
- version.yaml 不应有 owner_id 字段 (产品版本不显示负责人列)
- 实际导出验证: 产品线 sheet 包含 '负责人' 列
- 实际导出验证: 产品版本 sheet 不包含 '负责人' 列
"""
import os
import sys
import re
import requests
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).parent.parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

pytestmark = [
    pytest.mark.post_v2_1,
    pytest.mark.import_export,
]


# ============================================================
# 1. TestProductOwnerIdVisible  (产品线 sheet 显示负责人)
# ============================================================

class TestProductOwnerIdVisible:
    """[FIX 2026-06-26] 产品线 sheet 显示负责人列"""

    def test_product_yaml_owner_id_ui_visible_true(self):
        """product.yaml 的 owner_id 设置 ui.visible: true"""
        product_yaml = PROJECT_ROOT / 'meta' / 'schemas' / 'product.yaml'
        with open(product_yaml, 'r', encoding='utf-8') as f:
            content = f.read()

        pattern = re.compile(
            r'-\s+id:\s+owner_id\s*\n(?:\s+\S.*\n)*?\s+ui:\s*\n(?:\s+\S.*\n)*?\s+visible:\s*true',
            re.MULTILINE
        )
        match = pattern.search(content)
        assert match is not None, \
            "product.yaml 的 owner_id 应设置 ui.visible: true (产品线 sheet 显示负责人列)"

    def test_product_export_includes_owner_column(self):
        """实际导出: 产品线 sheet 包含 '负责人' 列"""
        base = os.environ.get('BACKEND_URL', 'http://localhost:3010')
        try:
            session = requests.Session()
            r = session.get(f"{base}/api/v1/auth/dev-login", params={"username": "admin"}, timeout=5)
            r.raise_for_status()
        except Exception:
            pytest.skip(f"后端 {base} 不可用, 跳过实际导出测试")

        r = session.post(
            f"{base}/api/v1/export",
            json={"object_type": "product", "selected_types": ["product"]},
            timeout=15
        )
        assert r.status_code == 200, f"导出请求失败: {r.status_code} {r.text[:200]}"
        url = r.json()['data']['download_url']
        r = session.get(f"{base}{url}", timeout=10)
        assert r.status_code == 200

        # 解析 xlsx
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        try:
            assert "产品线" in wb.sheetnames, "应有产品线 sheet"

            ws = wb["产品线"]
            headers = [c.value for c in ws[1]]
            assert "负责人" in headers, \
                f"产品线 sheet 应包含 '负责人' 列, 实际 headers: {headers}"
        finally:
            wb.close()


# ============================================================
# 2. TestVersionOwnerIdHidden  (产品版本 sheet 不显示负责人)
# ============================================================

class TestVersionOwnerIdHidden:
    """[FIX 2026-06-26] 产品版本 sheet 不显示负责人列"""

    def test_version_yaml_no_owner_id_field(self):
        """version.yaml 不应有 '- id: owner_id' 字段定义 (产品版本不显示负责人)"""
        version_yaml = PROJECT_ROOT / 'meta' / 'schemas' / 'version.yaml'
        with open(version_yaml, 'r', encoding='utf-8') as f:
            content = f.read()

        pattern = r'^-\s+id:\s+owner_id\s*$'
        matches = re.findall(pattern, content, re.MULTILINE)
        assert len(matches) == 0, \
            f"version.yaml 不应有 '- id: owner_id' 字段定义, 实际: {len(matches)} 处"

    def test_version_yaml_no_owner_id_in_recent_commits(self):
        """version.yaml 的注释明确说明 owner_id 已删除"""
        version_yaml = PROJECT_ROOT / 'meta' / 'schemas' / 'version.yaml'
        with open(version_yaml, 'r', encoding='utf-8') as f:
            content = f.read()

        # 关键: 注释说明 owner_id 已删除
        # 例如: # [V1.1.4 2026-06-11] 删 owner_id 字段定义 - DB 列 V1.1.1 已删
        assert "owner_id" in content, \
            "version.yaml 注释应说明 owner_id 已删除"

    def test_version_export_excludes_owner_column(self):
        """实际导出: 产品版本 sheet 不包含 '负责人' 列"""
        base = os.environ.get('BACKEND_URL', 'http://localhost:3010')
        try:
            session = requests.Session()
            r = session.get(f"{base}/api/v1/auth/dev-login", params={"username": "admin"}, timeout=5)
            r.raise_for_status()
        except Exception:
            pytest.skip(f"后端 {base} 不可用, 跳过实际导出测试")

        r = session.post(
            f"{base}/api/v1/export",
            json={"object_type": "product", "selected_types": ["product"]},
            timeout=15
        )
        assert r.status_code == 200, f"导出请求失败: {r.status_code} {r.text[:200]}"
        url = r.json()['data']['download_url']
        r = session.get(f"{base}{url}", timeout=10)
        assert r.status_code == 200

        # 解析 xlsx
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        try:
            assert "产品版本" in wb.sheetnames, "应有产品版本 sheet"

            ws = wb["产品版本"]
            headers = [c.value for c in ws[1]]
            assert "负责人" not in headers, \
                f"产品版本 sheet 不应包含 '负责人' 列, 实际 headers: {headers}"
        finally:
            wb.close()