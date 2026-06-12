"""
Excel导出设计规范配置
遵循SAP Fiori、Salesforce等头部产品的Excel设计最佳实践

设计原则：
1. 简洁至上 (Less is More) - 颜色、边框都应有存在的理由
2. 一致性 (Consistency) - 同类信息使用同一种格式
3. 突出重点 (Highlighting) - 用颜色引导注意力
4. 对齐是灵魂 (Alignment) - 文本左对齐，数字右对齐
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelDesignSystem:
    """Excel设计系统 - 统一的样式配置"""
    
    # 主色调 - 蓝色系（参考Salesforce #00A1E0）
    PRIMARY_COLOR = "1565C0"
    PRIMARY_LIGHT = "E3F2FD"
    
    # 功能色
    SUCCESS_COLOR = "2E7D32"
    SUCCESS_LIGHT = "E6F7E6"
    WARNING_COLOR = "F57C00"
    WARNING_LIGHT = "FFF2CC"
    ERROR_COLOR = "C62828"
    ERROR_LIGHT = "FFEBEE"
    INFO_COLOR = "0288D1"
    INFO_LIGHT = "E1F5FE"
    
    # 中性色
    GRAY_50 = "FAFAFA"
    GRAY_100 = "F5F5F5"
    GRAY_200 = "EEEEEE"
    GRAY_300 = "E0E0E0"
    GRAY_500 = "9E9E9E"
    GRAY_700 = "616161"
    GRAY_900 = "212121"
    
    # 表头样式
    HEADER_FILL = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    # 标签样式
    LABEL_FONT = Font(bold=True, size=10, color=GRAY_700)
    VALUE_FONT = Font(size=10, color=GRAY_900)
    
    # 区域标题样式
    SECTION_FILL = PatternFill(start_color=PRIMARY_LIGHT, end_color=PRIMARY_LIGHT, fill_type="solid")
    SECTION_FONT = Font(bold=True, size=11, color=PRIMARY_COLOR)
    
    # 字段状态样式
    READONLY_FILL = PatternFill(start_color=GRAY_300, end_color=GRAY_300, fill_type="solid")
    REQUIRED_FILL = PatternFill(start_color=WARNING_LIGHT, end_color=WARNING_LIGHT, fill_type="solid")
    BUSINESS_KEY_FILL = PatternFill(start_color=SUCCESS_LIGHT, end_color=SUCCESS_LIGHT, fill_type="solid")
    CREATE_NEW_FILL = PatternFill(start_color=INFO_LIGHT, end_color=INFO_LIGHT, fill_type="solid")
    # [NEW v1.1 2026-06-11] 自动/可手动模式底色（浅蓝）
    AUTO_GEN_OR_MANUAL_FILL = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")
    
    # 边框样式
    THIN_BORDER = Border(
        left=Side(style='thin', color=GRAY_300),
        right=Side(style='thin', color=GRAY_300),
        top=Side(style='thin', color=GRAY_300),
        bottom=Side(style='thin', color=GRAY_300)
    )
    
    MEDIUM_BORDER = Border(
        left=Side(style='medium', color=GRAY_500),
        right=Side(style='medium', color=GRAY_500),
        top=Side(style='medium', color=GRAY_500),
        bottom=Side(style='medium', color=GRAY_500)
    )
    
    # 对齐方式
    TEXT_LEFT = Alignment(horizontal="left", vertical="center")
    TEXT_CENTER = Alignment(horizontal="center", vertical="center")
    TEXT_RIGHT = Alignment(horizontal="right", vertical="center")
    
    @classmethod
    def apply_header_style(cls, cell):
        """应用表头样式"""
        cell.fill = cls.HEADER_FILL
        cell.font = cls.HEADER_FONT
        cell.alignment = cls.HEADER_ALIGNMENT
        cell.border = cls.THIN_BORDER
    
    @classmethod
    def apply_label_style(cls, cell):
        """应用标签样式"""
        cell.font = cls.LABEL_FONT
        cell.alignment = cls.TEXT_LEFT
        cell.border = cls.THIN_BORDER
    
    @classmethod
    def apply_value_style(cls, cell):
        """应用值样式"""
        cell.font = cls.VALUE_FONT
        cell.alignment = cls.TEXT_LEFT
        cell.border = cls.THIN_BORDER
    
    @classmethod
    def apply_section_style(cls, cell):
        """应用区域标题样式"""
        cell.fill = cls.SECTION_FILL
        cell.font = cls.SECTION_FONT
        cell.border = cls.THIN_BORDER
    
    @classmethod
    def apply_readonly_style(cls, cell):
        """应用只读字段样式"""
        cell.fill = cls.READONLY_FILL
        cell.border = cls.THIN_BORDER
    
    @classmethod
    def apply_required_style(cls, cell):
        """应用必填字段样式"""
        cell.fill = cls.REQUIRED_FILL
        cell.border = cls.THIN_BORDER
    
    @classmethod
    def apply_business_key_style(cls, cell):
        """应用业务键字段样式"""
        cell.fill = cls.BUSINESS_KEY_FILL
        cell.border = cls.THIN_BORDER
    
    @classmethod
    def auto_column_width(cls, ws, min_width=10, max_width=50):
        """自动调整列宽"""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width
    
    @classmethod
    def create_data_validation_list(cls, values, allow_blank=True):
        """创建数据验证下拉列表"""
        from openpyxl.worksheet.datavalidation import DataValidation
        formula = '"' + ','.join(str(v) for v in values) + '"'
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=allow_blank
        )
        return dv
