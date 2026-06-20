from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional
from datetime import datetime
import os
import re
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from meta.core.models import MetaObject, registry, ImportExportConfig, QueryOperator, FieldStorage
from meta.core.datasource import DataSource
from meta.services.manage_service import ManageService, BatchOperationResult, CreateRequest, UpdateRequest, DeleteRequest
from meta.services.query_service import QueryService, SearchRequest, QueryCondition
from meta.services.hierarchy_filter_service import HierarchyFilterService
from meta.services.cascade_service import HierarchyConfigLoader
from meta.services.excel_design_system import ExcelDesignSystem
from meta.services.import_export_types import (
    get_type_order,
    _sanitize_xml_string,
    _safe_cell_value,
    _has_cud_actions,
    ExportResult,
    ImportPreview,
    ImportResult,
)


# [FIX 2026-06-08] 统一 logger（SSOT）
# 之前全文 20+ 处 `import logging; logger = logging.getLogger(__name__)` 散落各处。
# 现在一次性 hoist 到模块顶部，所有调用方直接用 `logger` 即可。
logger = logging.getLogger(__name__)



class ImportExportService:

    def __init__(self, data_source: DataSource, manage_service: Optional[ManageService] = None,
                 query_service: Optional[QueryService] = None):
        self.data_source = data_source
        self.manage_service = manage_service or ManageService(data_source)
        self.query_service = query_service or QueryService(data_source)
        self.hierarchy_filter = HierarchyFilterService(self.query_service, data_source)

    def _get_current_trace_id(self) -> Optional[str]:
        """获取当前请求的 trace_id（M.1 规范）

        用于在 ExportResult / ImportResult / ImportPreview 中携带 trace_id，
        让前端可以直接显示在错误提示中，便于用户报障时定位。

        Returns:
            当前 Flask 请求的 trace_id，无请求上下文时返回 None
        """
        try:
            from meta.services.trace_service import get_trace_id
            return get_trace_id()
        except Exception:
            return None

    def import_from_excel(self, object_type: str, file_path: str,
                          mapping: Optional[Dict[str, str]] = None) -> BatchOperationResult:
        """[DEPRECATED 2026-06-08] 旧的单表导入入口

        ⚠️ 此方法不推荐使用，功能不完整：
        - 不支持 operation_mode（"操作模式"列）
        - 不支持 parent_key_headers（父对象 FK 列被静默忽略）
        - 不支持 resolve_from_field / resolve_to_object（多态 FK 解析失败）
        - 不支持 conflict_strategy
        - 不注入 version_id context
        - 不走 _filter_import_record（readonly_always / immutable 字段被错误写入）

        推荐使用 import_cascade(file_path, mode, conflict_strategy)：
        - mode: 'preview' / 'execute' / 'validate'
        - conflict_strategy: 'skip' / 'upsert' / 'replace'
        - 支持 operation_mode、parent_key、resolve_from_field、context_field

        保留此方法仅用于向后兼容，内部已无任何调用方（API 全面走 import_cascade）。
        建议在 v3.20 移除。
        """
        import warnings
        warnings.warn(
            "import_from_excel is deprecated since v3.18, use import_cascade instead",
            DeprecationWarning,
            stacklevel=2
        )
        meta_obj = registry.get(object_type)
        if meta_obj is None:
            return BatchOperationResult(
                failed_count=1,
                errors=["Meta object not found: {0}".format(object_type)]
            )

        if not os.path.exists(file_path):
            return BatchOperationResult(
                failed_count=1,
                errors=["File not found: {0}".format(file_path)]
            )

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            return BatchOperationResult(
                failed_count=1,
                errors=["Failed to read Excel file: {0}".format(str(e))]
            )

        if len(rows) < 2:
            return BatchOperationResult(
                failed_count=1,
                errors=["Excel file is empty or has no data rows"]
            )

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]

        if mapping:
            field_map = mapping
        else:
            field_map = self._auto_map_fields(meta_obj, headers)

        data_list = []
        for row in rows[1:]:
            record = {}
            for col_idx, header in enumerate(headers):
                if header and header in field_map:
                    field_id = field_map[header]
                    value = row[col_idx] if col_idx < len(row) else None
                    meta_field = meta_obj.get_field(field_id)
                    if meta_field and value is not None:
                        value = self._convert_value(value, meta_field)
                    record[field_id] = value
            if record:
                data_list.append(record)

        if not data_list:
            return BatchOperationResult(
                failed_count=1,
                errors=["No valid data rows found after mapping"]
            )

        return self.manage_service.batch_create(object_type, data_list)

    def export_to_excel(self, object_type: str, filters: Optional[Dict[str, Any]] = None,
                        fields: Optional[List[str]] = None) -> str:
        """[DEPRECATED 2026-06-08] 旧的单表导出入口

        ⚠️ 此方法不推荐使用，功能不完整：
        - 不走 _get_export_headers_with_editable → 没有 parent FK 编码/名称列
        - 不写"操作模式"列 → 用户无法区分 create/update
        - 不加 BUSINESS_KEY_FILL / REQUIRED_FILL 等底色 → 视觉无提示
        - 不注入层级路径
        - 不支持 protect_sheet

        推荐使用 export_selected_types([object_type], filters, options)：
        - 自动生成 parent FK 列（编码+名称）
        - 自动添加操作模式列 + 颜色提示
        - 自动注入层级路径（如果 include_hierarchy_path=True）
        - 支持 protect_sheet 选项

        保留此方法仅用于向后兼容，内部已无任何调用方（API 全面走 export_selected_types）。
        建议在 v3.20 移除。
        """
        import warnings
        warnings.warn(
            "export_to_excel is deprecated since v3.18, use export_selected_types instead",
            DeprecationWarning,
            stacklevel=2
        )
        meta_obj = registry.get(object_type)
        if meta_obj is None:
            raise ValueError("Meta object not found: {0}".format(object_type))

        conditions = []
        if filters:
            for key, value in filters.items():
                conditions.append(QueryCondition(field=key, operator="eq", value=value))

        search_request = SearchRequest(
            object_type=object_type,
            conditions=conditions,
            page=1,
            page_size=100000,
        )

        search_result = self.query_service.search(search_request)
        data = search_result.data

        if not data:
            data = []

        if fields:
            export_fields = []
            for f_id in fields:
                mf = meta_obj.get_field(f_id)
                if mf:
                    export_fields.append(mf)
        else:
            export_fields = meta_obj.fields

        wb = Workbook()
        ws = wb.active
        ws.title = meta_obj.name

        header_row = []
        for mf in export_fields:
            header_row.append(mf.name or mf.id)
        ws.append(header_row)

        for record in data:
            row_data = []
            for mf in export_fields:
                value = record.get(mf.id) or record.get(mf.db_column)
                row_data.append(_safe_cell_value(value))
            ws.append(row_data)

        output_dir = os.path.join(os.getcwd(), "exports")
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = self._build_export_filename([object_type], timestamp)
        file_path = os.path.join(output_dir, file_name)

        wb.save(file_path)
        wb.close()

        return file_path

    def _auto_map_fields(self, meta_object: MetaObject, headers: List[str]) -> Dict[str, str]:
        field_map = {}
        field_by_id = {f.id: f for f in meta_object.fields}
        # [FIX 2026-06-16 BMRD] field_by_name 改为 list[(name, field)] 以避免
        # 多个字段共享同一中文名时 dict 互相覆盖 (例如 annotation 的
        # category 和 category_label 都叫 "备注分类", 之前 dict 写法会
        # 随机保留最后一个, 导致 "备注分类" 列被映射到 virtual 的
        # category_label 而非 stored 的 category, 数据丢失).
        field_by_name_list = []
        for f in meta_object.fields:
            field_by_name_list.append((f.name, f))
        field_by_alias = {}

        def _is_pure_display_virtual(f) -> bool:
            """判断是否为"纯展示虚拟字段" (类似 *_bo_name / *_bo_code)
            特征: storage=virtual, 有 redundancy.type=virtual
            """
            if not (f.storage and f.storage.value == 'virtual') and not getattr(f.semantics, 'virtual', False):
                return False
            redundancy = getattr(f.semantics, 'redundancy', None)
            if not redundancy:
                return False
            # [FIX 2026-06-16 BMRD] redundancy 是 dict, 必须用 .get() 而非 getattr()
            # getattr(dict, 'type', None) 返回 None (dict 没有 .type 属性)
            rtype = redundancy.get('type') if isinstance(redundancy, dict) else getattr(redundancy, 'type', None)
            if rtype != 'virtual':
                return False
            return True

        # [FIX 2026-06-16 BMRD] 预计算"resolve_from_field 源字段"集合
        # 这些字段存储的是用于解析 FK 的编码值, 导入时优先选它们
        # 场景: '源业务对象编码' 匹配到 virtual 的 source_bo_code (display)
        #       但应映射到 stored 的 source_code (FK 解析源)
        resolve_from_sources: set = set()
        # virtual_field_id -> related_stored_code_field_id 映射
        # 用途: 当按 name 匹配到 virtual display 字段时, 自动改用 stored code 字段
        # 例: source_bo_code (virtual display) -> source_code (stored resolve_from source)
        virtual_display_to_stored_code: Dict[str, str] = {}
        for f in meta_object.fields:
            resolve_from = getattr(f.semantics, 'resolve_from_field', None)
            if resolve_from:
                resolve_from_sources.add(resolve_from)
                # f 是 FK 字段 (如 source_bo_id), resolve_from 是其解析源 (如 source_code)
                # 找到所有跟 f 同 source_field 的 virtual display 字段
                fk_id = f.id
                stored_code_id = resolve_from
                for vf in meta_object.fields:
                    if vf.id == stored_code_id or vf.id == fk_id:
                        continue
                    if _is_pure_display_virtual(vf):
                        vf_redundancy = getattr(vf.semantics, 'redundancy', None)
                        # [FIX 2026-06-16 BMRD] vf_redundancy 是 dict, 用 .get() 而非 getattr()
                        vf_source_field = vf_redundancy.get('source_field') if isinstance(vf_redundancy, dict) else (getattr(vf_redundancy, 'source_field', None) if vf_redundancy else None)
                        # [FIX 2026-06-16 BMRD] 只映射 derived_from 以 .code 结尾的虚拟字段
                        # 避免把 source_sub_domain_id (derived_from=sub_domain.name)
                        # 也映射到 source_code (只有 source_bo_code derived_from=business_object.code 才应映射)
                        vf_derived_from = vf_redundancy.get('derived_from', '') if isinstance(vf_redundancy, dict) else ''
                        is_code_display = isinstance(vf_derived_from, str) and vf_derived_from.endswith('.code')
                        if vf_source_field == fk_id and is_code_display:
                            virtual_display_to_stored_code[vf.id] = stored_code_id

        for f in meta_object.fields:
            if f.semantics and f.semantics.aliases:
                for alias in f.semantics.aliases:
                    field_by_alias[alias] = f

        def _is_stored_code_field(f) -> bool:
            """判断是否为"stored code field" (类似 source_code)
            特征: 不是 virtual, 且被某个 FK 字段用作 resolve_from_field 源
            """
            if _is_pure_display_virtual(f):
                return False
            return f.id in resolve_from_sources

        def _pick_best_field(candidates, header: str = ''):
            """从同名字段中选最优先的:
            1. resolve_from_field 源字段 (stored, 用于 FK 解析) > 最高优先
            2. stored 字段 (非 virtual)
            3. import_visible=True
            4. field.id 字典序在前
            """
            if not candidates:
                return None
            def _score(f):
                is_virtual = (
                    (f.storage and f.storage.value == 'virtual')
                    or getattr(f.semantics, 'virtual', False)
                )
                import_visible = getattr(f.semantics, 'import_visible', True)
                is_resolve_source = f.id in resolve_from_sources
                # 数字越小越优先
                # resolve_from_source 字段给 -1 (最高优先)
                # stored + import_visible 给 0,0
                # stored + not import_visible 给 0,1
                # virtual + import_visible 给 1,0
                # virtual + not import_visible 给 1,1
                resolve_score = -1 if is_resolve_source else 0
                return (resolve_score, 1 if is_virtual else 0, 1 if not import_visible else 0, f.id)
            return sorted(candidates, key=_score)[0]

        def _resolve_pure_display_to_stored(picked_field) -> 'Field | None':
            """如果 picked_field 是"纯展示虚拟字段", 找其相关的 stored code 字段
            例: source_bo_code (virtual display) -> source_code (stored resolve_from)
            """
            if picked_field.id in virtual_display_to_stored_code:
                stored_id = virtual_display_to_stored_code[picked_field.id]
                stored_field = field_by_id.get(stored_id)
                if stored_field:
                    return stored_field
            return None

        # [FIX 2026-06-17] 构建 list view title -> field_id 映射
        # 导出时优先用 list view 的 title 作为列标题 (L1505),
        # 导入时也必须能按 list title 反向映射到 field_id
        # 例: relation_direction 的 name="关系方向", 但 list title="方向"
        #     导出列标题是"方向", 导入时按 name 匹配不到, 需要按 list title 匹配
        list_title_to_field_id: Dict[str, str] = {}
        if (meta_object.ui_view_config and meta_object.ui_view_config.list
                and meta_object.ui_view_config.list.columns):
            for col in meta_object.ui_view_config.list.columns:
                col_key = getattr(col, 'key', None)
                col_title = getattr(col, 'title', None)
                if col_key and col_title:
                    list_title_to_field_id[col_title] = col_key

        for header in headers:
            if not header:
                continue

            if header in field_by_id:
                # [FIX 2026-06-16 BMRD] header 也是 field id 时, 检查是否纯展示虚拟字段
                # 例: header='source_bo_code' (virtual display) -> 改用 source_code
                picked = field_by_id[header]
                stored_alt = _resolve_pure_display_to_stored(picked)
                if stored_alt:
                    field_map[header] = stored_alt.id
                else:
                    field_map[header] = header
                continue

            # 按 name 查找所有候选
            name_candidates = [f for (n, f) in field_by_name_list if n == header]
            if name_candidates:
                best = _pick_best_field(name_candidates, header)
                if best:
                    # [FIX 2026-06-16 BMRD] 纯展示虚拟字段 → 改用 stored code 字段
                    # 例: '源业务对象编码' -> source_bo_code (virtual) -> source_code (stored)
                    stored_alt = _resolve_pure_display_to_stored(best)
                    if stored_alt:
                        field_map[header] = stored_alt.id
                    else:
                        field_map[header] = best.id
                    continue

            if header in field_by_alias:
                field_map[header] = field_by_alias[header].id
                continue

            # [FIX 2026-06-17] 按 list view title 查找
            # 导出用 list title 作列标题, 导入必须能反向映射
            if header in list_title_to_field_id:
                field_id = list_title_to_field_id[header]
                field = field_by_id.get(field_id)
                if field:
                    stored_alt = _resolve_pure_display_to_stored(field)
                    if stored_alt:
                        field_map[header] = stored_alt.id
                    else:
                        field_map[header] = field_id
                    continue

            header_lower = header.lower().replace(" ", "_").replace("-", "_")
            for f in meta_object.fields:
                fid_lower = f.id.lower().replace(" ", "_").replace("-", "_")
                if header_lower == fid_lower:
                    field_map[header] = f.id
                    break
                fname_lower = (f.name or "").lower().replace(" ", "_").replace("-", "_")
                if header_lower == fname_lower:
                    field_map[header] = f.id
                    break

        return field_map

    def _get_enum_value_map_from_value_help(self, meta_field) -> Optional[Dict[str, str]]:
        """[FR-006] 委托到 enum_resolver.get_enum_map"""
        from meta.core.enum_resolver import get_enum_map
        return get_enum_map(meta_field, self.data_source)

    def _get_enum_type_id_from_value_help(self, meta_field) -> Optional[str]:
        """[FR-006] 委托到 enum_resolver.get_enum_type_id"""
        from meta.core.enum_resolver import get_enum_type_id
        return get_enum_type_id(meta_field)

    def _get_bo_display_map_from_value_help(self, meta_field, record_ids: List[Any]) -> Optional[Dict[Any, str]]:
        vh = self._get_value_help(meta_field)
        if not vh:
            return None

        source = getattr(vh, 'source', None)
        if not source or getattr(source, 'type', None) != 'bo':
            return None

        target_bo = getattr(source, 'target_bo', None)
        display_field = getattr(source, 'display_field', 'name') or 'name'

        if not target_bo or not record_ids:
            return None

        try:
            # [FR-010] 批量 SQL 查询替代逐条 BOEngine.get_record
            from meta import get_meta_object
            meta_obj = get_meta_object(target_bo)
            if not meta_obj:
                return None

            target_table = meta_obj.table_name or target_bo + 's'
            valid_ids = [rid for rid in record_ids if rid is not None]
            if not valid_ids:
                return None

            placeholders = ','.join(['?'] * len(valid_ids))
            sql = f"SELECT id, {display_field} FROM {target_table} WHERE id IN ({placeholders})"
            cursor = self.data_source.execute(sql, valid_ids)

            result_map = {}
            for row in cursor.fetchall():
                result_map[row[0]] = str(row[1]) if row[1] else ''
            return result_map if result_map else None
        except Exception as e:
            logger.warning(f"[ValueHelp] 获取BO显示名映射失败: {target_bo} - {e}")
        return None

    def _convert_value(self, value: Any, meta_field) -> Any:
        if value is None:
            return None

        from meta.core.models import FieldType

        try:
            enum_map = None
            if meta_field.enum_values:
                enum_map = {ev.get('value'): ev.get('label', ev.get('value')) for ev in meta_field.enum_values}
            elif not meta_field.enum_values:
                enum_map = self._get_enum_value_map_from_value_help(meta_field)

            if enum_map:
                valid_keys = set(enum_map.keys())
                label_to_key = {v: k for k, v in enum_map.items()}

                if isinstance(value, str):
                    if ' - ' in value:
                        key_part = value.split(' - ')[0].strip()
                        if key_part in valid_keys:
                            return key_part

                    if value in valid_keys:
                        return value

                    if value in label_to_key:
                        return label_to_key[value]

            if isinstance(value, str):
                vh_bo = self._get_value_help(meta_field)
                if vh_bo:
                    vh_source = getattr(vh_bo, 'source', None)
                    if vh_source and getattr(vh_source, 'type', None) == 'bo':
                        import re
                        # [FIX v1.2.15 2026-06-19] 支持多种 BO display 格式:
                        # 1) "name (id)"  e.g. "客户 (16)" → 16
                        m = re.search(r'\((\d+)\)\s*$', value)
                        if m:
                            return int(m.group(1))
                        # 2) "CODE - NAME"  e.g. "BO_CUSTOMER - 客户" → BO_CUSTOMER
                        #    返回 code (string), FK resolve loop 会用 lookup_index 转 id
                        if ' - ' in value:
                            code_part = value.split(' - ', 1)[0].strip()
                            if code_part:
                                return code_part
                        # 3) 纯 name / 纯 code → 原样返回, FK resolve loop 会查 lookup_index

            if meta_field.field_type == FieldType.INTEGER:
                return int(value)
            elif meta_field.field_type == FieldType.FLOAT:
                return float(value)
            elif meta_field.field_type == FieldType.BOOLEAN:
                if isinstance(value, str):
                    return value.lower() in ("true", "1", "yes")
                return bool(value)
            elif meta_field.field_type == FieldType.DATETIME:
                if isinstance(value, datetime):
                    return value.isoformat()
                return str(value)
            return value
        except (ValueError, TypeError):
            return value

    def export_template(self, selected_types: List[str], options: Optional[Dict[str, Any]] = None) -> ExportResult:
        """
        生成导入模板（只包含表头，不包含数据）
        
        Args:
            selected_types: 选定的对象类型列表
            options: 导出选项
        
        Returns:
            ExportResult: 导出结果
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.comments import Comment
        from meta.core.models import registry

        options = options or {}
        include_operation_mode = options.get("include_operation_mode", True)

        # [FR-008] 使用 HierarchyConfigLoader.sort_by_hierarchy 替代 get_type_order
        from meta.services.cascade_service import HierarchyConfigLoader
        all_types = [level.get('object') for level in HierarchyConfigLoader.get_levels('biz_hierarchy') if level.get('object')]
        type_order = HierarchyConfigLoader.sort_by_hierarchy(all_types)
        type_order = self._ensure_association_types_in_order(type_order)

        # 检查selected_types是否在registry中注册
        from meta.core.models import registry
        valid_types = [t for t in selected_types if registry.get(t)]

        # 如果selected_types不在type_order中，直接使用valid_types
        ordered_types = [t for t in type_order if t in valid_types]
        if not ordered_types:
            # 如果没有在type_order中，直接使用valid_types
            ordered_types = valid_types

        if not ordered_types:
            return ExportResult(success=False, errors=["No valid object types selected"])

        # 检查是否有任何对象支持 CUD 操作
        has_cud = any(_has_cud_actions(registry.get(ot)) for ot in ordered_types if registry.get(ot))
        
        # 检查是否所有对象都是只读的
        all_readonly = all(not _has_cud_actions(registry.get(ot)) for ot in ordered_types if registry.get(ot))
        
        wb = Workbook()
        ws_meta = wb.active
        ws_meta.title = "说明"

        ds = ExcelDesignSystem

        # [SSOT 2026-06-08] 使用 _write_meta_sheet_header / _write_meta_sheet_operations / _finalize_meta_sheet
        included_names = [registry.get(ot).name for ot in ordered_types if registry.get(ot)]
        self._write_meta_sheet_header(
            ws_meta,
            title="导入模板说明",
            time_str=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            included_names=included_names,
        )
        # export_template 没有上下文信息，直接从 row 5 开始操作说明
        last_row = self._write_meta_sheet_operations(
            ws_meta, has_cud=has_cud, all_readonly=all_readonly, start_row=5,
            is_cascade=False, has_child_sheets=False,
            objects=ordered_types,
        )
        self._finalize_meta_sheet(ws_meta, last_row)

        header_fill = ds.HEADER_FILL
        header_font = ds.HEADER_FONT
        
        sheets_info = []
        
        for ot in ordered_types:
            obj = registry.get(ot)
            if obj is None:
                continue
            
            if not obj.import_export.export_enabled:
                continue
            
            ws = wb.create_sheet(title=obj.name)
            
            # [NEW v1.2.3 2026-06-17] 主导出按 'create' 模式处理 (模板可填父对象)
            headers, editable_columns, readonly_columns, parent_key_columns, create_required_columns, header_comments, header_to_field, enum_value_maps, bo_display_fields, fk_display_code_columns = self._get_export_headers_with_editable(obj, options, mode='create')
            
            bo_display_maps = {}
            
            obj_has_cud = _has_cud_actions(obj)
            
            if obj_has_cud:
                headers.insert(0, "操作模式")
                header_comments.insert(0, "create - 新增/update - 更新/delete - 删除，留空默认为update")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = ds.TEXT_CENTER
                cell.border = ds.THIN_BORDER
                if col_idx - 1 < len(header_comments) and header_comments[col_idx - 1]:
                    cell.comment = Comment(_sanitize_xml_string(header_comments[col_idx - 1]), "系统")
            
            if obj_has_cud:
                op_mode_validation = self._create_operation_mode_dv(ws)
            
            col_offset = 1 if include_operation_mode else 0
            enum_validations = {}
            for col_idx, header in enumerate(headers[col_offset:], 1 + col_offset):
                actual_col_idx = col_idx - 1 - col_offset
                field_id = header_to_field.get(header, header)
                
                if field_id in enum_value_maps:
                    enum_map = enum_value_maps[field_id]
                    display_values = [f"{k} - {v}" for k, v in enum_map.items()]
                    if display_values:
                        col_letter = get_column_letter(col_idx)
                        dv = DataValidation(
                            type="list",
                            formula1='"' + ','.join(display_values) + '"',
                            allow_blank=True
                        )
                        dv.error = f"请从下拉列表中选择有效的{header}"
                        dv.errorTitle = "无效输入"
                        dv.prompt = f"请选择{header} (Key - Label)"
                        dv.promptTitle = header
                        ws.add_data_validation(dv)
                        enum_validations[actual_col_idx] = (dv, display_values, list(enum_map.keys()))
            
            empty_rows_count = options.get("empty_rows_for_new", 5)
            for row_idx in range(2, 2 + empty_rows_count):
                if obj_has_cud:
                    op_cell = ws.cell(row=row_idx, column=1, value="create - 新增")
                    # [REMOVED 2026-06-14 BMRD] CREATE_NEW_FILL 已删除
                    # 替代: 保留"create - 新增"文字作为视觉提示
                    op_cell.fill = ds.READONLY_FILL  # 复用灰色 (只读, 表示空行待填)
                    op_cell.border = ds.THIN_BORDER
                    op_cell.alignment = ds.TEXT_CENTER
                    op_mode_validation.add(op_cell)
                    col_offset = 1
                else:
                    col_offset = 0
                
                for col_idx, header in enumerate(headers[col_offset:], 1 + col_offset):
                    cell = ws.cell(row=row_idx, column=col_idx, value="")
                    cell.border = ds.THIN_BORDER
                    
                    actual_col_idx = col_idx - 1 - col_offset
                    if actual_col_idx in enum_validations:
                        dv, display_values, keys = enum_validations[actual_col_idx]
                        dv.add(cell)
                    
                    if actual_col_idx in parent_key_columns:
                        self._apply_classification_fill(cell, 'parent_key')
                    elif actual_col_idx in fk_display_code_columns:
                        # [NEW 2026-06-16 BMRD] FK 编码显示字段 - 浅绿色
                        self._apply_classification_fill(cell, 'fk_display_code')
                    elif actual_col_idx in create_required_columns:
                        self._apply_classification_fill(cell, 'create_required')
                    elif actual_col_idx in readonly_columns:
                        self._apply_classification_fill(cell, 'readonly')
                    # [NEW v1.1 2026-06-11] auto_or_manual_code 差异化底色
                    # [FIX 2026-06-16 BMRD] meta_obj 不存在, 用当前循环的 obj 替代 (L448 registry.get 返回)
                    auto_cols = (getattr(self, '_auto_or_manual_code_columns', {}) or {}).get(obj.name, [])
                    if actual_col_idx in auto_cols:
                        self._apply_classification_fill(cell, 'auto_or_manual_code')
            
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                max_length = len(str(header))
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            sheets_info.append({
                "name": obj.name,
                "row_count": 0
            })
        
        file_name = f"import_template_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        export_dir = os.path.join(os.getcwd(), "exports")
        os.makedirs(export_dir, exist_ok=True)
        file_path = os.path.join(export_dir, file_name)
        
        wb.save(file_path)
        
        return ExportResult(
            success=True,
            file_path=file_path,
            sheets=sheets_info,
            total_rows=0,
            trace_id=self._get_current_trace_id(),
        )

    def export_selected_types(self, selected_types: List[str], filters: Optional[Dict[str, Any]] = None,
                              options: Optional[Dict[str, Any]] = None, sort_by: str = None, 
                              sort_order: str = 'asc',
                              progress_callback: Optional[callable] = None,
                              page: int = None, page_size: int = None) -> ExportResult:
        """
        导出选定的对象类型
        
        Args:
            selected_types: 选定的对象类型列表
            filters: 筛选条件
            options: 导出选项
            sort_by: 排序字段
            sort_order: 排序方向 (asc/desc)
            progress_callback: 进度回调函数，接收 dict:
                {progress, current_type, current_type_name, total_types, current_index, message}
            page: 页码（分页导出时使用）
            page_size: 每页数量（分页导出时使用）
        
        Returns:
            ExportResult: 导出结果
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Protection, Border, Side
        from openpyxl.worksheet.protection import SheetProtection
        from openpyxl.comments import Comment
        from meta.core.models import registry

        options = options or {}
        include_readonly = options.get("include_readonly", True)
        include_operation_mode = options.get("include_operation_mode", True)
        protect_sheet = options.get("protect_sheet", False)
        include_annotations = options.get("include_child_objects",
                                           options.get("include_annotations", True))

        # [FR-008] 使用 HierarchyConfigLoader.sort_by_hierarchy 替代 get_type_order
        from meta.services.cascade_service import HierarchyConfigLoader
        all_types = [level.get('object') for level in HierarchyConfigLoader.get_levels('biz_hierarchy') if level.get('object')]
        type_order = HierarchyConfigLoader.sort_by_hierarchy(all_types)
        type_order = self._ensure_association_types_in_order(type_order)

        # 检查selected_types是否在registry中注册
        from meta.core.models import registry
        valid_types = [t for t in selected_types if registry.get(t)]

        # 如果selected_types不在type_order中，直接使用valid_types
        ordered_types = [t for t in type_order if t in valid_types]
        if not ordered_types:
            # 如果没有在type_order中，直接使用valid_types
            ordered_types = valid_types

        if not ordered_types:
            return ExportResult(success=False, errors=["No valid object types selected"])

        wb = Workbook()
        ws_meta = wb.active
        ws_meta.title = "说明"

        ds = ExcelDesignSystem

        # [SSOT 2026-06-08] 标题（export_selected_types 用较小字号 11）
        ws_meta.cell(row=1, column=1, value="导出信息").font = Font(bold=True, size=11, color=ds.PRIMARY_COLOR)
        ws_meta.row_dimensions[1].height = 20

        product_code, version_code = self._get_product_version_codes(filters)
        product_name, version_name = self._get_product_version_info(filters)

        try:
            from flask import request as flask_request
            export_user = flask_request.headers.get('X-User-Name', '') if flask_request else ''
        except Exception:
            export_user = ''

        # [SSOT 2026-06-08] 自定义 header（包含 导出用户 + 导出范围 + 包含对象）
        # 替代了 _write_meta_sheet_header 的标准 3 行，加入 export_user 行
        ws_meta.cell(row=2, column=1, value="导出时间").font = ds.LABEL_FONT
        ws_meta.cell(row=2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")).font = ds.VALUE_FONT
        ws_meta.cell(row=3, column=1, value="导出用户").font = ds.LABEL_FONT
        ws_meta.cell(row=3, column=2, value=export_user).font = ds.VALUE_FONT
        ws_meta.cell(row=4, column=1, value="导出范围").font = ds.LABEL_FONT
        ws_meta.cell(row=4, column=2, value="选定对象导出").font = ds.VALUE_FONT
        ws_meta.cell(row=5, column=1, value="包含对象").font = ds.LABEL_FONT
        included_names = [registry.get(ot).name for ot in ordered_types if registry.get(ot)]
        ws_meta.cell(row=5, column=2, value=", ".join(included_names)).font = ds.VALUE_FONT

        # 上下文信息（仅 selected_types 模式有）
        ws_meta.cell(row=7, column=1, value="上下文信息").font = ds.SECTION_FONT
        ws_meta.cell(row=7, column=1).fill = ds.SECTION_FILL
        ws_meta.cell(row=8, column=1, value="产品编码").font = ds.LABEL_FONT
        ws_meta.cell(row=8, column=2, value=product_code).font = ds.VALUE_FONT
        ws_meta.cell(row=9, column=1, value="产品名称").font = ds.LABEL_FONT
        ws_meta.cell(row=9, column=2, value=product_name).font = ds.VALUE_FONT
        ws_meta.cell(row=10, column=1, value="版本编码").font = ds.LABEL_FONT
        ws_meta.cell(row=10, column=2, value=version_code).font = ds.VALUE_FONT
        ws_meta.cell(row=11, column=1, value="版本名称").font = ds.LABEL_FONT
        ws_meta.cell(row=11, column=2, value=version_name).font = ds.VALUE_FONT
        ws_meta.cell(row=12, column=1, value="版本ID").font = ds.LABEL_FONT
        ws_meta.cell(row=12, column=2, value=str(filters.get('version_id', '')) if filters else '').font = ds.VALUE_FONT

        # 检查是否有任何对象支持 CUD 操作
        has_cud = any(_has_cud_actions(registry.get(ot)) for ot in ordered_types if registry.get(ot))

        # 检查是否所有对象都是只读的
        all_readonly = all(not _has_cud_actions(registry.get(ot)) for ot in ordered_types if registry.get(ot))

        # [SSOT 2026-06-08] 使用 _write_meta_sheet_operations（从 row 14 开始）
        last_row = self._write_meta_sheet_operations(
            ws_meta, has_cud=has_cud, all_readonly=all_readonly, start_row=14,
            is_cascade=False, has_child_sheets=True,
            objects=selected_types,
        )
        self._finalize_meta_sheet(ws_meta, last_row)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        sheets_info = []
        total_rows = 0
        
        total_types = len(ordered_types)
        if progress_callback:
            progress_callback({
                'progress': 0,
                'current_type': '',
                'current_type_name': '',
                'total_types': total_types,
                'current_index': 0,
                'message': '开始导出，共 {0} 种对象类型'.format(total_types)
            })
        
        for i, ot in enumerate(ordered_types):
            obj = registry.get(ot)
            if obj is None:
                continue
            
            if not obj.import_export.export_enabled:
                continue
            
            type_name = obj.name or ot
            current_index = i + 1
            if progress_callback:
                progress_callback({
                    'progress': int((i / total_types) * 100),
                    'current_type': ot,
                    'current_type_name': type_name,
                    'total_types': total_types,
                    'current_index': current_index,
                    'message': '正在导出: {0} ({1}/{2})'.format(type_name, current_index, total_types)
                })
            
            sheet_data = self._query_with_hierarchy(ot, filters, options, sort_by=sort_by, sort_order=sort_order, page=page, page_size=page_size)
            
            ws = wb.create_sheet(title=obj.name)
            
            # [NEW v1.2.3 2026-06-17] 选中对象导出按 'create' 模式处理 (模板可填父对象)
            headers, editable_columns, readonly_columns, parent_key_columns, create_required_columns, header_comments, header_to_field, enum_value_maps, bo_display_fields, fk_display_code_columns = self._get_export_headers_with_editable(obj, options, mode='create')
            
            bo_display_maps = self._build_bo_display_maps(sheet_data, bo_display_fields)

            # 根据 include_operation_mode 选项决定是否包含操作模式列
            if include_operation_mode:
                headers.insert(0, "操作模式")
                header_comments.insert(0, "新增/更新/删除/跳过，留空默认为更新")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = ds.THIN_BORDER
                if col_idx - 1 < len(header_comments) and header_comments[col_idx - 1]:
                    cell.comment = Comment(_sanitize_xml_string(header_comments[col_idx - 1]), "系统")
            
            if include_operation_mode:
                op_mode_validation = self._create_operation_mode_dv(ws)
            
            col_offset = 1 if include_operation_mode else 0

            sheet_row_count = len(sheet_data)
            progress_interval = max(100, sheet_row_count // 10) if sheet_row_count > 0 else 100

            for row_idx, record in enumerate(sheet_data, 2):
                if sheet_row_count > 500 and (row_idx - 2) % progress_interval == 0:
                    progress = (row_idx - 2) * 100 // sheet_row_count
                    print(f"[Export] 正在导出 {obj.name}，进度 {progress}% ({row_idx - 2}/{sheet_row_count})")

                if include_operation_mode:
                    op_cell = ws.cell(row=row_idx, column=1, value="update - 更新")
                    op_cell.fill = ds.READONLY_FILL
                    op_cell.border = ds.THIN_BORDER
                    if protect_sheet:
                        op_cell.protection = Protection(locked=False)
                    op_mode_validation.add(op_cell)

                for col_idx, header in enumerate(headers[col_offset:], 1 + col_offset):
                    field_id = header_to_field.get(header, header)
                    value = record.get(field_id) or record.get(header)

                    if field_id in enum_value_maps and value is not None:
                        label = enum_value_maps[field_id].get(value, value)
                        value = f"{value} - {label}" if label != value else value
                    
                    if field_id in bo_display_maps and value is not None:
                        display_name = bo_display_maps[field_id].get(value)
                        if display_name:
                            value = f"{display_name} ({value})"

                    cell = ws.cell(row=row_idx, column=col_idx, value=_safe_cell_value(value))
                    cell.border = ds.THIN_BORDER

                    actual_col_idx = col_idx - 1 - col_offset
                    # [NEW v1.1 2026-06-11] auto_or_manual_code 差异化底色
                    auto_cols = (getattr(self, '_auto_or_manual_code_columns', {}) or {}).get(obj.name, [])
                    if actual_col_idx in auto_cols:
                        self._apply_classification_fill(cell, 'auto_or_manual_code')
                        if protect_sheet:
                            cell.protection = Protection(locked=False)
                    elif actual_col_idx in parent_key_columns:
                        self._apply_classification_fill(cell, 'parent_key')
                        if protect_sheet:
                            cell.protection = Protection(locked=False)
                    elif actual_col_idx in fk_display_code_columns:
                        # [NEW 2026-06-16 BMRD] FK 编码显示字段 - 浅绿色, 不锁
                        self._apply_classification_fill(cell, 'fk_display_code')
                        if protect_sheet:
                            cell.protection = Protection(locked=True)
                    elif actual_col_idx in create_required_columns:
                        self._apply_classification_fill(cell, 'create_required')
                        if protect_sheet:
                            cell.protection = Protection(locked=False)
                    elif actual_col_idx in readonly_columns:
                        self._apply_classification_fill(cell, 'readonly')
                        if protect_sheet:
                            cell.protection = Protection(locked=True)
                    elif actual_col_idx in editable_columns:
                        if protect_sheet:
                            cell.protection = Protection(locked=False)

            print(f"[Export] {obj.name} 数据写入完成，共 {sheet_row_count} 行")
            
            export_enum_validations = {}  # key: actual_col_idx (不含offset)，用于新增行
            export_enum_validations_by_col = {}  # key: col_idx (含offset)，用于数据行
            for col_idx, header in enumerate(headers[col_offset:] if include_operation_mode else headers, 1 + col_offset if include_operation_mode else 1):
                actual_col_idx = col_idx - 1 - col_offset if include_operation_mode else col_idx - 1
                field_id = header_to_field.get(header, header)

                if field_id in enum_value_maps:
                    enum_map = enum_value_maps[field_id]
                    display_values = [f"{k} - {v}" for k, v in enum_map.items()]
                    if display_values:
                        dv = DataValidation(
                            type="list",
                            formula1='"' + ','.join(display_values) + '"',
                            allow_blank=True
                        )
                        dv.error = f"请从下拉列表中选择有效的{header}"
                        dv.errorTitle = "无效输入"
                        dv.prompt = f"请选择{header} (Key - Label)"
                        dv.promptTitle = header
                        ws.add_data_validation(dv)
                        # [FIX v3.18] 双索引: actual_col_idx (用于新增行) + col_idx (用于数据行)
                        export_enum_validations[actual_col_idx] = dv
                        export_enum_validations_by_col[col_idx] = dv

            for row_idx in range(2, len(sheet_data) + 2):
                for col_idx in export_enum_validations_by_col:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    export_enum_validations_by_col[col_idx].add(cell)

            # [FIX] 有数据时不追加新增行，无数据时添加空白新增行 (v3.18)
            if len(sheet_data) > 0:
                empty_rows_count = 0
            else:
                empty_rows_count = options.get("empty_rows_for_new", 5)
            next_row_idx = len(sheet_data) + 2
            for empty_row in range(empty_rows_count):
                col_idx = 1

                if include_operation_mode:
                    cell = ws.cell(row=next_row_idx, column=col_idx, value="create - 新增")
                    # [REMOVED 2026-06-14 BMRD] CREATE_NEW_FILL 已删除
                    cell.fill = ds.READONLY_FILL  # 复用灰色
                    cell.border = ds.THIN_BORDER
                    cell.alignment = Alignment(horizontal="center")
                    if protect_sheet:
                        cell.protection = Protection(locked=False)
                    if op_mode_validation:
                        op_mode_validation.add(cell)
                    col_idx += 1

                for header in headers[1:] if include_operation_mode else headers:
                    cell = ws.cell(row=next_row_idx, column=col_idx, value="")
                    cell.border = ds.THIN_BORDER

                    original_col_idx = col_idx - 2 if include_operation_mode else col_idx - 1
                    # [NEW v1.1 2026-06-11] auto_or_manual_code 差异化底色
                    auto_cols = (getattr(self, '_auto_or_manual_code_columns', {}) or {}).get(meta_obj.name, [])
                    if original_col_idx in auto_cols:
                        self._apply_classification_fill(cell, 'auto_or_manual_code')
                    elif original_col_idx in parent_key_columns:
                        self._apply_classification_fill(cell, 'parent_key')
                    elif original_col_idx in fk_display_code_columns:
                        # [NEW 2026-06-16 BMRD] FK 编码显示字段 - 浅绿色
                        self._apply_classification_fill(cell, 'fk_display_code')
                    elif original_col_idx in create_required_columns:
                        self._apply_classification_fill(cell, 'create_required')
                    elif original_col_idx in readonly_columns:
                        self._apply_classification_fill(cell, 'readonly')

                    if original_col_idx in export_enum_validations:
                        export_enum_validations[original_col_idx].add(cell)
                    col_idx += 1
                next_row_idx += 1

            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                max_length = len(str(header))
                for row in range(2, len(sheet_data) + 2):
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            if protect_sheet:
                ws.protection = SheetProtection(
                    sheet=True,
                    autoFilter=False,
                    deleteColumns=False,
                    deleteRows=False,
                    formatCells=True,
                    formatColumns=True,
                    formatRows=True,
                    insertColumns=False,
                    insertRows=True,
                    sort=True,
                    objects=False,
                    scenarios=False,
                    pivotTables=False
                )
            
            sheets_info.append({
                "name": obj.name,
                "object_type": ot,
                "row_count": len(sheet_data) if sheet_data else 0
            })
            total_rows += len(sheet_data) if sheet_data else 0
        
        if progress_callback:
            progress_callback({
                'progress': 100,
                'current_type': '',
                'current_type_name': '',
                'total_types': total_types,
                'current_index': total_types,
                'message': '导出完成'
            })
        
        if include_annotations:
            child_parent_map = self._collect_child_object_types(ordered_types)
            if child_parent_map:
                total_child_types = len(child_parent_map)
                for idx, (child_type_name, parent_list) in enumerate(child_parent_map.items()):
                    # 跳过已在主导出中处理的类型，避免重复 sheet
                    if child_type_name in ordered_types:
                        continue
                    if progress_callback:
                        try:
                            child_reg_meta = registry.get(child_type_name)
                            child_display = child_reg_meta.name if child_reg_meta else child_type_name
                        except Exception:
                            child_display = child_type_name
                        progress_callback({
                            'progress': 95 + int((idx / total_child_types) * 5),
                            'current_type': child_type_name,
                            'current_type_name': child_display,
                            'total_types': total_child_types,
                            'current_index': idx + 1,
                            'message': '正在导出子对象: {0}'.format(child_display)
                        })

                    try:
                        child_data = self._query_child_object(child_type_name, parent_list, filters)
                        child_meta = registry.get(child_type_name)
                        if child_meta:
                            self._write_child_sheet(wb, child_type_name, child_meta, child_data or [], sheets_info, options)
                            total_rows += len(child_data) if child_data else 0
                    except Exception as e:
                        logger.warning(
                            f"[Export] 子对象 {child_type_name} 导出失败: {e}"
                        )
        
        output_dir = os.path.join(os.getcwd(), "exports")
        os.makedirs(output_dir, exist_ok=True)
        
        product_name, version_name = self._get_product_version_info(filters)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = self._build_export_filename([product_name, version_name], timestamp)
        file_path = os.path.join(output_dir, file_name)
        
        wb.save(file_path)
        wb.close()
        
        return ExportResult(
            success=True,
            file_path=file_path,
            sheets=sheets_info,
            total_rows=total_rows,
            trace_id=self._get_current_trace_id(),
        )

    def export_cascade(self, object_type: str, filters: Optional[Dict[str, Any]] = None,
                       options: Optional[Dict[str, Any]] = None, sort_by: str = None, 
                       sort_order: str = 'asc',
                       page: int = None, page_size: int = None) -> ExportResult:
        """
        级联导出：导出指定对象及其所有子级对象
        
        参考SAP SuccessFactors导入模板方案：
        1. 单元格级别保护：只读字段使用灰色背景+锁定
        2. 操作模式列：使用下拉列表数据验证
        3. 工作表保护：允许编辑可编辑单元格
        
        Args:
            object_type: 起始对象类型
            filters: 筛选条件
            options: 导出选项
                - include_hierarchy_path: 是否包含层级路径列
                - include_hierarchy_ids: 是否包含层级ID列
                - include_metadata_sheet: 是否包含元数据Sheet
                - include_readonly: 是否标记只读字段
                - include_operation_mode: 是否包含操作模式列
                - protect_sheet: 是否保护工作表
            sort_by: 排序字段
            sort_order: 排序方向 (asc/desc)
            page: 页码（分页导出时使用）
            page_size: 每页数量（分页导出时使用）
        
        Returns:
            ExportResult: 导出结果
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Protection, Border, Side
        from openpyxl.worksheet.protection import SheetProtection
        from openpyxl.comments import Comment

        options = options or {}
        include_readonly = options.get("include_readonly", True)
        include_operation_mode = options.get("include_operation_mode", True)
        protect_sheet = options.get("protect_sheet", False)

        ds = ExcelDesignSystem

        meta_obj = registry.get(object_type)
        if meta_obj is None:
            return ExportResult(success=False, errors=["Meta object not found: {0}".format(object_type)])
        
        if meta_obj.import_export and not meta_obj.import_export.cascade_export:
            return self.export_selected_types([object_type], filters, options, page=page, page_size=page_size)

        object_types = self._get_cascade_object_types(object_type)
        ordered_types = self._sort_by_hierarchy(object_types)
        
        wb = Workbook()
        ws_meta = wb.active
        ws_meta.title = "说明"

        meta_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        section_font = Font(bold=True, size=11, color="1565C0")

        # [SSOT 2026-06-08] 自定义 header（cascade 模式：导出范围 + 起始对象 + 包含对象）
        ws_meta.cell(row=1, column=1, value="导出信息").font = Font(bold=True, size=14, color="1565C0")
        ws_meta.row_dimensions[1].height = 25
        ws_meta.cell(row=2, column=1, value="导出时间").font = ds.LABEL_FONT
        ws_meta.cell(row=2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")).font = ds.VALUE_FONT
        ws_meta.cell(row=3, column=1, value="导出范围").font = ds.LABEL_FONT
        ws_meta.cell(row=3, column=2, value="级联导出").font = ds.VALUE_FONT
        ws_meta.cell(row=4, column=1, value="起始对象").font = ds.LABEL_FONT
        ws_meta.cell(row=4, column=2, value=meta_obj.name).font = ds.VALUE_FONT
        ws_meta.cell(row=5, column=1, value="包含对象").font = ds.LABEL_FONT
        exclude_object_types = {'version', 'product'}
        included_objects = [registry.get(ot).name for ot in ordered_types
                           if registry.get(ot) and ot not in exclude_object_types]
        ws_meta.cell(row=5, column=2, value=", ".join(included_objects)).font = ds.VALUE_FONT

        # [SSOT 2026-06-08] 使用 _write_meta_sheet_operations（从 row 7 开始，cascade 模式）
        last_row = self._write_meta_sheet_operations(
            ws_meta, has_cud=True, all_readonly=False, start_row=7,
            is_cascade=True, has_child_sheets=False,
            objects=ordered_types,
        )
        self._finalize_meta_sheet(ws_meta, last_row)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        sheets_info = []
        total_rows = 0
        
        for ot in ordered_types:
            if ot in exclude_object_types:
                continue
            
            obj = registry.get(ot)
            if obj is None:
                continue
            
            if not obj.import_export.export_enabled:
                continue
            
            sheet_data = self._query_with_hierarchy(ot, filters, options, sort_by=sort_by, sort_order=sort_order, page=page, page_size=page_size)
            
            ws = wb.create_sheet(title=obj.name)
            
            # [NEW v1.2.3 2026-06-17] 模板导出按 'create' 模式处理 (模板就是用来填的)
            headers, editable_columns, readonly_columns, parent_key_columns, create_required_columns, header_comments, header_to_field, enum_value_maps, bo_display_fields, fk_display_code_columns = self._get_export_headers_with_editable(obj, options, mode='create')
            
            bo_display_maps = self._build_bo_display_maps(sheet_data, bo_display_fields)
            
            if include_operation_mode:
                headers.insert(0, "操作模式")
                header_comments.insert(0, "新增/更新/删除/跳过，留空默认为更新")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = ds.THIN_BORDER
                if col_idx - 1 < len(header_comments) and header_comments[col_idx - 1]:
                    cell.comment = Comment(_sanitize_xml_string(header_comments[col_idx - 1]), "系统")
            
            if include_operation_mode:
                operation_dv = self._create_operation_mode_dv(ws, with_prompt=True, verbose_error=True)
            
            row_idx = 2
            for record in (sheet_data or []):
                col_idx = 1
                
                if include_operation_mode:
                    cell = ws.cell(row=row_idx, column=col_idx, value="update - 更新")
                    cell.fill = light_blue_fill
                    cell.border = ds.THIN_BORDER
                    cell.alignment = Alignment(horizontal="center")
                    if protect_sheet:
                        cell.protection = Protection(locked=False)
                    operation_dv.add(cell)
                    col_idx += 1
                
                for header in headers[1:] if include_operation_mode else headers:
                    field_id = header_to_field.get(header, header)
                    value = record.get(field_id) or record.get(header)
                    
                    if field_id in enum_value_maps and value is not None:
                        label = enum_value_maps[field_id].get(value, value)
                        value = f"{value} - {label}" if label != value else value
                    
                    if field_id in bo_display_maps and value is not None:
                        display_name = bo_display_maps[field_id].get(value)
                        if display_name:
                            value = f"{display_name} ({value})"
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = ds.THIN_BORDER
                    
                    original_col_idx = col_idx - 2 if include_operation_mode else col_idx - 1

                    # [NEW v1.1 2026-06-11] auto_or_manual_code 差异化底色
                    auto_cols = (getattr(self, '_auto_or_manual_code_columns', {}) or {}).get(meta_obj.name, [])
                    if original_col_idx in auto_cols:
                        self._apply_classification_fill(cell, 'auto_or_manual_code')
                        if protect_sheet:
                            cell.protection = Protection(locked=False)
                    elif original_col_idx in parent_key_columns:
                        self._apply_classification_fill(cell, 'parent_key')
                        if protect_sheet:
                            cell.protection = Protection(locked=False)
                    elif original_col_idx in fk_display_code_columns:
                        # [NEW 2026-06-16 BMRD] FK 编码显示字段
                        self._apply_classification_fill(cell, 'fk_display_code')
                        if protect_sheet:
                            cell.protection = Protection(locked=True)
                    elif original_col_idx in create_required_columns:
                        self._apply_classification_fill(cell, 'create_required')
                        if protect_sheet:
                            cell.protection = Protection(locked=False)
                    elif original_col_idx in readonly_columns:
                        self._apply_classification_fill(cell, 'readonly')
                        if protect_sheet:
                            cell.protection = Protection(locked=True)
                    else:
                        if protect_sheet:
                            cell.protection = Protection(locked=False)

                    col_idx += 1
                row_idx += 1

            export_enum_validations = {}  # key: actual_col_idx (不含offset)，用于新增行
            export_enum_validations_by_col = {}  # key: col_idx (含offset)，用于数据行
            for col_idx, header in enumerate(headers[1:] if include_operation_mode else headers, 2 if include_operation_mode else 1):
                # [FIX v3.18] actual_col_idx 不含操作模式列偏移，与新增行循环保持一致
                actual_col_idx = col_idx - 1 - (1 if include_operation_mode else 0)
                field_id = header_to_field.get(header, header)

                if field_id in enum_value_maps:
                    enum_map = enum_value_maps[field_id]
                    display_values = [f"{k} - {v}" for k, v in enum_map.items()]
                    if display_values:
                        dv = DataValidation(
                            type="list",
                            formula1='"' + ','.join(display_values) + '"',
                            allow_blank=True
                        )
                        dv.error = f"请从下拉列表中选择有效的{header}"
                        dv.errorTitle = "无效输入"
                        dv.prompt = f"请选择{header} (Key - Label)"
                        dv.promptTitle = header
                        ws.add_data_validation(dv)
                        export_enum_validations[actual_col_idx] = dv
                        export_enum_validations_by_col[col_idx] = dv

            for row in range(2, row_idx):
                for col_idx in export_enum_validations_by_col:
                    cell = ws.cell(row=row, column=col_idx)
                    export_enum_validations_by_col[col_idx].add(cell)
            
            # [FIX] 有数据时不追加新增行，无数据时添加空白新增行 (v3.18)
            if len(sheet_data) > 0:
                empty_rows_count = 0
            else:
                empty_rows_count = options.get("empty_rows_for_new", 5)
            for empty_row in range(empty_rows_count):
                col_idx = 1
                
                if include_operation_mode:
                    cell = ws.cell(row=row_idx, column=col_idx, value="create - 新增")
                    # [REMOVED 2026-06-14 BMRD] CREATE_NEW_FILL 已删除
                    cell.fill = ds.READONLY_FILL  # 复用灰色
                    cell.border = ds.THIN_BORDER
                    cell.alignment = Alignment(horizontal="center")
                    if protect_sheet:
                        cell.protection = Protection(locked=False)
                    operation_dv.add(cell)
                    col_idx += 1
                
                for header in headers[1:] if include_operation_mode else headers:
                    cell = ws.cell(row=row_idx, column=col_idx, value="")
                    cell.border = ds.THIN_BORDER
                    
                    original_col_idx = col_idx - 2 if include_operation_mode else col_idx - 1

                    # [NEW v1.1 2026-06-11] auto_or_manual_code 差异化底色
                    auto_cols = (getattr(self, '_auto_or_manual_code_columns', {}) or {}).get(meta_obj.name, [])
                    if original_col_idx in auto_cols:
                        self._apply_classification_fill(cell, 'auto_or_manual_code')
                        if protect_sheet:
                            cell.protection = Protection(locked=False)
                    elif original_col_idx in parent_key_columns:
                        self._apply_classification_fill(cell, 'parent_key')
                        cell.comment = Comment(_sanitize_xml_string("新增时必填：请填写父对象的业务键编码"), "System")
                        if protect_sheet:
                            cell.protection = Protection(locked=False)
                    elif original_col_idx in fk_display_code_columns:
                        # [NEW 2026-06-16 BMRD] FK 编码显示字段
                        self._apply_classification_fill(cell, 'fk_display_code')
                        if protect_sheet:
                            cell.protection = Protection(locked=True)
                    elif original_col_idx in create_required_columns:
                        self._apply_classification_fill(cell, 'create_required')
                        if protect_sheet:
                            cell.protection = Protection(locked=False)
                    elif original_col_idx in readonly_columns:
                        self._apply_classification_fill(cell, 'readonly')
                        if protect_sheet:
                            cell.protection = Protection(locked=True)
                    else:
                        if protect_sheet:
                            cell.protection = Protection(locked=False)
                    
                    col_idx += 1
                row_idx += 1
            
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row in range(1, min(row_idx, 100)):
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            if protect_sheet:
                ws.protection = SheetProtection(
                    sheet=True,
                    autoFilter=False,
                    deleteColumns=False,
                    deleteRows=False,
                    formatCells=True,
                    formatColumns=True,
                    formatRows=True,
                    insertColumns=False,
                    insertRows=True,
                    sort=True,
                    objects=False,
                    scenarios=False,
                    pivotTables=False
                )
            
            sheets_info.append({
                "name": obj.name,
                "object_type": ot,
                "row_count": len(sheet_data) if sheet_data else 0
            })
            total_rows += len(sheet_data) if sheet_data else 0
        
        output_dir = os.path.join(os.getcwd(), "exports")
        os.makedirs(output_dir, exist_ok=True)
        
        product_name, version_name = self._get_product_version_info(filters)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = self._build_export_filename([product_name, version_name], timestamp)
        file_path = os.path.join(output_dir, file_name)
        
        wb.save(file_path)
        wb.close()
        
        return ExportResult(
            success=True,
            file_path=file_path,
            sheets=sheets_info,
            total_rows=total_rows,
            trace_id=self._get_current_trace_id(),
        )

    def _get_export_headers_with_editable(self, meta_obj: MetaObject, options: Optional[Dict[str, Any]], mode: str = 'update') -> tuple:
        """获取导出表头及可编辑列索引

        默认导出规则（参考 SAP Fiori "所见即所导" 原则）：
        1. 列表中可见的字段（ui.visible: true）默认可导出
        2. 显式设置 export_visible: true 的字段额外导出（不在列表中但需要导出）
        3. 显式设置 export_visible: false 的字段排除（在列表中但不导出）
        4. 默认排除系统字段：id, created_at, updated_at, created_by, updated_by
        5. 默认排除层级外键字段：version_id, product_id 等

        [NEW v1.2.3 2026-06-17] mode 参数:
        - mode='create' (默认 for 模板导出): 父对象编码列可填（批量新增/重新分类）
        - mode='update' (默认 for 主调用): 父对象编码列只读
        由调用方根据场景传入。

        返回：
        - headers: 表头列表
        - editable_columns: 可编辑列索引（普通可编辑字段）
        - readonly_columns: 只读列索引（ID、系统字段等）
        - parent_key_columns: 父对象业务键列索引（新增时需要填写）
        - create_required_columns: 新增必填列索引（business_key、parent_key等）
        """
        options = options or {}
        include_hierarchy_path = options.get("include_hierarchy_path", True)

        hierarchy_fields = self._get_hierarchy_field_names(meta_obj)

        default_exclude_fields = self._build_default_exclude_fields(meta_obj)

        headers = []
        header_comments = []
        header_to_field = {}
        editable_columns = []
        readonly_columns = []
        parent_key_columns = []
        # [NEW 2026-06-16 BMRD] FK 编码显示字段列 (parent_key_display)
        fk_display_code_columns = []
        create_required_columns = []
        enum_value_maps = {}
        bo_display_fields = {}
        col_idx = 0

        def should_export_field(f):
            return self._should_export_field(meta_obj, f)

        # 调试日志
        all_fields = list(meta_obj.fields)
        excluded_by_default = [f for f in all_fields if f.id in default_exclude_fields]
        excluded_by_storage = [f for f in all_fields if f.storage.value == "virtual" and not hasattr(f, 'ui')]
        passed_fields = [f for f in all_fields if should_export_field(f)]
        
        hierarchy_fields = self._get_hierarchy_field_names(meta_obj)
        
        logger.info(f"[Export] 对象类型: {meta_obj.id}, 总字段数: {len(all_fields)}")
        logger.info(f"[Export] hierarchy_fields: {hierarchy_fields}")
        logger.info(f"[Export] 因默认排除而过滤的字段: {[f.id for f in excluded_by_default]}")
        logger.info(f"[Export] 因virtual且无ui而过滤的字段: {[f.id for f in excluded_by_storage]}")
        logger.info(f"[Export] 通过 should_export_field 的字段: {[f.id for f in passed_fields]}")
        logger.info(f"[Export] 最终导出的字段 (排除hierarchy): {[f.name or f.id for f in passed_fields if (f.name or f.id) not in hierarchy_fields]}")
        
        # [REWRITE 2026-06-16 BMRD] sort key: business_key 排第一, 其余按 import_order 排序
        # 之前 storage.value != "virtual" 强制让 virtual 字段排到 physical 之后,
        # 导致 source_bo_code (import_order=3) 被排到 source_bo_id (no import_order) 之后,
        # 不符合"按用户规范排序"的需求.
        # 新策略: import_order 是用户/PM 指定的列顺序, 应优先于 storage 类型
        export_fields = sorted(
            passed_fields,
            key=lambda f: (
                0 if getattr(f.semantics, 'business_key', False) else 1,
                f.semantics.import_order if f.semantics.import_order else 999
            )
        )

        # [FIX 2026-06-11] 优先用 list columns 的 title 作为 export header
        # 保持 list 视图与 export 列名一致
        list_title_map = {}
        if (meta_obj.ui_view_config and meta_obj.ui_view_config.list
                and meta_obj.ui_view_config.list.columns):
            for col in meta_obj.ui_view_config.list.columns:
                key = getattr(col, 'key', None)
                title = getattr(col, 'title', None)
                if key and title:
                    list_title_map[key] = title

        for f in export_fields:
            field_name = f.name or f.id
            if field_name not in hierarchy_fields:
                # 优先用 list title, 保持 list/excel 一致
                header_name = list_title_map.get(f.id, field_name)
                headers.append(header_name)
                header_to_field[header_name] = f.id
                
                if f.enum_values:
                    enum_map = {ev.get('value'): ev.get('label', ev.get('value')) for ev in f.enum_values}
                    enum_value_maps[f.id] = enum_map
                else:
                    vh_enum_map = self._get_enum_value_map_from_value_help(f)
                    if vh_enum_map:
                        enum_value_maps[f.id] = vh_enum_map

                vh = self._get_value_help(f)
                if vh:
                    vh_source = getattr(vh, 'source', None)
                    if vh_source and getattr(vh_source, 'type', None) == 'bo':
                        bo_display_fields[f.id] = {
                            'target_bo': getattr(vh_source, 'target_bo', ''),
                            'display_field': getattr(vh_source, 'display_field', 'name') or 'name',
                            'value_field': getattr(vh_source, 'value_field', 'id') or 'id',
                            'code_field': getattr(vh_source, 'code_field', 'code') or 'code',
                        }
                
                comment_parts = []
                if f.description:
                    comment_parts.append(f.description)

                # 只在有控制信息时添加标识
                has_control_info = False
                # [NEW v1.1 2026-06-11] KeyTemplate user_editable 差异化标识
                kt_user_editable = self._get_key_template_user_editable(meta_obj)
                is_auto_or_manual_code = (
                    f.semantics.business_key
                    and kt_user_editable == 'auto_or_manual'
                )
                # [NEW 2026-06-14 BMRD] 关键修复: key_template 完整信息 (格式 + 示例)
                # 用于业务编码列头 comment, 用户已反馈 5 次以上
                kt_info = self._get_key_template_info(meta_obj)

                if f.semantics.business_key:
                    if is_auto_or_manual_code:
                        # [REWRITE 2026-06-14 BMRD] 业务化重写, 去掉【】底色等术语
                        bk_comment = "业务编码（必填）\n留空由系统自动生成；填写则使用填入值（系统会校验是否重复）"
                    else:
                        # [REWRITE 2026-06-14 BMRD] 业务化重写
                        bk_comment = "业务编码（必填，每行唯一标识）"
                    # [NEW 2026-06-14 BMRD] 关键: 追加 key_template 格式 + 示例
                    if kt_info.get('pattern'):
                        bk_comment += f"\n编码规则: {kt_info['pattern']}"
                    if kt_info.get('preview'):
                        bk_comment += f"\n示例: {kt_info['preview']}"
                    # [FIX 2026-06-16 BMRD] 跳过 f.description (避免与 bk_comment 内容重复)
                    # 之前 f.description = "业务对象编码" / "关系实例编码" 跟 "业务编码（必填）" 重复
                    comment_parts = [bk_comment]
                    has_control_info = True
                    create_required_columns.append(col_idx)
                elif f.required or getattr(f.semantics, 'mandatory', False):
                    comment_parts.append("【必填】")
                    has_control_info = True
                elif getattr(f.semantics, 'parent_key', False) and hasattr(f, 'ui') and hasattr(f.ui, 'relation') and f.ui.relation:
                    comment_parts.append("【父对象外键】新增必填")
                    has_control_info = True
                    create_required_columns.append(col_idx)

                if not self._is_field_editable(f):
                    comment_parts.append("【只读】")
                    has_control_info = True

                # [NEW 2026-06-16 BMRD] FK 编码显示字段 (parent_key_display)
                # 例: relationship.source_bo_code / target_bo_code
                # 用途: 跟随 source_bo_id / target_bo_id 自动带出编码
                # 视觉: 浅绿色 (与父对象编码一致), 但语义是"自动带出, 无需填写"
                if getattr(f.semantics, 'parent_key_display', False):
                    fk_display_code_columns.append(col_idx)
                    if "FK 编码显示字段" not in "；".join(comment_parts):
                        comment_parts.append("FK 编码显示字段（自动带出，无需填写）")
                    has_control_info = True

                # [NEW v1.1 2026-06-11] 记录 auto_or_manual code 列索引（用于差异化底色）
                if is_auto_or_manual_code:
                    if not hasattr(self, '_auto_or_manual_code_columns'):
                        self._auto_or_manual_code_columns = {}
                    self._auto_or_manual_code_columns.setdefault(meta_obj.name, []).append(col_idx)

                header_comments.append("；".join(comment_parts) if has_control_info else "")
                
                if self._is_field_editable(f):
                    editable_columns.append(col_idx)
                else:
                    readonly_columns.append(col_idx)
                col_idx += 1

        if include_hierarchy_path:
            headers.append("层级路径")
            header_to_field["层级路径"] = "层级路径"
            header_comments.append("对象的完整层级路径，如：领域/子领域/服务模块")
            readonly_columns.append(col_idx)
            col_idx += 1

        # [FIX 2026-06-08] 重构：使用 _build_parent_fk_columns SSOT 替代原本的 inline 循环
        # 行为保持完全一致：
        #   - 第一个非 context_field 父对象的编码列 → parent_key（可填写）
        #   - 其他编码列 + 所有名称列 → readonly
        # [NEW v1.2.3 2026-06-17] 传 mode 参数
        #   - mode='create' → 父对象编码列可填（用户重新分类场景）
        #   - mode='update' → 父对象编码列只读（向后兼容）
        parent_fk_columns = self._build_parent_fk_columns(meta_obj, mode=mode)
        for col_def in parent_fk_columns:
            header_name = col_def['header_name']
            headers.append(header_name)
            header_to_field[header_name] = header_name

            if col_def['kind'] == '编码':
                if col_def['classification'] == 'parent_key':
                    comment_msg = "【父对象编码】新增必填；编辑时可切换到其他父对象"
                    parent_key_columns.append(col_idx)
                else:
                    comment_msg = "父对象编码，只读"
                    readonly_columns.append(col_idx)
            else:  # 名称
                comment_msg = "父对象名称，只读"
                readonly_columns.append(col_idx)

            header_comments.append(comment_msg)
            col_idx += 1

        return headers, editable_columns, readonly_columns, parent_key_columns, create_required_columns, header_comments, header_to_field, enum_value_maps, bo_display_fields, fk_display_code_columns

    def _compute_list_computed_fields_for_export(self, meta_obj, data):
        """为导出数据计算 list 中配置的 computed 字段（如 child_count）

        [FR-005] 使用 computation_service.collect_computed_columns SSOT，
        与 query_service._compute_list_computed_fields 统一。
        [FIX 2026-06-14 BMRD] 同时调用 compute_by_semantics,
        修复 bug: 导出 relationship 时 category_label/category_type 字段为空
        (因为 computed_by: hierarchy_scope 字段在 collect_computed_columns 路径中不计算,
        需要单独走 compute_by_semantics)
        """
        from meta.services.computation_service import computation_service

        # 1. 计算 ui_view_config + rules 中的 computed 字段 (如 child_count)
        computed_cols = computation_service.collect_computed_columns(meta_obj)

        if computed_cols:
            computation_service.compute_batch(self.data_source, meta_obj.id, data, computed_cols)

        # 2. [FIX 2026-06-14 BMRD] 计算 semantics.computed_by 字段 (如 hierarchy_scope)
        # 与 query_service.search() 路径保持一致
        try:
            computation_service.compute_by_semantics(meta_obj.id, data, self.data_source)
        except Exception as e:
            logger = __import__('logging').getLogger(__name__)
            logger.warning(
                f"[Export] compute_by_semantics failed for {meta_obj.id}: {e}"
            )

    def _compute_hierarchy_scope_for_export(self, meta_obj, data):
        """[NEW 2026-06-16 BMRD] 为导出数据计算 hierarchy_scope 虚拟字段 (category_label / category_type)

        复刻 query_service._compute_hierarchy_scope_for_export 的逻辑.
        之前 import_export 路径没调用 compute_by_semantics, 导致 category_label 在此路径上为 None,
        Excel 导出后该列永远是空.
        """
        if not data:
            return data
        try:
            from meta.services.computation_service import computation_service
            computation_service.compute_by_semantics(meta_obj.id, data, self.data_source)
        except Exception as e:
            logger = __import__('logging').getLogger(__name__)
            logger.warning(f"[Export] compute_by_semantics failed: {e}")
        return data

    def _ensure_hierarchy_scope_computed(self, object_type: str, data: list) -> list:
        """[FIX 2026-06-14 BMRD] 确保 hierarchy_scope 虚拟字段已计算

        _query_association_with_hierarchy_filters / _query_association_by_version /
        _query_association_by_level 是 direct SQL 路径, 不走 search() 因此不调用
        compute_by_semantics. 此方法显式调用, 使 category_label/category_type
        在导出数据中非空.

        [SSOT 2026-06-14] query_service.search() 内部已包含此逻辑,
        本方法保证 direct SQL 路径也具备同样能力.
        """
        if not data:
            return data
        try:
            from meta.services.computation_service import computation_service
            computation_service.compute_by_semantics(object_type, data, self.data_source)
        except Exception as e:
            logger = __import__('logging').getLogger(__name__)
            logger.warning(f"[Export] compute_by_semantics failed for {object_type}: {e}")
        return data

    # ============================================================
    # [FIX 2026-06-08] Parent FK 列生成 SSOT (Single Source of Truth)
    # ============================================================
    # 背景：之前主导出（_get_export_headers_with_editable）和子对象 sheet
    # （_write_child_sheet）各自维护 parent FK 列生成逻辑，导致：
    #   1. 子对象 sheet 之前完全缺失该逻辑
    #   2. 两边代码独立演进，行为可能漂移
    # 重构：抽离到 _build_parent_fk_columns + _query_parent_fk_value_maps
    # 两个 helper，所有调用方统一使用。
    # ============================================================

    def _build_parent_fk_columns(self, meta_obj, mode: str = 'update'):
        """统一的 parent FK 列定义（SSOT）

        沿 meta_obj.parent_object 链向上追溯，生成 (编码, 名称) 列定义。

        [FR-009] 使用 _iter_parent_chain 替代 inline while 循环。
        [NEW v1.2.3 2026-06-17] 增加 mode 参数，区分 create/update 场景
        - mode='create': 模板可填父对象编码（适用于批量新增/重新分类）
        - mode='update'（默认）: 父对象编码只读（运行时保护 FK 引用/审计链）
        - parent_key_template_editable='always' 强制 parent_key 列可填（无视 mode）
        - parent_key_template_editable='create_only' 在 mode='create' 时可填
        - parent_key_template_editable='never' 始终只读

        返回 list[dict]，每个 dict 包含：
        - parent_obj: 父对象的 MetaObject
        - parent_fk_field_id: 例如 "product_id" / "domain_id"（用于从子数据取 parent_id）
        - kind: '编码' | '名称'
        - lookup_field: 'code' | 'name'（用于从父表取显示值）
        - header_name: 如 "产品线编码" / "产品线名称"
        - classification: 'parent_key'（最近一级可填写的 FK）| 'readonly'
        - is_first_parent: True 表示最近的父对象
        - readonly_always: 父对象的 FK 字段是否声明 readonly_always

        设计要点：
        - 第一个非 context_field 的父对象编码列标记为 'parent_key'（可填写），
          其他编码列 + 所有名称列标记为 'readonly'
        - 遇到 context_field 字段时停止向上追溯（上下文边界，之外通过导入界面选择）
        """
        columns = []
        is_first_parent = True

        for current_obj, parent_obj, parent_fk_field_id, is_context_boundary in \
                self._iter_parent_chain(meta_obj, stop_at_context_field=True):
            if is_context_boundary:
                break

            parent_key_field = current_obj.get_field(parent_fk_field_id)
            readonly_always = parent_key_field and getattr(
                parent_key_field.semantics, 'readonly_always', False
            )

            # [NEW v1.2.3 2026-06-17] 优先看 parent_key_template_editable
            template_editable = parent_key_field and getattr(
                parent_key_field.semantics, 'parent_key_template_editable', None
            )
            is_template_editable = False
            if template_editable == 'always':
                is_template_editable = True
            elif template_editable == 'create_only' and mode == 'create':
                is_template_editable = True
            elif template_editable == 'never':
                is_template_editable = False
            elif template_editable is None:
                # fallback: 旧的 readonly_always 逻辑（保持向后兼容）
                is_template_editable = (is_first_parent and parent_key_field and not readonly_always)

            # 编码列：可填写（parent_key）或只读（readonly）
            code_classification = 'parent_key' if is_template_editable else 'readonly'
            columns.append({
                'parent_obj': parent_obj,
                'parent_fk_field_id': parent_fk_field_id,
                'kind': '编码',
                'lookup_field': 'code',
                'header_name': "{0}编码".format(parent_obj.name),
                'classification': code_classification,
                'is_first_parent': is_first_parent,
                'readonly_always': readonly_always,
            })
            # 名称列：始终只读
            columns.append({
                'parent_obj': parent_obj,
                'parent_fk_field_id': parent_fk_field_id,
                'kind': '名称',
                'lookup_field': 'name',
                'header_name': "{0}名称".format(parent_obj.name),
                'classification': 'readonly',
                'is_first_parent': is_first_parent,
                'readonly_always': True,
            })

            is_first_parent = False

        return columns

    def _query_parent_fk_value_maps(self, data, parent_fk_columns):
        """根据 parent_fk_columns 预查询父对象编码/名称映射

        用法：先用 _build_parent_fk_columns(meta_obj) 获取列定义，
        然后用本方法批量预查询所有需要的父对象 code/name 映射。

        Args:
            data: 子对象数据列表
            parent_fk_columns: _build_parent_fk_columns 返回的列定义列表

        Returns:
            dict[列索引 -> dict[parent_id -> value]]
            例如：{0: {1: 'TEST15', 2: 'TEST14'}, 1: {1: '产品A', 2: '产品B'}}
            列索引 0 对应"产品线编码"，列索引 1 对应"产品线名称"
        """
        value_maps = {}
        for col_idx, col_def in enumerate(parent_fk_columns):
            parent_obj = col_def['parent_obj']
            parent_fk_field_id = col_def['parent_fk_field_id']
            lookup_field = col_def['lookup_field']

            # 收集所有出现的 parent ID
            parent_ids = list({
                r.get(parent_fk_field_id)
                for r in data
                if r.get(parent_fk_field_id) is not None
            })
            if not parent_ids:
                value_maps[col_idx] = {}
                continue

            try:
                placeholders = ','.join(['?'] * len(parent_ids))
                sql = "SELECT id, {0} FROM {1} WHERE id IN ({2})".format(
                    lookup_field, parent_obj.table_name, placeholders
                )
                cursor = self.data_source.execute(sql, tuple(parent_ids))
                value_maps[col_idx] = {row[0]: row[1] for row in cursor.fetchall()}
            except Exception as e:
                logger.warning(
                    f"[Export] 查询父对象 {parent_obj.id}.{lookup_field} 失败: {e}"
                )
                value_maps[col_idx] = {}

        return value_maps

    def _get_cascade_object_types(self, object_type: str) -> List[str]:
        """获取级联导出的对象类型列表
        
        级联导出包含：
        1. 当前对象的所有父对象（向上追溯）
        2. 当前对象的所有子对象（向下追溯）
        3. 如果包含业务对象，则自动包含关系对象
        
        这样可以导出完整的层级数据。
        """
        result = set()
        result.add(object_type)
        
        parent_types = self._get_parent_object_types(object_type)
        result.update(parent_types)
        
        child_types = self._get_child_object_types(object_type)
        result.update(child_types)
        
        result.update(self._get_association_bound_types(result))
        
        return list(result)
    
    def _get_parent_object_types(self, object_type: str) -> set:
        """向上获取所有父对象类型"""
        result = set()
        current_type = object_type
        
        while current_type:
            obj = registry.get(current_type)
            if obj and obj.parent_object:
                parent_obj = registry.get(obj.parent_object)
                if parent_obj:
                    result.add(obj.parent_object)
                    current_type = obj.parent_object
                else:
                    break
            else:
                break
        
        return result
    
    def _get_child_object_types(self, object_type: str) -> set:
        """向下获取所有子对象类型"""
        result = set()
        
        for obj_id in registry.list_objects():
            obj = registry.get(obj_id)
            if obj and obj.parent_object == object_type:
                result.add(obj_id)
                result.update(self._get_child_object_types(obj_id))
        
        return result

    def _sort_by_hierarchy(self, object_types: List[str]) -> List[str]:
        """按层级排序（父对象在前，子对象在后）

        [FR-008] 委托到 HierarchyConfigLoader.sort_by_hierarchy，
        统一 4 个导出/导入入口的排序逻辑。
        """
        from meta.services.cascade_service import HierarchyConfigLoader
        return HierarchyConfigLoader.sort_by_hierarchy(object_types)

    def _query_with_hierarchy(self, object_type: str, filters: Optional[Dict[str, Any]],
                              options: Optional[Dict[str, Any]], sort_by: str = None, 
                              sort_order: str = 'asc',
                              page: int = None, page_size: int = None) -> List[Dict[str, Any]]:
        """查询数据并添加层级信息
        
        使用与列表查询相同的 resolve_conditions 方法，确保导出和列表数据一致性
        
        Args:
            object_type: 对象类型
            filters: 筛选条件
            options: 导出选项
            sort_by: 排序字段
            sort_order: 排序方向 (asc/desc)
            page: 页码（分页导出时使用，优先级高于 MAX_EXPORT_LIMIT）
            page_size: 每页数量（分页导出时使用）
        """
        conditions = []
        normalized_filters = {}
        if filters:
            # 防御性规范化：version_id / product_id 等上下文字段应为标量
            normalized_filters = {
                key: self._normalize_scalar_id(value) if key in ('version_id', 'product_id') else value
                for key, value in filters.items()
            }
            args_dict = {}
            for key, value in normalized_filters.items():
                if isinstance(value, list):
                    args_dict[key] = [str(v) for v in value]
                else:
                    args_dict[key] = [str(value)] if value is not None else []

            conditions = self.hierarchy_filter.resolve_conditions(object_type, args_dict)

        if self._is_association_type(object_type) and filters:
            return self._query_association_with_hierarchy_filters(object_type, normalized_filters)
        
        MAX_EXPORT_LIMIT = self._get_export_limit(object_type)

        if page is not None and page_size is not None:
            actual_page = page
            actual_page_size = page_size
            print(f"[Export] 使用分页参数: page={actual_page}, page_size={actual_page_size}")
        else:
            actual_page = 1
            actual_page_size = MAX_EXPORT_LIMIT
            print(f"[Export] 开始查询 {object_type}，限制最大 {actual_page_size} 条")

        search_request = SearchRequest(
            object_type=object_type,
            conditions=conditions,
            page=actual_page,
            page_size=actual_page_size,
            sort_by=sort_by or '',
            sort_order=sort_order or 'asc',
        )

        search_result = self.query_service.search(search_request)
        data = search_result.data or []

        print(f"[Export] 查询完成，获取 {len(data)} 条数据")

        data = self._inject_hierarchy_info(data, object_type, filters, options)

        print(f"[Export] 层级信息注入完成")

        return data
    
    def _query_association_with_hierarchy_filters(self, object_type: str,
                                                     filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """配置驱动的 association 层级过滤查询

        从 hierarchies.yaml 的 association_filter_config 读取配置，
        按优先级（从细粒度到粗粒度）匹配过滤器，使用通用 SQL 查询 + 回退机制。

        支持任意 association 类型和任意层级结构，无需硬编码。
        """
        from meta.services.config_driven_hierarchy_filter import HierarchyConfigLoader

        filter_config = HierarchyConfigLoader.get_association_filter_config(object_type)
        if not filter_config:
            return []

        meta_obj = registry.get(object_type)
        table_name = meta_obj.table_name if meta_obj else object_type + 's'
        version_id = self._normalize_scalar_id(filters.get('version_id'))
        relation_codes = filters.get('relation_codes', [])
        if not isinstance(relation_codes, list):
            relation_codes = [relation_codes] if relation_codes else []

        filter_levels = filter_config.get('hierarchy_filter_levels', [])

        for level_config in filter_levels:
            level_object = level_config.get('object')
            filter_key = level_object + '_id'
            level_ids = filters.get(filter_key, [])

            if not level_ids:
                continue
            if not isinstance(level_ids, list):
                level_ids = [level_ids]
            # 防御性拍平：确保 level_ids 内部元素不是列表
            level_ids = [
                self._normalize_scalar_id(v) for v in level_ids
                if self._normalize_scalar_id(v) is not None
            ]

            return self._query_association_by_level(
                object_type, table_name, version_id, level_ids,
                level_config, filter_config, relation_codes
            )

        if version_id:
            return self._query_association_by_version(object_type, table_name, version_id, relation_codes)

        return []

    def _query_association_by_level(self, object_type: str, table_name: str,
                                      version_id: int, level_ids: List[int],
                                      level_config: Dict[str, Any],
                                      filter_config: Dict[str, Any],
                                      relation_codes: List[str]) -> List[Dict[str, Any]]:
        """通用层级关联查询：直接查 association 表的 source/target 列，无结果则回退到子级"""
        from meta.core.enrichment_engine import enrich_records

        if not level_ids:
            return []

        source_col = level_config.get('source_column')
        target_col = level_config.get('target_column')
        if not source_col or not target_col:
            return []

        placeholders = ','.join(['?'] * len(level_ids))
        # [FIX FR-001] 软删除过滤 + 默认排序
        soft_delete_filter = ""
        meta_obj = registry.get(object_type)
        if meta_obj and any(f.id == 'is_deleted' for f in meta_obj.fields):
            soft_delete_filter = " AND (r.is_deleted IS NULL OR r.is_deleted = 0)"

        # [FIX FR-002] 数据权限过滤
        perm_filter_sql, perm_params = self._build_permission_filter(object_type, 'r')

        sql = f"""
            SELECT r.* FROM {table_name} r
            WHERE r.version_id = ?
            AND (r.{source_col} IN ({placeholders}) OR r.{target_col} IN ({placeholders}))
            {soft_delete_filter}{perm_filter_sql}
            ORDER BY r.id ASC
        """
        params = [version_id, *level_ids, *level_ids] + perm_params

        if relation_codes:
            code_placeholders = ','.join(['?'] * len(relation_codes))
            sql += f" AND r.relation_code IN ({code_placeholders})"
            params.extend(relation_codes)

        cursor = self.data_source.execute(sql, tuple(params))
        columns = [desc[0] for desc in cursor.description]
        data = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # [FIX 2026-06-14 BMRD] direct SQL 路径 - 先计算 hierarchy_scope
        data = self._ensure_hierarchy_scope_computed(object_type, data)

        if not data:
            fallback = level_config.get('fallback')
            if fallback:
                child_object = fallback.get('child_object')
                child_fk_field = fallback.get('child_fk_field')
                if child_object and child_fk_field:
                    child_level = self._get_table_name_for_object(child_object)
                    child_placeholders = ','.join(['?'] * len(level_ids))
                    child_sql = f"SELECT id FROM {child_level} WHERE {child_fk_field} IN ({child_placeholders})"
                    cursor2 = self.data_source.execute(child_sql, tuple(level_ids))
                    child_ids = [row[0] for row in cursor2.fetchall()]
                    if child_ids:
                        child_level_config = self._find_filter_level_config(
                            object_type, child_object
                        )
                        if child_level_config:
                            return self._query_association_by_level(
                                object_type, table_name, version_id, child_ids,
                                child_level_config, filter_config, relation_codes
                            )

        data = enrich_records(object_type, data)

        # [FIX 2026-06-14 BMRD] direct SQL 路径必须显式计算 hierarchy_scope
        return self._ensure_hierarchy_scope_computed(object_type, data)

    def _get_table_name_for_object(self, object_type: str) -> str:
        """获取对象类型对应的表名"""
        meta_obj = registry.get(object_type)
        if meta_obj and meta_obj.table_name:
            return meta_obj.table_name
        from meta.services.config_driven_hierarchy_filter import HierarchyConfigLoader
        level = HierarchyConfigLoader.get_level_by_object(object_type)
        return level.get('table_name', object_type + 's')

    def _find_filter_level_config(self, object_type: str, target_object: str) -> Optional[Dict[str, Any]]:
        """在 association_filter_config 中查找指定 object 的层级配置"""
        from meta.services.config_driven_hierarchy_filter import HierarchyConfigLoader
        filter_levels = HierarchyConfigLoader.get_association_filter_levels(object_type)
        for level in filter_levels:
            if level.get('object') == target_object:
                return level
        return None

    def _query_association_by_version(self, object_type: str, table_name: str,
                                        version_id: int,
                                        relation_codes: List[str]) -> List[Dict[str, Any]]:
        """按 version_id 查询 association 数据（无层级过滤时使用）"""
        from meta.core.enrichment_engine import enrich_records

        # [FIX FR-001] 软删除过滤 + 默认排序
        soft_delete_filter = ""
        meta_obj = registry.get(object_type)
        if meta_obj and any(f.id == 'is_deleted' for f in meta_obj.fields):
            soft_delete_filter = " AND (is_deleted IS NULL OR is_deleted = 0)"

        # [FIX FR-002] 数据权限过滤
        perm_filter_sql, perm_params = self._build_permission_filter(object_type)

        sql = f"SELECT * FROM {table_name} WHERE version_id = ?"
        params = [version_id]

        if relation_codes:
            code_placeholders = ','.join(['?'] * len(relation_codes))
            sql += f" AND relation_code IN ({code_placeholders})"
            params.extend(relation_codes)

        sql += soft_delete_filter + perm_filter_sql + " ORDER BY id ASC"
        params.extend(perm_params)

        cursor = self.data_source.execute(sql, tuple(params))
        columns = [desc[0] for desc in cursor.description]
        data = [dict(zip(columns, row)) for row in cursor.fetchall()]

        data = enrich_records(object_type, data)

        # [FIX 2026-06-14 BMRD] direct SQL 路径必须显式计算 hierarchy_scope,
        # 否则 category_label/category_type 在 export 时为 None
        return self._ensure_hierarchy_scope_computed(object_type, data)

    def _is_association_type(self, object_type: str) -> bool:
        """检查对象类型是否是 association 类型"""
        from meta.services.config_driven_hierarchy_filter import HierarchyConfigLoader
        level = HierarchyConfigLoader.get_level_by_object(object_type)
        return level.get('kind') == 'association'

    def _ensure_association_types_in_order(self, type_order: List[str]) -> List[str]:
        """确保 association 类型出现在 type_order 末尾

        从 hierarchies.yaml 的 levels 中查找 kind=association 的类型，
        自动追加到 type_order 末尾（如果尚未存在）。
        """
        from meta.services.config_driven_hierarchy_filter import HierarchyConfigLoader
        levels = HierarchyConfigLoader.get_levels()
        for level in levels:
            if level.get('kind') == 'association':
                obj = level.get('object')
                if obj and obj not in type_order:
                    type_order = type_order + [obj]
        return type_order

    def _get_export_limit(self, object_type: str) -> int:
        """获取导出数量限制

        优先从 meta_obj.import_export.max_export_limit 读取，
        默认 10000。
        """
        meta_obj = registry.get(object_type)
        if meta_obj and meta_obj.import_export:
            limit = getattr(meta_obj.import_export, 'max_export_limit', None)
            if limit:
                return limit
        return 10000

    def _build_type_labels(self) -> Dict[str, str]:
        """从 hierarchies.yaml 的 levels 配置构建类型标签映射"""
        from meta.services.config_driven_hierarchy_filter import HierarchyConfigLoader
        labels = {}
        levels = HierarchyConfigLoader.get_levels()
        for level in levels:
            obj = level.get('object')
            display_name = level.get('display_name', '')
            if obj and display_name:
                labels[obj] = display_name
        return labels

    def _build_full_data_sheets(self) -> set:
        """从 hierarchies.yaml 的 levels 配置构建需要完整数据的 sheet 集合

        entity 类型（kind=entity）的层级对象需要完整数据（可能被其他 sheet 引用）。
        """
        from meta.services.config_driven_hierarchy_filter import HierarchyConfigLoader
        sheets = set()
        levels = HierarchyConfigLoader.get_levels()
        for level in levels:
            if level.get('kind', 'entity') == 'entity':
                obj = level.get('object')
                table = level.get('table_name', '')
                if obj:
                    sheets.add(obj)
                    if table:
                        sheets.add(table)
        return sheets

    def _get_association_bound_types(self, existing_types: set) -> set:
        """获取与已有类型绑定的 association 类型

        从 hierarchies.yaml 的 levels 中查找 kind=association 的类型，
        如果其 source_entity 在 existing_types 中，则该 association 也应被包含。
        """
        from meta.services.config_driven_hierarchy_filter import HierarchyConfigLoader
        result = set()
        levels = HierarchyConfigLoader.get_levels()
        for level in levels:
            if level.get('kind') == 'association':
                source_entity = level.get('source_entity', '')
                if source_entity in existing_types:
                    result.add(level.get('object'))
        return result
    
    def _inject_hierarchy_info(self, data: List[Dict[str, Any]], object_type: str, 
                               filters: Optional[Dict[str, Any]], 
                               options: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        
        meta_obj = registry.get(object_type)
        
        for record in data:
            if options and options.get("include_hierarchy_path", True):
                record["层级路径"] = self._build_hierarchy_path(record, meta_obj)

            self._add_hierarchy_fields(record, meta_obj, options)

            if object_type == 'relationship':
                self._enrich_relationship_record(record)
        
        return data

    def _build_hierarchy_path(self, record: Dict[str, Any], meta_obj: MetaObject) -> str:
        """构建层级路径"""
        parts = []
        current_obj = meta_obj
        current_record = record
        
        while current_obj:
            name = current_record.get("name") or current_record.get(current_obj.id + "_name") or ""
            if name:
                parts.insert(0, name)
            
            if current_obj.parent_object:
                parent_obj = registry.get(current_obj.parent_object)
                parent_id = current_record.get(current_obj.parent_object + "_id")
                if parent_obj and parent_id:
                    parent_record = self._get_parent_record(current_obj.parent_object, parent_id)
                    current_obj = parent_obj
                    current_record = parent_record
                else:
                    break
            else:
                break
        
        return "/".join(parts)

    @staticmethod
    def _normalize_scalar_id(value: Any) -> Any:
        """将可能是列表的标量 ID 规范化为单个标量值。

        前端某些场景会把单个 version_id / product_id 以数组形式下发
        （如 {"version_id": [764]}），直接绑定到 SQL 的 `= ?` 会报
        "Error binding parameter 1: type 'list' is not supported"。
        本方法在保留 None / 空值语义的前提下，取列表第一个元素。
        """
        if value is None:
            return None
        if isinstance(value, list):
            if not value:
                return None
            return value[0]
        return value

    def _get_product_version_info(self, filters: Optional[Dict[str, Any]]) -> tuple:
        """获取产品线和版本名称"""
        product_name = ""
        version_name = ""
        
        if not filters:
            return product_name, version_name
        
        version_id = self._normalize_scalar_id(filters.get("version_id"))
        if version_id:
            try:
                version_obj = registry.get("version")
                if version_obj:
                    search_request = SearchRequest(
                        object_type="version",
                        conditions=[QueryCondition(field="id", operator="eq", value=version_id)],
                        page=1,
                        page_size=1,
                    )
                    result = self.query_service.search(search_request)
                    if result.data:
                        version_data = result.data[0]
                        version_name = version_data.get("name", "")
                        product_id = version_data.get("product_id")
                        if product_id:
                            product_search = SearchRequest(
                                object_type="product",
                                conditions=[QueryCondition(field="id", operator="eq", value=product_id)],
                                page=1,
                                page_size=1,
                            )
                            product_result = self.query_service.search(product_search)
                            if product_result.data:
                                product_name = product_result.data[0].get("name", "")
            except Exception:
                pass
        
        return product_name, version_name

    def _get_product_version_codes(self, filters: Optional[Dict[str, Any]]) -> tuple:
        """根据 filters 中的 version_id 获取 product_code 和 version_code
        
        Args:
            filters: 包含 version_id 的过滤条件
            
        Returns:
            tuple: (product_code, version_code)
        """
        if not filters:
            return ('', '')
        
        version_id = self._normalize_scalar_id(filters.get('version_id'))
        if not version_id:
            return ('', '')

        try:
            query = """
                SELECT p.code as product_code, v.code as version_code
                FROM versions v
                LEFT JOIN products p ON v.product_id = p.id
                WHERE v.id = ?
                LIMIT 1
            """
            cursor = self.data_source.execute(query, (version_id,))
            row = cursor.fetchone()
            if row:
                return (row[0] or '', row[1] or '')
        except Exception as e:
            logger.warning(f"Failed to get product/version codes: {e}")
        
        return ('', '')

    def _collect_child_object_types(self, selected_types: List[str]) -> Dict[str, List[str]]:
        """从 child_sections 配置收集子对象类型及其父对象映射

        扫描每个选中类型的 ui_view_config.child_sections 配置，
        收集所有 child_object 类型并记录父对象映射。

        Returns:
            Dict[str, List[str]]: {child_object_type: [parent_type_1, parent_type_2, ...]}
        """
        from meta.core.models import registry
        
        child_parent_map = {}
        for obj_type in selected_types:
            meta_obj = registry.get(obj_type)
            if not meta_obj:
                continue
            child_sections = getattr(meta_obj.ui_view_config, 'child_sections', [])
            for section in child_sections:
                child_type = section.get('child_object')
                if child_type:
                    if child_type not in child_parent_map:
                        child_parent_map[child_type] = []
                    if obj_type not in child_parent_map[child_type]:
                        child_parent_map[child_type].append(obj_type)
        return child_parent_map

    def _query_child_object(self, child_type: str, parent_types: List[str],
                            filters: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """通用子对象查询

        自动判断关联类型：
        - 多态关联（annotation）：通过 associations 的 polymorphic_type_field 查询
        - 直接 FK 关联（enum_value、version 等）：通过子对象 parent_object 字段反查父对象数据

        Args:
            child_type: 子对象类型 ID（如 'annotation', 'enum_value', 'version'）
            parent_types: 父对象类型 ID 列表
            filters: 导出过滤条件

        Returns:
            List[Dict]: 增强后的子对象数据列表
        """
        from meta.core.models import registry

        child_meta = registry.get(child_type)
        if not child_meta:
            return []

        if child_type == 'annotation':
            return self._query_annotations_impl(parent_types, filters)

        parent_object = getattr(child_meta, 'parent_object', None)
        if parent_object and parent_object in parent_types:
            return self._query_direct_fk_child(child_type, child_meta, parent_object, filters)

        return []

    def _classify_field(self, field) -> str:
        """对单个字段进行导出分类（与主导出保持一致）

        返回值：
        - 'parent_key'：父对象外键（多态 FK / 直接 FK），新增必填、编辑时只读
        - 'fk_display_code'：FK 编码显示字段（parent_key_display），自动带出，无需填写
        - 'create_required'：新增必填字段（业务关键字、必填字段等）
        - 'readonly'：始终只读字段
        - 'editable'：普通可编辑字段
        """
        if not field:
            return 'readonly'

        if getattr(field.semantics, 'parent_key', False):
            if not getattr(field.semantics, 'readonly_always', False):
                return 'parent_key'

        # [NEW 2026-06-16 BMRD] FK 编码显示字段 - 与 parent_key 视觉一致 (浅绿),
        # 但语义是"自动带出" (无需填写)
        if getattr(field.semantics, 'parent_key_display', False):
            return 'fk_display_code'

        if getattr(field.semantics, 'business_key', False):
            return 'create_required'

        if (field.required or getattr(field.semantics, 'mandatory', False)) and \
                not self._is_field_editable(field, mode='create'):
            return 'readonly'

        if not self._is_field_editable(field, mode='edit'):
            return 'readonly'

        return 'editable'

    def _write_child_sheet(self, wb, child_type: str, child_meta, data: List[Dict],
                           sheets_info: List[Dict], options: Optional[Dict[str, Any]] = None) -> None:
        """通用子对象 Sheet 写入

        使用 child_meta 的 export_visible 语义动态构建表头和数据行。

        对于需要 round-trip（导出后修改再导入）的场景，子对象 Sheet 采用
        export_visible UNION import_visible 的联合规则，确保导出的 Excel 包含
        导入所需的关键字段（如 annotation 的 target_type）。

        当子对象支持 CUD 操作时，自动添加"操作模式"列、数据验证、
        单元格级别的只读/可编辑高亮、列注释，以及空行用于新增记录。
        此行为是通用的：任何定义了 crud 类型 action 的子对象
        （无论多态关联如 annotation，还是直接 FK 关联如 enum_value）
        都会自动获得 CUD 支持。

        字段控制逻辑（与主导出保持一致）：
        - 父对象外键（parent_key）：现有行用 BUSINESS_KEY_FILL（浅绿），
          新增行同样用 BUSINESS_KEY_FILL（表示必填但可切换）
        - 业务关键字（business_key）：现有行用 REQUIRED_FILL（浅黄），
          新增行同样用 REQUIRED_FILL（新增必填、编辑时只读）
        - 普通必填（required / mandatory）：现有行不加底色，新增行用 REQUIRED_FILL
        - 始终只读（immutable、readonly_always、virtual、ui.editable=false）：
          所有行用 READONLY_FILL（灰色）
        - 可编辑字段：不加底色

        Args:
            wb: openpyxl Workbook 对象
            child_type: 子对象类型 ID
            child_meta: 子对象的 MetaObject
            data: 子对象数据列表
            sheets_info: Sheet 信息列表（会被修改）
            options: 导出选项（支持 include_operation_mode, empty_rows_for_new）
        """
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.comments import Comment
        from openpyxl.utils import get_column_letter

        options = options or {}
        sheet_name = child_meta.name or child_type

        has_cud = _has_cud_actions(child_meta)
        include_operation_mode = options.get("include_operation_mode", True) and has_cud

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ds = ExcelDesignSystem

        default_exclude_fields = self._build_default_exclude_fields(child_meta)

        cud_required_fields = {'id'} if has_cud else set()

        candidates = []
        seen_names = set()
        for f in child_meta.fields:
            if f.id in default_exclude_fields and f.id not in cud_required_fields:
                continue
            if f.storage.value == "virtual" and not hasattr(f, 'ui'):
                continue

            export_vis = getattr(f.semantics, 'export_visible', None)
            import_vis = getattr(f.semantics, 'import_visible', None)

            is_cud_required = f.id in cud_required_fields

            if not is_cud_required and export_vis is False and import_vis is False:
                continue

            is_export = export_vis is True or is_cud_required
            is_import = import_vis is True

            if is_export or is_import or (hasattr(f, 'ui') and hasattr(f.ui, 'visible') and f.ui.visible is True):
                candidates.append((f, is_export, is_import))

        candidates.sort(key=lambda x: (
            0 if (x[1] and x[2]) else 1,
            0 if (x[2] and not (x[0].storage.value == 'virtual' or getattr(x[0].semantics, 'virtual', False))) else 1,
            0 if x[1] else 1
        ))

        export_fields = []
        for f, is_export, is_import in candidates:
            if f.name and f.name in seen_names:
                continue
            if f.name:
                seen_names.add(f.name)
            export_fields.append(f)

        # [REWRITE 2026-06-16 BMRD] sort key 简化为 (business_key, import_order)
        # 详见 L1361-1372 注释
        export_fields.sort(key=lambda f: (
            0 if getattr(f.semantics, 'business_key', False) else 1,
            f.semantics.import_order if f.semantics.import_order else 999
        ))

        if has_cud:
            id_in_export = any(f.id == 'id' for f in export_fields)
            if not id_in_export:
                for f in child_meta.fields:
                    if f.id == 'id':
                        export_fields.insert(0, f)
                        break

        ws = wb.create_sheet(title=sheet_name)
        col_offset = 0

        if include_operation_mode:
            col_offset = 1
            cell = ws.cell(row=1, column=1, value="操作模式")
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = ds.THIN_BORDER
            ws.column_dimensions['A'].width = 18
            cell.comment = Comment("新增/更新/删除/跳过，留空默认为更新", "系统")

            operation_dv = self._create_operation_mode_dv(ws, with_prompt=True, verbose_error=True)

        field_ids = [f.id for f in export_fields]
        field_classifications = {
            f.id: self._classify_field(f) for f in export_fields
        }
        enum_validations = {}
        value_help_map = {}

        # [FIX 2026-06-08] 为子对象添加 parent FK 列（SSOT）
        # 使用 _build_parent_fk_columns + _query_parent_fk_value_maps，
        # 与 _get_export_headers_with_editable 保持完全一致的列定义
        # [NEW v1.2.3 2026-06-17] 子对象 sheet 也按 'create' 模式 (模板可填父对象)
        parent_fk_columns = self._build_parent_fk_columns(child_meta, mode='create')
        parent_fk_value_maps = self._query_parent_fk_value_maps(data, parent_fk_columns) if data else {}

        for col_idx, f in enumerate(export_fields):
            actual_col = col_idx + 1 + col_offset
            header_name = f.name
            cell = ws.cell(row=1, column=actual_col, value=header_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = ds.THIN_BORDER

            classification = field_classifications[f.id]
            is_editable_create = self._is_field_editable(f, mode='create')
            is_editable_edit = self._is_field_editable(f, mode='edit')

            comment_parts = []
            if f.description:
                comment_parts.append(f.description)
            if classification == 'parent_key':
                if getattr(f.semantics, 'parent_key', False) and getattr(f.semantics, 'readonly_always', False):
                    comment_parts.append("【父对象外键】始终只读（上下文带入）")
                else:
                    comment_parts.append("【父对象外键 / 必填】新增必填，编辑时只读（可切换到其他父对象）")
            elif classification == 'create_required':
                if getattr(f.semantics, 'business_key', False):
                    comment_parts.append("【业务关键字】新增必填，编辑时只读")
                else:
                    comment_parts.append("【必填】")
            if not is_editable_create and not (classification == 'parent_key' or classification == 'create_required'):
                comment_parts.append("【只读】")
            if comment_parts:
                cell.comment = Comment("；".join(comment_parts), "系统")

            enum_dv_values = self._build_enum_dv_values(f)
            if enum_dv_values:
                dv = DataValidation(
                    type="list",
                    formula1=enum_dv_values,
                    allow_blank=True
                )
                dv.error = "请从下拉列表中选择有效值"
                dv.errorTitle = "无效输入"
                if classification == 'parent_key' or classification == 'create_required':
                    dv.error = "新增时该字段必填，且必须从下拉列表中选择"
                ws.add_data_validation(dv)
                enum_validations[actual_col] = dv
                value_help_map[f.id] = enum_dv_values

        # [FIX 2026-06-08] 添加 parent FK 列（编码 + 名称）的表头
        for pfk_idx, col_def in enumerate(parent_fk_columns):
            actual_col = len(export_fields) + 1 + pfk_idx + col_offset
            cell = ws.cell(row=1, column=actual_col, value=col_def['header_name'])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = ds.THIN_BORDER
            # 与 _get_export_headers_with_editable 的 comment 保持一致
            if col_def['kind'] == '编码':
                if col_def['classification'] == 'parent_key':
                    cell.comment = Comment("【父对象编码】新增必填；编辑时可切换到其他父对象", "系统")
                else:
                    cell.comment = Comment("父对象编码，只读", "系统")
            else:
                cell.comment = Comment("父对象名称，只读", "系统")

        for row_idx, record in enumerate(data, 2):
            if include_operation_mode:
                cell = ws.cell(row=row_idx, column=1, value="update - 更新")
                cell.fill = ds.READONLY_FILL
                cell.border = ds.THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                operation_dv.add(cell)

            for col_idx, field_id in enumerate(field_ids):
                actual_col = col_idx + 1 + col_offset
                value = record.get(field_id, "")

                # Resolve enum label for display
                # [FIX 2026-06-11] 修复 child sheet enum 解析：覆盖 value_help → DB 路径
                # 原逻辑只检查静态 enum_values，导致仅有 value_help 的字段
                # (annotation.target_type/category, relationship.relation_type/direction)
                # 在 child sheet 中只显示原始 key
                field_obj = child_meta.get_field(field_id)
                if field_obj and value not in (None, ""):
                    enum_map = None
                    if field_obj.enum_values:
                        enum_map = {ev.get('value'): ev.get('label', ev.get('value'))
                                    for ev in field_obj.enum_values}
                    if not enum_map:
                        enum_map = self._get_enum_value_map_from_value_help(field_obj)
                    if enum_map and value in enum_map:
                        label = enum_map[value]
                        if label != value:
                            value = f"{value} - {label}"

                cell = ws.cell(row=row_idx, column=actual_col, value=_safe_cell_value(value))
                cell.border = ds.THIN_BORDER

                classification = field_classifications[field_id]
                self._apply_classification_fill(cell, classification)
                if actual_col in enum_validations:
                    enum_validations[actual_col].add(cell)

            # [FIX 2026-06-08] 填充 parent FK 列（SSOT）
            for pfk_idx, col_def in enumerate(parent_fk_columns):
                actual_col = len(export_fields) + 1 + pfk_idx + col_offset
                parent_fk_field_id = col_def['parent_fk_field_id']
                parent_id = record.get(parent_fk_field_id)
                value_map = parent_fk_value_maps.get(pfk_idx, {})
                value = value_map.get(parent_id, "")
                cell = ws.cell(row=row_idx, column=actual_col, value=_safe_cell_value(value))
                cell.border = ds.THIN_BORDER
                # 与主导出保持一致：parent_key 用 BUSINESS_KEY_FILL（浅绿色），readonly 用 READONLY_FILL（灰色）
                if col_def['classification'] == 'parent_key':
                    cell.fill = ds.BUSINESS_KEY_FILL
                else:
                    cell.fill = ds.READONLY_FILL

        # [FIX] 有数据时不追加新增行，无数据时添加空白新增行 (v3.18)
        if len(data) > 0:
            empty_rows_count = 0
        elif include_operation_mode:
            empty_rows_count = options.get("empty_rows_for_new", 5)
        else:
            empty_rows_count = 0
        for empty_row in range(empty_rows_count):
                row_idx = len(data) + 2 + empty_row

                cell = ws.cell(row=row_idx, column=1, value="create - 新增")
                # [REMOVED 2026-06-14 BMRD] CREATE_NEW_FILL 已删除
                cell.fill = ds.READONLY_FILL  # 复用灰色
                cell.border = ds.THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                operation_dv.add(cell)

                for col_idx, field_id in enumerate(field_ids):
                    actual_col = col_idx + 1 + col_offset
                    cell = ws.cell(row=row_idx, column=actual_col, value="")
                    cell.border = ds.THIN_BORDER

                    classification = field_classifications[field_id]
                    self._apply_classification_fill(cell, classification)
                    if classification == 'parent_key':
                        cell.comment = Comment("新增时必填：请填写父对象的业务键编码", "System")
                    if actual_col in enum_validations:
                        enum_validations[actual_col].add(cell)

        total_rows_in_sheet = 2 + len(data) + empty_rows_count
        for col_idx, f in enumerate(export_fields):
            actual_col = col_idx + 1 + col_offset
            col_letter = get_column_letter(actual_col)
            max_length = len(f.name or f.id)
            for row in range(2, total_rows_in_sheet):
                try:
                    cell_val = ws.cell(row=row, column=actual_col).value
                    if cell_val is not None:
                        max_length = max(max_length, len(str(cell_val)))
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, 8), 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        sheets_info.append({
            "name": sheet_name,
            "object_type": child_type,
            "row_count": len(data) + (empty_rows_count if include_operation_mode else 0)
        })

    def _apply_classification_fill(self, cell, classification: str) -> None:
        """统一根据 classification 设置单元格底色（SSOT）

        7 处数据写入块原本各自 inline 同样的 if-elif 链设置底色：
        - 主导出（4 处）：用列索引 in (parent_key_columns / create_required_columns / readonly_columns)
        - _write_child_sheet（2 处）：用 _classify_field() 返回的 classification 字符串

        本 helper 统一底色规则：
        - 'parent_key' → BUSINESS_KEY_FILL（浅绿，父对象业务键）
        - 'auto_or_manual_code' → AUTO_GEN_OR_MANUAL_FILL（浅蓝，自动/可手动，v1.1 NEW）
        - 'create_required' → REQUIRED_FILL（浅黄，新增必填）
        - 'readonly' → READONLY_FILL（灰色，只读）
        - 'editable' 或其他 → 不动（保持默认）

        注意：Protection（locked/unlocked）和 comment 不在本 helper 范围内，
        仍由调用方根据 context 决定（不同出口对 protect_sheet 行为不一致）。
        """
        if classification == 'parent_key':
            cell.fill = ExcelDesignSystem.BUSINESS_KEY_FILL
        elif classification == 'fk_display_code':
            # [NEW 2026-06-16 BMRD] FK 编码显示字段 - 浅绿色 (与 parent_key 同色, 表示"自动带出")
            cell.fill = ExcelDesignSystem.BUSINESS_KEY_FILL
        elif classification == 'auto_or_manual_code':
            # [NEW v1.1 2026-06-11] 自动/可手动模式底色（浅蓝）
            cell.fill = ExcelDesignSystem.AUTO_GEN_OR_MANUAL_FILL
        elif classification == 'create_required':
            cell.fill = ExcelDesignSystem.REQUIRED_FILL
        elif classification == 'readonly':
            cell.fill = ExcelDesignSystem.READONLY_FILL

    def _build_bo_display_maps(self, sheet_data: List[Dict[str, Any]],
                                bo_display_fields: Dict[str, Dict[str, str]]) -> Dict[str, Dict[Any, str]]:
        """统一批量构建 BO 显示名映射（SSOT）

        2 处原 inline 重复（export_selected_types L786, export_cascade L1148）：
        遍历 bo_display_fields，对每个 field_id 收集 record_ids 后
        调 BOEngine.get_record 查 target_bo 的 display_field。

        Args:
            sheet_data: 当前 sheet 的数据
            bo_display_fields: {field_id: {target_bo, display_field, value_field}}

        Returns:
            {field_id: {record_id: display_string}} 字典
        """
        bo_display_maps = {}
        if not bo_display_fields or not sheet_data:
            return bo_display_maps
        # [FR-010] 批量 SQL 查询替代逐条 BOEngine.get_record（N+1 → 1 次 SQL）
        from meta import get_meta_object
        for field_id, vh_info in bo_display_fields.items():
            record_ids = list(set(
                r.get(field_id) for r in sheet_data
                if r.get(field_id) is not None
            ))
            if not record_ids:
                continue
            try:
                target_meta = get_meta_object(vh_info['target_bo'])
                if not target_meta:
                    continue
                target_table = target_meta.table_name or vh_info['target_bo'] + 's'
                display_field = vh_info['display_field']

                placeholders = ','.join(['?'] * len(record_ids))
                sql = f"SELECT id, {display_field} FROM {target_table} WHERE id IN ({placeholders})"
                cursor = self.data_source.execute(sql, list(record_ids))

                display_map = {}
                for row in cursor.fetchall():
                    display_map[row[0]] = str(row[1]) if row[1] else ''
                if display_map:
                    bo_display_maps[field_id] = display_map
            except Exception:
                # 整个 field 失败不影响其他 field
                continue
        return bo_display_maps

    def _build_export_filename(self, prefix_parts: List[str], timestamp: Optional[str] = None) -> str:
        """统一生成导出文件名（SSOT）

        3 处原 inline 重复（export_to_excel L278 / export_selected_types L1033 /
        export_cascade L1360）漂移已开始（不同地方转义规则可能不一致）。
        本 helper 统一规则。

        Args:
            prefix_parts: 文件名前缀部分（产品名、版本名等），按顺序用 _ 连接
            timestamp: 时间戳字符串（默认用当前时间 YYYYMMDD_HHMMSS）

        Returns:
            文件名字符串（不含路径），如 "TEST15_V1_20260608_230000.xlsx"
        """
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 安全字符：仅保留字母数字 + _ -
        safe_parts = [
            "".join(c if c.isalnum() or c in '_-' else '_' for c in part)
            for part in prefix_parts
            if part  # 跳过空值
        ]
        if safe_parts:
            return "{0}_{1}.xlsx".format("_".join(safe_parts), timestamp)
        return "{0}.xlsx".format(timestamp)

    # 全局导出唯一特殊菜单（"架构数据管理"），固定前缀"架构数据"
    # 2026-06-19 规范：导出的 Excel / 导入模版文件名按 objectname 拼前缀，
    # 全局导出（arch-data 菜单的级联导出）前缀固定为"架构数据"，其他场景用 objectname
    GLOBAL_MENU_PREFIX_MAP = {
        'arch-data': '架构数据',
    }

    def _resolve_cascade_prefix(self, object_type: str, menu_code: Optional[str] = None) -> str:
        """解析级联导出的文件名前缀（SSOT）

        优先级：
        1. menu_code 在 GLOBAL_MENU_PREFIX_MAP 中 → 用映射的固定前缀（如 'arch-data' → '架构数据'）
        2. 否则 → registry.get(object_type).name（中文 objectname）
        3. 都没有 → object_type（英文 id 兜底）

        Args:
            object_type: 级联导出起始对象类型
            menu_code: 发起导出的菜单编码（可选，前端从 route 推导）

        Returns:
            文件名前缀字符串
        """
        if menu_code and menu_code in self.GLOBAL_MENU_PREFIX_MAP:
            return self.GLOBAL_MENU_PREFIX_MAP[menu_code]
        meta_obj = registry.get(object_type)
        if meta_obj and getattr(meta_obj, 'name', None):
            return meta_obj.name
        return object_type

    def _resolve_object_names(self, object_types: List[str]) -> List[str]:
        """根据 object_type 列表解析为中文名列表（SSOT）

        每个 type 取 registry.get(type).name，找不到则用原 type 作为兜底。
        用于 export_selected_types / export_template 的多对象文件名拼接。

        Args:
            object_types: object_type id 列表，如 ['business_object', 'relationship']

        Returns:
            中文名列表，如 ['业务对象', '关系']
        """
        names: List[str] = []
        for ot in object_types:
            if not ot:
                continue
            meta_obj = registry.get(ot)
            if meta_obj and getattr(meta_obj, 'name', None):
                names.append(meta_obj.name)
            else:
                names.append(ot)
        return names

    def _get_value_help(self, meta_field):
        """统一获取字段的 value_help 配置（SSOT）

        [FR-007] 委托到 meta.core.value_help_accessor.get_value_help，
        与 manage_service / EnrichmentEngine 统一。
        """
        from meta.core.value_help_accessor import get_value_help
        return get_value_help(meta_field)

    def _write_meta_sheet_header(self, ws_meta, title: str, time_str: str,
                                  included_names: List[str]) -> None:
        """统一写入说明 sheet 头部（标题 + 时间 + 包含对象）

        SSOT: 3 处 export 入口（export_template / export_selected_types /
        export_cascade）原本各自 inline 同样的 header 写入。
        """
        ws_meta.cell(row=1, column=1, value=title).font = Font(
            bold=True, size=14, color=ExcelDesignSystem.PRIMARY_COLOR
        )
        ws_meta.row_dimensions[1].height = 24
        ws_meta.cell(row=2, column=1, value="生成时间").font = ExcelDesignSystem.LABEL_FONT
        ws_meta.cell(row=2, column=2, value=time_str).font = ExcelDesignSystem.VALUE_FONT
        ws_meta.cell(row=3, column=1, value="包含对象").font = ExcelDesignSystem.LABEL_FONT
        ws_meta.cell(row=3, column=2, value=", ".join(included_names)).font = ExcelDesignSystem.VALUE_FONT

    def _write_meta_sheet_operations(self, ws_meta, has_cud: bool,
                                     all_readonly: bool, start_row: int,
                                     is_cascade: bool = False,
                                     has_child_sheets: bool = False,
                                     objects=None) -> int:
        """统一写入操作说明 section（SSOT）

        3 种模式：
        - has_cud=True: 完整操作说明（操作模式、颜色示例、业务关键字、父对象编码、删除/新增/冲突/注意）
        - all_readonly=True: 只读说明（导出说明 + 使用说明）
        - 混合（部分对象支持 CUD）: 简短注意

        Args:
            ws_meta: openpyxl Worksheet
            has_cud: 是否有任何对象支持 CUD
            all_readonly: 是否所有对象都只读
            start_row: 起始行
            is_cascade: 是否级联导出（影响"注意事项"文案）
            has_child_sheets: 是否包含子对象 sheet（selected_types 模式）

        Returns:
            最后一个写入的 row 编号（用于 caller 决定 border 范围）
        """
        ds = ExcelDesignSystem
        if has_cud:
            row = start_row
            # [REWRITE 2026-06-16 BMRD] 操作说明 section 化, 整合所有操作相关内容
            ws_meta.cell(row=row, column=1, value="操作说明").font = ds.SECTION_FONT
            ws_meta.cell(row=row, column=1).fill = ds.SECTION_FILL
            row += 1

            # 操作模式
            ws_meta.cell(row=row, column=1, value="操作模式").font = ds.LABEL_FONT
            ws_meta.cell(row=row, column=2, value="create - 新增/update - 更新/delete - 删除，留空默认为 create").font = ds.VALUE_FONT
            row += 1

            # [MERGE 2026-06-16 BMRD] 之前分散在"业务说明"中的"删除操作"内容合并到这里
            ws_meta.cell(row=row, column=1, value="更新删除").font = ds.LABEL_FONT
            ws_meta.cell(row=row, column=2, value="在操作模式列选择'删除' 或'更新'，系统按业务编码找到对应记录并删除，或更新").font = ds.VALUE_FONT
            row += 1

            # [MERGE 2026-06-16 BMRD] 之前"业务说明"中的"冲突处理策略"合并到这里
            ws_meta.cell(row=row, column=1, value="冲突处理").font = ds.LABEL_FONT
            ws_meta.cell(row=row, column=2, value="导入时如记录已存在，可选择'更新'或'跳过'").font = ds.VALUE_FONT
            row += 1

            # 注意事项（cascade 模式特殊文案）
            if is_cascade:
                notice = "请勿修改灰色背景单元格的值，否则导入时会忽略这些字段。级联导出不含子对象Sheet"
            else:
                notice = "请勿修改灰色背景单元格的值，否则导入时会忽略这些字段"
            ws_meta.cell(row=row, column=1, value="注意事项").font = ds.LABEL_FONT
            ws_meta.cell(row=row, column=2, value=notice).font = ds.VALUE_FONT
            row += 1

            # [REWRITE 2026-06-16 BMRD] 单元格颜色 section 化 (独立 section, 蓝色填充)
            # 之前 "单元格颜色" 只是 LABEL_FONT 的描述行, 不像 "自动编码规则" 那样是独立 section
            ws_meta.cell(row=row, column=1, value="单元格颜色").font = ds.SECTION_FONT
            ws_meta.cell(row=row, column=1).fill = ds.SECTION_FILL
            ws_meta.cell(row=row, column=2, value="不同颜色背景表示不同的字段控制：").font = ds.VALUE_FONT
            row += 1

            # 颜色示例
            # [FIX 2026-06-16 BMRD] 浅绿色描述补充 "FK 编码显示字段（自动带出）"
            # 之前只说"父对象编码", 漏掉了 source_bo_code / target_bo_code 等 parent_key_display 字段
            # [FIX 2026-06-16 BMRD] 移除"浅蓝色 - 新增操作行"项
            # 原因: 2026-06-14 BMRD 已删除 CREATE_NEW_FILL (新增操作行不再使用底色)
            # 替代: 保留"create - 新增"文字 + 底部空白位置作为视觉提示
            color_examples = [
                ("  灰色", ds.READONLY_FILL, "只读字段，不可编辑"),
                ("  浅绿色", ds.BUSINESS_KEY_FILL,
                 "父对象编码字段（必填或自动带出）;FK 编码显示字段（自动带出，无需填写）"),
                ("  浅黄色", ds.REQUIRED_FILL, "业务关键字，新增必填，编辑时只读"),
                ("  浅蓝灰", ds.AUTO_GEN_OR_MANUAL_FILL, "自动/可手动编码，留空由系统生成（v1.1）"),
            ]
            for label, fill, desc in color_examples:
                ws_meta.cell(row=row, column=1, value=label).font = ds.LABEL_FONT
                ws_meta.cell(row=row, column=1).fill = fill
                ws_meta.cell(row=row, column=2, value=desc).font = ds.VALUE_FONT
                row += 1

            # [REMOVED 2026-06-16 BMRD] 删除"自动编码规则" section
            # 原因: 编码规则和示例已在 sheet 列头 hover 即可看到 (B1/G1 编码列 comment)
            # 在说明页中重复显示既冗余又容易脱节
            # [REMOVED 2026-06-16 BMRD] 删除"业务说明" section
            # 原因: 业务关键字/父对象编码的描述也已合并到列头 comment 中
            # 重复的"删除操作/新增操作/冲突处理"已合并到"操作说明"section
            # [REMOVED 2026-06-16 BMRD] 删除"子对象Sheet" 区块
            # 原因: 已在"操作说明"中通过"更新删除"行覆盖了子对象的操作方式

            return row - 1
        elif all_readonly:
            row = start_row
            ws_meta.cell(row=row, column=1, value="说明").font = ds.SECTION_FONT
            ws_meta.cell(row=row, column=1).fill = ds.SECTION_FILL
            row += 1
            ws_meta.cell(row=row, column=1, value="导出说明").font = ds.LABEL_FONT
            ws_meta.cell(row=row, column=2, value="当前导出的对象为只读数据，不支持导入修改。").font = ds.VALUE_FONT
            row += 1
            ws_meta.cell(row=row, column=1, value="使用说明").font = ds.LABEL_FONT
            ws_meta.cell(row=row, column=2, value="此模板仅用于查看和导出数据，如需修改请通过系统界面操作。").font = ds.VALUE_FONT
            row += 1
            return row - 1
        else:
            row = start_row
            ws_meta.cell(row=row, column=1, value="说明").font = ds.SECTION_FONT
            ws_meta.cell(row=row, column=1).fill = ds.SECTION_FILL
            row += 1
            ws_meta.cell(row=row, column=1, value="注意").font = ds.LABEL_FONT
            ws_meta.cell(row=row, column=2, value="部分对象不支持导入修改，仅支持导出。").font = ds.VALUE_FONT
            row += 1
            return row - 1

    def _finalize_meta_sheet(self, ws_meta, last_row: int) -> None:
        """统一收尾说明 sheet：边框 + 列宽"""
        ds = ExcelDesignSystem
        for row in range(1, last_row + 1):
            for col in range(1, 3):
                ws_meta.cell(row=row, column=col).border = ds.THIN_BORDER
        ws_meta.column_dimensions['A'].width = 18
        ws_meta.column_dimensions['B'].width = 60

    def _iter_parent_chain(self, meta_obj, stop_at_context_field: bool = False):
        """统一遍历 parent_object 链（SSOT）

        4 处原 inline 循环独立实现，漂移已开始：
        - L1763 _build_parent_fk_columns: 已 SSOT 化但 inline 写法
        - L3055 _add_hierarchy_fields: 在 context_field 停止
        - L3245 _get_export_headers: 简单遍历（dead code）
        - L3512 _get_hierarchy_field_names: 简单遍历

        本 helper 统一链遍历规则，让调用方决定每层做什么。

        Args:
            meta_obj: 起始对象
            stop_at_context_field: 是否在 context_field 边界停止
                - True: _add_hierarchy_fields 模式（context_field 之外的数据通过导入界面选择）
                - False: 向上完整遍历（_get_hierarchy_field_names 模式）

        Yields:
            (current_obj, parent_obj, parent_fk_field_id, is_context_boundary)
            - current_obj: 当前层对象（meta_obj 起，第一层是 meta_obj 本身）
            - parent_obj: registry 查到的父对象
            - parent_fk_field_id: 如 'product_id' / 'domain_id'
            - is_context_boundary: True 表示这是 context_field 边界
        """
        current_obj = meta_obj
        while current_obj and getattr(current_obj, 'parent_object', None):
            parent_obj = registry.get(current_obj.parent_object)
            if not parent_obj:
                break
            parent_fk_field_id = "{0}_id".format(current_obj.parent_object)
            parent_key_field = current_obj.get_field(parent_fk_field_id)
            is_context_field = parent_key_field and getattr(
                parent_key_field.semantics, 'context_field', False
            )
            is_context_boundary = is_context_field and stop_at_context_field
            yield (current_obj, parent_obj, parent_fk_field_id, is_context_boundary)
            if is_context_boundary:
                break
            current_obj = parent_obj

    def _parse_operation_mode_from_label(self, cell_value: Any) -> Optional[str]:
        """统一解析"操作模式" cell 值为 operation_mode key（SSOT）

        2 处 import 路径（_validate_sheets L3759, _import_sheet L4357）原本 inline
        同样的 if-elif 链，漂移已开始（_import_sheet 多 1 个 else 分支、
        _validate_sheets 多 skip 分支）。本 helper 统一。

        支持的格式:
        - "create - 新增" / "update - 更新" / "delete - 删除" / "skip - 跳过"（含 label）
        - "create" / "新增"（纯 key）
        - 大小写不敏感（"Create" / "CREATE" 都接受）
        - 中英文 key 都接受（"create" / "新增" 都映射到 "create"）

        Args:
            cell_value: Excel 单元格值（任意类型，自动转 str）

        Returns:
            operation_mode key: "create" / "update" / "delete" / "skip"
            如果无法识别返回 None（调用方决定 fallback）
        """
        if cell_value is None:
            return None
        s = str(cell_value).strip()
        if not s:
            return None
        # 解析 "Key - Label" 格式
        if ' - ' in s:
            key_part = s.split(' - ')[0].strip().lower()
        else:
            key_part = s.strip().lower()

        if key_part in ["新增", "插入", "create", "insert"]:
            return "create"
        if key_part in ["删除", "delete"]:
            return "delete"
        if key_part in ["跳过", "skip"]:
            return "skip"
        if key_part in ["更新", "update"]:
            return "update"
        return None  # 无法识别

    def _create_operation_mode_dv(self, ws, with_prompt: bool = False,
                                  verbose_error: bool = False) -> DataValidation:
        """统一创建"操作模式" DataValidation（SSOT）

        4 处 export 入口（export_template / export_selected_types /
        export_cascade / _write_child_sheet）原本各自 inline 同样的 DV 创建，
        漂移已开始（error 文案不一致、prompt 配置不一致）。本 helper 统一。

        Args:
            ws: openpyxl Worksheet
            with_prompt: 是否添加 prompt（鼠标悬停提示）
                - export_cascade / _write_child_sheet 用 True
                - export_template / export_selected_types 用 False
            verbose_error: error 文案是否详细（列出 create/update/delete 三个选项）
                - 默认 False（简略），与历史行为对齐

        Returns:
            DataValidation（同时已 add_data_validation 到 ws）
        """
        operation_dv = DataValidation(
            type="list",
            formula1='"create - 新增,update - 更新,delete - 删除"',
            allow_blank=True,
        )
        if verbose_error:
            operation_dv.error = "请从下拉列表中选择：create - 新增/update - 更新/delete - 删除"
        else:
            operation_dv.error = "请从下拉列表中选择操作模式"
        operation_dv.errorTitle = "无效输入"
        if with_prompt:
            operation_dv.prompt = "选择操作模式"
            operation_dv.promptTitle = "操作模式"
        ws.add_data_validation(operation_dv)
        return operation_dv

    def _build_enum_dv_values(self, field) -> Optional[str]:
        """构造字段的 DataValidation 下拉值（统一优先：value_help > enum_values > ui.options）

        返回 openpyxl DataValidation 的 formula1 字符串（如 '"a,b,c"'），
        若字段没有可枚举的值则返回 None。
        """
        candidates = []

        vh = self._get_value_help(field)
        if vh:
            source = getattr(vh, 'source', None)
            if source and getattr(source, 'type', None) == 'enum':
                enum_type_id = getattr(source, 'enum_type_id', None)
                if enum_type_id:
                    enum_map = self._get_enum_value_map_from_value_help(field)
                    if enum_map:
                        candidates.extend(f"{k} - {v}" for k, v in enum_map.items())

        if not candidates and getattr(field, 'enum_values', None):
            candidates.extend(
                f"{ev.get('value')} - {ev.get('label', ev.get('value'))}"
                for ev in field.enum_values
            )

        if not candidates:
            ui = getattr(field, 'ui', None)
            if ui and getattr(ui, 'options', None):
                candidates.extend(
                    f"{opt.get('value')} - {opt.get('label', opt.get('value'))}"
                    if isinstance(opt, dict) else str(opt)
                    for opt in ui.options
                )

        if not candidates:
            return None

        safe_values = []
        for v in candidates:
            v = str(v).replace('"', "'")
            safe_values.append(v)

        return '"' + ','.join(safe_values) + '"'

    def _query_direct_fk_child(self, child_type: str, child_meta, parent_type: str,
                                filters: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """查询直接 FK 关联的子对象数据

        根据子对象的 parent_object 字段，反查父对象数据并获取其子对象。

        Args:
            child_type: 子对象类型 ID
            child_meta: 子对象的 MetaObject
            parent_type: 父对象类型 ID
            filters: 导出过滤条件

        Returns:
            List[Dict]: 子对象数据列表
        """
        parent_meta = registry.get(parent_type)
        if not parent_meta:
            return []

        parent_table = parent_meta.table_name or parent_type + 's'
        child_table = child_meta.table_name or child_type + 's'

        version_id = (filters or {}).get('version_id')
        parent_fk_field = child_meta.parent_object + '_id'

        # [FIX FR-001] 软删除过滤：排除已删除的父记录
        soft_delete_filter = ""
        if any(f.id == 'is_deleted' for f in parent_meta.fields):
            soft_delete_filter = " AND (p.is_deleted IS NULL OR p.is_deleted = 0)"

        # [FIX FR-002] 数据权限过滤
        perm_filter_sql, perm_params = self._build_permission_filter(child_type, 'c')

        try:
            if version_id:
                sql = (f"SELECT c.* FROM {child_table} c "
                       f"INNER JOIN {parent_table} p ON c.{parent_fk_field} = p.id "
                       f"WHERE p.version_id = ?{soft_delete_filter}{perm_filter_sql} "
                       f"ORDER BY c.id ASC")
                params = [version_id] + perm_params
            else:
                sql = (f"SELECT c.* FROM {child_table} c "
                       f"INNER JOIN {parent_table} p ON c.{parent_fk_field} = p.id"
                       f"{soft_delete_filter}{perm_filter_sql} "
                       f"ORDER BY c.id ASC")
                params = perm_params

            cursor = self.data_source.execute(sql, tuple(params))
            columns = [desc[0] for desc in cursor.description]
            data = [dict(zip(columns, row)) for row in cursor.fetchall()]

            # [FIX 2026-06-08] 计算 list 中配置的 computed 字段（如 child_count）
            # _query_direct_fk_child 走直接 SQL 路径，没经过 query_service.search()
            # 所以需要手动调用 _compute_list_computed_fields
            if data:
                try:
                    self._compute_list_computed_fields_for_export(child_meta, data)
                    # [NEW 2026-06-16 BMRD] 计算 hierarchy_scope 虚拟字段 (category_label/category_type)
                    self._compute_hierarchy_scope_for_export(child_meta, data)
                except Exception as e:
                    logger.warning(
                        f"[Export] 子对象 {child_type} 计算字段失败: {e}"
                    )

            return data
        except Exception as e:
            logger.warning(
                f"[Export] 查询子对象 {child_type} 失败: {e}"
            )
            return []

    def _build_permission_filter(self, object_type: str, table_alias: str = '') -> tuple:
        """[FR-002] 构建数据权限过滤 SQL 片段

        复用 query_service._apply_data_permission 的逻辑，
        获取当前用户的 allowed_ids 并转为 SQL WHERE 条件。

        Returns:
            (sql_fragment, params) — sql_fragment 如 " AND c.id IN (?, ?, ?)"
            无权限限制时返回 ("", [])
        """
        try:
            from meta.services.auth_middleware import get_current_user, is_admin
            from meta.services.data_permission_service import DataPermissionService

            user = get_current_user()
            if not user or is_admin(user):
                return "", []

            user_id = user.get('user_id')
            if not user_id:
                return "", []

            perm_service = DataPermissionService(self.data_source)
            allowed_ids = perm_service.get_allowed_resource_ids(user_id, object_type)

            if allowed_ids is None:
                # None 表示无权限配置，允许全部
                return "", []

            if not allowed_ids:
                # 空列表表示无任何权限，返回不可能匹配的条件
                prefix = f"{table_alias}." if table_alias else ""
                return f" AND {prefix}id = -1", []

            prefix = f"{table_alias}." if table_alias else ""
            placeholders = ','.join(['?'] * len(allowed_ids))
            return f" AND {prefix}id IN ({placeholders})", list(allowed_ids)

        except Exception as e:
            logger.warning(f"[FR-002] 数据权限过滤构建失败: {e}")
            return "", []

    def _query_annotations_impl(self, parent_types: List[str],
                                 filters: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """查询备注信息（内部实现，供 _query_child_object 调用）"""
        from meta.core.models import registry

        category_labels = {
            "important": "重要",
            "warning": "警告",
            "info": "信息",
            "tip": "提示"
        }

        type_labels = self._build_type_labels()

        try:
            conditions = [QueryCondition(field="target_type", operator=QueryOperator.IN, values=parent_types)]

            search_request = SearchRequest(
                object_type="annotation",
                conditions=conditions,
                page=1,
                page_size=100000,
            )
            result = self.query_service.search(search_request)
            annotations = result.data or []

            enriched_annotations = []
            for ann in annotations:
                target_type = ann.get("target_type", "")
                target_id = ann.get("target_id")

                target_code = ""
                target_name = ""

                if target_type and target_id:
                    try:
                        target_obj = registry.get(target_type)
                        if target_obj:
                            target_search = SearchRequest(
                                object_type=target_type,
                                conditions=[QueryCondition(field="id", operator=QueryOperator.EQ, value=target_id)],
                                page=1,
                                page_size=1,
                            )
                            target_result = self.query_service.search(target_search)
                            if target_result.data:
                                target_data = target_result.data[0]

                                bk_fields = [f for f in target_obj.fields
                                            if getattr(f.semantics, 'business_key', False)
                                            and not getattr(f.semantics, 'virtual', False)]
                                name_field = next((f for f in target_obj.fields if f.id == "name" or "name" in f.id.lower()), None)

                                if target_type == "relationship":
                                    target_code = target_data.get("relation_code", "")
                                    target_name = target_data.get("relation_desc", "") or " -> ".join(filter(None, [target_data.get("source_code", ""), target_data.get("target_code", "")]))
                                elif bk_fields:
                                    target_code = target_data.get(bk_fields[0].id, "")
                                    target_name = target_data.get(name_field.id, "") if name_field else ""
                                else:
                                    target_code = target_data.get("code", "")
                                    target_name = target_data.get("name", "")
                    except Exception:
                        pass

                created_at_raw = ann.get("created_at", "")
                created_at_formatted = ""
                if created_at_raw:
                    try:
                        from datetime import datetime
                        if isinstance(created_at_raw, datetime):
                            created_at_formatted = created_at_raw.strftime("%Y-%m-%d %H:%M")
                        elif isinstance(created_at_raw, str):
                            try:
                                dt = datetime.fromisoformat(created_at_raw.replace("Z", "+00:00"))
                                created_at_formatted = dt.strftime("%Y-%m-%d %H:%M")
                            except ValueError:
                                created_at_formatted = created_at_raw
                        else:
                            created_at_formatted = str(created_at_raw)
                    except Exception:
                        created_at_formatted = str(created_at_raw)

                enriched_annotations.append({
                    "id": ann.get("id"),
                    "target_type": target_type,
                    "target_type_label": type_labels.get(target_type, target_type),
                    "target_code": target_code,
                    "target_name": target_name,
                    "category": ann.get("category", ""),
                    "category_label": category_labels.get(ann.get("category", ""), ann.get("category", "")),
                    "content": ann.get("content", ""),
                    "created_at": created_at_formatted,
                    "created_by": ann.get("created_by", "")
                })

            return enriched_annotations
        except Exception:
            return []

    def _query_annotations(self, object_types: List[str], filters: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """查询备注信息（已废弃，请使用 _query_child_object）

        保留此方法用于向后兼容。
        """
        return self._query_annotations_impl(object_types, filters)

    def _get_parent_record(self, object_type: str, record_id: int) -> Dict[str, Any]:
        """获取父记录"""
        try:
            search_request = SearchRequest(
                object_type=object_type,
                conditions=[QueryCondition(field="id", operator="eq", value=record_id)],
                page=1,
                page_size=1,
            )
            result = self.query_service.search(search_request)
            return result.data[0] if result.data else {}
        except Exception:
            return {}

    def _add_hierarchy_fields(self, record: Dict[str, Any], meta_obj: MetaObject, options: Dict[str, Any]):
        """添加层级字段（编码和名称）- 向上追溯所有父对象

        注意：遇到 context_field 时停止追溯，这是上下文边界
        """
        current_record = record
        for current_obj, parent_obj, parent_fk_field_id, is_context_boundary in self._iter_parent_chain(
            meta_obj, stop_at_context_field=True
        ):
            # helper 已在 is_context_boundary=True 时停止
            parent_id = current_record.get(parent_fk_field_id)
            if not parent_id:
                break
            parent_record = self._get_parent_record(parent_obj.id, parent_id)
            bk_fields = [f for f in parent_obj.fields
                        if getattr(f.semantics, 'business_key', False)
                        and not getattr(f.semantics, 'virtual', False)]
            code = parent_record.get(bk_fields[0].id, "") if bk_fields else parent_record.get("code", "")
            record["{0}编码".format(parent_obj.name)] = code
            name = parent_record.get("name", "")
            record["{0}名称".format(parent_obj.name)] = name
            current_record = parent_record

    def _add_hierarchy_ids(self, record: Dict[str, Any], meta_obj: MetaObject):
        """添加层级ID列（保留兼容性）"""
        pass

    def _add_hierarchy_names(self, record: Dict[str, Any], meta_obj: MetaObject):
        """添加层级名称列（保留兼容性）"""
        pass

    def _enrich_relationship_record(self, record: Dict[str, Any]):
        """填充关系记录的 virtual 字段 - 兼容入口"""
        self._enrich_association_record('relationship', record)

    def _enrich_association_record(self, object_type: str, record: Dict[str, Any]):
        """填充 association 记录的 virtual 字段（配置驱动）

        从 hierarchies.yaml 的 association_filter_config 读取 source/target 列名和实体类型，
        动态填充层级名称字段。
        """

        from meta.services.config_driven_hierarchy_filter import HierarchyConfigLoader
        filter_config = HierarchyConfigLoader.get_association_filter_config(object_type)
        if not filter_config:
            return

        source_prefix = filter_config.get('source_prefix', 'source')
        target_prefix = filter_config.get('target_prefix', 'target')
        entity_id_field = filter_config.get('entity_id_field', 'bo_id')
        source_entity = HierarchyConfigLoader.get_level_by_object(object_type).get('source_entity', '')

        source_id = record.get(source_prefix + '_' + entity_id_field)
        target_id = record.get(target_prefix + '_' + entity_id_field)
        relation_code = record.get('relation_code')

        if relation_code:
            enum_info = self._get_enum_value_info('relation_type', relation_code)
            if enum_info:
                record['relation_type_name'] = enum_info.get('name', relation_code)
                record['relation_type_name_en'] = enum_info.get('name_en', '')
                if enum_info.get('dimensions'):
                    record['relation_type_dimensions'] = enum_info['dimensions']
            else:
                record['relation_type_name'] = relation_code

        if source_id and source_entity:
            source_entity_data = self._get_entity_with_hierarchy(source_entity, source_id)
            if source_entity_data:
                record[source_prefix + '_bo_name'] = source_entity_data.get('name', '')
                record[source_prefix + '_bo_code'] = source_entity_data.get('code', '')
                record[source_prefix + '_code'] = record.get(source_prefix + '_code') or source_entity_data.get('code', '')
                for key, value in source_entity_data.items():
                    if key.endswith('_name') and key != 'name':
                        record[source_prefix + '_' + key] = value

        if target_id and source_entity:
            target_entity_data = self._get_entity_with_hierarchy(source_entity, target_id)
            if target_entity_data:
                record[target_prefix + '_bo_name'] = target_entity_data.get('name', '')
                record[target_prefix + '_bo_code'] = target_entity_data.get('code', '')
                record[target_prefix + '_code'] = record.get(target_prefix + '_code') or target_entity_data.get('code', '')
                for key, value in target_entity_data.items():
                    if key.endswith('_name') and key != 'name':
                        record[target_prefix + '_' + key] = value

    def _get_bo_by_id(self, bo_id: int) -> Dict[str, Any]:
        """根据 ID 获取业务对象（包含层级信息）- 配置驱动动态 JOIN"""
        return self._get_entity_with_hierarchy('business_object', bo_id)

    def _get_entity_with_hierarchy(self, object_type: str, entity_id: int) -> Dict[str, Any]:
        """根据 ID 获取实体对象（包含层级名称信息）

        从 hierarchies.yaml 的层级链动态构建 LEFT JOIN SQL，
        而非硬编码 bo → sm → sd → d 四表 JOIN。
        """
        from meta.services.config_driven_hierarchy_filter import HierarchyConfigLoader

        meta_obj = registry.get(object_type)
        if not meta_obj:
            return {}

        table_name = meta_obj.table_name
        levels = HierarchyConfigLoader.get_levels()

        object_to_level = {l.get('object'): l for l in levels}
        current = object_type
        join_parts = []
        select_parts = ["t0.id", "t0.code", "t0.name"]
        result_keys = ['id', 'code', 'name']
        table_idx = 0

        while current:
            level = object_to_level.get(current, {})
            parent_object = level.get('parent_object')
            if not parent_object or parent_object == 'version':
                break

            parent_level = object_to_level.get(parent_object, {})
            parent_table = parent_level.get('table_name')
            fk_field = level.get('foreign_key_field')

            if not parent_table or not fk_field:
                break

            table_idx += 1
            parent_alias = "t{0}".format(table_idx)
            child_alias = "t{0}".format(table_idx - 1)
            join_parts.append(
                "LEFT JOIN {0} {1} ON {2}.{3} = {1}.id".format(
                    parent_table, parent_alias, child_alias, fk_field
                )
            )
            name_key = parent_object + '_name'
            select_parts.append("{0}.name as {1}".format(parent_alias, name_key))
            result_keys.append(name_key)

            current = parent_object

        sql = "SELECT {0} FROM {1} t0 {2} WHERE t0.id = ? LIMIT 1".format(
            ', '.join(select_parts), table_name, ' '.join(join_parts)
        )

        try:
            cursor = self.data_source.execute(sql, (entity_id,))
            row = cursor.fetchone()
            if row:
                result = {}
                for i, key in enumerate(result_keys):
                    result[key] = row[i] if row[i] is not None else ''
                return result
        except Exception as e:
            logger.warning("[Enrichment] Failed to get %s by id %s: %s", object_type, entity_id, e)
        return {}

    def _get_export_headers(self, meta_obj: MetaObject, options: Optional[Dict[str, Any]]) -> List[str]:
        """获取导出表头"""
        options = options or {}
        include_hierarchy_path = options.get("include_hierarchy_path", True)
        include_readonly = options.get("include_readonly", True)

        hierarchy_fields = self._get_hierarchy_field_names(meta_obj)

        headers = []

        export_fields = sorted(
            [f for f in meta_obj.fields if f.semantics.export_visible and f.storage.value != "virtual"],
            key=lambda f: f.semantics.import_order
        )

        for f in export_fields:
            field_name = f.name or f.id
            if field_name not in hierarchy_fields:
                if include_readonly and not self._is_field_editable(f):
                    headers.append("{0}(只读)".format(field_name))
                else:
                    headers.append(field_name)

        if include_hierarchy_path:
            headers.append("层级路径")

        # [FR-009] 使用 _iter_parent_chain 替代 inline while 循环
        for current_obj, parent_obj, parent_fk_field_id, _ in \
                self._iter_parent_chain(meta_obj, stop_at_context_field=False):
            headers.append("{0}编码".format(parent_obj.name))
            headers.append("{0}名称".format(parent_obj.name))

        return headers

    def _get_key_template_user_editable(self, meta_obj) -> str:
        """[NEW v1.1 2026-06-11] 获取对象的 key_template user_editable 模式

        返回 'auto_only' / 'auto_or_manual' / 'manual_only'，未配置时返回空字符串。
        参数兼容 MetaObject / dict / str（object_type）。
        """
        if meta_obj is None:
            return ''
        if isinstance(meta_obj, str):
            try:
                meta_obj = registry.get(meta_obj)
                if meta_obj is None:
                    return ''
            except Exception:
                return ''
        if isinstance(meta_obj, dict):
            kt = meta_obj.get('key_template', {}) or {}
        else:
            kt = getattr(meta_obj, 'key_template', None) or {}
        if isinstance(kt, dict):
            return kt.get('user_editable', '')
        return ''

    def _get_key_template_info(self, meta_obj) -> dict:
        """[NEW 2026-06-14 BMRD] 获取对象的完整 key_template 信息, 用于表头 comment

        返回 dict 包含:
        - enabled: 是否启用 key_template
        - user_editable: 'auto_only' / 'auto_or_manual' / 'manual_only'
        - pattern: 编码格式模板 (如 '{source_code}-{target_code}-{SEQ:2}')
        - preview: 示例 (如 'PUM01')
        未配置时返回空 dict.

        [REWRITE 2026-06-16 BMRD] 5次反馈未生效的根因是: 此函数在 stash 恢复时丢失,
        重新实现并明确绑定到字段 comment.
        """
        if meta_obj is None:
            return {}
        if isinstance(meta_obj, str):
            try:
                meta_obj = registry.get(meta_obj)
                if meta_obj is None:
                    return {}
            except Exception:
                return {}
        if isinstance(meta_obj, dict):
            kt = meta_obj.get('key_template', {}) or {}
        else:
            kt = getattr(meta_obj, 'key_template', None) or {}
        if isinstance(kt, dict):
            return {
                'enabled': kt.get('enabled', False),
                'user_editable': kt.get('user_editable', ''),
                'auto_suggest': kt.get('auto_suggest', False),
                'pattern': kt.get('pattern', ''),
                'preview': kt.get('preview', ''),
            }
        return {}

    def _auto_generate_code_from_key_template(self, object_type: str, record: Dict[str, Any]) -> Optional[str]:
        """[NEW 2026-06-17] 使用 key_template 自动生成 code

        action_executor._do_create 不走 bo_framework 拦截器链，
        KeyTemplateInterceptor 不会执行，需要在导入时手动生成。

        Args:
            object_type: 对象类型
            record: 记录数据

        Returns:
            生成的 code，如果无法生成则返回 None
        """
        try:
            meta_obj = registry.get(object_type)
            if not meta_obj:
                return None

            kt = getattr(meta_obj, 'key_template', None) or {}
            if not isinstance(kt, dict) or not kt.get('enabled') or not kt.get('auto_suggest'):
                return None

            from meta.core.key_template_engine import KeyTemplateEngine, KeyTemplateConfig
            config = KeyTemplateConfig.from_dict(object_type, kt)
            if not config.enabled or not config.auto_suggest:
                return None

            engine = KeyTemplateEngine(self.data_source)

            # 准备 field_values（从 record 中提取）
            field_values = dict(record)
            logger.info(f"[KeyTemplate] Auto-generate code for {object_type}, field_values keys={list(field_values.keys())}, service_module_code={field_values.get('service_module_code')}, source_code={field_values.get('source_code')}")

            # 解析 pattern 中的父字段引用
            import re
            field_refs = re.findall(r'\{([a-zA-Z_][a-zA-Z0-9_]*)\}', config.pattern)
            for ref in field_refs:
                if ref.upper().startswith('SEQ'):
                    continue
                if ref not in field_values or not field_values[ref]:
                    # 尝试从 FK 解析获取
                    # 例如 source_code → source_bo_id → business_object.code
                    resolve_to = None
                    for field in meta_obj.fields:
                        if field.id == ref:
                            resolve_to = getattr(field.semantics, 'resolve_to_object', None)
                            break
                    if resolve_to:
                        # 从 resolve_from_field 获取值
                        for field in meta_obj.fields:
                            rff = getattr(field.semantics, 'resolve_from_field', None)
                            if rff == ref:
                                resolve_from_val = record.get(field.id)
                                if resolve_from_val:
                                    # 查询目标对象获取 code
                                    target_obj = registry.get(resolve_to)
                                    if target_obj:
                                        target_record = self.data_source.find_by_business_key(
                                            target_obj.table_name, resolve_from_val
                                        )
                                        if target_record:
                                            field_values[ref] = target_record.get('code') or target_record.get(ref)

            # 解析物理表名
            physical_table = meta_obj.table_name
            tokens = engine._parser.parse(config.pattern)
            prefix_filter = engine._parser.resolve_prefix(tokens, field_values)

            code = engine.generate_code(
                config, field_values, object_type,
                table_name=physical_table,
                prefix_filter=prefix_filter
            )
            return code
        except Exception as e:
            logger.warning(f"[Import] Key template auto-generate failed for {object_type}: {e}")
            return None

    def _is_field_editable(self, field, mode: str = 'edit') -> bool:
        """判断字段是否可编辑（用于导入导出）

        参考 SAP CDS View 字段控制逻辑：
        1. 系统字段：始终只读
        2. readonly_always 字段：始终只读（新建+编辑）
        3. 计算字段（computation.formula 或 semantics.computed）：始终只读
        4. virtual 字段（无 computation）：如果是搜索帮助则可编辑，否则只读
        5. immutable 字段：编辑时只读
        6. parent_key 字段：可编辑（SAP One Model 允许移动层级）
        7. ui.editable=false：始终只读
        8. [NEW v1.2.3 2026-06-17] parent_key + parent_key_template_editable 优先级最高
           - 区分"运行时不可改"(readonly_always) 与 "Excel 模板是否可填"(parent_key_template_editable)
           - 用户场景：service_module_id 是 parent_key + readonly_always，但批量重新分类需可填
        """
        readonly_field_ids = {'id', 'created_at', 'updated_at', 'created_by', 'updated_by'}

        if field.id in readonly_field_ids:
            return False

        # 计算字段（computation.formula）始终只读
        if hasattr(field, 'computation') and getattr(field.computation, 'formula', None):
            return False

        # 计算字段（semantics.computed）始终只读
        if getattr(field.semantics, 'computed', False):
            return False

        is_parent_key = getattr(field.semantics, 'parent_key', False)

        # [NEW v1.2.3 2026-06-17] parent_key_template_editable 优先级最高
        # 明确声明"Excel 模板可填"，与 readonly_always 完全解耦
        if is_parent_key:
            template_editable = getattr(field.semantics, 'parent_key_template_editable', None)
            if template_editable == 'always':
                return True
            if template_editable == 'create_only' and mode == 'create':
                return True
            if template_editable == 'never':
                return False
            # 未显式声明，fallback 到 readonly_always 判断
            if getattr(field.semantics, 'readonly_always', False):
                return False
            # parent_key 字段默认可编辑（SAP One Model 允许移动层级）
            if mode == 'create':
                return True
            # edit 模式: 再看 immutable
            if getattr(field.semantics, 'immutable', False):
                return False
            return True

        # readonly_always 始终只读（非 parent_key 字段）
        if getattr(field.semantics, 'readonly_always', False):
            return False

        # virtual 字段：如果是计算字段则只读，如果是外键/搜索帮助则可编辑
        if field.storage.value == 'virtual' or getattr(field.semantics, 'virtual', False):
            # 有 ui.relation 的 virtual 字段是外键/搜索帮助，可编辑
            if hasattr(field, 'ui') and hasattr(field.ui, 'relation') and field.ui.relation:
                pass  # 可编辑
            else:
                return False  # 计算字段，只读

        if mode == 'edit':
            if getattr(field.semantics, 'immutable', False):
                return False

        if hasattr(field, 'ui') and hasattr(field.ui, 'editable') and field.ui.editable is False:
            return False

        if hasattr(field.semantics, 'import_editable') and field.semantics.import_editable is False:
            return False

        return True
    
    def _filter_import_record(self, record: Dict[str, Any], meta_obj, operation_mode: str = 'update') -> Dict[str, Any]:
        """过滤导入记录，只保留可导入的字段
        
        导入规则（参考 _is_field_editable）：
        1. 系统字段（id, created_at 等）：新增时忽略，编辑时保留
        2. readonly_always 字段：始终忽略
        3. virtual 字段：始终忽略（除非有 ui.relation 作为搜索帮助字段，且操作是新增）
        4. immutable 字段：编辑模式下忽略（业务键不可修改）
        5. ui.editable=false：始终忽略
        """
        filtered = {}
        system_fields = {'id', 'created_at', 'updated_at', 'created_by', 'updated_by'}
        
        is_create = operation_mode in ["新增", "插入", "create", "insert", "Create", "Insert"]
        is_update = operation_mode in ["更新", "update", "Update"]
        # 删除和更新都需要保留业务键来定位记录
        is_need_bk = operation_mode in ["更新", "update", "Update", "删除", "delete", "Delete"]
        
        # [FIX v1.2.18j 2026-06-20] 预计算 FK 解析来源字段集合
        # 这些字段（如 relationship.source_code / target_code）即使 ui.editable=false
        # 也必须在过滤后保留，否则 FK resolve 拿不到值，导致 source_bo_id 无法解析。
        resolve_from_sources: set = set()
        for f in meta_obj.fields:
            rf = getattr(f.semantics, 'resolve_from_field', None)
            if rf:
                resolve_from_sources.add(rf)

        for field_id, value in record.items():
            if field_id in system_fields:
                if is_create:
                    continue  # 新增时忽略 id 等系统字段
                else:
                    filtered[field_id] = value  # 编辑时保留 id 等系统字段
                    continue

            field = meta_obj.get_field(field_id)
            if not field:
                continue

            # 业务键字段始终保留（用于定位记录）
            is_bk = getattr(field.semantics, 'business_key', False)
            # FK 解析来源字段同样必须保留（即使 ui.editable=false / virtual）
            is_resolve_source = field_id in resolve_from_sources
            is_parent_key = getattr(field.semantics, 'parent_key', False)

            # readonly_always 字段忽略（业务键 / FK 解析来源字段除外）
            if getattr(field.semantics, 'readonly_always', False) and not is_bk and not is_resolve_source:
                continue

            # ui.editable=false 字段忽略（业务键 / FK 解析来源字段除外）
            if hasattr(field, 'ui') and hasattr(field.ui, 'editable') and field.ui.editable is False and not is_bk and not is_resolve_source:
                continue

            # import_editable=false 字段忽略（业务键 / FK 解析来源字段除外）
            if hasattr(field.semantics, 'import_editable') and field.semantics.import_editable is False and not is_bk and not is_resolve_source:
                continue

            # virtual 字段的处理
            is_virtual = field.storage.value == 'virtual' or getattr(field.semantics, 'virtual', False)
            has_relation = hasattr(field, 'ui') and hasattr(field.ui, 'relation') and field.ui.relation

            if is_virtual:
                if (has_relation or is_parent_key) and is_create:
                    filtered[field_id] = value
                elif is_parent_key and is_need_bk:
                    filtered[field_id] = value
                elif is_resolve_source and value is not None:
                    # resolve_from_field 的虚拟字段:保留到 FK 解析后由后续步骤清掉
                    filtered[field_id] = value
                else:
                    continue

            # immutable 字段在编辑/更新/删除时忽略（业务键不可修改）
            # 但 business_key 字段必须保留，用于查找记录
            # [SYMBOL] 关键修复：parent_key 字段也必须保留，用于建立父子关系
            if is_update and getattr(field.semantics, 'immutable', False):
                is_bk = getattr(field.semantics, 'business_key', False)
                if not is_bk and not is_parent_key:
                    continue

            # 删除模式下，非 business_key 的 immutable 字段也忽略
            if operation_mode in ["删除", "delete", "Delete"] and getattr(field.semantics, 'immutable', False):
                is_bk = getattr(field.semantics, 'business_key', False)
                if not is_bk:
                    continue

            filtered[field_id] = value
        
        return filtered
    
    def _is_business_key(self, field) -> bool:
        """判断字段是否为业务键"""
        return hasattr(field.semantics, 'business_key') and field.semantics.business_key is True

    def _build_default_exclude_fields(self, meta_obj: MetaObject) -> set:
        """从 hierarchy 配置自动推导默认排除字段

        替代硬编码的 default_exclude_fields 集合。
        系统字段 + 层级 FK/名称字段 + association 虚拟字段均从配置推导。
        """
        from meta.services.config_driven_hierarchy_filter import HierarchyConfigLoader

        exclude = {
            'id', 'created_at', 'updated_at', 'created_by', 'updated_by',
            'version_id', 'product_id', 'version_name', 'product_name',
            'product_line', 'product_line_id', 'version_code',
            '版本编码', '版本名称', '产品线编码', '产品线名称',
            'ID'
        }

        levels = HierarchyConfigLoader.get_levels()
        for level in levels:
            obj = level.get('object')
            kind = level.get('kind', 'entity')
            display_name = level.get('display_name', '')

            if kind == 'entity':
                exclude.add(obj + '_id')
                exclude.add(obj + '_name')
                exclude.add(display_name)
                exclude.add(display_name + '编码')
                exclude.add(display_name + '名称')

        level_config = HierarchyConfigLoader.get_level_by_object(meta_obj.id)
        if level_config.get('kind') == 'association':
            filter_config = HierarchyConfigLoader.get_association_filter_config(meta_obj.id)
            if filter_config:
                for fl in filter_config.get('hierarchy_filter_levels', []):
                    src_col = fl.get('source_column', '')
                    tgt_col = fl.get('target_column', '')
                    if not fl.get('direct_filter'):
                        exclude.add(src_col)
                        exclude.add(tgt_col)

            for f in meta_obj.fields:
                if f.storage.value == 'virtual' and not getattr(f.semantics, 'export_visible', False):
                    exclude.add(f.id)
                    if f.name:
                        exclude.add(f.name)

        return exclude

    def _object_has_field(self, object_type: str, field_id: str) -> bool:
        """检查对象类型是否有指定字段"""
        obj = registry.get(object_type)
        if not obj:
            return False
        return any(f.id == field_id for f in obj.fields)

    def _should_export_field(self, meta_obj: MetaObject, field) -> bool:
        """判断字段是否应导出（统一规则）

        导出规则：
        1. 在默认排除列表中 → 不导出
        2. 敏感字段（sensitivity: restricted/confidential） → 不导出
        3. virtual 且无 ui → 不导出
        4. ui.hidden_in_list: true → 不导出（与列表显示保持一致）
        5. ui.hidden_in_form: true → 不导出（表单中隐藏，导出也应隐藏）
        6. 显式 export_visible: false → 不导出
        7. 显式 export_visible: true → 导出
        8. ui.visible: true → 导出
        9. 以上都不满足 → 导出（默认导出）
        """
        default_exclude_fields = self._build_default_exclude_fields(meta_obj)

        if field.id in default_exclude_fields:
            return False

        sensitivity = getattr(field.semantics, 'sensitivity', None)
        if sensitivity in ('restricted', 'confidential'):
            return False

        if field.storage.value == "virtual" and not hasattr(field, 'ui'):
            return False

        if hasattr(field, 'ui'):
            # 检查 ui.hidden_in_list - 与列表显示保持一致
            if hasattr(field.ui, 'hidden_in_list') and field.ui.hidden_in_list:
                return False
            # 检查 ui.hidden_in_form - 表单中隐藏，导出也应隐藏
            if hasattr(field.ui, 'hidden_in_form') and field.ui.hidden_in_form:
                return False
            # 检查 ui.visible - 显式设置为不可见
            if hasattr(field.ui, 'visible') and field.ui.visible is False:
                return False

        if hasattr(field.semantics, 'export_visible'):
            if field.semantics.export_visible is False:
                return False
            if field.semantics.export_visible is True:
                return True

        # [FIX 2026-06-11] ui.export_visible 作为 fallback（activity_label 等字段在此声明）
        if hasattr(field, 'ui') and hasattr(field.ui, 'export_visible'):
            if field.ui.export_visible is False:
                return False
            if field.ui.export_visible is True:
                return True

        if hasattr(field, 'ui') and hasattr(field.ui, 'visible'):
            return field.ui.visible is True

        return True

    def _get_hierarchy_field_names(self, meta_obj: MetaObject) -> set:
        """获取与层级关联的字段名集合（排除这些字段避免重复）

        只排除父对象的业务键字段，因为它们已通过层级编码/名称列显示。
        当前对象自己的业务键字段（如编码）应该保留。
        """
        hierarchy_fields = set()

        # [SSOT 2026-06-08] 使用 _iter_parent_chain helper
        for _current_obj, parent_obj, _parent_fk_field_id, _is_context_boundary in self._iter_parent_chain(meta_obj):
            hierarchy_fields.add("{0}_id".format(parent_obj.id))
            hierarchy_fields.add("{0}_name".format(parent_obj.id))
            hierarchy_fields.add("{0}ID".format(parent_obj.name))
            hierarchy_fields.add("{0}名".format(parent_obj.name))
            hierarchy_fields.add("{0}编码".format(parent_obj.name))
            hierarchy_fields.add("{0}名称".format(parent_obj.name))
            hierarchy_fields.add("{0}_id".format(parent_obj.name.lower()))
            hierarchy_fields.add("{0}_name".format(parent_obj.name.lower()))
            hierarchy_fields.add(parent_obj.name)

        return hierarchy_fields

    def _get_parent_key_headers(self, meta_obj: MetaObject, headers: List[str]) -> Dict[str, Dict[str, str]]:
        """获取父对象业务键表头映射
        
        返回格式：{ "父对象编码": { "parent_type": "parent_object_id", "id_field": "parent_id" } }
        例如：{ "服务模块编码": { "parent_type": "service_module", "id_field": "service_module_id" } }
        """
        result = {}
        
        if not meta_obj.parent_object:
            return result
        
        parent_obj = registry.get(meta_obj.parent_object)
        if not parent_obj:
            return result
        
        parent_code_header = "{0}编码".format(parent_obj.name)
        if parent_code_header in headers:
            result[parent_code_header] = {
                "parent_type": meta_obj.parent_object,
                "id_field": "{0}_id".format(meta_obj.parent_object)
            }
        
        return result

    def _header_to_field_id(self, meta_obj: MetaObject, header: str) -> str:
        """将表头转换为字段ID"""
        for f in meta_obj.fields:
            if f.name == header or f.id == header:
                return f.id
        return header

    def import_cascade(self, file_path: str, mode: str = "execute",
                       conflict_strategy: str = "upsert",
                       context: Optional[Dict[str, Any]] = None,
                       progress_callback: Optional[callable] = None) -> Any:
        """
        级联导入：从Excel文件导入多个对象类型的数据

        Args:
            file_path: Excel文件路径
            mode: 导入模式 (preview | execute)
            conflict_strategy: 冲突处理策略 (upsert | skip | replace)
            context: 导入上下文，包含version_id和product_id等
            progress_callback: 进度回调函数，接收 dict: {progress, current_type, current_type_name, total_types, current_index, message}

        Returns:
            ImportPreview 或 ImportResult
        """

        context = context or {}

        if not os.path.exists(file_path):
            trace_id = self._get_current_trace_id()
            if mode == "preview":
                return ImportPreview(trace_id=trace_id)
            return ImportResult(
                success=False,
                errors=[{"message": "File not found: {0}".format(file_path)}],
                trace_id=trace_id,
            )
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            return ImportPreview() if mode == "preview" else ImportResult(
                success=False, errors=[{"message": "Failed to read Excel: {0}".format(str(e))}]
            )
        
        sheets = []
        # [SYMBOL] 收集所有需要完整数据的Sheet（可能被其他Sheet引用）
        full_data_sheets = self._build_full_data_sheets()

        for sheet_name in wb.sheetnames:
            if sheet_name == "元数据":
                continue

            object_type = self._sheet_name_to_object_type(sheet_name)
            if not object_type:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) < 2:
                continue

            headers = [str(h).strip() if h is not None else "" for h in rows[0]]

            # [SYMBOL] 预览时：
            # - 业务对象等可能被引用的Sheet：收集所有数据（用于引用完整性检查）
            # - 其他Sheet：只取前10行
            if object_type in full_data_sheets:
                data_rows = [list(row) for row in rows[1:]]  # 收集所有行
                logger.info(f"[Import] 收集完整数据: {sheet_name} ({len(data_rows)} 行)")
            else:
                data_rows = [list(row) for row in rows[1:10]]  # 预览只取前10行

            sheets.append({
                "name": sheet_name,
                "object_type": object_type,
                "row_count": len(rows) - 1,
                "columns": headers,
                "preview_rows": data_rows
            })
        
        wb.close()

        import_order = self._sort_by_hierarchy([s["object_type"] for s in sheets])

        logger.info(f"[Import] sheets: {[s['object_type'] for s in sheets]}")
        logger.info(f"[Import] import_order: {import_order}")

        if mode == "preview":
            validation = self._validate_sheets(sheets, context)
            return ImportPreview(
                sheets=sheets,
                validation=validation,
                import_order=import_order,
                trace_id=self._get_current_trace_id(),
            )
        
        results = {}
        all_errors = []

        enabled_types = [ot for ot in import_order if registry.get(ot) and registry.get(ot).import_export.import_enabled]
        total_types = len(enabled_types)
        completed_count = 0
        # [FIX v1.2.16 2026-06-20] 用 enabled_types (而不是 import_order) 计算进度,
        # 否则 current_index/total_types 数字不匹配 (例如业务关系 0/5)
        # 保留 original_index 给 progress_callback 显示真实导入顺序 (供 log 用)
        enabled_with_idx = [(i, ot) for i, ot in enumerate(import_order)
                            if registry.get(ot) and registry.get(ot).import_export.import_enabled]

        if progress_callback:
            progress_callback({
                'progress': 0,
                'current_type': '',
                'current_type_name': '',
                'total_types': total_types,
                'current_index': 0,
                'message': '开始导入，共 {0} 种对象类型'.format(total_types)
            })

        for enabled_pos, (orig_i, ot) in enumerate(enabled_with_idx):
            sheet_info = next((s for s in sheets if s["object_type"] == ot), None)
            if not sheet_info:
                continue

            obj = registry.get(ot)
            if not obj or not obj.import_export.import_enabled:
                continue

            type_name = obj.name or ot
            current_index = enabled_pos + 1  # [FIX v1.2.16] 用 enabled 位置计算, 而不是 orig_i+1

            type_progress_base = int((enabled_pos / total_types) * 100) if total_types > 0 else 0
            type_progress_weight = int(100 / total_types) if total_types > 0 else 100

            if progress_callback:
                progress_callback({
                    'progress': type_progress_base,
                    'current_type': ot,
                    'current_type_name': type_name,
                    'total_types': total_types,
                    'current_index': current_index,
                    'message': '开始导入: {0} ({1}/{2})'.format(type_name, current_index, total_types)
                })

            obj_conflict_strategy = conflict_strategy
            if obj and obj.import_export and obj.import_export.conflict_strategy:
                obj_conflict_strategy = obj.import_export.conflict_strategy

            sheet_result = self._import_sheet(file_path, sheet_info, obj_conflict_strategy, context,
                                              progress_callback, type_progress_base, type_progress_weight)
            results[ot] = sheet_result
            completed_count += 1

            if sheet_result.get("errors"):
                all_errors.extend([
                    {"object_type": ot, **e} for e in sheet_result["errors"]
                ])

            if progress_callback:
                progress_callback({
                    'progress': min(99, int(((enabled_pos + 1) / total_types) * 100)) if total_types > 0 else 100,
                    'current_type': ot,
                    'current_type_name': type_name,
                    'total_types': total_types,
                    'current_index': current_index,
                    'message': '已完成: {0} ({1}/{2})'.format(type_name, completed_count, total_types)
                })

        if progress_callback:
            progress_callback({
                'progress': 100,
                'current_type': '',
                'current_type_name': '',
                'total_types': total_types,
                'current_index': total_types,
                'message': '导入完成'
            })
        
        return ImportResult(
            success=len(all_errors) == 0,
            results=results,
            errors=all_errors,
            trace_id=self._get_current_trace_id(),
        )

    def _sheet_name_to_object_type(self, sheet_name: str) -> Optional[str]:
        """将Sheet名称转换为对象类型"""
        for obj_id in registry.list_objects():
            obj = registry.get(obj_id)
            if obj and obj.name == sheet_name:
                return obj_id
        return None

    def _validate_sheets(self, sheets: List[Dict[str, Any]], context: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """校验导入数据

        校验规则：
        1. business_key 字段必填（新增时）
        2. business_key 值不能重复（在当前版本内）
        3. 父对象 business_key（如 code）必须存在（数据库或当前Excel中）
        """

        context = context or {}
        version_id = context.get('version_id')
        logger.info(f"[Validate] 开始校验，版本ID: {version_id}")

        valid_count = 0
        invalid_count = 0
        errors = []
        warnings = []  # [NEW v1.2.16 2026-06-20] 与 errors 区分

        # [SYMBOL] 第一步：收集所有即将导入的业务键，建立索引
        # 格式: {object_type: {bk_field_id: set of values}}
        importing_bk_values: Dict[str, Dict[str, set]] = {}

        for sheet in sheets:
            obj = registry.get(sheet["object_type"])
            if not obj:
                continue

            # 获取业务键字段
            bk_fields = []
            for f in obj.fields:
                if getattr(f.semantics, 'business_key', False):
                    is_virtual = f.storage.value == 'virtual' or getattr(f.semantics, 'virtual', False)
                    import_visible = getattr(f.semantics, 'import_visible', True)
                    if not is_virtual and import_visible:
                        bk_fields.append(f)

            if not bk_fields:
                continue

            # 收集该对象类型的所有业务键值
            if obj.id not in importing_bk_values:
                importing_bk_values[obj.id] = {}

            for row in sheet.get("preview_rows", []):
                record = dict(zip(sheet["columns"], row))

                if len(bk_fields) == 1:
                    # 单业务键
                    # [SYMBOL] 尝试多种方式获取值：中文列名、英文列名、字段ID
                    bk_value = None
                    bk_field = bk_fields[0]

                    # 首先尝试字段的 name 属性
                    if bk_field.name and bk_field.name in record:
                        bk_value = record[bk_field.name]
                    # 然后尝试字段的 ID
                    elif bk_field.id and bk_field.id in record:
                        bk_value = record[bk_field.id]
                    # 最后遍历 record 的 key 查找中文列名
                    else:
                        for key in record.keys():
                            if key in ['编码', '名称', 'code', bk_field.name, bk_field.id]:
                                bk_value = record[key]
                                break

                    if bk_value and str(bk_value).strip():
                        bk_value_str = str(bk_value).strip()
                        if bk_field.id not in importing_bk_values[obj.id]:
                            importing_bk_values[obj.id][bk_field.id] = set()
                        importing_bk_values[obj.id][bk_field.id].add(bk_value_str)
                        # [SYMBOL] 添加调试日志
                        logger.info(f"[Validate] 索引收集: {obj.id}.{bk_field.id} = {bk_value_str} (key={bk_field.name}/{bk_field.id})")
                else:
                    # 组合业务键
                    bk_values = []
                    all_empty = True
                    for bk_field in bk_fields:
                        bk_value = record.get(bk_field.name) or record.get(bk_field.id)
                        if bk_value and str(bk_value).strip():
                            all_empty = False
                            bk_values.append(str(bk_value).strip())
                        else:
                            bk_values.append("")

                    if not all_empty:
                        composite_key = "||".join(bk_values)
                        if 'composite' not in importing_bk_values[obj.id]:
                            importing_bk_values[obj.id]['composite'] = {}
                        if obj.id not in importing_bk_values[obj.id]['composite']:
                            importing_bk_values[obj.id]['composite'][obj.id] = set()
                        importing_bk_values[obj.id]['composite'][obj.id].add(composite_key)

        logger.info(f"[Validate] 收集到的待导入业务键: {importing_bk_values.keys()}")

        for sheet in sheets:
            obj = registry.get(sheet["object_type"])
            if not obj:
                continue
            
            # 获取可用于导入的业务关键字字段（支持组合键）
            # 条件：business_key=true, NOT virtual, import_visible != false
            bk_fields = []
            for f in obj.fields:
                if getattr(f.semantics, 'business_key', False):
                    is_virtual = f.storage.value == 'virtual' or getattr(f.semantics, 'virtual', False)
                    import_visible = getattr(f.semantics, 'import_visible', True)
                    if not is_virtual and import_visible:
                        bk_fields.append(f)
            
            existing_composite_keys = set()
            
            # 检查是否有操作模式列
            op_mode_idx = None
            columns = sheet.get("columns", [])
            if "操作模式" in columns:
                op_mode_idx = columns.index("操作模式")
            
            for idx, row in enumerate(sheet.get("preview_rows", [])):
                row_num = idx + 2
                record = dict(zip(sheet["columns"], row))
                
                # 确定当前行的操作模式
                # [FIX 2026-06-08] 默认值统一为 "create"，与 _import_sheet (L4350) 对齐。
                # 之前默认 "update" 会导致：用户留空时 _validate_sheets 走 update 路径（跳过 BK 必填检查），
                # 而 _import_sheet 走 create 路径（调用 create() 但 BK 没填）→ 预览通过但执行失败。
                operation_mode = "create"  # 默认是新增模式
                if op_mode_idx is not None and op_mode_idx < len(row):
                    op_value = row[op_mode_idx]
                    if op_value:
                        # [SSOT 2026-06-08] 使用 _parse_operation_mode_from_label 统一解析
                        parsed = self._parse_operation_mode_from_label(op_value)
                        if parsed is not None:
                            operation_mode = parsed
                
                # 只有新增和更新模式才验证必填字段
                should_validate_required = operation_mode in ["create", "update"]
                
                # 构建列名到字段的映射
                column_to_field = {}
                for field in obj.fields:
                    column_to_field[field.name] = field
                    column_to_field[field.id] = field
                
                # 层级字段的列名映射（如"领域编码" -> domain_id 字段）
                parent_object = obj.parent_object
                current_for_hierarchy = obj
                while parent_object:
                    parent_obj = registry.get(parent_object)
                    if parent_obj:
                        parent_key_field_id = "{0}_id".format(parent_object)
                        parent_key_field = current_for_hierarchy.get_field(parent_key_field_id)
                        if parent_key_field:
                            # 层级字段的列名是 "{父对象名称}编码" 和 "{父对象名称}名称"
                            code_header = "{0}编码".format(parent_obj.name)
                            column_to_field[code_header] = parent_key_field
                            name_header = "{0}名称".format(parent_obj.name)
                            column_to_field[name_header] = parent_key_field
                    parent_obj_instance = registry.get(parent_object)
                    if parent_obj_instance:
                        parent_object = parent_obj_instance.parent_object
                        current_for_hierarchy = parent_obj_instance
                    else:
                        break
                
                for field in obj.fields:
                    # 跳过系统字段和上下文字段（这些由系统自动填充）
                    if field.id == 'id' and field.required and getattr(field.semantics, 'meaning', '').find('自增主键') >= 0:
                        continue  # ID 字段是自增主键，导入时自动生成，不验证
                    
                    if getattr(field.semantics, 'readonly_always', False):
                        continue  # readonly_always 字段不验证（从上下文带入）
                    
                    if getattr(field.semantics, 'context_field', False):
                        continue  # 上下文字段不验证（从上下文带入）
                    
                    # 跳过可以通过 resolve_from_field 解析的字段
                    # 借鉴 SAP @ObjectModel.foreignKey.association 注解
                    resolve_from = getattr(field.semantics, 'resolve_from_field', None)
                    resolve_to = getattr(field.semantics, 'resolve_to_object', None)
                    if not resolve_from or not resolve_to:
                        vh = self._get_value_help(field)
                        if vh:
                            vh_source = getattr(vh, 'source', None)
                            if vh_source and getattr(vh_source, 'type', None) == 'bo':
                                resolve_to = getattr(vh_source, 'target_bo', None)
                                code_field = getattr(vh_source, 'code_field', None) or 'code'
                                if code_field and code_field != field.id:
                                    resolve_from = code_field
                    if resolve_from and resolve_to:
                        # 检查源字段是否有值
                        # record 的 key 可能是中文字段名或字段 ID
                        source_field = obj.get_field(resolve_from)
                        source_value = None
                        if source_field:
                            source_value = record.get(source_field.name) or record.get(resolve_from)
                        else:
                            source_value = record.get(resolve_from)
                        if source_value:
                            continue  # 可以通过源字段解析，跳过必填验证
                    
                    # 检查是否必填：mandatory / business_key / parent_key
                    # 注意：更新模式下 parent_key 字段不验证必填（保留原值）
                    is_parent_key_field = getattr(field.semantics, 'parent_key', False)
                    is_business_key_field = getattr(field.semantics, 'business_key', False)
                    is_required = (
                        getattr(field.semantics, 'mandatory', False)
                        or (is_parent_key_field and operation_mode == "create")
                        or (is_business_key_field and operation_mode == "create")
                    )
                    # [FIX 2026-06-17] 如果 business_key 字段为空但对象有 key_template
                    # 且 user_editable 不是 'manual'（即允许自动生成），则不应报必填错误
                    # 例: relationship 的 code 字段，key_template pattern="{source_code}-{target_code}-{SEQ:2}"
                    # 用户留空 code 时，KeyTemplateInterceptor 会自动生成
                    if is_required and is_business_key_field and operation_mode == "create":
                        kt_info = self._get_key_template_info(obj)
                        if kt_info.get('enabled') and kt_info.get('user_editable') != 'manual':
                            is_required = False
                    # [FIX 2026-06-16 BMRD] 如果 context 中已经传入了 product_id / version_id,
                    # 则产品→版本链路上所有 parent_key / business_key 字段的必填验证都应跳过.
                    # 因为 context 已经确定了顶层范围, 整条链路会从数据库自动填充.
                    if operation_mode == "create" and context:
                        # 顶层 product + version 都已确定: 跳过整条链路
                        if context.get('product_id') and context.get('version_id'):
                            is_required = False
                        # 只有 version_id: 跳过 product 链路上版本及其下游所有字段
                        elif context.get('version_id') and (
                            field.id in ('version_id', 'version_code', 'version_name',
                                         'domain_id', 'domain_code', 'domain_name',
                                         'sub_domain_id', 'sub_domain_code', 'sub_domain_name',
                                         'service_module_id', 'service_module_code', 'service_module_name',
                                         'product_id', 'product_code', 'product_name',
                                         'source_bo_code', 'target_bo_code',
                                         'source_code', 'target_code',
                                         'source_bo_id', 'target_bo_id',
                                         'source_bo_name', 'target_bo_name')
                        ):
                            is_required = False
                    
                    if should_validate_required and is_required:
                        # 尝试多种方式获取值：直接用列名或通过 column_to_field 映射
                        value = None
                        # 先尝试 field.name 和 field.id
                        value = record.get(field.name) or record.get(field.id)
                        # 如果没找到，遍历 record 的 key 看是否映射到当前 field
                        if value is None or value == "":
                            for col_name in record.keys():
                                mapped_field = column_to_field.get(col_name)
                                if mapped_field == field:
                                    value = record[col_name]
                                    break
                        
                        if value is None or value == "":
                            field_label = field.name or field.id
                            meaning = getattr(field.semantics, 'meaning', field_label)
                            errors.append({
                                "sheet": sheet["name"],
                                "row": row_num,
                                "operation": operation_mode,  # [FIX v1.2.18f] 前端 UI 显示需要操作模式
                                "field": field_label,
                                "value": "",
                                "error": f"{meaning} 不能为空"
                            })
                            invalid_count += 1
                    
                    enum_type_ref = getattr(field.semantics, 'enum_type_ref', None) or (
                        hasattr(field, 'ui') and getattr(field.ui, 'enum_type', None)
                    )
                    if not enum_type_ref:
                        enum_type_ref = self._get_enum_type_id_from_value_help(field)
                    if enum_type_ref and should_validate_required:
                        field_value = record.get(field.name) or record.get(field.id)
                        if field_value and str(field_value).strip():
                            field_value_str = str(field_value).strip()
                            # _validate_enum_value 内部会拆解 "CODE - LABEL" 格式
                            if not self._validate_enum_value(enum_type_ref, field_value_str):
                                field_label = field.name or field.id
                                errors.append({
                                    "sheet": sheet["name"],
                                    "row": row_num,
                                    "operation": operation_mode,  # [FIX v1.2.18f]
                                    "field": field_label,
                                    "value": field_value_str,
                                    "error": f"【枚举值无效】'{field_value_str}' 不是有效的 {field_label}，请检查枚举值配置"
                                })
                                invalid_count += 1
                
                if bk_fields:
                    # 构建组合业务键值
                    bk_values = []
                    all_bk_empty = True
                    for bk_field in bk_fields:
                        bk_value = record.get(bk_field.name) or record.get(bk_field.id)
                        if bk_value is not None and str(bk_value).strip() != "":
                            all_bk_empty = False
                            bk_values.append(str(bk_value).strip())
                        else:
                            bk_values.append("")
                    
                    composite_key = "||".join(bk_values)
                    
                    if all_bk_empty:
                        if operation_mode == "create":
                            # [FIX 2026-06-17] 如果对象有 key_template 且 enabled + auto_suggest，
                            # 则 business_key 为空时不应报必填错误，改为 warning
                            # 例: business_object 的 code 字段，pattern="{service_module_code}{SEQ:2}"
                            kt_info = self._get_key_template_info(obj)
                            if kt_info.get('enabled') and kt_info.get('auto_suggest'):
                                # [FIX v1.2.16 2026-06-20] key_template 可以自动生成, 推到 warnings 数组
                                # 而不是混进 errors (前端只显示 errors, 会让用户以为是错误)
                                bk_field_names = "、".join([f.name for f in bk_fields])
                                warnings.append({
                                    "sheet": sheet["name"],
                                    "row": row_num,
                                    "field": bk_field_names,
                                    "value": "",
                                    "error": f"【业务关键字】{bk_field_names}为空，将由编码模板自动生成",
                                    "severity": "warning"
                                })
                                # warning 不计入 invalid_count，不阻止导入
                            else:
                                bk_field_names = "、".join([f.name for f in bk_fields])
                                errors.append({
                                    "sheet": sheet["name"],
                                    "row": row_num,
                                    "operation": operation_mode,  # [FIX v1.2.18f]
                                    "field": bk_field_names,
                                    "value": "",
                                    "error": "【业务关键字】新增必填"
                                })
                                invalid_count += 1
                    else:
                        if composite_key in existing_composite_keys:
                            bk_field_names = "、".join([f.name for f in bk_fields])
                            composite_value = composite_key.replace("||", " + ")
                            # [NEW v1.2.14 2026-06-19] 单字段时不显示"组合"
                            if len(bk_fields) == 1:
                                bk_value = bk_values[0] if bk_values else ""
                                error_msg = f"【业务关键字】{bk_field_names} 值重复：{bk_value}"
                            else:
                                error_msg = f"【业务关键字】组合值重复：{composite_value}"
                            errors.append({
                                "sheet": sheet["name"],
                                "row": row_num,
                                "operation": operation_mode,  # [FIX v1.2.18f]
                                "field": bk_field_names,
                                "value": composite_value,
                                "error": error_msg
                            })
                            invalid_count += 1
                        else:
                            existing_composite_keys.add(composite_key)
                            # 检查组合键是否已存在于数据库（考虑版本）
                            if len(bk_fields) == 1:
                                existing_record = self._find_by_key(sheet["object_type"], bk_fields[0].id, bk_values[0], version_id)
                            else:
                                existing_record = self._find_by_composite_key(sheet["object_type"], bk_fields, bk_values, version_id)
                            if existing_record:
                                if operation_mode == "create":
                                    bk_field_names = "、".join([f.name for f in bk_fields])
                                    version_hint = f" (版本ID: {version_id})" if version_id else " (所有版本)"
                                    composite_value = composite_key.replace("||", " + ")
                                    # [NEW v1.2.14 2026-06-19] 单字段时不显示"组合"
                                    if len(bk_fields) == 1:
                                        bk_value = bk_values[0] if bk_values else ""
                                        error_msg = f"【业务关键字】{bk_field_names} 值已存在：{bk_value}{version_hint}"
                                    else:
                                        error_msg = f"【业务关键字】组合值已存在：{composite_value}{version_hint}"
                                    errors.append({
                                        "sheet": sheet["name"],
                                        "row": row_num,
                                        "operation": operation_mode,  # [FIX v1.2.18f]
                                        "field": bk_field_names,
                                        "value": composite_value,
                                        "error": error_msg
                                    })
                                    invalid_count += 1
                                    logger.warning(f"[Validate] 业务键冲突: {composite_key} (版本ID: {version_id})")

                # [SYMBOL] 引用完整性检查 - 检查所有外键引用
                # 包括通过 resolve_from_field 引用其他对象的字段
                for field in obj.fields:
                    # 检查是否有 resolve_from_field 和 resolve_to_object 语义
                    resolve_to = getattr(field.semantics, 'resolve_to_object', None)
                    resolve_from = getattr(field.semantics, 'resolve_from_field', None)

                    # 如果字段本身有 resolve_to_object，说明它是通过 ID 引用
                    # 如果字段有 resolve_from_field，说明它是引用源字段（如 source_code 引用 business_object）
                    if resolve_to:
                        # 获取源字段的值
                        source_field = None
                        if resolve_from:
                            # 通过 resolve_from_field 获取源字段
                            source_field = obj.get_field(resolve_from)
                            if not source_field:
                                source_field = obj.get_field_by_name(resolve_from)
                        else:
                            # 直接使用当前字段
                            source_field = field

                        if source_field:
                            # 从 record 中获取值
                            source_value = record.get(source_field.name) or record.get(source_field.id)

                            if source_value and str(source_value).strip():
                                # 检查被引用的对象是否存在
                                source_value_str = str(source_value).strip()

                                # 获取引用对象的业务键字段
                                ref_obj = registry.get(resolve_to)
                                if ref_obj:
                                    # [SYMBOL] 收集用于导入的 business_key 字段（与索引收集的逻辑一致）
                                    ref_bk_fields = []
                                    for rf in ref_obj.fields:
                                        if getattr(rf.semantics, 'business_key', False):
                                            # [SYMBOL] 过滤 virtual 和 import_visible 字段（与索引收集逻辑一致）
                                            is_virtual = getattr(rf, 'storage', None) and rf.storage.value == 'virtual' or getattr(rf.semantics, 'virtual', False)
                                            import_visible = getattr(rf.semantics, 'import_visible', True)
                                            if not is_virtual and import_visible:
                                                ref_bk_fields.append(rf)

                                    if ref_bk_fields:
                                        # [SYMBOL] 在数据库中查找被引用的对象
                                        ref_record = None
                                    if len(ref_bk_fields) == 1:
                                        ref_record = self._find_by_key(resolve_to, ref_bk_fields[0].id, source_value_str, version_id)
                                    else:
                                        ref_record = self._find_by_composite_key(resolve_to, ref_bk_fields, [source_value_str], version_id)

                                    # [SYMBOL] 如果数据库中没有，检查是否在即将导入的数据中
                                    in_importing = False
                                    if not ref_record:
                                        logger.info(f"[Validate] 数据库中未找到 {source_value_str}，检查即将导入的索引...")
                                        logger.info(f"[Validate] importing_bk_values keys: {list(importing_bk_values.keys())}")

                                        if resolve_to in importing_bk_values:
                                            ref_obj_data = importing_bk_values[resolve_to]
                                            logger.info(f"[Validate] 找到引用对象的索引, ref_obj_data keys: {list(ref_obj_data.keys())}")

                                            if len(ref_bk_fields) == 1:
                                                bk_field_id = ref_bk_fields[0].id
                                                logger.info(f"[Validate] 查询字段: {bk_field_id}")

                                                if bk_field_id in ref_obj_data:
                                                    values = ref_obj_data[bk_field_id]
                                                    logger.info(f"[Validate] 索引中的值数量: {len(values)}")
                                                    if source_value_str in values:
                                                        in_importing = True
                                                        logger.info(f"[Validate] [OK] 在即将导入的数据中找到 {source_value_str}")
                                                    else:
                                                        logger.info(f"[Validate] [X] 索引中不包含 {source_value_str}")
                                                        logger.info(f"[Validate] 索引中的部分值: {list(values)[:5]}")
                                                else:
                                                    logger.info(f"[Validate] 引用 {source_value_str} 在即将导入的数据中找到")

                                        if not ref_record and not in_importing:
                                            field_label = source_field.name or source_field.id
                                            version_info = f"(版本ID: {version_id})" if version_id else ""
                                            error_msg = f"【引用完整性】引用的 {ref_obj.name} '{source_value_str}' 不存在 {version_info}"
                                            hint = f"请先导入 {ref_obj.name} 数据，或检查业务键 '{source_value_str}' 是否正确"
                                            errors.append({
                                                "sheet": sheet["name"],
                                                "row": row_num,
                                                "operation": operation_mode,  # [FIX v1.2.18f]
                                                "field": field_label,
                                                "value": source_value_str,
                                                "error": error_msg,
                                                "hint": hint
                                            })
                                            invalid_count += 1
                                            logger.warning(f"[Validate] 引用完整性错误: {obj.name}.{field_label} -> {resolve_to}.{source_value_str} (版本ID: {version_id})")
                
                # [FIX FR-004] addability 检查：新增模式下验证 addability 条件
                # 避免预览通过但执行时被 manage_service.create() 的 addability 拒绝
                if operation_mode == "create":
                    try:
                        if not self.manage_service.check_can_add(sheet["object_type"], record):
                            obj_name = obj.name or sheet["object_type"]
                            errors.append({
                                "sheet": sheet["name"],
                                "row": row_num,
                                "field": "操作模式",
                                "value": operation_mode,
                                "error": f"【新增限制】{obj_name} 当前不满足新增条件（addability 规则），无法新增"
                            })
                            invalid_count += 1
                            continue  # 跳过后续 valid_count
                    except Exception as e:
                        logger.warning(f"[FR-004] addability 检查异常: {e}")

                valid_count += 1
        
        return {
            "valid_count": valid_count,
            "invalid_count": invalid_count,
            "errors": errors[:20],
            "warnings": warnings[:20]  # [NEW v1.2.16 2026-06-20] 区分 warnings
        }

    def _validate_enum_value(self, enum_type_id: str, code: str) -> bool:
        """验证枚举值是否有效

        [FR-006] 委托到 enum_resolver.validate_enum_value
        [FIX 2026-06-16 BMRD] 如果 value 是 "CODE - LABEL" 格式 (从下拉框选择的值),
        先拆解成 CODE 再验证, 避免 "REFERENCES - 引用" 整个字符串被传过去.
        """
        from meta.core.enum_resolver import validate_enum_value
        code = self._parse_enum_display_to_code(code)
        return validate_enum_value(enum_type_id, code, self.data_source)

    def _parse_enum_display_to_code(self, value) -> str:
        """[NEW 2026-06-16 BMRD] 拆解 "CODE - LABEL" 格式为 CODE

        用户在前端下拉选择的是 "CODE - LABEL" (例如 "REFERENCES - 引用"),
        但数据库 enum_values.code 列只存 CODE (例如 "REFERENCES").
        验证时如果直接传 "REFERENCES - 引用" 会失败.

        Returns: 拆解后的 code, 如果 value 不是 str 或不含 " - " 则原样返回.
        """
        if not value:
            return value
        if not isinstance(value, str):
            value = str(value)
        value = value.strip()
        if not value:
            return value
        if ' - ' in value:
            return value.split(' - ')[0].strip()
        return value

    def _parse_bo_cell_value(self, value: Any) -> List[str]:
        """[NEW v1.2.15 2026-06-19] 拆解 BO 字段的 Excel 单元格值

        Excel 中 BO 字段导出格式多样 (取决于 display_format 配置):
          - "BO_CUSTOMER - 客户"   (display_format: "{code} - {name}")
          - "客户 (16)"            (display_format: "{name} ({id})")
          - "客户"                 (仅 name, 用户手填)
          - "BO_CUSTOMER"          (仅 code, 用户手填)
          - 16                     (int id)

        本方法返回**所有可能的 lookup key** (code/name/原值),
        让 _preload_references 能批量加载, _convert_value 能逐个尝试匹配.

        Returns: list of candidate keys (去除空字符串)
        """
        if value is None:
            return []
        if not isinstance(value, str):
            value = str(value)
        v = value.strip()
        if not v:
            return []

        candidates: List[str] = []

        # 1) "CODE - NAME"  (e.g. "BO_CUSTOMER - 客户")
        if ' - ' in v:
            parts = v.split(' - ', 1)
            code_part = parts[0].strip()
            name_part = parts[1].strip()
            if code_part:
                candidates.append(code_part)
            if name_part:
                candidates.append(name_part)
            # 整串也加入兜底
            candidates.append(v)
            return [c for c in candidates if c]

        # 2) "name (id)"  (e.g. "客户 (16)")
        import re as _re
        m = _re.search(r'^(.+?)\s*\((\d+)\)\s*$', v)
        if m:
            name_part = m.group(1).strip()
            id_part = m.group(2).strip()
            if name_part:
                candidates.append(name_part)
            if id_part:
                candidates.append(id_part)
            return [c for c in candidates if c]

        # 3) 纯字符串: 可能 name 也可能 code
        candidates.append(v)
        return [c for c in candidates if c]

    def _resolve_bo_lookup(self, value: Any, target_bo: str,
                            lookup_index: Dict[tuple, Dict],
                            ds=None) -> Optional[Dict]:
        """[NEW v1.2.15 2026-06-19] 从 lookup_index 多 key 查找 BO 记录

        尝试顺序: id (int) → code (业务编码) → name (中文名)
        返回首个命中的 record (含 id/code/name).

        Args:
            value: 单元格原始值 (可能是 "客户"、"BO_CUSTOMER"、16、"客户 (16)"、"BO_CUSTOMER - 客户")
            target_bo: 目标对象类型 (e.g. "business_object", "service_module")
            lookup_index: _preload_references 返回的索引
            ds: 备用 data_source (lookup 未命中时回退到直接 DB 查询)

        Returns: 命中的 record dict {id, code, name, ...} 或 None
        """
        if value is None:
            return None
        candidates = self._parse_bo_cell_value(value)

        for cand in candidates:
            # 1) 优先查 lookup_index
            rec = lookup_index.get((target_bo, cand))
            if rec:
                return rec

        # 2) 兜底: 直接查 DB
        if ds is not None:
            try:
                from meta import get_meta_object
            except ImportError:
                try:
                    from meta.core.models import get_meta_object
                except ImportError:
                    get_meta_object = None
            target_meta = get_meta_object(target_bo) if get_meta_object else None
            if target_meta:
                target_fields = getattr(target_meta, 'fields', [])
                code_field = None
                name_field = None
                for tf in target_fields:
                    if getattr(tf.semantics, 'business_key', False) and not code_field:
                        code_field = tf.db_column
                    if tf.id == 'name' and not name_field:
                        name_field = tf.db_column
                for cand in candidates:
                    # 数字 → id
                    if str(cand).isdigit():
                        try:
                            cur = ds.execute(f"SELECT id, code, name FROM {target_meta.table_name} WHERE id = ? LIMIT 1", (int(cand),))
                            row = cur.fetchone()
                            if row:
                                return {'id': row[0], 'code': row[1], 'name': row[2]}
                        except Exception:
                            pass
                    # 业务编码
                    if code_field:
                        try:
                            cur = ds.execute(f"SELECT id, code, name FROM {target_meta.table_name} WHERE {code_field} = ? LIMIT 1", (cand,))
                            row = cur.fetchone()
                            if row:
                                return {'id': row[0], 'code': row[1], 'name': row[2]}
                        except Exception:
                            pass
                    # 名称
                    if name_field and name_field != code_field:
                        try:
                            cur = ds.execute(f"SELECT id, code, name FROM {target_meta.table_name} WHERE {name_field} = ? LIMIT 1", (cand,))
                            row = cur.fetchone()
                            if row:
                                return {'id': row[0], 'code': row[1], 'name': row[2]}
                        except Exception:
                            pass
        return None

    def _get_enum_value_info(self, enum_type_id: str, code: str) -> Optional[Dict[str, Any]]:
        """获取枚举值详细信息（包括名称和维度）
        
        Args:
            enum_type_id: 枚举类型ID
            code: 枚举值编码
            
        Returns:
            Dict: 枚举值信息，包含 name, dimensions 等
        """
        try:
            sql = """
                SELECT code, name, name_en, dimensions 
                FROM enum_values 
                WHERE enum_type_id = ? AND code = ? AND is_active = 1
            """
            cursor = self.data_source.execute(sql, [enum_type_id, code])
            row = cursor.fetchone()
            if row:
                import json
                result = {
                    'code': row[0],
                    'name': row[1],
                    'name_en': row[2] or ''
                }
                if row[3]:
                    try:
                        result['dimensions'] = json.loads(row[3])
                    except:
                        result['dimensions'] = {}
                return result
        except Exception as e:
            logger.warning(f"[Validate] 获取枚举值信息失败: {enum_type_id}.{code} - {e}")
        return None

    def _preload_references(self, rows: List[tuple], headers: List[str],
                           parent_key_headers: Dict, obj, version_id: Optional[int] = None,
                           field_map: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
        """批量预加载外键引用，避免 N+1 查询问题

        优化策略：
        1. 收集所有需要查询的 (object_type, code/name) 组合
        2. 批量查询数据库
        3. 建立内存索引 {(object_type, code): record} 和 {(object_type, name): record}

        Args:
            rows: Excel 数据行
            headers: 列名列表
            parent_key_headers: 父对象键信息
            obj: 元对象
            version_id: 版本ID
            field_map: 表头到字段ID的映射 (header -> field_id)

        Returns:
            Dict: {(object_type, code_or_name): record} 内存索引
            同时支持 code 和 name 反查 (Excel 中可能填中文名而非编码)
        """

        lookup_index: Dict[tuple, Dict] = {}
        if not rows or not headers:
            return lookup_index

        object_codes: Dict[str, set] = {}

        for row in rows:
            for parent_header, parent_info in parent_key_headers.items():
                parent_code_idx = headers.index(parent_header) if parent_header in headers else -1
                if parent_code_idx >= 0 and parent_code_idx < len(row):
                    parent_code = row[parent_code_idx]
                    if parent_code:
                        parent_type = parent_info["parent_type"]
                        if parent_type not in object_codes:
                            object_codes[parent_type] = set()
                        object_codes[parent_type].add(parent_code)

            # [NEW v1.2.15 2026-06-19] 同时收集 name 作为 lookup key
            # Excel 中用户可能填中文名（如 "客户"、"采购管理"）而非业务编码
            # _convert_value 会优先尝试按 id、code 转换；失败时按 name 查 lookup_index
            for field in obj.fields:
                vh = self._get_value_help(field)
                if not vh:
                    continue
                vh_source = getattr(vh, 'source', None)
                if not vh_source or getattr(vh_source, 'type', None) != 'bo':
                    continue
                target_bo = getattr(vh_source, 'target_bo', None)
                if not target_bo:
                    continue
                # 找这个 field 对应的 header
                field_id = field.id
                field_name = field.name
                # 先按 field_id 找 header
                target_header = None
                if field_map:
                    for hdr, fid in field_map.items():
                        if fid == field_id:
                            target_header = hdr
                            break
                if not target_header:
                    for h in headers:
                        if h == field_name:
                            target_header = h
                            break
                if not target_header:
                    continue
                try:
                    idx = headers.index(target_header)
                except ValueError:
                    continue
                if idx >= len(row):
                    continue
                cell = row[idx]
                if cell is None or (isinstance(cell, str) and not cell.strip()):
                    continue
                # 拆解 "CODE - LABEL" 或 "name (id)" 或纯 name / code
                parts = self._parse_bo_cell_value(cell)
                if target_bo not in object_codes:
                    object_codes[target_bo] = set()
                for p in parts:
                    object_codes[target_bo].add(p)

            for field in obj.fields:
                resolve_from = getattr(field.semantics, 'resolve_from_field', None)
                resolve_to_object = getattr(field.semantics, 'resolve_to_object', None)
                resolve_to_field = getattr(field.semantics, 'resolve_to_field', None)

                if not resolve_from and not resolve_to_object and not resolve_to_field:
                    vh = self._get_value_help(field)
                    if vh:
                        vh_source = getattr(vh, 'source', None)
                        if vh_source and getattr(vh_source, 'type', None) == 'bo':
                            resolve_to_object = getattr(vh_source, 'target_bo', None)
                            code_field = getattr(vh_source, 'code_field', None) or 'code'
                            if code_field and code_field != field.id:
                                resolve_from = code_field

                if resolve_from and (resolve_to_object or resolve_to_field):
                    # [FIX 2026-06-16] resolve_from 是字段ID(如 "domain_code"),
                    # 但 headers 中是字段的显示名(如 "领域编码")。
                    # 需要通过 field_map 找到对应的 header,再取列值。
                    resolve_from_idx = -1
                    if field_map:
                        for hdr, fid in field_map.items():
                            if fid == resolve_from:
                                resolve_from_idx = headers.index(hdr) if hdr in headers else -1
                                break
                    if resolve_from_idx < 0:
                        resolve_from_idx = headers.index(resolve_from) if resolve_from in headers else -1
                    if resolve_from_idx >= 0 and resolve_from_idx < len(row):
                        source_value = row[resolve_from_idx]
                        if source_value:
                            if resolve_to_object:
                                target_type = resolve_to_object
                            elif resolve_to_field:
                                # [FIX 2026-06-16 BMRD] resolve_to_field 也是字段ID
                                # (如 annotation.target_id 的 resolve_to_field='target_type'),
                                # 需要先通过 field_map 找到 header, 再取列值.
                                # 之前直接 headers.index(resolve_to_field) 找不到
                                # (因为 header 是 '关联对象类型' 不是 'target_type').
                                target_type_idx = -1
                                if field_map:
                                    for hdr, fid in field_map.items():
                                        if fid == resolve_to_field:
                                            target_type_idx = headers.index(hdr) if hdr in headers else -1
                                            break
                                if target_type_idx < 0:
                                    target_type_idx = headers.index(resolve_to_field) if resolve_to_field in headers else -1
                                if target_type_idx >= 0 and target_type_idx < len(row):
                                    target_type = row[target_type_idx]
                                    # [FIX 2026-06-16 BMRD] target_type 是从 Excel 取的,
                                    # 可能带 " - 中文" 标签 (例如 "service_module - 服务模块").
                                    # 在做 _preload_references 预加载 key 前先拆解为 code,
                                    # 否则 object_codes 用错 key 查询, 导致 target_id 解析失败.
                                    if target_type and isinstance(target_type, str) and ' - ' in target_type:
                                        target_type = target_type.split(' - ')[0].strip()
                                else:
                                    continue
                            else:
                                continue

                            if target_type:
                                if target_type not in object_codes:
                                    object_codes[target_type] = set()
                                object_codes[target_type].add(source_value)

        logger.info(f"[Preload] 需要预加载的对象类型: {list(object_codes.keys())}")
        logger.info(f"[Preload] 各类型编码数量: {[(k, len(v)) for k, v in object_codes.items()]}")

        for object_type, codes in object_codes.items():
            if not codes:
                continue
            try:
                # [NEW v1.2.15 2026-06-19] 按 code 和 name 分别查, 同时建 name 索引
                # 用户可能填中文名 (如 "客户"、"采购管理") 而非编码, 需要按 name 反查
                # 1) 按 code 查
                code_conditions = [
                    QueryCondition(field="code", operator="in", values=list(codes))
                ]
                if version_id is not None:
                    code_conditions.append(
                        QueryCondition(field="version_id", operator="eq", value=version_id)
                    )

                search_request = SearchRequest(
                    object_type=object_type,
                    conditions=code_conditions,
                    page=1,
                    page_size=len(codes) * 2,
                )
                result = self.query_service.search(search_request)

                for record in result.data:
                    code_value = record.get("code")
                    if code_value:
                        lookup_index[(object_type, code_value)] = record
                    # [NEW v1.2.15 2026-06-19] 同时建 name 索引
                    name_value = record.get("name")
                    if name_value:
                        lookup_index[(object_type, name_value)] = record

                logger.info(f"[Preload] 预加载 {object_type} (by code): 查询到 {len(result.data)} 条记录")

                # 2) 按 name 查 (补充: Excel 中用户可能填了 name 但 code 没匹配)
                #    提取 codes 中的非数字字符串作为 name 查询
                try:
                    from meta import get_meta_object as _get_meta
                except ImportError:
                    try:
                        from meta.core.models import get_meta_object as _get_meta
                    except ImportError:
                        _get_meta = None
                target_meta = _get_meta(object_type) if _get_meta else None
                if target_meta and hasattr(target_meta, 'table_name'):
                    target_fields = getattr(target_meta, 'fields', [])
                    name_db_col = None
                    for tf in target_fields:
                        if tf.id == 'name':
                            name_db_col = tf.db_column
                            break
                    if name_db_col:
                        # 收集 candidates 中非数字、非已按 code 命中的字符串
                        candidate_names = []
                        for c in codes:
                            if isinstance(c, str) and c and not c.isdigit():
                                # 跳过已找到的 (按 code 查的)
                                if (object_type, c) not in lookup_index:
                                    candidate_names.append(c)
                        if candidate_names:
                            try:
                                name_conditions = [
                                    QueryCondition(field=name_db_col, operator="in", values=candidate_names)
                                ]
                                if version_id is not None:
                                    name_conditions.append(
                                        QueryCondition(field="version_id", operator="eq", value=version_id)
                                    )
                                name_request = SearchRequest(
                                    object_type=object_type,
                                    conditions=name_conditions,
                                    page=1,
                                    page_size=len(candidate_names) * 2,
                                )
                                name_result = self.query_service.search(name_request)
                                for record in name_result.data:
                                    code_value = record.get("code")
                                    name_value = record.get("name")
                                    if code_value:
                                        lookup_index[(object_type, code_value)] = record
                                    if name_value:
                                        lookup_index[(object_type, name_value)] = record
                                logger.info(f"[Preload] 预加载 {object_type} (by name): 候选 {len(candidate_names)} 条, 命中 {len(name_result.data)}")
                            except Exception as e2:
                                logger.warning(f"[Preload] 按 name 预加载 {object_type} 失败: {e2}")

            except Exception as e:
                logger.warning(f"[Preload] 预加载 {object_type} 失败: {e}")

        logger.info(f"[Preload] 预加载完成，总索引条目: {len(lookup_index)}")
        return lookup_index

    def _find_from_index(self, lookup_index: Dict[tuple, Dict], 
                         object_type: str, code: str) -> Optional[Dict]:
        """从内存索引中查找记录

        Args:
            lookup_index: 预加载的内存索引
            object_type: 对象类型
            code: 编码值

        Returns:
            Optional[Dict]: 找到的记录，或 None
        """
        return lookup_index.get((object_type, code))

    def _import_sheet(self, file_path: str, sheet_info: Dict[str, Any],
                      conflict_strategy: str, context: Optional[Dict[str, Any]] = None,
                      progress_callback: Optional[callable] = None,
                      type_progress_base: int = 0,
                      type_progress_weight: int = 100) -> Dict[str, Any]:
        """导入单个Sheet的数据

        Args:
            file_path: Excel文件路径
            sheet_info: Sheet信息
            conflict_strategy: 冲突处理策略
            context: 导入上下文，包含version_id和product_id等
            progress_callback: 进度回调函数
            type_progress_base: 该类型的基础进度（0-100中的起始值）
            type_progress_weight: 该类型的进度权重（占100的比例）
        """

        context = context or {}

        frontend_version_id = context.get('version_id')
        logger.info(f"[Import] 前端传入的版本ID: {frontend_version_id}, context keys: {list(context.keys())}")

        meta = self._read_meta_sheet(file_path)
        excel_product_code = meta.get('product_code')
        excel_version_code = meta.get('version_code')

        logger.info(f"[Import] Excel元数据表信息: product_code={excel_product_code}, version_code={excel_version_code}")

        if frontend_version_id:
            logger.info(f"[Import] 使用前端传入的version_id={frontend_version_id}，忽略Excel元数据表中的版本")
            if excel_product_code or excel_version_code:
                logger.warning(f"[Import] [WARNING] 前端传入的版本ID({frontend_version_id})与Excel元数据表版本({excel_version_code})可能不一致，将使用前端版本")
        elif excel_product_code and excel_version_code:
            resolved_version_id = self._resolve_version_id(excel_product_code, excel_version_code)
            if resolved_version_id:
                logger.info(f"[Import] 从Excel元数据表解析出version_id={resolved_version_id}")
                context['version_id'] = resolved_version_id
                context['product_code'] = excel_product_code
                context['version_code'] = excel_version_code
            else:
                logger.warning(f"[Import] 无法从Excel元数据表(product={excel_product_code}, version={excel_version_code})解析出version_id")
        else:
            logger.info(f"[Import] 前端未传入version_id，Excel元数据表也无可用版本信息，将不指定版本导入")
        
        object_type = sheet_info["object_type"]
        obj = registry.get(object_type)
        if not obj:
            return {"success": 0, "failed": 0, "errors": [{"message": "Object not found"}]}
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_info["name"]]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            return {"success": 0, "failed": 0, "errors": [{"message": str(e)}]}
        
        if len(rows) < 2:
            return {"success": 0, "failed": 0, "errors": []}

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        field_map = self._auto_map_fields(obj, headers)

        parent_key_headers = self._get_parent_key_headers(obj, headers)
        if object_type == 'relationship':
            logger.info(f"[DEBUG IMPORT] {object_type} field_map={field_map}")
            logger.info(f"[DEBUG IMPORT] {object_type} parent_key_headers={parent_key_headers}")

        # [SYMBOL] 调试日志：打印 parent_key_headers
        if parent_key_headers:
            logger.info(f"[Import] parent_key_headers for {object_type}: {parent_key_headers}")

        has_operation_mode = "操作模式" in headers
        operation_mode_idx = headers.index("操作模式") if has_operation_mode else -1
        # [NEW v1.2.16 2026-06-19] 找到 code/name 列下标, 用于成功/跳过明细兜底
        # 即便 record 在 update 时被清理了 (parent_key 移除), 也能从原 Excel 行取到 code/name
        # 不同 sheet 用不同列名: 编码 (默认) / 关系编码 (relationship) / 关联对象编码 (annotation)
        code_col_candidates = ["编码", "关系编码", "关联对象编码", "业务编码"]
        name_col_candidates = ["名称", "关联对象名称", "业务名称", "业务对象名称"]
        code_col_idx = -1
        name_col_idx = -1
        for cn in code_col_candidates:
            if cn in headers:
                code_col_idx = headers.index(cn)
                break
        for nn in name_col_candidates:
            if nn in headers:
                name_col_idx = headers.index(nn)
                break

        def _get_row_code(row, record=None):
            """从原 Excel 行取业务编码 (兜底). 优先级: record.code > row[code_col]"""
            if record:
                rec_code = record.get("code") or record.get("id_code")
                if rec_code and isinstance(rec_code, str) and rec_code.strip():
                    return rec_code.strip()
            if code_col_idx >= 0 and code_col_idx < len(row):
                v = row[code_col_idx]
                return str(v).strip() if v is not None else ""
            return ""
        def _get_row_name(row, record=None):
            """从原 Excel 行取名称 (兜底). 优先级: record.name > row[name_col]"""
            if record:
                rec_name = record.get("name") or record.get("display_name")
                if rec_name and isinstance(rec_name, str) and rec_name.strip():
                    return rec_name.strip()
            if name_col_idx >= 0 and name_col_idx < len(row):
                v = row[name_col_idx]
                return str(v).strip() if v is not None else ""
            return ""

        logger.info(f"[Import] {object_type}导入 - headers: {headers[:10]}, has_operation_mode={has_operation_mode}")
        logger.info(f"[Import] {object_type}导入 - field_map: {field_map}")

        # [SYMBOL] 性能优化：批量预加载外键引用
        lookup_version_id = context.get('version_id')
        lookup_index = self._preload_references(rows[1:], headers, parent_key_headers, obj, lookup_version_id, field_map)

        success_count = 0
        failed_count = 0
        skipped_count = 0
        deleted_count = 0
        # [NEW v1.2.3 2026-06-17] 拆分 created / updated 统计，给前端第四步骤用
        created_count = 0
        updated_count = 0
        errors = []
        warnings = []
        # [NEW v1.2.14 2026-06-19] 成功/跳过明细 (前端第 4 步成功/跳过 subtab 用)
        # 限制最多 100 条 (避免响应体过大), 超出部分用 +N 表示
        successes = []
        skipped_items = []
        _MAX_DETAIL = 100

        # [NEW v1.2.14 2026-06-19] Helper: 记录成功项明细
        def _record_success_item(items, row_num, operation, record, max_count, code_override=None, name_override=None):
            if len(items) < max_count:
                # [FIX v1.2.16 2026-06-20] 优先级:
                # 1) code_override 非空 (来自 row) 优先
                # 2) record.code/name (来自 import 解析)
                # 3) 用空字符串兜底
                if code_override:
                    code_val = code_override
                else:
                    code_val = record.get("code") or record.get("id_code") or ""
                if name_override:
                    name_val = name_override
                else:
                    name_val = record.get("name") or record.get("display_name") or ""
                items.append({
                    "row": row_num,
                    "operation": operation,
                    "code": code_val,
                    "name": name_val
                })

        # [NEW v1.2.14 2026-06-19] Helper: 记录跳过项明细
        def _record_skipped_item(items, row_num, operation, record, reason, max_count, code_override=None, name_override=None):
            if len(items) < max_count:
                if code_override:
                    code_val = code_override
                else:
                    code_val = record.get("code") or record.get("id_code") or ""
                if name_override:
                    name_val = name_override
                else:
                    name_val = record.get("name") or record.get("display_name") or ""
                items.append({
                    "row": row_num,
                    "operation": operation,
                    "code": code_val,
                    "name": name_val,
                    "reason": reason
                })

        total_rows = len(rows) - 1
        type_name = obj.name or object_type
        
        progress_stages = [0.2, 0.4, 0.6, 0.8, 1.0]
        progress_stage_index = 0
        
        for idx, row in enumerate(rows[1:]):
            row_num = idx + 2

            # [FIX 2026-06-16 BMRD] 空行过滤: Excel 中常出现整行除"操作模式"列外全空的
            # 残留行 (例如用户复制粘贴模板后未填数据). 之前会触发 create() 失败,
            # 整 sheet 事务回滚, 真正的数据 (R1) 也跟着丢失.
            # 判定标准: 除"操作模式"列外, 其它列值全为 None 或空白字符串.
            if not row or all(
                (cell is None) or (isinstance(cell, str) and not cell.strip())
                for col_idx, cell in enumerate(row)
                if col_idx != operation_mode_idx
            ):
                logger.info(f"[Import] {object_type} row={row_num} 跳过空行 (除操作模式外全空)")
                skipped_count += 1
                continue

            if progress_callback and total_rows > 0:
                current_progress = (idx + 1) / total_rows
                while progress_stage_index < len(progress_stages) and current_progress >= progress_stages[progress_stage_index]:
                    stage_percent = int(type_progress_base + progress_stages[progress_stage_index] * type_progress_weight)
                    progress_callback({
                        'progress': min(99, stage_percent),
                        'current_type': object_type,
                        'current_type_name': type_name,
                        'total_types': 0,
                        'current_index': 0,
                        'message': '正在导入 {0}: {1}/{2} 行 ({3}%)'.format(
                            type_name, idx + 1, total_rows, int(progress_stages[progress_stage_index] * 100))
                    })
                    progress_stage_index += 1

            record = {}

            operation_mode = "create"
            operation_mode_explicit = False
            if has_operation_mode and operation_mode_idx >= 0 and operation_mode_idx < len(row):
                mode_value = row[operation_mode_idx]
                if mode_value and str(mode_value).strip():
                    # [SSOT 2026-06-08] 使用 _parse_operation_mode_from_label 统一解析
                    parsed = self._parse_operation_mode_from_label(mode_value)
                    if parsed is not None:
                        operation_mode = parsed
                        operation_mode_explicit = True
                    # 无法识别时保留默认 "create"，不再用 key_part 兜底
                else:
                    operation_mode = "create"
            
            for col_idx, header in enumerate(headers):
                if header == "操作模式":
                    continue
                if object_type == 'relationship' and row_num == 35:
                    logger.info(f"[DEBUG IMPORT] {object_type} row=35 PRE-CHECK header={repr(header)} field_map_match={header in field_map}")
                if header and header in field_map:
                    field_id = field_map[header]
                    value = row[col_idx] if col_idx < len(row) else None
                    if object_type == 'relationship' and row_num == 35:
                        logger.info(f"[DEBUG IMPORT] {object_type} row=35 col_idx={col_idx} header={header} field_id={field_id} value={value}")
                    meta_field = obj.get_field(field_id)
                    if meta_field and value is not None:
                        value = self._convert_value(value, meta_field)
                        # [FIX 2026-06-17] 枚举字段: 拆解 "CODE - LABEL" 格式为 CODE
                        # Excel 导出的枚举值格式为 "PULL - 拉"，但数据库只存 CODE "PULL"
                        # _parse_enum_display_to_code 之前只在验证阶段调用，字段赋值时未调用
                        enum_type_ref = getattr(meta_field.semantics, 'enum_type_ref', None) or (
                            hasattr(meta_field, 'ui') and getattr(meta_field.ui, 'enum_type', None)
                        )
                        if not enum_type_ref:
                            enum_type_ref = self._get_enum_type_id_from_value_help(meta_field)
                        if enum_type_ref and isinstance(value, str):
                            value = self._parse_enum_display_to_code(value)
                    record[field_id] = value

            for parent_header, parent_info in parent_key_headers.items():
                parent_code_idx = headers.index(parent_header) if parent_header in headers else -1
                logger.debug(f"[Import] 检查父对象列: header={parent_header}, idx={parent_code_idx}")
                if parent_code_idx >= 0 and parent_code_idx < len(row):
                    parent_code = row[parent_code_idx]
                    logger.debug(f"[Import] 父对象编码: {parent_header}={parent_code}")
                    if parent_code:
                        parent_id_field = parent_info["id_field"]
                        if parent_id_field not in record or record.get(parent_id_field) is None:
                            parent_obj = registry.get(parent_info["parent_type"])
                            if parent_obj:
                                # [SYMBOL] 性能优化：使用预加载的内存索引替代数据库查询
                                parent_record = self._find_from_index(lookup_index, parent_info["parent_type"], parent_code)
                                if parent_record:
                                    record[parent_id_field] = parent_record.get("id")
                                    logger.info(f"[Import] 解析父对象成功: {parent_id_field}={parent_record.get('id')}")
                                else:
                                    logger.warning(f"[Import] 未找到父对象: {parent_info['parent_type']}.code={parent_code}")

            # [FIX 2026-06-16 BMRD] 先过滤, 再做动态 FK 解析
            # 之前 FK 解析在过滤之前, 导致解析出来的 target_id (import_visible=false) 被过滤掉
            # 导致创建时报 "关联对象ID 不能为空" 错
            record = self._filter_import_record(record, obj, operation_mode)
            logger.info(f"[DEBUG IMPORT] {object_type} row={row_num} operation_mode={operation_mode} record={record}")

            # 外键解析：根据 resolve_from_field 和 resolve_to_object/resolve_to_field 自动解析外键ID
            # 借鉴 SAP @ObjectModel.foreignKey.association 注解
            # 支持两种模式：
            #   1. 静态模式: resolve_to_object = "business_object" （硬编码目标对象类型）
            #   2. 动态模式: resolve_to_field = "target_type"  （从同记录另一字段取目标类型，用于多态外键）
            # [FIX 2026-06-16 BMRD] 在 filter 之后做 FK 解析, 这样解析出来的 FK ID 不会被 filter 掉
            # 同时确保 resolve_from 字段 (如 target_code, 是 parent_key + resolve_from_source) 已经被保留
            for field in obj.fields:
                resolve_from = getattr(field.semantics, 'resolve_from_field', None)
                resolve_to_object = getattr(field.semantics, 'resolve_to_object', None)
                resolve_to_field = getattr(field.semantics, 'resolve_to_field', None)

                if not resolve_from and not resolve_to_object and not resolve_to_field:
                    vh = self._get_value_help(field)
                    if vh:
                        vh_source = getattr(vh, 'source', None)
                        if vh_source and getattr(vh_source, 'type', None) == 'bo':
                            resolve_to_object = getattr(vh_source, 'target_bo', None)
                            code_field = getattr(vh_source, 'code_field', None) or 'code'
                            if code_field and code_field != field.id:
                                resolve_from = code_field

                if resolve_from and (resolve_to_object or resolve_to_field):
                    # [FIX 2026-06-16 BMRD] 不仅检查 None, 还检查值是否为有效整数 ID
                    # Excel 中 '源业务对象' 列的值是中文名 (如 '供应商'), 不是 ID
                    # 这种情况下应从 resolve_from_field (如 source_code) 重新解析
                    current_value = record.get(field.id)
                    needs_resolve = current_value is None
                    if current_value is not None and field.field_type.value in ('integer', 'int'):
                        try:
                            int(current_value)
                        except (ValueError, TypeError):
                            needs_resolve = True
                    # [NEW v1.2.15 2026-06-19] 如果 current_value 是字符串 (name 或 "code - name" 残留),
                    # 也需要重查; 因为此时字段 type=integer, 但 record 里塞了 string, validate 会失败
                    if current_value is not None and isinstance(current_value, str) and field.field_type.value in ('integer', 'int'):
                        needs_resolve = True
                    # [FIX v1.2.18g 2026-06-20] 支持多种 resolve_from 候选字段
                    # 旧逻辑只从 resolve_from_field (如 source_code) 取值, 但 record 实际可能
                    # 用 source_bo_code / source_bo_name / source_code 多种命名, 需依次回退
                    if needs_resolve:
                        # 候选 resolve_from 字段 (按常见命名顺序)
                        candidate_resolve_froms = []
                        if resolve_from:
                            candidate_resolve_froms.append(resolve_from)
                        # 从 schema 中找所有同名字段 (source_bo_code / source_code / source_bo_name 等)
                        for f in obj.fields:
                            for prefix in ['source_bo_code', 'source_code', 'source_bo_name',
                                          'target_bo_code', 'target_code', 'target_bo_name']:
                                if f.id == prefix and f.id not in candidate_resolve_froms:
                                    candidate_resolve_froms.append(f.id)
                        # 取第一个非空值
                        source_value = None
                        actual_resolve_from = None
                        for cand in candidate_resolve_froms:
                            v = record.get(cand)
                            if v is not None and str(v).strip():
                                source_value = v
                                actual_resolve_from = cand
                                break
                        # [FIX v1.2.18i 2026-06-20] 兜底: 候选字段都为空时,
                        # 直接用 FK 字段当前值 (如 source_bo_id='TEST885') 去查 code/name。
                        # 场景: Excel 只有 '源业务对象' 列且填的是编码, 没有 source_code/source_bo_code 列。
                        if not source_value and current_value is not None:
                            source_value = current_value
                            actual_resolve_from = field.id
                        if source_value:
                            if resolve_to_field:
                                dynamic_type = record.get(resolve_to_field)
                                # [FIX 2026-06-16 BMRD] dynamic_type (如 target_type) 可能带
                                # " - 中文" 标签 (例如 "service_module - 服务模块"),
                                # _find_from_index 用它作 key 查 lookup_index, 必须先拆解.
                                if dynamic_type and isinstance(dynamic_type, str) and ' - ' in dynamic_type:
                                    dynamic_type = dynamic_type.split(' - ')[0].strip()
                                if dynamic_type:
                                    # [NEW v1.2.15 2026-06-19] 多 key 查找 (id/code/name)
                                    target_record = self._resolve_bo_lookup(source_value, dynamic_type, lookup_index, self.data_source)
                                    if target_record:
                                        record[field.id] = target_record.get('id')
                                        logger.info(f"[Import] 动态外键解析成功: {field.id}={record[field.id]} ({dynamic_type}.code={source_value})")
                                    else:
                                        logger.warning(f"[Import] 未找到动态外键对象: {dynamic_type}.code={source_value}")
                                else:
                                    logger.warning(f"[Import] 动态外键类型字段为空: resolve_to_field={resolve_to_field}")
                            elif resolve_to_object:
                                # [NEW v1.2.15 2026-06-19] 多 key 查找 (id/code/name)
                                target_record = self._resolve_bo_lookup(source_value, resolve_to_object, lookup_index, self.data_source)
                                if target_record:
                                    record[field.id] = target_record.get('id')
                                    logger.info(f"[Import] 外键解析成功: {field.id}={record[field.id]} ({resolve_to_object}.code={source_value})")
                                else:
                                    logger.warning(f"[Import] 未找到外键对象: {resolve_to_object}.code={source_value} | row={row_num} | record keys={list(record.keys())} | source_value type={type(source_value)} | actual_resolve_from={actual_resolve_from}")

            logger.info(f"[DEBUG IMPORT] {object_type} row={row_num} after FK resolve record keys={list(record.keys())} sample={ {k: record.get(k) for k in ['source_bo_id','target_bo_id','source_code','target_code']} }")

            if context:
                valid_fields = set()
                for f in obj.fields:
                    valid_fields.add(f.id)

                logger.info(f"[Import] context before adding to record: version_id={context.get('version_id')}")
                logger.info(f"[Import] valid_fields includes 'version_id': {'version_id' in valid_fields}")

                for ctx_key, ctx_value in context.items():
                    if ctx_key in valid_fields and ctx_key not in record:
                        logger.info(f"[Import] 添加context字段到record: {ctx_key}={ctx_value}")
                        record[ctx_key] = ctx_value

            logger.info(f"[Import] 最终record中的version_id: {record.get('version_id')}")
            logger.info(f"[Import] operation_mode={operation_mode}, conflict_strategy={conflict_strategy}")

            # [SYMBOL] 关键修复：如果record中没有version_id但context中有，强制添加
            if record.get('version_id') is None and context.get('version_id') is not None:
                logger.info(f"[Import] [SYMBOL] 强制添加version_id={context.get('version_id')}到record")
                record['version_id'] = context.get('version_id')

            try:
                if operation_mode == "delete":
                    logger.info(f"[Import] 执行删除操作")
                    try:
                        self._delete_record(object_type, record, obj.import_export)
                        deleted_count += 1
                        success_count += 1
                        _record_success_item(successes, row_num, "delete", record, _MAX_DETAIL, code_override=_get_row_code(row), name_override=_get_row_name(row))
                    except ValueError as ve:
                        # [FIX v1.2.18 2026-06-20] 删除不存在的记录只推到 warnings (不再推 skipped)
                        # 原因: skipped 会让前端合并入 "成功" tab, 但删除失败不是成功也不是 skip, 是告警
                        logger.warning(f"[Import] 删除记录不存在，作为告警: {ve}")
                        # 不增加 skipped_count (避免前端把告警合并到成功 tab)
                        warnings.append({
                            "row": row_num,
                            "operation": operation_mode,
                            "field": "编码",
                            "value": record.get("code", ""),
                            "message": f"记录不存在: {ve}",
                            "severity": "warning"
                        })
                elif operation_mode == "skip":
                    logger.info(f"[Import] 跳过记录")
                    skipped_count += 1
                    _record_skipped_item(skipped_items, row_num, "skip", record, "操作模式为跳过", _MAX_DETAIL, code_override=_get_row_code(row), name_override=_get_row_name(row))
                    continue
                elif operation_mode == "create":
                    logger.info(f"[Import] 执行新增操作")
                    if 'version_id' not in record and context.get('version_id'):
                        record['version_id'] = context.get('version_id')
                    logger.info(f"[Import] 新增数据中version_id={record.get('version_id')}")
                    # [FIX v1.2.18l 2026-06-20] 当 Excel 中显式填写了 create 时，按 create 语义执行（不 upsert）
                    # conflict_strategy=upsert 只在未显式指定操作模式时生效
                    if conflict_strategy == "upsert" and not operation_mode_explicit:
                        logger.info(f"[Import] conflict_strategy=upsert，使用 upsert 处理可能已存在的记录")
                        upsert_result = self._upsert_record(object_type, record, obj.import_export)
                        if upsert_result["success"]:
                            success_count += 1
                            op = upsert_result.get("operation", "create")
                            # [NEW v1.2.3 2026-06-17] 拆分 created / updated 统计
                            if op == "create":
                                created_count += 1
                            else:
                                updated_count += 1
                            _record_success_item(successes, row_num, op, record, _MAX_DETAIL, code_override=_get_row_code(row), name_override=_get_row_name(row))
                        else:
                            failed_count += 1
                            errors.append({"row": row_num, "operation": operation_mode, "field": "编码", "value": record.get("code", ""), "message": upsert_result.get("error", "Upsert failed")})
                    else:
                        result = self.manage_service.create(CreateRequest(object_type=object_type, data=record))
                        if result.success:
                            success_count += 1
                            created_count += 1
                            _record_success_item(successes, row_num, "create", record, _MAX_DETAIL, code_override=_get_row_code(row), name_override=_get_row_name(row))
                        else:
                            failed_count += 1
                            errors.append({"row": row_num, "operation": operation_mode, "field": "编码", "value": record.get("code", ""), "message": result.message or "创建失败"})
                elif operation_mode == "update":
                    # [SYMBOL] 关键修复：如果 conflict_strategy=upsert，执行 upsert 而不是更新
                    # [FIX v1.2.18l 2026-06-20] Excel 中显式 update 时，按 update 语义执行（不 upsert）
                    if conflict_strategy == "upsert" and not operation_mode_explicit:
                        logger.info(f"[Import] 执行upsert操作 (operation_mode=update 但 conflict_strategy=upsert)")
                        if 'version_id' not in record and context.get('version_id'):
                            record['version_id'] = context.get('version_id')
                        upsert_result = self._upsert_record(object_type, record, obj.import_export)
                        if upsert_result["success"]:
                            success_count += 1
                            op = upsert_result.get("operation", "update")
                            if op == "create":
                                created_count += 1
                            else:
                                updated_count += 1
                            _record_success_item(successes, row_num, op, record, _MAX_DETAIL, code_override=_get_row_code(row), name_override=_get_row_name(row))
                        else:
                            failed_count += 1
                            errors.append({"row": row_num, "operation": operation_mode, "field": "编码", "value": record.get("code", ""), "message": upsert_result.get("error", "Upsert failed")})
                    else:
                        logger.info(f"[Import] 执行更新操作")
                        self._update_record(object_type, record, obj.import_export)
                        success_count += 1
                        updated_count += 1
                        _record_success_item(successes, row_num, "update", record, _MAX_DETAIL, code_override=_get_row_code(row), name_override=_get_row_name(row))
                else:
                    logger.info(f"[Import] 执行upsert操作 (conflict_strategy={conflict_strategy})")
                    # [SYMBOL] 确保version_id存在
                    if 'version_id' not in record and context.get('version_id'):
                        record['version_id'] = context.get('version_id')
                    logger.info(f"[Import] upsert数据中version_id={record.get('version_id')}")
                    if conflict_strategy == "upsert":
                        upsert_result = self._upsert_record(object_type, record, obj.import_export)
                        if upsert_result["success"]:
                            success_count += 1
                            op = upsert_result.get("operation", "upsert")
                            if op == "create":
                                created_count += 1
                            else:
                                updated_count += 1
                            _record_success_item(successes, row_num, op, record, _MAX_DETAIL, code_override=_get_row_code(row), name_override=_get_row_name(row))
                        else:
                            failed_count += 1
                            errors.append({"row": row_num, "operation": operation_mode, "field": "编码", "value": record.get("code", ""), "message": upsert_result.get("error", "Upsert failed")})
                    elif conflict_strategy == "skip":
                        if self._record_exists(object_type, record, obj.import_export):
                            logger.info(f"[Import] 记录已存在，跳过")
                            skipped_count += 1
                            _record_skipped_item(skipped_items, row_num, "skip", record, "记录已存在", _MAX_DETAIL, code_override=_get_row_code(row), name_override=_get_row_name(row))
                            continue
                        result = self.manage_service.create(CreateRequest(object_type=object_type, data=record))
                        if result.success:
                            success_count += 1
                            created_count += 1
                            _record_success_item(successes, row_num, "create", record, _MAX_DETAIL, code_override=_get_row_code(row), name_override=_get_row_name(row))
                        else:
                            failed_count += 1
                    else:
                        result = self.manage_service.create(CreateRequest(object_type=object_type, data=record))
                        if result.success:
                            success_count += 1
                            created_count += 1
                            _record_success_item(successes, row_num, "create", record, _MAX_DETAIL, code_override=_get_row_code(row), name_override=_get_row_name(row))
                        else:
                            failed_count += 1

            except Exception as e:
                logger.error(f"[Import] ERROR: Operation failed - {type(e).__name__}: {str(e)}")
                failed_count += 1
                errors.append({
                    "row": row_num,
                    "operation": operation_mode,
                    "field": "编码",
                    "value": record.get("code", ""),
                    "message": f"{type(e).__name__}: {str(e)}"
                })

        logger.info(f"[Import] 导入完成: success={success_count}, created={created_count}, updated={updated_count}, "
                    f"failed={failed_count}, skipped={skipped_count}, deleted={deleted_count}")
        logger.info(f"[Import] errors: {errors[:5]}")

        return {
            "success": success_count,
            "created": created_count,   # [NEW v1.2.3 2026-06-17] 拆分 created
            "updated": updated_count,   # [NEW v1.2.3 2026-06-17] 拆分 updated
            "failed": failed_count,
            "skipped": skipped_count,
            "deleted": deleted_count,
            "errors": errors[:20],
            "warnings": warnings[:20],  # [NEW v1.2.3 2026-06-17] 告警明细
            # [NEW v1.2.14 2026-06-19] 成功/跳过明细 (前端第 4 步 subtab 用)
            "successes": successes,
            "skipped_items": skipped_items,
        }

    def _get_business_key_fields(self, object_type: str) -> List:
        """获取对象的业务键字段列表（支持组合键）

        [FR-011] 委托到 MetaObject.get_business_key_fields()。
        """
        obj = registry.get(object_type)
        if not obj:
            return []
        return obj.get_business_key_fields()

    def _find_existing_record(self, object_type: str, record: Dict[str, Any],
                               version_id: Optional[int] = None) -> Optional[Dict[str, Any]]:
        """根据业务键或conflict_key查找现有记录（支持组合键）

        Args:
            object_type: 对象类型
            record: 记录数据
            version_id: 版本ID，如果指定则只在该版本内查找
        """
        obj = registry.get(object_type)
        conflict_key = ""
        if obj and obj.import_export and obj.import_export.conflict_key:
            conflict_key = obj.import_export.conflict_key

        if conflict_key:
            key_value = record.get(conflict_key)
            if key_value is not None:
                return self._find_by_key(object_type, conflict_key, key_value, version_id)
            return None

        bk_fields = self._get_business_key_fields(object_type)

        if not bk_fields:
            record_id = record.get('id')
            if record_id is not None:
                return self._find_by_id(object_type, record_id)
            return None

        if len(bk_fields) == 1:
            key_value = record.get(bk_fields[0].id)
            if key_value is None:
                return None
            # [SYMBOL] 传入 version_id，只在指定版本内查找
            return self._find_by_key(object_type, bk_fields[0].id, key_value, version_id)
        else:
            key_values = []
            for bk_field in bk_fields:
                key_values.append(record.get(bk_field.id))

            if all(v is None for v in key_values):
                return None

            # [SYMBOL] 传入 version_id，只在指定版本内查找
            return self._find_by_composite_key(object_type, bk_fields, key_values, version_id)

    def _delete_record(self, object_type: str, record: Dict[str, Any], config: ImportExportConfig):
        """删除记录"""
        # 从record中获取version_id
        version_id = record.get('version_id')
        existing = self._find_existing_record(object_type, record, version_id)

        if existing:
            self.manage_service.delete(DeleteRequest(object_type=object_type, id=existing["id"]))
        else:
            bk_fields = self._get_business_key_fields(object_type)
            obj = registry.get(object_type)
            obj_name = obj.name if obj else object_type
            if bk_fields:
                key_desc = ", ".join(["{0}={1}".format(f.id, record.get(f.id)) for f in bk_fields])
            elif record.get('id') is not None:
                key_desc = "id={0}".format(record.get('id'))
            else:
                # [FIX v1.2.18i 2026-06-20] 无业务键对象(如 annotation)删除时提示需要 ID
                key_desc = "无业务键，{0} 删除需要在 Excel 中填写 ID 列".format(obj_name)
            raise ValueError("要删除的记录不存在: {0}".format(key_desc))

    def _update_record(self, object_type: str, record: Dict[str, Any], config: ImportExportConfig):
        """更新记录（仅更新，不存在则报错）"""
        # 从record中获取version_id
        version_id = record.get('version_id')
        existing = self._find_existing_record(object_type, record, version_id)

        if existing:
            self.manage_service.update(UpdateRequest(object_type=object_type, id=existing["id"], data=record))
        else:
            bk_fields = self._get_business_key_fields(object_type)
            if bk_fields:
                key_desc = ", ".join(["{0}={1}".format(f.id, record.get(f.id)) for f in bk_fields])
            else:
                key_desc = "无业务键"
            raise ValueError("要更新的记录不存在: {0}".format(key_desc))

    def _upsert_record(self, object_type: str, record: Dict[str, Any],
                       config: ImportExportConfig) -> dict:
        """Upsert记录：存在则更新，不存在则插入

        Returns:
            dict: {"success": bool, "error": str or None, "operation": "create"|"update"}
        """

        record_version_id = record.get('version_id')
        logger.info(f"[Upsert] object_type={object_type}, record_version_id={record_version_id}")
        logger.info(f"[Upsert] record keys: {list(record.keys())}")

        # [SYMBOL] 关键修复：传入 version_id，只在当前版本查找
        existing = self._find_existing_record(object_type, record, record_version_id)
        logger.info(f"[Upsert] _find_existing_record -> existing={existing.get('id') if existing else None}, "
                    f"existing_version_id={existing.get('version_id') if existing else None}, "
                    f"existing_code={existing.get('code') if existing else None}")

        if existing:
            existing_version_id = existing.get('version_id')
            record_id = existing.get("id")
            logger.info(f"[Upsert] 找到已存在记录: id={record_id}, version_id={existing_version_id}")
            logger.info(f"[Upsert] 将执行更新操作...")

            # [SYMBOL] 关键修复：确保 UPDATE 时 version_id 也被设置
            if record_version_id is not None:
                record['version_id'] = record_version_id
                logger.info(f"[Upsert] 强制设置 record.version_id={record_version_id}")

            # [FIX v1.2.18 2026-06-20] UPDATE 时强制 parent_key 字段保持原值, 避免 PARENT_FIELD_IMMUTABLE
            # parent_key 字段（如 sub_domain_id）在 hierarchy_validation 中被标记为 immutable
            # 即便用户 Excel 里填了新领域 (PROCUREMENT), DB 原值是其他, 也以 DB 为准
            # [FIX] 不再删除 parent_key + resolve_from_field (会触发 action_executor 重新解析),
            #       而是直接用 existing 原值覆盖 record 中所有 parent_key 字段
            obj_meta = registry.get(object_type)
            if obj_meta:
                for field in obj_meta.fields:
                    is_pk = getattr(field.semantics, 'parent_key', False)
                    if is_pk:
                        # 用 existing 原值覆盖 record 中的 parent_key 值
                        original_value = existing.get(field.id)
                        if original_value is not None and field.id in record:
                            if str(record.get(field.id)) != str(original_value):
                                logger.info(f"[Upsert] parent_key {field.id} 强制使用 DB 原值: 原值={original_value}, Excel新值={record.get(field.id)}")
                            record[field.id] = original_value
            else:
                logger.warning(f"[Upsert] registry.get({object_type}) returned None, cannot normalize parent_key fields")

            result = self.manage_service.update(UpdateRequest(object_type=object_type, id=record_id, data=record))
            logger.info(f"[Upsert] update result: success={result.success}, error={result.error}, record_keys_after_remove={list(record.keys())}")
            if result.success:
                return {"success": True, "error": None, "operation": "update"}
            else:
                error_msg = f"更新失败: {result.error} - {result.message}"
                logger.warning(f"[Upsert] {error_msg}")
                return {"success": False, "error": error_msg, "operation": "update"}
        else:
            logger.info(f"[Upsert] 未找到已存在记录，将执行插入操作")
            if 'version_id' not in record:
                logger.warning(f"[Upsert] [WARNING] record中没有version_id！")
            else:
                logger.info(f"[Upsert] record.version_id={record_version_id}")

            # [FIX 2026-06-17] 如果 code 为空且对象有 key_template，自动生成 code
            # action_executor._do_create 不走 bo_framework 拦截器链，
            # KeyTemplateInterceptor 不会执行，需要手动生成
            code_value = record.get('code', '')
            if (not code_value or not str(code_value).strip()):
                kt_code = self._auto_generate_code_from_key_template(object_type, record)
                if kt_code:
                    record['code'] = kt_code
                    logger.info(f"[Upsert] Key template auto-generated code: {kt_code}")

            result = self.manage_service.create(CreateRequest(object_type=object_type, data=record))
            if result.success:
                return {"success": True, "error": None, "operation": "create"}
            else:
                error_msg = f"创建失败: {result.error} - {result.message}"
                logger.warning(f"[Upsert] {error_msg}")
                return {"success": False, "error": error_msg, "operation": "create"}

    def _record_exists(self, object_type: str, record: Dict[str, Any],
                       config: ImportExportConfig) -> bool:
        """检查记录是否存在

        [FIX 2026-06-08] 传递 record.get('version_id') 以保证 conflict_strategy='skip'
        不会跨版本误命中（之前漏传会导致同 business_key 的 v1 记录遮挡 v2 新增）。
        """
        return self._find_existing_record(
            object_type, record, record.get('version_id')
        ) is not None

    def _find_by_id(self, object_type: str, record_id: Any) -> Optional[Dict[str, Any]]:
        """根据 id 查找记录（用于无业务键的导入对象，如 annotation）

        [FIX FR-003] 改用 data_source.find_by_id，与 manage_service 保持一致，
        避免手写 SQL 漂移风险。
        """
        try:
            obj = registry.get(object_type)
            if not obj:
                return None
            table_name = obj.table_name or object_type
            return self.data_source.find_by_id(table_name, record_id)
        except Exception:
            return None

    def _find_by_key(self, object_type: str, key_field: str, key_value: Any,
                      version_id: Optional[int] = None) -> Optional[Dict[str, Any]]:
        """根据关键字段查找记录

        Args:
            object_type: 对象类型
            key_field: 关键字段名
            key_value: 关键字段值
            version_id: 版本ID，如果指定则只在该版本内查找
        """
        try:
            conditions = [QueryCondition(field=key_field, operator="eq", value=key_value)]

            if version_id is not None and self._object_has_field(object_type, "version_id"):
                conditions.append(QueryCondition(field="version_id", operator="eq", value=version_id))

            search_request = SearchRequest(
                object_type=object_type,
                conditions=conditions,
                page=1,
                page_size=1,
                # [FIX v1.2.16 2026-06-20] upsert 路径不应受数据权限过滤,
                # 否则非 admin 用户 (或维度 scope 限制) 会找不到已存在记录 → 误走 create → 创建失败
                # upsert 内部的 update 流程由 action_executor 做权限检查, 这里只需纯存在性查询
                skip_data_permission=True,
            )
            result = self.query_service.search(search_request)
            return result.data[0] if result.data else None
        except Exception:
            return None

    def _find_by_composite_key(self, object_type: str, key_fields: List, key_values: List[str],
                                version_id: Optional[int] = None) -> Optional[Dict[str, Any]]:
        """根据组合关键字段查找记录

        Args:
            object_type: 对象类型
            key_fields: 关键字段列表
            key_values: 关键字段值列表
            version_id: 版本ID，如果指定则只在该版本内查找
        """
        try:
            conditions = []
            for field, value in zip(key_fields, key_values):
                if value:
                    conditions.append(QueryCondition(field=field.id, operator="eq", value=value))

            if not conditions:
                return None

            if version_id is not None and self._object_has_field(object_type, "version_id"):
                conditions.append(QueryCondition(field="version_id", operator="eq", value=version_id))

            search_request = SearchRequest(
                object_type=object_type,
                conditions=conditions,
                page=1,
                page_size=1,
            )
            result = self.query_service.search(search_request)
            return result.data[0] if result.data else None
        except Exception:
            return None

    def _read_meta_sheet(self, file_path: str) -> Dict[str, str]:
        """从 Excel 文件的元数据 Sheet 读取上下文信息
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            dict: 包含 product_code, version_code, version_id 等信息的字典
        """
        meta = {}
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            
            if '元数据' in wb.sheetnames:
                ws = wb['元数据']
                for row in ws.iter_rows(min_row=1, max_row=10, max_col=2):
                    if row[0].value and row[1].value:
                        meta[str(row[0].value)] = str(row[1].value)
            
            wb.close()
        except Exception as e:
            logger.warning(f"Failed to read meta sheet: {e}")
        
        return meta

    def _resolve_version_id(self, product_code: str, version_code: str) -> Optional[int]:
        """根据 product_code 和 version_code 解析 version_id
        
        Args:
            product_code: 产品编码
            version_code: 版本编码
            
        Returns:
            int: version_id，如果未找到返回 None
        """
        if not product_code or not version_code:
            return None
        
        try:
            query = """
                SELECT v.id
                FROM versions v
                LEFT JOIN products p ON v.product_id = p.id
                WHERE p.code = ? AND v.code = ?
                LIMIT 1
            """
            cursor = self.data_source.execute(query, (product_code, version_code))
            row = cursor.fetchone()
            if row:
                return row[0]
        except Exception as e:
            logger.warning(f"Failed to resolve version_id: {e}")
        
        return None
