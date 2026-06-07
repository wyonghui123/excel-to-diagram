# -*- coding: utf-8 -*-
"""
Excel导出HTTP验证测试

通过HTTP API测试导出功能，验证生成的Excel内容：
1. 敏感字段不导出
2. 层级路径列控制
3. 分页导出控制
"""

import requests
import json
import os
import sys
from openpyxl import load_workbook

BASE_URL = "http://localhost:3010"
HEADERS = {
    'Content-Type': 'application/json'
}


def login_and_get_token():
    """登录获取token"""
    response = requests.post(
        f"{BASE_URL}/api/v1/auth/login",
        json={"username": "admin", "password": "admin123"},
        headers=HEADERS
    )
    if response.status_code == 200:
        data = response.json()
        if data.get('success'):
            token = data.get('data', {}).get('token', '')
            return {
                'Authorization': f'Bearer {token}',
                'Content-Type': 'application/json'
            }
    print(f"登录失败: {response.status_code} {response.text}")
    return None


def export_and_verify(object_type, options=None, page=None, page_size=None):
    """导出并验证Excel内容"""
    payload = {
        'object_type': object_type,
        'scope': 'single'
    }

    if page is not None and page_size is not None:
        payload['page'] = page
        payload['page_size'] = page_size

    if options:
        payload['options'] = options

    headers = login_and_get_token()
    if not headers:
        return None

    print(f"\n{'='*60}")
    print(f"测试导出: {object_type}")
    print(f"选项: {options}")
    print(f"分页: page={page}, page_size={page_size}")
    print(f"{'='*60}")

    response = requests.post(
        f"{BASE_URL}/api/v1/export",
        json=payload,
        headers=headers
    )

    if response.status_code != 200:
        print(f"导出请求失败: {response.status_code}")
        print(response.text)
        return None

    data = response.json()
    if not data.get('success'):
        print(f"导出失败: {data.get('message')}")
        return None

    file_path = data.get('data', {}).get('file_path')
    if not file_path:
        print("没有返回文件路径")
        return None

    # 确保文件存在
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return None

    print(f"导出成功，文件: {file_path}")
    return file_path


def verify_excel_content(file_path, expected_excludes=None, expected_includes=None, max_rows=None):
    """验证Excel内容"""
    if not file_path or not os.path.exists(file_path):
        return False

    print(f"\n验证Excel内容: {file_path}")

    wb = load_workbook(file_path, read_only=True)

    all_checks_passed = True

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n工作表: {sheet_name}")

        # 获取表头
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
        print(f"  列数: {len(headers)}")
        print(f"  表头: {headers[:10]}{'...' if len(headers) > 10 else ''}")

        # 验证排除列
        if expected_excludes:
            for excluded_col in expected_excludes:
                if excluded_col in headers:
                    print(f"  [X] 错误: 不应包含的列 '{excluded_col}' 出现了")
                    all_checks_passed = False
                else:
                    print(f"  [OK] 正确: 列 '{excluded_col}' 未出现")

        # 验证包含列
        if expected_includes:
            for included_col in expected_includes:
                if included_col in headers:
                    print(f"  [OK] 正确: 列 '{included_col}' 已包含")
                else:
                    print(f"  [X] 错误: 列 '{included_col}' 未出现")
                    all_checks_passed = False

        # 验证行数
        if max_rows:
            row_count = ws.max_row - 1
            print(f"  数据行数: {row_count} (期望 <= {max_rows})")
            if row_count > max_rows:
                print(f"  [X] 错误: 导出行数({row_count})超过限制({max_rows})")
                all_checks_passed = False
            else:
                print(f"  [OK] 正确: 导出行数符合限制")

    wb.close()
    return all_checks_passed


def main():
    print("="*60)
    print("Excel导出功能验证测试")
    print("="*60)

    # 测试1: 导出业务对象，不包含层级路径
    file_path = export_and_verify(
        'business_object',
        options={
            'include_hierarchy_path': False,
            'include_hierarchy_ids': False
        }
    )
    if file_path:
        verify_excel_content(
            file_path,
            expected_excludes=['层级路径', 'hierarchy_path', '路径', 'path']
        )

    # 测试2: 导出业务对象，包含层级路径
    file_path = export_and_verify(
        'business_object',
        options={
            'include_hierarchy_path': True,
            'include_hierarchy_ids': True
        }
    )
    if file_path:
        verify_excel_content(
            file_path,
            expected_includes=['层级路径']
        )

    # 测试3: 导出用户对象，验证敏感字段不导出
    file_path = export_and_verify('user')
    if file_path:
        verify_excel_content(
            file_path,
            expected_excludes=['password_hash', 'password']
        )

    # 测试4: 导出领域对象
    file_path = export_and_verify('domain')
    if file_path:
        verify_excel_content(file_path)

    print("\n" + "="*60)
    print("测试完成")
    print("="*60)


if __name__ == '__main__':
    main()
