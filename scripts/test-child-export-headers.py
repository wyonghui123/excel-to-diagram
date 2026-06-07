"""
测试子对象导出表头 — 验证使用 export_visible 而非 child_sections columns
"""
import urllib.request, urllib.error, json, sys, os, time

BASE = "http://localhost:3011"
TOKEN = None

def login():
    global TOKEN
    data = json.dumps({"username":"admin","password":"admin123"}).encode()
    req = urllib.request.Request(f"{BASE}/api/v1/auth/login", data=data,
        headers={"Content-Type":"application/json"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        r = json.loads(resp.read().decode())
        TOKEN = r["data"]["token"]
        print("登录成功")
        return TOKEN

def api(path, method="GET", body=None):
    url = f"{BASE}{path}"
    headers = {"Content-Type":"application/json"}
    if TOKEN: headers["Authorization"] = f"Bearer {TOKEN}"
    data = json.dumps(body).encode() if body else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                return json.loads(resp.read().decode())
        except urllib.error.HTTPError as e:
            error_body = e.read().decode()[:1000]
            return {"error": e.code, "body": error_body}
        except Exception as e:
            if attempt < 2:
                time.sleep(2)
                continue
            return {"error": str(e)}
    return {}

def do_export_and_check(label, options, expect_child_sheet):
    print(f"\n{'='*60}")
    print(f"  测试: {label}")
    print(f"{'='*60}")

    body = {
        "object_type": "domain",
        "scope": "selected",
        "selected_types": ["domain", "sub_domain", "service_module", "business_object", "relationship"],
        "filters": {"version_id": version_id},
        "options": options
    }

    r = api("/api/v1/export", method="POST", body=body)

    if r.get("error"):
        print(f"  ERROR: {r.get('error')} - {r.get('body','')[:500]}")
        return

    data = r.get("data", {})
    if data.get("success") is False:
        print(f"  FAILED: {data.get('message','')} {data.get('error','')}"[:500])
        return

    sheets = data.get("sheets", [])
    file_path = data.get("file_path", "")
    print(f"  file_path: {file_path}")

    if sheets:
        print(f"  Sheets from response:")
        for s in sheets:
            name = s.get("name", "?")
            count = s.get("row_count", s.get("count", 0))
            obj_type = s.get("object_type", "")
            print(f"    [{name}] ({obj_type}): {count} rows")

    has_annotation = any(s.get("name") == "备注信息" or s.get("object_type") == "annotation" for s in sheets)
    if expect_child_sheet and not has_annotation:
        print(f"  WARNING: 期望有子对象表但未找到")
    elif not expect_child_sheet and has_annotation:
        print(f"  WARNING: 不应有子对象表但找到了")

    if file_path and os.path.exists(file_path):
        print(f"\n  读取文件验证表头:")
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            row_count = sum(1 for _ in ws.iter_rows(min_row=2))
            print(f"    [{sn}]: row_count={row_count}")
            print(f"      表头: {header_row}")

            if sn == "备注信息" or sn == "annotation":
                print(f"      >>> 这是子对象 (annotation) 的导出表头")
                for h in header_row:
                    if h and ('target_type' in str(h).lower() or 'category' in str(h)):
                        print(f"      >>> 发现不应导出的字段: {h}")
        wb.close()

    return r

def main():
    print("="*60)
    print("  子对象导出表头测试")
    print("="*60)

    global version_id
    login()
    if not TOKEN: sys.exit(1)

    versions = api("/api/v2/bo/version?page_size=1")
    items = (versions.get("data") or {}).get("items", [])
    if not items:
        print("没有版本数据")
        sys.exit(1)
    version_id = items[0]["id"]
    print(f"版本 ID={version_id}")

    base_opts = {
        "include_hierarchy_path": False,
        "include_hierarchy_ids": True,
        "protect_sheet": False,
        "mark_readonly": False,
    }

    opts1 = dict(base_opts)
    opts1["include_child_objects"] = True
    do_export_and_check("include_child_objects=true", opts1, expect_child_sheet=True)

    opts2 = dict(base_opts)
    opts2["include_child_objects"] = False
    do_export_and_check("include_child_objects=false", opts2, expect_child_sheet=False)

    print(f"\n{'='*60}")
    print(f"  Done")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()
