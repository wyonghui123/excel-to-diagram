# -*- coding: utf-8 -*-
"""
用户导出字段验证测试
"""

import requests
import json
import os
from openpyxl import load_workbook

BASE_URL = "http://localhost:3010"


def login_and_get_token():
    """登录获取token"""
    response = requests.post(
        f"{BASE_URL}/api/v1/auth/login",
        json={"username": "admin", "password": "admin123"},
        headers={'Content-Type': 'application/json'}
    )
    if response.status_code == 200:
        data = response.json()
        if data.get('success'):
            token = data.get('data', {}).get('token', '')
            return {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    print(f"登录失败: {response.status_code} {response.text}")
    return None


def test_user_export_fields():
    """测试用户导出的字段"""
    headers = login_and_get_token()
    if not headers:
        return

    print("="*60)
    print("测试用户导出字段")
    print("="*60)

    # 导出用户
    response = requests.post(
        f"{BASE_URL}/api/v1/export",
        json={'object_type': 'user', 'scope': 'single'},
        headers=headers
    )

    if response.status_code != 200:
        print(f"导出请求失败: {response.status_code}")
        print(response.text)
        return

    data = response.json()
    if not data.get('success'):
        print(f"导出失败: {data.get('message')}")
        return

    file_path = data.get('data', {}).get('file_path')
    if not file_path or not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return

    print(f"导出成功，文件: {file_path}\n")

    wb = load_workbook(file_path, read_only=True)

    # 检查用户工作表
    user_headers = []
    for sheet_name in wb.sheetnames:
        if '用户' in sheet_name:
            ws = wb[sheet_name]
            print(f"工作表: {sheet_name}")
            print(f"列数: {ws.max_column}")
            print(f"数据行数: {ws.max_row - 1}")
            print("\n所有表头:")

            for i, cell in enumerate(ws[1], 1):
                col_name = cell.value
                user_headers.append(col_name)
                print(f"  {i}. {col_name}")

            # 验证不应出现的字段
            print("\n验证不应出现的字段:")
            excluded_fields = [
                '密码哈希', 'password_hash', '密码', 'password',
                '语言区域', 'locale', '时区', 'timezone',
                '日期格式长度', 'date_style', '时间格式长度', 'time_style',
                '时间制式', 'hour_cycle',
                '密码历史', 'password_history',
                '最后登录时间', 'last_login_at'  # 可选，有些系统可能需要保留
            ]

            for field in excluded_fields:
                if field in user_headers:
                    print(f"  [X] 错误: '{field}' 不应出现在导出结果中")
                else:
                    print(f"  [OK] 正确: '{field}' 未出现")

            break

    wb.close()


if __name__ == '__main__':
    test_user_export_fields()
