# -*- coding: utf-8 -*-
"""
导入模板HTTP测试
"""

import requests
import json
import os
from openpyxl import load_workbook

BASE_URL = "http://localhost:3010"


def login_and_get_token():
    """登录获取token"""
    response = requests.post(
        f"{BASE_URL}/api/v1/auth/login",
        json={"username": "admin", "password": "admin123"},
        headers={'Content-Type': 'application/json'}
    )
    if response.status_code == 200:
        data = response.json()
        if data.get('success'):
            token = data.get('data', {}).get('token', '')
            return {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    print(f"登录失败: {response.status_code} {response.text}")
    return None


def test_role_template():
    """测试角色导入模板"""
    headers = login_and_get_token()
    if not headers:
        return

    print("="*60)
    print("测试角色导入模板")
    print("="*60)

    # 下载角色模板
    response = requests.get(
        f"{BASE_URL}/api/v1/import/template/role",
        headers=headers
    )

    if response.status_code == 200:
        # 保存文件
        export_dir = r"d:\filework\excel-to-diagram\meta\tests\exports"
        os.makedirs(export_dir, exist_ok=True)
        file_path = os.path.join(export_dir, 'role_template.xlsx')

        with open(file_path, 'wb') as f:
                    f.write(response.content)

        print(f"模板文件保存到: {file_path}")
        print(f"文件大小: {os.path.getsize(file_path)} bytes")

        # 验证模板内容
        wb = load_workbook(file_path)
        ws = wb['角色']

        # 获取表头
        headers_list = []
        for cell in ws[1]:
            if cell.value:
                headers_list.append(cell.value)

        print(f"\n角色模板表头 ({len(headers_list)} 列):")
        for i, h in enumerate(headers_list, 1):
            print(f"  {i}. {h}")

        # 检查readonly字段
        readonly_fields = []
        editable_fields = []
        business_key_fields = []

        if ws.max_row > 1:
            for col_idx in range(1, ws.max_column + 1):
                header_name = ws.cell(row=1, column=col_idx).value
                if not header_name:
                    continue

                data_cell = ws.cell(row=2, column=col_idx)
                fill = data_cell.fill

                fill_color = None
                if hasattr(fill, 'start_color') and fill.start_color:
                    if hasattr(fill.start_color, 'rgb'):
                        fill_color = str(fill.start_color.rgb)

                if fill_color and 'E0E0E0' in fill_color:
                    readonly_fields.append(header_name)
                elif fill_color and 'FFF2CC' in fill_color:
                    business_key_fields.append(header_name)
                else:
                    editable_fields.append(header_name)

        print(f"\nreadonly字段: {readonly_fields}")
        print(f"业务键字段: {business_key_fields}")
        print(f"可编辑字段: {editable_fields}")

        # 验证统计字段是readonly
        stats = ['用户数', '菜单数', '权限数', '数据权限数']
        for stat in stats:
            if stat in headers_list:
                if stat in readonly_fields:
                    print(f"  [OK] {stat} 是readonly")
                else:
                    print(f"  [X] {stat} 应该是readonly")

        wb.close()
    else:
        print(f"下载失败: {response.status_code}")
        print(response.text)


def test_user_template():
    """测试用户导入模板"""
    headers = login_and_get_token()
    if not headers:
        return

    print("\n" + "="*60)
    print("测试用户导入模板")
    print("="*60)

    # 下载用户模板
    response = requests.get(
        f"{BASE_URL}/api/v1/import/template/user",
        headers=headers
    )

    if response.status_code == 200:
        export_dir = r"d:\filework\excel-to-diagram\meta\tests\exports"
        file_path = os.path.join(export_dir, 'user_template.xlsx')

        with open(file_path, 'wb') as f:
            f.write(response.data)

        print(f"模板文件保存到: {file_path}")

        wb = load_workbook(file_path)

        if '用户' in wb.sheetnames:
            ws = wb['用户']
            headers_list = [cell.value for cell in ws[1] if cell.value]

            print(f"\n用户模板表头 ({len(headers_list)} 列):")
            for i, h in enumerate(headers_list, 1):
                print(f"  {i}. {h}")

            # 检查敏感字段
            sensitive_fields = ['密码哈希', 'password_hash', '密码', 'password',
                              '语言区域', 'locale', '时区', 'timezone']
            found_sensitive = []
            for sf in sensitive_fields:
                if sf in headers_list:
                    found_sensitive.append(sf)

            if found_sensitive:
                print(f"\n[X] 敏感字段出现在模板中: {found_sensitive}")
            else:
                print(f"\n[OK] 敏感字段未出现在模板中")

        wb.close()
    else:
        print(f"下载失败: {response.status_code}")


if __name__ == '__main__':
    test_role_template()
    test_user_template()
    print("\n测试完成!")
