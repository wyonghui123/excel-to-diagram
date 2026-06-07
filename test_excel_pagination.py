# -*- coding: utf-8 -*-
"""
Excel分页导出验证测试
"""

import requests
import json
import os
from openpyxl import load_workbook

BASE_URL = "http://localhost:3010"


def login_and_get_token():
    """登录获取token"""
    response = requests.post(
        f"{BASE_URL}/api/v1/auth/login",
        json={"username": "admin", "password": "admin123"},
        headers={'Content-Type': 'application/json'}
    )
    if response.status_code == 200:
        data = response.json()
        if data.get('success'):
            token = data.get('data', {}).get('token', '')
            return {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    print(f"登录失败: {response.status_code} {response.text}")
    return None


def test_pagination_export():
    """测试分页导出功能"""
    headers = login_and_get_token()
    if not headers:
        return

    print("="*60)
    print("测试分页导出功能")
    print("="*60)

    # 测试1: 导出全部数据（无分页）
    print("\n【测试1】导出全部数据")
    response = requests.post(
        f"{BASE_URL}/api/v1/export",
        json={'object_type': 'domain', 'scope': 'all'},
        headers=headers
    )

    if response.status_code == 200:
        data = response.json()
        if data.get('success'):
            file_path = data.get('data', {}).get('file_path')
            if file_path and os.path.exists(file_path):
                wb = load_workbook(file_path, read_only=True)
                if '领域' in wb.sheetnames:
                    ws = wb['领域']
                    row_count = ws.max_row - 1
                    print(f"  总行数: {row_count}")
                wb.close()
            else:
                print("  [X] 没有返回文件路径")
        else:
            print(f"  [X] 导出失败: {data.get('message')}")
    else:
        print(f"  [X] 请求失败: {response.status_code}")

    # 测试2: 导出前5条（分页）
    print("\n【测试2】导出前5条数据 (page=1, page_size=5)")
    response = requests.post(
        f"{BASE_URL}/api/v1/export",
        json={'object_type': 'domain', 'scope': 'all', 'page': 1, 'page_size': 5},
        headers=headers
    )

    if response.status_code == 200:
        data = response.json()
        if data.get('success'):
            file_path = data.get('data', {}).get('file_path')
            if file_path and os.path.exists(file_path):
                wb = load_workbook(file_path, read_only=True)
                if '领域' in wb.sheetnames:
                    ws = wb['领域']
                    row_count = ws.max_row - 1
                    print(f"  导出行数: {row_count}")

                    if row_count <= 5:
                        print(f"  [OK] 正确: 导出行数({row_count}) <= 限制(5)")
                    else:
                        print(f"  [X] 错误: 导出行数({row_count}) > 限制(5)")
                wb.close()
            else:
                print("  [X] 没有返回文件路径")
        else:
            print(f"  [X] 导出失败: {data.get('message')}")
    else:
        print(f"  [X] 请求失败: {response.status_code}")

    # 测试3: 导出前2条
    print("\n【测试3】导出前2条数据 (page=1, page_size=2)")
    response = requests.post(
        f"{BASE_URL}/api/v1/export",
        json={'object_type': 'domain', 'scope': 'all', 'page': 1, 'page_size': 2},
        headers=headers
    )

    if response.status_code == 200:
        data = response.json()
        if data.get('success'):
            file_path = data.get('data', {}).get('file_path')
            if file_path and os.path.exists(file_path):
                wb = load_workbook(file_path, read_only=True)
                if '领域' in wb.sheetnames:
                    ws = wb['领域']
                    row_count = ws.max_row - 1
                    print(f"  导出行数: {row_count}")

                    if row_count <= 2:
                        print(f"  [OK] 正确: 导出行数({row_count}) <= 限制(2)")
                    else:
                        print(f"  [X] 错误: 导出行数({row_count}) > 限制(2)")
                wb.close()
            else:
                print("  [X] 没有返回文件路径")
        else:
            print(f"  [X] 导出失败: {data.get('message')}")
    else:
        print(f"  [X] 请求失败: {response.status_code}")

    print("\n" + "="*60)
    print("分页测试完成")
    print("="*60)


if __name__ == '__main__':
    test_pagination_export()
