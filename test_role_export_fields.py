# -*- coding: utf-8 -*-
"""
角色导出字段验证测试 - 检查readonly标记
"""

import requests
import json
import os
from openpyxl import load_workbook

BASE_URL = "http://localhost:3010"


def login_and_get_token():
    """登录获取token"""
    response = requests.post(
        f"{BASE_URL}/api/v1/auth/login",
        json={"username": "admin", "password": "admin123"},
        headers={'Content-Type': 'application/json'}
    )
    if response.status_code == 200:
        data = response.json()
        if data.get('success'):
            token = data.get('data', {}).get('token', '')
            return {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    print(f"登录失败: {response.status_code} {response.text}")
    return None


def test_role_export():
    """测试角色导出"""
    headers = login_and_get_token()
    if not headers:
        return

    print("="*60)
    print("测试角色导出字段和readonly标记")
    print("="*60)

    # 导出角色
    response = requests.post(
        f"{BASE_URL}/api/v1/export",
        json={'object_type': 'role', 'scope': 'single'},
        headers=headers
    )

    if response.status_code != 200:
        print(f"导出请求失败: {response.status_code}")
        print(response.text)
        return

    data = response.json()
    if not data.get('success'):
        print(f"导出失败: {data.get('message')}")
        return

    file_path = data.get('data', {}).get('file_path')
    if not file_path or not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return

    print(f"导出成功，文件: {file_path}\n")

    wb = load_workbook(file_path)

    # 检查角色工作表
    role_headers = []
    readonly_columns = []

    for sheet_name in wb.sheetnames:
        if '角色' in sheet_name:
            ws = wb[sheet_name]
            print(f"工作表: {sheet_name}")
            print(f"列数: {ws.max_column}")
            print(f"数据行数: {ws.max_row - 1}")
            print("\n所有表头和readonly状态:")

            for col_idx in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=1, column=col_idx)
                header_name = header_cell.value

                role_headers.append(header_name)

                # 检查数据行的背景色来判断readonly（第2行是第一条数据）
                if ws.max_row > 1:
                    data_cell = ws.cell(row=2, column=col_idx)
                    fill = data_cell.fill
                    fill_color = None
                    if hasattr(fill, 'start_color') and fill.start_color:
                        if hasattr(fill.start_color, 'rgb'):
                            fill_color = fill.start_color.rgb
                        elif hasattr(fill.start_color, 'type'):
                            fill_color = f"type:{fill.start_color.type}"
                    
                    # 灰色背景 (E0E0E0) 表示只读
                    is_readonly = fill_color and 'E0E0E0' in str(fill_color)
                    status = "【只读】" if is_readonly else "【可编辑】"
                else:
                    status = "【未知】"
                    fill_color = None

                print(f"  {col_idx}. {header_name} {status} (color={fill_color})")

                # 记录readonly列
                if is_readonly:
                    readonly_columns.append(header_name)

            # 验证应该只读的字段
            print("\n验证应该只读的字段:")
            readonly_fields = [
                '用户数', 'user_count',
                '菜单数', 'menu_count',
                '权限数', 'permission_count',
                '数据权限数', 'data_perm_count',
                'ID', 'id',
                '创建时间', 'created_at',
                '更新时间', 'updated_at',
            ]

            for field in readonly_fields:
                if field in role_headers:
                    if field in readonly_columns:
                        print(f"  [OK] 正确: '{field}' 标记为只读")
                    else:
                        print(f"  [X] 错误: '{field}' 应该标记为只读但没有")

            break

    wb.close()


if __name__ == '__main__':
    test_role_export()
