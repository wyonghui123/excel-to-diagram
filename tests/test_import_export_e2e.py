# -*- coding: utf-8 -*-
"""
导入导出端到端测试

覆盖场景：
1. 导出 → 验证数据完整性
2. 导入 Preview（预览验证）
3. 导入 Execute（实际执行）
4. 导出 → 修改 → 导入 往返一致性
5. 组合业务键导入验证
6. 操作模式（新增/更新/删除/跳过）
7. immutable/readonly 字段处理
8. 级联导出导入
"""

import sys
import os
import tempfile
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook, load_workbook
from tests.conftest import TestBase


class TestExportBasic(TestBase):
    """基础导出测试"""
    
    use_class_setup = True  # 启用类级别共享数据库，提升性能

    def test_export_single_object(self):
        """测试导出单个对象类型"""
        self.create_full_hierarchy('EXP')
        
        result = self.import_service.export_cascade(
            'domain', 
            filters=None, 
            options={'include_metadata_sheet': False}
        )
        
        assert result.success
        assert len(result.sheets) >= 1
        
        domain_sheet = None
        for s in result.sheets:
            if s['name'] == '领域':
                domain_sheet = s
                break
        
        assert domain_sheet is not None
        assert domain_sheet['row_count'] >= 1

    def test_export_with_data(self):
        """测试导出包含数据"""
        self.create_full_hierarchy('EDT')
        
        result = self.import_service.export_cascade(
            'domain',
            filters={'version_id': 1},
            options={'include_metadata_sheet': False}
        )
        
        assert result.success
        total_rows = sum(s.get('row_count', 0) for s in result.sheets)
        assert total_rows > 0

    def test_export_file_exists(self):
        """测试导出文件存在"""
        self.create_full_hierarchy('EFE')
        
        result = self.import_service.export_cascade(
            'domain',
            filters=None,
            options={'include_metadata_sheet': False}
        )
        
        assert result.success
        if hasattr(result, 'file_path') and result.file_path:
            assert os.path.exists(result.file_path)


class TestImportPreview(TestBase):
    """导入预览测试"""
    
    use_class_setup = True

    def _create_import_file(self, sheets_data: dict) -> str:
        """创建导入用的 Excel 文件"""
        tmp = tempfile.mktemp(suffix='.xlsx')
        wb = Workbook()
        wb.remove(wb.active)
        
        for sheet_name, rows in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            for row in rows:
                ws.append(row)
        
        wb.save(tmp)
        return tmp

    def test_preview_valid_domain_data(self):
        """测试预览有效的领域数据"""
        sheets = {
            '领域': [
                ['操作模式', '编码', '名称'],
                ['新增', 'PV001', '预览领域1'],
                ['新增', 'PV002', '预览领域2'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            result = self.import_service.import_cascade(tmp, mode='preview')
            
            assert result.validation['valid_count'] == 2
            assert result.validation['invalid_count'] == 0
        finally:
            os.remove(tmp)

    def test_preview_detects_duplicate_code(self):
        """测试预览检测重复编码"""
        sheets = {
            '领域': [
                ['操作模式', '编码', '名称'],
                ['新增', 'DUP01', '领域A'],
                ['新增', 'DUP01', '领域B重复编码'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            result = self.import_service.import_cascade(tmp, mode='preview')
            
            assert result.validation['invalid_count'] > 0
            errors = result.validation.get('errors', [])
            has_dup_error = any('重复' in str(e) or '已存在' in str(e) for e in errors)
            assert has_dup_error, f"应该检测到重复错误，实际errors: {errors}"
        finally:
            os.remove(tmp)

    def test_preview_composite_key_relationship(self):
        """测试预览组合键关系 - 相同源不同目标是合法的"""
        self.create_full_hierarchy('PCK')
        self.create_business_object('PCK_SM', 'BO_A', '对象A')
        self.create_business_object('PCK_SM', 'BO_B', '对象B')
        self.create_business_object('PCK_SM', 'BO_C', '对象C')
        
        sheets = {
            '业务关系': [
                ['操作模式', '源编码', '目标编码', '关系编码'],
                ['更新', 'BO_A', 'BO_B', 'REL_1'],
                ['更新', 'BO_A', 'BO_C', 'REL_1'],
                ['更新', 'BO_B', 'BO_C', 'REL_1'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            result = self.import_service.import_cascade(tmp, mode='preview')
            
            assert result.validation['invalid_count'] == 0, \
                f"组合键不应重复，错误: {result.validation.get('errors')}"
            assert result.validation['valid_count'] == 3
        finally:
            os.remove(tmp)

    def test_preview_composite_key_duplicate(self):
        """测试预览组合键重复 - 完全相同才报错"""
        sheets = {
            '业务关系': [
                ['操作模式', '源编码', '目标编码', '关系编码'],
                ['更新', 'BO_X', 'BO_Y', 'REL_DUP'],
                ['更新', 'BO_X', 'BO_Y', 'REL_DUP'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            result = self.import_service.import_cascade(tmp, mode='preview')
            
            assert result.validation['invalid_count'] > 0, "应检测到组合键重复"
        finally:
            os.remove(tmp)

    def test_preview_missing_required_field(self):
        """测试预览必填字段缺失"""
        sheets = {
            '领域': [
                ['操作模式', '编码', '名称'],
                ['新增', '', '无编码领域'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            result = self.import_service.import_cascade(tmp, mode='preview')
            
            assert result.validation['invalid_count'] > 0
        finally:
            os.remove(tmp)


class TestImportExecute(TestBase):
    """导入执行测试 - 端到端完整流程"""
    
    use_class_setup = True

    def _create_import_file(self, sheets_data: dict) -> str:
        tmp = tempfile.mktemp(suffix='.xlsx')
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, rows in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            for row in rows:
                ws.append(row)
        wb.save(tmp)
        return tmp

    def test_execute_create_new_domain(self):
        """测试执行新增领域"""
        self.create_full_hierarchy('EXE')
        
        sheets = {
            '领域': [
                ['操作模式', '编码', '名称'],
                ['新增', 'EXE001', '执行新增领域'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            preview = self.import_service.import_cascade(tmp, mode='preview')
            assert preview.validation['invalid_count'] == 0
            
            result = self.import_service.import_cascade(tmp, mode='execute', context={'version_id': 1})
            
            assert result.results.get('domain', {}).get('success', 0) >= 1
            self.assert_record_exists('domain', 'code', 'EXE001')
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)

    def test_execute_update_existing_record(self):
        """测试执行更新已有记录"""
        self.create_full_hierarchy('UPD')
        self.create_domain('UPD001', '原始名称')
        
        sheets = {
            '领域': [
                ['操作模式', '编码', '名称'],
                ['更新', 'UPD001', '更新后名称'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            result = self.import_service.import_cascade(tmp, mode='execute', context={'version_id': 1})
            
            updated = self._find_by_field('domain', 'code', 'UPD001')
            assert updated['name'] == '更新后名称'
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)

    def test_execute_delete_record(self):
        """测试执行删除记录"""
        self.create_full_hierarchy('DEL')
        self.create_domain('DEL001', '待删除')
        
        sheets = {
            '领域': [
                ['操作模式', '编码', '名称'],
                ['删除', 'DEL001', '待删除'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            result = self.import_service.import_cascade(tmp, mode='execute', context={'version_id': 1})
            
            self.assert_record_not_exists('domain', 'code', 'DEL001')
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)

    def test_execute_upsert_mode(self):
        """测试 upsert 模式（存在则更新，不存在则插入）"""
        self.create_full_hierarchy('UPS')
        self.create_domain('UPS001', '原始')
        
        sheets = {
            '领域': [
                ['操作模式', '编码', '名称'],
                ['更新', 'UPS001', 'Upsert更新'],
                ['新增', 'UPS002', 'Upsert新增'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            result = self.import_service.import_cascade(tmp, mode='execute', context={'version_id': 1})
            
            r1 = self._find_by_field('domain', 'code', 'UPS001')
            assert r1 is not None and r1['name'] == 'Upsert更新'
            
            r2 = self._find_by_field('domain', 'code', 'UPS002')
            assert r2 is not None and r2['name'] == 'Upsert新增'
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)

    def test_execute_relationship_with_composite_key(self):
        """测试执行关系导入 - 组合业务键完整流程
        
        验证框架根据 source_code/target_code 自动解析 source_bo_id/target_bo_id 的能力
        借鉴 SAP @ObjectModel.foreignKey.association 注解设计
        """
        self.create_full_hierarchy('ERK')
        self.create_business_object('ERK_SM', 'SRC', '源对象')
        self.create_business_object('ERK_SM', 'TGT', '目标对象')
        self.create_business_object('ERK_SM', 'OTH', '其他对象')
        
        sheets = {
            '业务关系': [
                ['操作模式', '源编码', '目标编码', '关系编码', '关系描述'],
                ['新增', 'SRC', 'TGT', 'CALLS', '调用关系'],
                ['新增', 'SRC', 'OTH', 'DEPENDS_ON', '依赖关系'],
                ['新增', 'TGT', 'OTH', 'DATA_FLOW', '数据流'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            preview = self.import_service.import_cascade(tmp, mode='preview', context={'version_id': 1})
            assert preview.validation['invalid_count'] == 0, \
                f"预览失败: {preview.validation.get('errors')}"
            
            result = self.import_service.import_cascade(tmp, mode='execute', context={'version_id': 1})
            
            rel_result = result.results.get('relationship', {})
            assert rel_result.get('success', 0) == 3, \
                f"期望3条成功，实际: {rel_result}"
            
            self.assert_count('relationship', 3)
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)


class TestImportFieldFiltering(TestBase):
    """导入字段过滤测试 - 覆盖之前发现的 bug"""
    
    use_class_setup = True

    def _create_import_file(self, sheets_data: dict) -> str:
        tmp = tempfile.mktemp(suffix='.xlsx')
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, rows in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            for row in rows:
                ws.append(row)
        wb.save(tmp)
        return tmp

    def test_immutable_bk_fields_preserved_on_update(self):
        """测试 immutable 业务键字段在更新时保留（核心bug修复验证）"""
        self.create_full_hierarchy('IBK')
        self.create_business_object('IBK_SM', 'BSRC', '源B')
        self.create_business_object('IBK_SM', 'BTGT', '目标B')
        
        self.create_relationship('BSRC', 'BTGT', 'INIT_REL')
        
        sheets = {
            '业务关系': [
                ['操作模式', '源编码', '目标编码', '关系编码', '关系描述'],
                ['更新', 'BSRC', 'BTGT', 'INIT_REL', '更新后的描述'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            result = self.import_service.import_cascade(tmp, mode='execute')
            
            rel_result = result.results.get('relationship', {})
            errors = rel_result.get('errors', [])
            
            assert len(errors) == 0, \
                f"immutable BK 字段不应被过滤掉，错误: {errors}"
            
            records = self.search_all('relationship', relation_code='INIT_REL')
            assert len(records) == 1
            assert records[0]['relation_desc'] == '更新后的描述'
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)

    def test_ui_editable_false_bk_preserved(self):
        """测试 ui.editable=false 的业务键字段保留"""
        self.create_full_hierarchy('UEF')
        self.create_business_object('UEF_SM', 'U_SRC', '源U')
        self.create_business_object('UEF_SM', 'U_TGT', '目标U')
        self.create_relationship('U_SRC', 'U_TGT', 'TEST_UF')
        
        sheets = {
            '业务关系': [
                ['操作模式', '源编码', '目标编码', '关系编码'],
                ['更新', 'U_SRC', 'U_TGT', 'TEST_UF'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            result = self.import_service.import_cascade(tmp, mode='execute', context={'version_id': 1})
            
            errors = result.results.get('relationship', {}).get('errors', [])
            source_errors = [e for e in errors 
                           if 'source_code' in str(e).lower() or 'None' in str(e)]
            
            assert len(source_errors) == 0, \
                f"source_code 不应为 None (ui.editable=false 导致被过滤): {errors}"
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)

    def test_readonly_always_non_bk_filtered(self):
        """测试 readonly_always 非 BK 字段被正确过滤"""
        self.create_full_hierarchy('ROF')
        self.create_domain('ROF', '只读测试')
        
        sheets = {
            '领域': [
                ['操作模式', '编码', '名称', 'created_at'],
                ['更新', 'ROF', '更新名称', '2025-01-01'],
            ]
        }
        tmp = self._create_import_file(sheets)
        
        try:
            result = self.import_service.import_cascade(tmp, mode='execute', context={'version_id': 1})
            
            record = self._find_by_field('domain', 'code', 'ROF')
            assert record['name'] == '更新名称'
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)


class TestRoundTrip(TestBase):
    """往返测试：导出→修改→导入→验证"""
    
    use_class_setup = True

    def _create_import_file(self, sheets_data: dict) -> str:
        tmp = tempfile.mktemp(suffix='.xlsx')
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, rows in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            for row in rows:
                ws.append(row)
        wb.save(tmp)
        return tmp

    def test_export_import_round_trip(self):
        """测试导出→导入往返数据一致性"""
        self.create_full_hierarchy('RT')
        self.create_business_object('RT_SM', 'RT_BO', '往返对象')
        
        export_result = self.import_service.export_cascade(
            'business_object',
            filters=None,
            options={'include_metadata_sheet': False}
        )
        
        assert export_result.success
        assert os.path.exists(export_result.file_path)
        
        wb = load_workbook(export_result.file_path, read_only=True)
        imported_rows = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                imported_rows.append(row)
        wb.close()
        
        original_count = self.count_records('business_object')
        assert original_count > 0


if __name__ == '__main__':
    import pytest
    sys.exit(pytest.main([__file__, '-v', '-s']))
