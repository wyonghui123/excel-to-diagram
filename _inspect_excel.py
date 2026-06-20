import openpyxl
wb = openpyxl.load_workbook(r'D:\download\export_2026-06-18.xlsx', data_only=True)
ws = None
for sname in wb.sheetnames:
    if '关' in sname:
        ws = wb[sname]
        break
with open(r'd:\filework\excel-to-diagram\_excel_inspect.txt', 'w', encoding='utf-8') as f:
    f.write(f'Sheet names: {wb.sheetnames}\n')
if ws:
    print(f'关系 sheet rows: {ws.max_row}, cols: {ws.max_column}')
    # Print headers
    headers = [c.value for c in ws[1]]
    print('Headers:', headers)
    # Print row 35 (index 34 + headers offset)
    if ws.max_row >= 35:
        row_35 = [c.value for c in ws[35]]
        print('Row 35:', row_35)
    # Print row 2 (first data row)
    row_2 = [c.value for c in ws[2]]
    print('Row 2:', row_2)

with open(r'd:\filework\excel-to-diagram\_excel_inspect.txt', 'w', encoding='utf-8') as f:
    if ws:
        f.write(f'关系 sheet rows: {ws.max_row}, cols: {ws.max_column}\n')
        headers = [c.value for c in ws[1]]
        f.write(f'Headers: {headers}\n')
        if ws.max_row >= 35:
            f.write(f'Row 35: {[c.value for c in ws[35]]}\n')
        f.write(f'Row 2: {[c.value for c in ws[2]]}\n')
    else:
        f.write('关系 sheet NOT FOUND\n')